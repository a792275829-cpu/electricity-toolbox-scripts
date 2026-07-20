from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

if __package__ in (None, ""):
    sys.path.append(str(Path(__file__).resolve().parents[1]))

from src.analyze_500kv_section_gaps import normalize_price_node_name, parse_topology_edges


TIME_HEADER_RE = re.compile(r"^\d{2}:\d{2}$")
KV_500_RE = re.compile(r"500\s*[_-]?\s*k\s*v", re.IGNORECASE)
TRAILING_HASH_BRANCH_RE = re.compile(r"\s*#\d+[A-Za-z]?\s*$")
TRAILING_KV_BRANCH_RE = re.compile(r"((?:500|220)\s*[_-]?\s*k\s*v)\d+[A-Za-z]?\s*$", re.IGNORECASE)
KV_TEXT_RE = re.compile(r"(?P<voltage>500|220)\s*[_-]?\s*k\s*v", re.IGNORECASE)
DEFAULT_TOPOLOGY_PATH = Path(__file__).resolve().parents[1] / "500kV节点拓扑.md"


@dataclass(frozen=True)
class ProcessResult:
    input_path: Path
    output_path: Path
    kept_rows: int
    removed_rows: int
    processed_sheets: int


def normalize_node_name(name: object) -> str:
    text = "" if name is None else str(name).strip()
    text = TRAILING_HASH_BRANCH_RE.sub("", text).strip()
    text = TRAILING_KV_BRANCH_RE.sub(r"\1", text).strip()
    return KV_TEXT_RE.sub(lambda match: f"{match.group('voltage')}kV", text).strip()


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_500kV去重{input_path.suffix}")


def load_topology_nodes(topology_path: Path) -> list[str]:
    edges, _ = parse_topology_edges(Path(topology_path))
    return sorted({edge.a for edge in edges} | {edge.b for edge in edges})


def match_topology_node(node_name: object, topology_nodes: Iterable[str]) -> str | None:
    raw_name = "" if node_name is None else str(node_name).strip()
    normalized_name = normalize_price_node_name(node_name)
    topology_nodes = tuple(topology_nodes)
    exact_matches = [node for node in topology_nodes if normalize_price_node_name(node) == normalized_name]
    if exact_matches:
        return sorted(exact_matches, key=lambda node: (-len(node), node))[0]

    contained_matches = [node for node in topology_nodes if node in raw_name]
    if contained_matches:
        return sorted(contained_matches, key=lambda node: (-len(node), node))[0]
    return None


def process_many(
    input_paths: Iterable[Path],
    topology_path: Path | None = None,
) -> list[ProcessResult]:
    results = []
    for input_path in input_paths:
        input_path = Path(input_path)
        results.append(
            process_workbook(
                input_path,
                default_output_path(input_path),
                topology_path=topology_path,
            )
        )
    return results


def process_workbook(
    input_path: Path,
    output_path: Path | None = None,
    topology_path: Path | None = None,
) -> ProcessResult:
    input_path = Path(input_path)
    output_path = Path(output_path) if output_path is not None else default_output_path(input_path)
    topology_nodes = load_topology_nodes(topology_path) if topology_path is not None else None

    source = load_workbook(input_path, read_only=True, data_only=True)
    output = Workbook()
    output.remove(output.active)

    kept_rows = 0
    removed_rows = 0
    processed_sheets = 0
    seen: set[tuple[object, ...]] = set()
    topology_groups: dict[tuple[str, object], dict[str, object]] = {}
    topology_sheet_title: str | None = None
    topology_time_headers: list[str] | None = None

    for sheet in source.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue

        column_map = _column_map(header)
        if column_map is None:
            continue

        processed_sheets += 1
        output_rows = []
        for row in rows:
            if _is_empty_row(row):
                continue

            node_name = _cell(row, column_map["node"])
            data_item = _cell(row, column_map["data_item"])
            price_values = tuple(_cell(row, index) for index in column_map["time_columns"])

            topology_node = (
                match_topology_node(node_name, topology_nodes)
                if topology_nodes is not None
                else None
            )
            if topology_nodes is not None:
                if topology_node is None:
                    removed_rows += 1
                    continue

                current_time_headers = list(column_map["time_headers"])
                if topology_time_headers is None:
                    topology_time_headers = current_time_headers
                    topology_sheet_title = sheet.title
                elif topology_time_headers != current_time_headers:
                    raise ValueError("包含拓扑节点的工作表分时时间列不一致，无法合并")

                group_key = (topology_node, _normalize_key_value(data_item))
                group = topology_groups.get(group_key)
                if group is None:
                    topology_groups[group_key] = {
                        "node": topology_node,
                        "data_item": data_item,
                        "price_rows": [price_values],
                    }
                    kept_rows += 1
                else:
                    group["price_rows"].append(price_values)
                    removed_rows += 1
                continue

            if not _is_500kv_node(node_name):
                removed_rows += 1
                continue

            normalized_name = normalize_node_name(node_name)
            dedupe_key = (
                _normalize_key_value(data_item),
                *(_normalize_key_value(value) for value in price_values),
            )
            if dedupe_key in seen:
                removed_rows += 1
                continue

            seen.add(dedupe_key)
            output_rows.append([normalized_name, data_item, *price_values])
            kept_rows += 1

        if output_rows:
            output_sheet = output.create_sheet(_safe_sheet_title(sheet.title, output.sheetnames))
            output_sheet.append(["节点名称", "数据项", *column_map["time_headers"]])
            for output_row in output_rows:
                output_sheet.append(output_row)

    if topology_groups:
        output_sheet = output.create_sheet(
            _safe_sheet_title(topology_sheet_title or "500kV节点电价", output.sheetnames)
        )
        output_sheet.append(["节点名称", "数据项", *(topology_time_headers or [])])
        for group in topology_groups.values():
            output_sheet.append(
                [
                    group["node"],
                    group["data_item"],
                    *_average_price_rows(group["price_rows"]),
                ]
            )

    if kept_rows == 0:
        raise ValueError(f"没有找到可输出的500kV节点行: {input_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output.save(output_path)
    return ProcessResult(input_path, output_path, kept_rows, removed_rows, processed_sheets)


def _column_map(header: tuple[object, ...]) -> dict[str, object] | None:
    cleaned = ["" if value is None else str(value).strip() for value in header]
    try:
        node_index = cleaned.index("节点名称")
        data_item_index = cleaned.index("数据项")
    except ValueError:
        return None

    time_columns = [index for index, value in enumerate(cleaned) if TIME_HEADER_RE.match(value)]
    if not time_columns:
        return None

    return {
        "node": node_index,
        "data_item": data_item_index,
        "time_columns": time_columns,
        "time_headers": [cleaned[index] for index in time_columns],
    }


def _is_500kv_node(value: object) -> bool:
    return bool(KV_500_RE.search("" if value is None else str(value)))


def _cell(row: tuple[object, ...], index: int) -> object:
    return row[index] if index < len(row) else None


def _is_empty_row(row: tuple[object, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in row)


def _normalize_key_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, str):
        text = value.strip()
        try:
            return Decimal(text).normalize()
        except InvalidOperation:
            return text
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).normalize()
    return str(value).strip()


def _average_price_rows(price_rows: list[tuple[object, ...]]) -> tuple[object, ...]:
    if len(price_rows) == 1:
        return price_rows[0]

    averaged = []
    for values in zip(*price_rows):
        numeric_values = []
        for value in values:
            normalized = _normalize_key_value(value)
            if normalized == "":
                continue
            if not isinstance(normalized, Decimal):
                raise ValueError(f"同一拓扑节点存在无法求平均的分时价格: {value!r}")
            numeric_values.append(normalized)

        if not numeric_values:
            averaged.append(None)
        else:
            averaged.append(float(sum(numeric_values, Decimal()) / len(numeric_values)))
    return tuple(averaged)


def _safe_sheet_title(title: str, existing_titles: list[str]) -> str:
    base = title[:31] or "Sheet"
    if base not in existing_titles:
        return base

    for suffix_number in range(2, 1000):
        suffix = f"_{suffix_number}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate not in existing_titles:
            return candidate
    raise ValueError(f"无法生成唯一工作表名称: {title}")


def main() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox, scrolledtext

    root = tk.Tk()
    root.title("500kV节点电价去重")
    root.geometry("760x460")

    selected_files: list[Path] = []

    frame = tk.Frame(root, padx=14, pady=14)
    frame.pack(fill=tk.BOTH, expand=True)

    toolbar = tk.Frame(frame)
    toolbar.pack(fill=tk.X)

    status = tk.StringVar(value="未选择文件")
    status_label = tk.Label(frame, textvariable=status, anchor="w")
    status_label.pack(fill=tk.X, pady=(10, 6))

    log = scrolledtext.ScrolledText(frame, height=18, wrap=tk.WORD)
    log.pack(fill=tk.BOTH, expand=True)

    def write_log(message: str) -> None:
        log.insert(tk.END, message + "\n")
        log.see(tk.END)
        root.update_idletasks()

    def choose_files() -> None:
        paths = filedialog.askopenfilenames(
            title="选择实时节点电价Excel文件",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not paths:
            return
        selected_files.clear()
        selected_files.extend(Path(path) for path in paths)
        status.set(f"已选择 {len(selected_files)} 个文件")
        log.delete("1.0", tk.END)
        for path in selected_files:
            write_log(f"已选择: {path}")

    def process_selected() -> None:
        if not selected_files:
            messagebox.showwarning("未选择文件", "请先选择一个或多个 .xlsx 文件。")
            return

        write_log("")
        write_log("开始处理...")
        success_count = 0
        for path in selected_files:
            try:
                result = process_workbook(path, topology_path=DEFAULT_TOPOLOGY_PATH)
            except Exception as exc:
                write_log(f"失败: {path}")
                write_log(f"  {exc}")
                continue

            success_count += 1
            write_log(f"完成: {result.output_path}")
            write_log(
                f"  保留 {result.kept_rows} 行，删除 {result.removed_rows} 行，处理 {result.processed_sheets} 个工作表"
            )

        status.set(f"处理完成: 成功 {success_count} / {len(selected_files)} 个文件")

    choose_button = tk.Button(toolbar, text="选择Excel文件", width=16, command=choose_files)
    choose_button.pack(side=tk.LEFT)

    process_button = tk.Button(toolbar, text="开始处理", width=12, command=process_selected)
    process_button.pack(side=tk.LEFT, padx=(8, 0))

    root.mainloop()


if __name__ == "__main__":
    main()
