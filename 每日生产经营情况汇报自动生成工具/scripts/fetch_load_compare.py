#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font
from playwright.sync_api import sync_playwright

from export_online_energy import (
    BASE_URL,
    OUTPUT_DIR,
    ensure_login,
    interactive_login,
    launch_context,
    load_config,
    parse_json_response,
    validate_date,
)


SCRIPT_DIR = Path(__file__).resolve().parent
SUBJECT_MAP_PATH = SCRIPT_DIR / "load_compare_subject_map.json"
DEFAULT_SUBJECT_MAP = {
    "统调负荷": "1",
    "B类竞价空间": "6",
    "西电东送电力": "2",
    "A类电源出力": "4",
    "地方电源出力": "18",
    "发电总出力": "8",
    "粤港联络线": "20",
}
POINT_TO_TIME_SEGMENT = {
    24: 1,
    96: 2,
}


def log(message: str) -> None:
    print(message, flush=True)


def load_subject_map() -> dict[str, str]:
    if SUBJECT_MAP_PATH.exists():
        raw = json.loads(SUBJECT_MAP_PATH.read_text(encoding="utf-8"))
        return {str(key): str(value) for key, value in raw.items()}
    return DEFAULT_SUBJECT_MAP.copy()


def normalize_subject(value: str) -> tuple[str, str]:
    raw = value.strip()
    if not raw:
        raise ValueError("负荷类型不能为空")

    subject_map = load_subject_map()
    if raw in subject_map.values():
        name = next((key for key, code in subject_map.items() if code == raw), raw)
        return name, raw
    if raw in subject_map:
        return raw, subject_map[raw]

    raise ValueError(
        f"暂不支持的负荷类型：{raw}。当前支持：{', '.join(subject_map)}，或直接传 loadType 编码。"
    )


def parse_subjects_arg(subject: str | None, subjects: str | None) -> list[str]:
    if subjects:
        if subjects.strip().lower() == "all":
            return list(load_subject_map().keys())
        items = [item.strip() for item in subjects.split(",") if item.strip()]
        if not items:
            raise ValueError("--subjects 不能为空")
        return items
    return [subject or "统调负荷"]


def safe_float(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def format_timestamp(day_text: str, index: int, points_per_day: int) -> str:
    day_start = datetime.strptime(day_text, "%Y-%m-%d %H:%M:%S")
    minutes_per_point = 1440 // points_per_day
    timestamp = day_start + timedelta(minutes=index * minutes_per_point)
    return timestamp.strftime("%Y/%m/%d %H:%M")


def fetch_load_dataset(
    context,
    start_date: str,
    end_date: str,
    *,
    data_type: int,
    point: int,
    province_area_id: str,
) -> dict[str, Any]:
    response = context.request.get(
        f"{BASE_URL}/gdfire/api/data/net/load",
        params={
            "startDate": start_date,
            "endDate": end_date,
            "type": data_type,
            "provinceAreaId": province_area_id,
            "timeSegment": POINT_TO_TIME_SEGMENT[point],
        },
        timeout=30000,
    )
    return parse_json_response(response, "读取负荷对比数据")


def select_load_type(dataset: dict[str, Any], load_type: str) -> dict[str, Any]:
    items = dataset.get("data", {}).get("dataNetLoadDTOList") or []
    for item in items:
        if str(item.get("loadType") or "") == load_type:
            return item
    available = ", ".join(sorted({str(item.get("loadType")) for item in items if item.get("loadType") is not None}))
    raise RuntimeError(f"接口返回中没有 loadType={load_type} 的数据，可用 loadType: {available or '无'}")


def build_summary(series: dict[str, Any], start_date: str, end_date: str, *, point: int) -> dict[str, Any]:
    periods = series.get("forecastPeriodList") or []
    if not periods:
        raise RuntimeError(f"{start_date} 到 {end_date} 没有返回负荷数据")

    max_value = float("-inf")
    min_value = float("inf")
    max_time = ""
    min_time = ""
    weighted_total = 0.0
    total_points = 0
    total_energy = 0.0
    version_date = ""

    for period in periods:
        day_text = str(period.get("date") or "")
        values = period.get("loadVal") or []
        if not isinstance(values, list) or not day_text:
            continue

        version_date = max(version_date, str(period.get("versionDate") or ""))
        total_energy += safe_float(period.get("sum"))

        for index, raw_value in enumerate(values):
            value = safe_float(raw_value)
            weighted_total += value
            total_points += 1
            timestamp = format_timestamp(day_text, index, point)
            if value > max_value:
                max_value = value
                max_time = timestamp
            if value < min_value:
                min_value = value
                min_time = timestamp

    if total_points == 0:
        raise RuntimeError(f"{start_date} 到 {end_date} 没有有效负荷点位数据")

    return {
        "startDate": start_date,
        "endDate": end_date,
        "versionDate": version_date,
        "dayCount": len(periods),
        "pointCount": total_points,
        "maxValue": max_value,
        "maxTime": max_time,
        "minValue": min_value,
        "minTime": min_time,
        "avgValue": weighted_total / total_points,
        "totalEnergy": total_energy,
    }


def build_comparison(selected: dict[str, Any], compare: dict[str, Any]) -> dict[str, dict[str, Any]]:
    metrics = {
        "最大值(MW)": ("maxValue", "maxTime"),
        "最小值(MW)": ("minValue", "minTime"),
        "平均值(MW)": ("avgValue", None),
        "总电量(亿kWh)": ("totalEnergy", None),
    }

    result: dict[str, dict[str, Any]] = {}
    for label, (value_key, time_key) in metrics.items():
        selected_value = safe_float(selected.get(value_key))
        compare_value = safe_float(compare.get(value_key))
        diff_value = selected_value - compare_value
        yoy_percent = (diff_value / compare_value * 100.0) if compare_value else None
        item: dict[str, Any] = {
            "selected": selected_value,
            "compare": compare_value,
            "difference": diff_value,
            "yoyPercent": yoy_percent,
        }
        if time_key:
            item["selectedTime"] = selected.get(time_key) or ""
            item["compareTime"] = compare.get(time_key) or ""
        result[label] = item
    return result


def sanitize_filename(value: str) -> str:
    cleaned = value
    for ch in '\\/:*?"<>|':
        cleaned = cleaned.replace(ch, "_")
    return cleaned


def build_subject_payload(subject_name: str, load_type: str, selected: dict[str, Any], compare: dict[str, Any]) -> dict[str, Any]:
    return {
        "subject": subject_name,
        "loadType": load_type,
        "selectedRange": {
            "startDate": selected["startDate"],
            "endDate": selected["endDate"],
            "dayCount": selected["dayCount"],
            "pointCount": selected["pointCount"],
            "versionDate": selected["versionDate"],
        },
        "compareRange": {
            "startDate": compare["startDate"],
            "endDate": compare["endDate"],
            "dayCount": compare["dayCount"],
            "pointCount": compare["pointCount"],
            "versionDate": compare["versionDate"],
        },
        "comparison": build_comparison(selected, compare),
    }


def build_output_payload(subject_payloads: list[dict[str, Any]]) -> dict[str, Any]:
    if not subject_payloads:
        raise RuntimeError("没有可输出的负荷指标")

    first = subject_payloads[0]
    return {
        "source": "api",
        "note": "该脚本抓取的是接口原始值。若页面卡片展示值与此不同，通常是页面展示口径或版本处理不同。",
        "selectedRange": first["selectedRange"],
        "compareRange": first["compareRange"],
        "subjects": subject_payloads,
    }


def metric_decimals(metric_name: str) -> int:
    if metric_name == "总电量(亿kWh)":
        return 4
    return 2


def save_output(payload: dict[str, Any], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)

    selected_start = payload["selectedRange"]["startDate"]
    selected_end = payload["selectedRange"]["endDate"]
    compare_start = payload["compareRange"]["startDate"]
    compare_end = payload["compareRange"]["endDate"]
    subject_label = (
        sanitize_filename(payload["subjects"][0]["subject"])
        if len(payload["subjects"]) == 1
        else f"{len(payload['subjects'])}项指标"
    )
    path = output_dir / (
        f"负荷信息对比_{subject_label}_{selected_start}_{selected_end}_vs_{compare_start}_{compare_end}.xlsx"
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对比汇总"

    ws["A1"] = "指标数量"
    ws["B1"] = len(payload["subjects"])
    ws["A2"] = "数据来源"
    ws["B2"] = "接口原始值"
    ws["A3"] = "说明"
    ws["B3"] = payload["note"]
    ws["A4"] = "选择区间"
    ws["B4"] = f"{selected_start} ~ {selected_end}"
    ws["A5"] = "选择版本时间"
    ws["B5"] = payload["selectedRange"]["versionDate"]
    ws["A6"] = "对比区间"
    ws["B6"] = f"{compare_start} ~ {compare_end}"
    ws["A7"] = "对比版本时间"
    ws["B7"] = payload["compareRange"]["versionDate"]

    header_row = 9
    headers = ["负荷类型", "指标", "选择值", "对比值", "差值", "同比(%)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = header_row + 1
    for subject_payload in payload["subjects"]:
        for metric_name, item in subject_payload["comparison"].items():
            decimals = metric_decimals(metric_name)
            ws.cell(row, 1, subject_payload["subject"])
            ws.cell(row, 2, metric_name)
            ws.cell(row, 3, round(item["selected"], decimals))
            ws.cell(row, 4, round(item["compare"], decimals))
            ws.cell(row, 5, round(item["difference"], decimals))
            if item.get("yoyPercent") is not None:
                ws.cell(row, 6, round(item["yoyPercent"], 4))
            row += 1

    for column, width in {"A": 18, "B": 18, "C": 16, "D": 16, "E": 16, "F": 14}.items():
        ws.column_dimensions[column].width = width

    for current_row in ws.iter_rows():
        for cell in current_row:
            cell.alignment = Alignment(vertical="center")

    for subject_payload in payload["subjects"]:
        sheet_name = sanitize_filename(subject_payload["subject"])[:31] or "Sheet"
        subject_ws = wb.create_sheet(sheet_name)
        subject_ws["A1"] = "负荷类型"
        subject_ws["B1"] = subject_payload["subject"]
        subject_ws["A2"] = "loadType"
        subject_ws["B2"] = subject_payload["loadType"]
        subject_ws["A3"] = "选择区间"
        subject_ws["B3"] = f"{selected_start} ~ {selected_end}"
        subject_ws["A4"] = "对比区间"
        subject_ws["B4"] = f"{compare_start} ~ {compare_end}"
        subject_ws["A5"] = "选择版本时间"
        subject_ws["B5"] = subject_payload["selectedRange"]["versionDate"]
        subject_ws["A6"] = "对比版本时间"
        subject_ws["B6"] = subject_payload["compareRange"]["versionDate"]

        detail_headers = ["指标", "选择值", "对比值", "差值", "同比(%)"]
        for col, header in enumerate(detail_headers, start=1):
            cell = subject_ws.cell(8, col, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        detail_row = 9
        for metric_name, item in subject_payload["comparison"].items():
            decimals = metric_decimals(metric_name)
            subject_ws.cell(detail_row, 1, metric_name)
            subject_ws.cell(detail_row, 2, round(item["selected"], decimals))
            subject_ws.cell(detail_row, 3, round(item["compare"], decimals))
            subject_ws.cell(detail_row, 4, round(item["difference"], decimals))
            if item.get("yoyPercent") is not None:
                subject_ws.cell(detail_row, 5, round(item["yoyPercent"], 4))
            detail_row += 1

        for column, width in {"A": 18, "B": 16, "C": 16, "D": 16, "E": 14}.items():
            subject_ws.column_dimensions[column].width = width

    raw_ws = wb.create_sheet("原始JSON")
    raw_ws["A1"] = json.dumps(payload, ensure_ascii=False, indent=2)
    raw_ws.column_dimensions["A"].width = 120

    try:
        wb.save(path)
    except PermissionError:
        path = output_dir / (
            f"负荷信息对比_{subject_label}_{selected_start}_{selected_end}_vs_{compare_start}_{compare_end}_{datetime.now().strftime('%H%M%S')}.xlsx"
        )
        wb.save(path)
    return path


def print_comparison(payload: dict[str, Any]) -> None:
    log(f"负荷类型数量: {len(payload['subjects'])}")
    log(
        "选择区间: "
        f"{payload['selectedRange']['startDate']} ~ {payload['selectedRange']['endDate']}  "
        f"版本时间: {payload['selectedRange']['versionDate'] or '未知'}"
    )
    log(
        "对比区间: "
        f"{payload['compareRange']['startDate']} ~ {payload['compareRange']['endDate']}  "
        f"版本时间: {payload['compareRange']['versionDate'] or '未知'}"
    )
    log("")
    for subject_payload in payload["subjects"]:
        log(f"{subject_payload['subject']} (loadType={subject_payload['loadType']})")
        for label, item in subject_payload["comparison"].items():
            log(
                f"  {label}: "
                f"选择={item['selected']} "
                f"对比={item['compare']} "
                f"差值={item['difference']} "
                f"同比(%)={item['yoyPercent']}"
            )


def run_compare_one(
    *,
    subject: str,
    selected_start: str,
    selected_end: str,
    compare_start: str,
    compare_end: str,
    data_type: int,
    point: int,
    province_area_id: str,
) -> dict[str, Any]:
    subject_name, load_type = normalize_subject(subject)
    config = load_config()

    with sync_playwright() as playwright:
        context = launch_context(playwright, headless=config.headless)
        try:
            ensure_login(context, config)
            selected_dataset = fetch_load_dataset(
                context,
                selected_start,
                selected_end,
                data_type=data_type,
                point=point,
                province_area_id=province_area_id,
            )
            compare_dataset = fetch_load_dataset(
                context,
                compare_start,
                compare_end,
                data_type=data_type,
                point=point,
                province_area_id=province_area_id,
            )
        finally:
            context.close()

    selected_series = select_load_type(selected_dataset, load_type)
    compare_series = select_load_type(compare_dataset, load_type)
    selected_summary = build_summary(selected_series, selected_start, selected_end, point=point)
    compare_summary = build_summary(compare_series, compare_start, compare_end, point=point)
    return build_subject_payload(subject_name, load_type, selected_summary, compare_summary)


def run_compare_many(
    *,
    subjects: list[str],
    selected_start: str,
    selected_end: str,
    compare_start: str,
    compare_end: str,
    data_type: int,
    point: int,
    province_area_id: str,
) -> dict[str, Any]:
    subject_payloads = [
        run_compare_one(
            subject=subject,
            selected_start=selected_start,
            selected_end=selected_end,
            compare_start=compare_start,
            compare_end=compare_end,
            data_type=data_type,
            point=point,
            province_area_id=province_area_id,
        )
        for subject in subjects
    ]
    return build_output_payload(subject_payloads)


def main() -> int:
    parser = argparse.ArgumentParser(description="抓取负荷信息对比接口的指定日期/区间数据")
    parser.add_argument("selected_date", nargs="?", help="选择日期，格式 YYYY-MM-DD")
    parser.add_argument("compare_date", nargs="?", help="对比日期，格式 YYYY-MM-DD")
    parser.add_argument("--selected-end", help="选择结束日期，默认与选择日期相同")
    parser.add_argument("--compare-end", help="对比结束日期，默认与对比日期相同")
    parser.add_argument("--subject", default="统调负荷", help="单个负荷类型名称或 loadType 编码")
    parser.add_argument("--subjects", help="多个负荷类型，逗号分隔；传 all 表示映射表里的全部指标")
    parser.add_argument("--data-type", type=int, default=1, help="数据类型，页面“日前”对应 1")
    parser.add_argument("--point", type=int, choices=sorted(POINT_TO_TIME_SEGMENT), default=96, help="点位数，支持 24 或 96")
    parser.add_argument("--province-area-id", default="044", help="省区编码，默认 044")
    parser.add_argument("--output-dir", default=str(OUTPUT_DIR), help="输出目录")
    parser.add_argument("--login", action="store_true", help="仅登录并保存会话")
    args = parser.parse_args()

    config = load_config()
    output_dir = Path(args.output_dir).expanduser()

    try:
        if args.login:
            with sync_playwright() as playwright:
                context = launch_context(playwright, headless=False)
                try:
                    interactive_login(context, config)
                finally:
                    context.close()
            return 0

        if not args.selected_date or not args.compare_date:
            raise ValueError("请至少提供 选择日期 和 对比日期，例如: python fetch_load_compare.py 2026-05-13 2026-05-12")

        selected_start = validate_date(args.selected_date)
        selected_end = validate_date(args.selected_end or args.selected_date)
        compare_start = validate_date(args.compare_date)
        compare_end = validate_date(args.compare_end or args.compare_date)
        subjects = parse_subjects_arg(args.subject, args.subjects)

        payload = run_compare_many(
            subjects=subjects,
            selected_start=selected_start,
            selected_end=selected_end,
            compare_start=compare_start,
            compare_end=compare_end,
            data_type=args.data_type,
            point=args.point,
            province_area_id=args.province_area_id,
        )
        print_comparison(payload)
        saved_path = save_output(payload, output_dir)
        log(f"\n结果已保存: {saved_path}")
        return 0
    except KeyboardInterrupt:
        print("\n已取消。")
        return 130
    except Exception as exc:
        print(f"\n失败：{exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
