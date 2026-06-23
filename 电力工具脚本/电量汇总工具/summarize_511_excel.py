#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import sys
import traceback
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except Exception:  # pragma: no cover
    tk = None
    filedialog = None
    messagebox = None
    ttk = None

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FILE_TYPE_PATTERNS = {
    "realtime": "实时交易结果查询",
    "dayahead": "日前交易结果查询",
}

MARKET_TITLES = {
    "dayahead": ("日前电量", "日前电价"),
    "realtime": ("实时电量", "实时电价"),
}

MARKET_ORDER = ["dayahead", "realtime"]
TIME_SLOTS = [f"{hour:02d}:00" for hour in range(24)]

STATION_DISPLAY_NAMES = {
    "汕头": "汕头",
    "海门": "海门",
    "东莞": "东莞",
    "海风": "海风",
    "鮀莲": "金平",
    "归湖": "潮安",
}

STATION_ORDER = ["汕头", "海门", "东莞", "海风", "鮀莲", "归湖"]

THIN = Side(style="thin", color="000000")
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
UNIT_FILL = PatternFill("solid", fgColor="F2F2F2")
SUMMARY_FILL = PatternFill("solid", fgColor="FFF200")


@dataclass
class UnitData:
    station_key: str
    station_name: str
    unit_name: str
    values: dict[str, dict[str, list["NumericCell"]]] = field(default_factory=dict)


@dataclass(frozen=True)
class NumericCell:
    value: float
    number_format: str = "General"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="汇总交易结果 Excel。")
    parser.add_argument("--input-dir", help="原始数据文件夹。")
    parser.add_argument("--output", help="输出 Excel 路径。")
    parser.add_argument(
        "--no-gui",
        action="store_true",
        help="即使未传参数，也直接用命令行模式运行。",
    )
    return parser.parse_args()


def configure_warnings() -> None:
    warnings.filterwarnings(
        "ignore",
        message="Workbook contains no default style, apply openpyxl's default",
        category=UserWarning,
    )


def default_output_path(input_dir: Path) -> Path:
    return input_dir / f"{input_dir.name}_汇总.xlsx"


def run_summary(
    input_dir: Path,
    output_path: Path,
    logger: Callable[[str], None] | None = None,
) -> Path:
    log = logger or (lambda message: None)

    if not input_dir.exists():
        raise FileNotFoundError(f"输入文件夹不存在：{input_dir}")
    if not input_dir.is_dir():
        raise NotADirectoryError(f"输入路径不是文件夹：{input_dir}")

    log(f"开始扫描文件夹：{input_dir}")
    units = collect_units(input_dir, log)
    if not units:
        raise ValueError("没有找到可汇总的 Excel 数据，请确认文件夹结构和 5-11 类似。")

    log(f"共识别 {len(units)} 台机组，开始生成汇总表。")
    final_output = write_workbook(units, output_path, log)
    log(f"汇总完成：{final_output}")
    return final_output


def collect_units(
    input_dir: Path,
    logger: Callable[[str], None],
) -> list[UnitData]:
    units: dict[tuple[str, str], UnitData] = {}

    for station_dir in sorted(path for path in input_dir.iterdir() if path.is_dir()):
        station_key = station_dir.name
        station_name = STATION_DISPLAY_NAMES.get(station_key, station_key)
        logger(f"处理场站：{station_name}")

        for market, pattern in FILE_TYPE_PATTERNS.items():
            files = sorted(station_dir.glob(f"*{pattern}*.xlsx"))
            for file_path in files:
                logger(f"  读取文件：{file_path.name}")
                for unit_name, energy, price in parse_workbook(file_path):
                    key = (station_key, unit_name)
                    unit = units.setdefault(
                        key,
                        UnitData(
                            station_key=station_key,
                            station_name=station_name,
                            unit_name=unit_name,
                        ),
                    )
                    unit.values[market] = {
                        "energy": energy,
                        "price": price,
                    }

    return sort_units(list(units.values()))


def parse_workbook(file_path: Path) -> list[tuple[str, list[NumericCell], list[NumericCell]]]:
    workbook = load_workbook(file_path, data_only=True)
    result: list[tuple[str, list[float], list[float]]] = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        headers = [str(worksheet.cell(1, col).value or "").strip() for col in range(1, worksheet.max_column + 1)]
        time_col = find_column(headers, "时刻")
        energy_col = find_column(headers, "电量")
        price_col = find_price_column(headers)

        if time_col is None or energy_col is None or price_col is None:
            continue

        energy_by_hour = {slot: NumericCell(0.0) for slot in TIME_SLOTS}
        price_by_hour = {slot: NumericCell(0.0) for slot in TIME_SLOTS}

        for row in range(2, worksheet.max_row + 1):
            slot = normalize_time(worksheet.cell(row, time_col).value)
            if slot not in energy_by_hour:
                continue
            energy_by_hour[slot] = read_numeric_cell(worksheet.cell(row, energy_col))
            price_by_hour[slot] = read_numeric_cell(worksheet.cell(row, price_col))

        result.append(
            (
                extract_unit_name(sheet_name),
                [energy_by_hour[slot] for slot in TIME_SLOTS],
                [price_by_hour[slot] for slot in TIME_SLOTS],
            )
        )

    return result


def find_column(headers: list[str], keyword: str) -> int | None:
    for index, header in enumerate(headers, start=1):
        if keyword in header:
            return index
    return None


def find_price_column(headers: list[str]) -> int | None:
    return find_column(headers, "结算电价") or find_column(headers, "出清电价")


def normalize_time(value: object) -> str:
    text = str(value or "").strip()
    if ":" not in text:
        return text
    hour, minute = text.split(":", 1)
    if not hour.isdigit():
        return text
    return f"{int(hour):02d}:{minute[:2]}"


def extract_unit_name(sheet_name: str) -> str:
    import re

    match = re.search(r"#(\d+)", sheet_name)
    if match:
        return f"{int(match.group(1))}号机"
    return sheet_name.strip()


def to_float(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return float(str(value).replace(",", "").strip())


def read_numeric_cell(cell) -> NumericCell:
    number_format = cell.number_format or "General"
    if number_format == "General" and isinstance(cell.value, str):
        number_format = infer_text_number_format(cell.value)
    return NumericCell(to_float(cell.value), number_format)


def infer_text_number_format(value: str) -> str:
    text = value.replace(",", "").strip()
    if "." not in text:
        return "General"
    decimal_part = text.split(".", 1)[1]
    if decimal_part.isdigit():
        return f"0.{''.join('0' for _ in decimal_part)}"
    return "General"


def sort_units(units: list[UnitData]) -> list[UnitData]:
    station_rank = {name: index for index, name in enumerate(STATION_ORDER)}

    def unit_rank(name: str) -> tuple[int, str]:
        import re

        match = re.match(r"^(\d+)号机$", name)
        if match:
            return (int(match.group(1)), name)
        return (9999, name)

    return sorted(
        units,
        key=lambda unit: (
            station_rank.get(unit.station_key, 9999),
            unit_rank(unit.unit_name),
        ),
    )


def write_workbook(
    units: list[UnitData],
    output_path: Path,
    logger: Callable[[str], None],
) -> Path:
    workbook = Workbook()
    grouped: dict[str, list[UnitData]] = {}
    for unit in units:
        grouped.setdefault(unit.station_name, []).append(unit)

    station_names = ordered_station_names(grouped)
    first_sheet = True
    for station_name in station_names:
        if first_sheet:
            worksheet = workbook.active
            first_sheet = False
        else:
            worksheet = workbook.create_sheet()

        worksheet.title = safe_sheet_title(station_name, workbook)
        worksheet.freeze_panes = "B5"
        write_station_block(worksheet, 1, station_name, grouped[station_name], include_station_title=False)
        set_widths(worksheet)
        logger(f"已生成工作表：{worksheet.title}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    final_path = save_workbook_with_fallback(workbook, output_path)
    logger(f"输出文件：{final_path}")
    return final_path


def ordered_station_names(groups: dict[str, list[UnitData]]) -> list[str]:
    ordered: list[str] = []
    for station_key in STATION_ORDER:
        station_name = STATION_DISPLAY_NAMES.get(station_key, station_key)
        if station_name in groups:
            ordered.append(station_name)
    for station_name in groups:
        if station_name not in ordered:
            ordered.append(station_name)
    return ordered


def write_station_block(
    worksheet,
    start_row: int,
    station_name: str,
    units: list[UnitData],
    include_station_title: bool = True,
) -> int:
    total_cols = 1 + len(units) * 5
    header_row = start_row

    if include_station_title:
        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
        title_cell = worksheet.cell(start_row, 1, station_name)
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = UNIT_FILL
        add_border_row(worksheet, start_row, total_cols)
        header_row = start_row + 1

    worksheet.cell(header_row, 1, "时刻")
    style_header(worksheet.cell(header_row, 1))
    worksheet.merge_cells(start_row=header_row, start_column=1, end_row=header_row + 2, end_column=1)

    for market_index, market in enumerate(MARKET_ORDER):
        base_col = 2 + market_index * (len(units) * 2 + 1)
        end_col = base_col + len(units) * 2 - 1

        worksheet.merge_cells(
            start_row=header_row,
            start_column=base_col,
            end_row=header_row,
            end_column=end_col,
        )
        market_cell = worksheet.cell(
            header_row,
            base_col,
            "日前数据" if market == "dayahead" else "实时数据",
        )
        market_cell.font = Font(bold=True, size=11)
        market_cell.alignment = Alignment(horizontal="center", vertical="center")
        market_cell.fill = UNIT_FILL
        for merge_col in range(base_col, end_col + 1):
            worksheet.cell(header_row, merge_col).border = full_border()

        col = base_col
        for unit in units:
            worksheet.merge_cells(
                start_row=header_row + 1,
                start_column=col,
                end_row=header_row + 1,
                end_column=col + 1,
            )
            unit_cell = worksheet.cell(header_row + 1, col, unit.unit_name)
            unit_cell.font = Font(bold=True, size=11)
            unit_cell.alignment = Alignment(horizontal="center", vertical="center")
            unit_cell.fill = UNIT_FILL
            for merge_col in range(col, col + 2):
                worksheet.cell(header_row + 1, merge_col).border = full_border()

            sub_headers = [
                f"{MARKET_TITLES[market][0]}\n（MWh）",
                f"{MARKET_TITLES[market][1]}\n（元/MWh）",
            ]
            for offset, title in enumerate(sub_headers):
                style_header(worksheet.cell(header_row + 2, col + offset, title))
            col += 2

        if market_index < len(MARKET_ORDER) - 1:
            gap_col = end_col + 1
            for row_index in range(header_row, header_row + 3):
                gap_cell = worksheet.cell(row_index, gap_col, "")
                gap_cell.border = full_border()
                gap_cell.fill = UNIT_FILL if row_index < header_row + 2 else HEADER_FILL

    data_start = header_row + 3
    for index, slot in enumerate(TIME_SLOTS):
        style_body(worksheet.cell(data_start + index, 1, slot), bold=False)

    for market_index, market in enumerate(MARKET_ORDER):
        base_col = 2 + market_index * (len(units) * 2 + 1)
        col = base_col
        for unit in units:
            energy = get_series(unit, market, "energy")
            price = get_series(unit, market, "price")

            for offset, values in enumerate([energy, price]):
                for row_offset, source_cell in enumerate(values):
                    body_cell = worksheet.cell(data_start + row_offset, col + offset, source_cell.value)
                    style_body(body_cell, bold=False)
                    body_cell.number_format = source_cell.number_format
            col += 2

        if market_index < len(MARKET_ORDER) - 1:
            gap_col = base_col + len(units) * 2
            for row_index in range(data_start, data_start + len(TIME_SLOTS)):
                worksheet.cell(row_index, gap_col).border = full_border()

    return data_start + len(TIME_SLOTS) - 1


def safe_sheet_title(station_name: str, workbook: Workbook) -> str:
    invalid_chars = set('[]:*?/\\')
    cleaned = "".join(char for char in station_name if char not in invalid_chars).strip() or "Sheet"
    base = cleaned[:31]
    title = base
    index = 2
    existing = set(workbook.sheetnames)
    while title in existing:
        suffix = f"_{index}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    return title


def get_series(unit: UnitData, market: str, kind: str) -> list[NumericCell]:
    market_values = unit.values.get(market)
    if not market_values:
        return [NumericCell(0.0) for _ in TIME_SLOTS]
    return market_values[kind]


def style_header(cell) -> None:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = HEADER_FILL
    cell.border = full_border()


def style_body(cell, bold: bool) -> None:
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = full_border()


def add_border_row(worksheet, row: int, total_cols: int) -> None:
    for col in range(1, total_cols + 1):
        worksheet.cell(row, col).border = full_border()


def full_border() -> Border:
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_widths(worksheet) -> None:
    worksheet.column_dimensions["A"].width = 10
    for col in range(2, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = 12
    for row in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 24


def save_workbook_with_fallback(workbook: Workbook, output_path: Path) -> Path:
    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        fallback_path = output_path.with_name(f"{output_path.stem}_new{output_path.suffix}")
        workbook.save(fallback_path)
        return fallback_path


def launch_gui() -> int:
    if tk is None or ttk is None or filedialog is None or messagebox is None:
        print("当前 Python 环境不可用 tkinter，无法打开图形界面。", file=sys.stderr)
        return 1

    configure_warnings()
    app = SummaryApp()
    app.mainloop()
    return 0


class SummaryApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("电量汇总工具")
        self.geometry("760x430")
        self.minsize(720, 400)

        self.input_var = tk.StringVar()
        self.output_var = tk.StringVar()
        self.status_var = tk.StringVar(value="请选择一个原始数据文件夹。")

        self._build_ui()

    def _build_ui(self) -> None:
        container = ttk.Frame(self, padding=14)
        container.pack(fill="both", expand=True)
        container.columnconfigure(1, weight=1)

        ttk.Label(container, text="原始数据文件夹").grid(row=0, column=0, sticky="w", pady=(0, 8))
        ttk.Entry(container, textvariable=self.input_var).grid(row=0, column=1, sticky="ew", pady=(0, 8))
        ttk.Button(container, text="选择文件夹", command=self.choose_input_dir).grid(row=0, column=2, padx=(8, 0), pady=(0, 8))

        ttk.Label(container, text="输出文件").grid(row=1, column=0, sticky="w", pady=(0, 8))
        ttk.Entry(container, textvariable=self.output_var).grid(row=1, column=1, sticky="ew", pady=(0, 8))
        ttk.Button(container, text="选择位置", command=self.choose_output_file).grid(row=1, column=2, padx=(8, 0), pady=(0, 8))

        button_row = ttk.Frame(container)
        button_row.grid(row=2, column=0, columnspan=3, sticky="w", pady=(4, 10))
        ttk.Button(button_row, text="开始汇总", command=self.run_summary_from_ui).pack(side="left")
        ttk.Button(button_row, text="打开输出目录", command=self.open_output_folder).pack(side="left", padx=(8, 0))

        ttk.Label(container, textvariable=self.status_var, foreground="#1f4e79").grid(
            row=3, column=0, columnspan=3, sticky="w", pady=(0, 8)
        )

        self.log_text = tk.Text(container, height=16, wrap="word")
        self.log_text.grid(row=4, column=0, columnspan=3, sticky="nsew")
        container.rowconfigure(4, weight=1)

        scrollbar = ttk.Scrollbar(container, orient="vertical", command=self.log_text.yview)
        scrollbar.grid(row=4, column=3, sticky="ns")
        self.log_text.configure(yscrollcommand=scrollbar.set)

    def choose_input_dir(self) -> None:
        selected = filedialog.askdirectory(title="选择原始数据文件夹")
        if not selected:
            return
        self.input_var.set(selected)
        self.output_var.set(str(default_output_path(Path(selected))))
        self.status_var.set("已选择文件夹，可以直接开始汇总。")

    def choose_output_file(self) -> None:
        initial = self.output_var.get().strip()
        initial_path = Path(initial) if initial else None
        selected = filedialog.asksaveasfilename(
            title="选择输出文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialdir=str(initial_path.parent) if initial_path else "",
            initialfile=initial_path.name if initial_path else "汇总.xlsx",
        )
        if selected:
            self.output_var.set(selected)

    def append_log(self, message: str) -> None:
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.update_idletasks()

    def run_summary_from_ui(self) -> None:
        input_text = self.input_var.get().strip()
        output_text = self.output_var.get().strip()

        if not input_text:
            messagebox.showwarning("缺少文件夹", "先选择原始数据文件夹。")
            return

        input_dir = Path(input_text)
        output_path = Path(output_text) if output_text else default_output_path(input_dir)
        self.output_var.set(str(output_path))

        self.log_text.delete("1.0", "end")
        self.status_var.set("正在汇总，请稍等...")
        self.append_log("开始执行汇总任务。")

        try:
            final_output = run_summary(
                input_dir=input_dir,
                output_path=output_path,
                logger=self.append_log,
            )
        except Exception as exc:
            self.append_log("")
            self.append_log("执行失败：")
            self.append_log(str(exc))
            self.append_log(traceback.format_exc())
            self.status_var.set("汇总失败，请查看下方日志。")
            messagebox.showerror("汇总失败", str(exc))
            return

        self.status_var.set("汇总完成。")
        self.output_var.set(str(final_output))
        messagebox.showinfo("完成", f"汇总文件已生成：\n{final_output}")

    def open_output_folder(self) -> None:
        output_text = self.output_var.get().strip()
        if not output_text:
            messagebox.showinfo("提示", "还没有输出文件路径。")
            return
        target_dir = Path(output_text).parent
        if not target_dir.exists():
            messagebox.showwarning("目录不存在", f"目录不存在：\n{target_dir}")
            return
        import os

        os.startfile(str(target_dir))


def run_cli(args: argparse.Namespace) -> int:
    configure_warnings()
    if not args.input_dir:
        print("命令行模式下需要传 --input-dir", file=sys.stderr)
        return 1

    input_dir = Path(args.input_dir)
    output_path = Path(args.output) if args.output else default_output_path(input_dir)
    final_output = run_summary(
        input_dir=input_dir,
        output_path=output_path,
        logger=print,
    )
    print(f"已生成：{final_output}")
    return 0


def main() -> int:
    args = parse_args()
    use_cli = args.no_gui or bool(args.input_dir or args.output)
    if use_cli:
        return run_cli(args)
    return launch_gui()


if __name__ == "__main__":
    raise SystemExit(main())
