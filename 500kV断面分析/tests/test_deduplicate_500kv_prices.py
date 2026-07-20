import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.deduplicate_500kv_prices import (
    load_topology_nodes,
    match_topology_node,
    normalize_node_name,
    process_workbook,
)


class Deduplicate500kVPricesTest(unittest.TestCase):
    def _write_topology(self, path: Path) -> None:
        path.write_text(
            """# test

## 2. 已确认的接线拓扑

### 测试片区
穗东换流站-宝安换流站
宝安换流站-凤城站
徐闻开关站-柘林电厂
柘林电厂-贤令山站

## 3. 程序读取约定
""",
            encoding="utf-8",
        )

    def test_normalize_node_name_removes_trailing_branch_markers(self):
        self.assertEqual(normalize_node_name("其他500kV水城变500kV#1M"), "其他500kV水城变500kV")
        self.assertEqual(normalize_node_name("其他500kV水城变500kV#2"), "其他500kV水城变500kV")
        self.assertEqual(normalize_node_name("其他海门电厂500kV1M"), "其他海门电厂500kV")
        self.assertEqual(normalize_node_name("其他海门电厂500kV4M"), "其他海门电厂500kV")
        self.assertEqual(normalize_node_name("其他500kV水城变500kV"), "其他500kV水城变500kV")

    def test_process_workbook_keeps_only_unique_500kv_price_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "实时节点电价查询(2026-06-22).xlsx"
            output_path = Path(tmp) / "cleaned.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "实时节点电价查询(2026-06-22)"
            ws.append(["节点名称", "数据项", "00:00", "00:15", None])
            ws.append(["其他500kV水城变500kV#1M", "电价(元/MWh)", "351", "663.41", None])
            ws.append(["其他500kV水城变500kV#2M", "电价(元/MWh)", "351", "663.41", None])
            ws.append(["其他500kV水城变500kV#3M", "电价(元/MWh)", "352", "663.41", None])
            ws.append(["其他一总降站220_KV#1", "电价(元/MWh)", "370.61", "717.56", None])

            summary = wb.create_sheet("全省-实时节点电价查询(2026-06-22)")
            summary.append(["节点名称", "数据项", "00:00", "00:15", None])
            summary.append(["全省", "电价(元/MWh)", "373.45", "700.71", None])
            wb.save(input_path)

            result = process_workbook(input_path, output_path)

            self.assertEqual(result.kept_rows, 2)
            self.assertEqual(result.removed_rows, 3)
            self.assertEqual(result.output_path, output_path)

            out_wb = load_workbook(output_path, data_only=True)
            self.assertEqual(out_wb.sheetnames, ["实时节点电价查询(2026-06-22)"])
            out_ws = out_wb.active
            rows = list(out_ws.iter_rows(values_only=True))
            self.assertEqual(rows[0], ("节点名称", "数据项", "00:00", "00:15"))
            self.assertEqual(
                rows[1:],
                [
                    ("其他500kV水城变500kV", "电价(元/MWh)", "351", "663.41"),
                    ("其他500kV水城变500kV", "电价(元/MWh)", "352", "663.41"),
                ],
            )

    def test_process_workbook_deduplicates_different_names_with_same_price_series(self):
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "实时节点电价查询.xlsx"
            output_path = Path(tmp) / "cleaned.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "实时节点电价查询"
            ws.append(["节点名称", "数据项", "00:00", "00:15"])
            ws.append(["其他穗东换流站500kV", "电价(元/MWh)", "351", "663.41"])
            ws.append(["其他穗东换流站500kVACF1母", "电价(元/MWh)", "351", "663.41"])
            ws.append(["其他穗东换流站500kVACF2母", "电价(元/MWh)", "351", "663.41"])
            ws.append(["其他宝安换流站500kv1母", "电价(元/MWh)", "400", "500"])
            ws.append(["其他宝安换流站500kv2母", "电价(元/MWh)", "400", "500"])
            wb.save(input_path)

            result = process_workbook(input_path, output_path)

            self.assertEqual(result.kept_rows, 2)
            self.assertEqual(result.removed_rows, 3)

            out_wb = load_workbook(output_path, data_only=True)
            rows = list(out_wb.active.iter_rows(values_only=True))
            self.assertEqual(
                rows[1:],
                [
                    ("其他穗东换流站500kV", "电价(元/MWh)", "351", "663.41"),
                    ("其他宝安换流站500kV1母", "电价(元/MWh)", "400", "500"),
                ],
            )

    def test_loads_and_matches_nodes_from_confirmed_topology_edges(self):
        with tempfile.TemporaryDirectory() as tmp:
            topology_path = Path(tmp) / "topology.md"
            self._write_topology(topology_path)

            nodes = load_topology_nodes(topology_path)

            self.assertEqual(
                nodes,
                ["凤城站", "宝安换流站", "徐闻开关站", "柘林电厂", "穗东换流站", "贤令山站"],
            )
            self.assertEqual(match_topology_node("其他穗东换流站500kVACF1母", nodes), "穗东换流站")
            self.assertIsNone(match_topology_node("其他非拓扑站500kV", nodes))

    def test_topology_filter_keeps_distinct_nodes_with_identical_prices(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / "实时节点电价查询.xlsx"
            output_path = tmp_path / "cleaned.xlsx"
            topology_path = tmp_path / "topology.md"
            self._write_topology(topology_path)

            wb = Workbook()
            ws = wb.active
            ws.title = "实时节点电价查询"
            ws.append(["节点名称", "数据项", "00:00", "00:15"])
            ws.append(["其他穗东换流站500kVACF1母", "电价(元/MWh)", 351, 663.41])
            ws.append(["其他穗东换流站500kVACF2母", "电价(元/MWh)", 351, 663.41])
            ws.append(["其他宝安换流站500kV1母", "电价(元/MWh)", 351, 663.41])
            ws.append(["其他非拓扑站500kV", "电价(元/MWh)", 400, 500])
            wb.save(input_path)

            result = process_workbook(input_path, output_path, topology_path=topology_path)

            self.assertEqual(result.kept_rows, 2)
            self.assertEqual(result.removed_rows, 2)
            out_wb = load_workbook(output_path, data_only=True)
            rows = list(out_wb.active.iter_rows(values_only=True))
            self.assertEqual(
                rows[1:],
                [
                    ("穗东换流站", "电价(元/MWh)", 351, 663.41),
                    ("宝安换流站", "电价(元/MWh)", 351, 663.41),
                ],
            )

    def test_topology_aliases_without_voltage_text_are_kept(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / "日前节点电价查询.xlsx"
            output_path = tmp_path / "cleaned.xlsx"
            topology_path = tmp_path / "topology.md"
            self._write_topology(topology_path)

            wb = Workbook()
            ws = wb.active
            ws.append(["节点名称", "数据项", "00:00"])
            ws.append(["其他徐闻开关站徐闻开关站任一个", "电价(元/MWh)", 100])
            ws.append(["其他柘林电厂#1M", "电价(元/MWh)", 200])
            ws.append(["其他贤令山站#1M", "电价(元/MWh)", 300])
            wb.save(input_path)

            result = process_workbook(input_path, output_path, topology_path=topology_path)

            self.assertEqual(result.kept_rows, 3)
            out_wb = load_workbook(output_path, data_only=True)
            rows = list(out_wb.active.iter_rows(values_only=True))
            self.assertEqual([row[0] for row in rows[1:]], ["徐闻开关站", "柘林电厂", "贤令山站"])

    def test_topology_rows_for_same_node_are_averaged_by_time(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / "日前节点电价查询.xlsx"
            output_path = tmp_path / "cleaned.xlsx"
            topology_path = tmp_path / "topology.md"
            self._write_topology(topology_path)

            wb = Workbook()
            ws = wb.active
            ws.append(["节点名称", "数据项", "00:00", "00:15"])
            ws.append(["其他木棉站500kV1母", "电价(元/MWh)", 100, 200])
            ws.append(["其他木棉站500kV2母", "电价(元/MWh)", 300, None])
            ws.append(["其他木棉站500kV3母", "电价(元/MWh)", 500, 400])
            wb.save(input_path)

            topology_path.write_text(
                """## 2. 已确认的接线拓扑
### 测试片区
木棉站-凤城站
## 3. 程序读取约定
""",
                encoding="utf-8",
            )
            result = process_workbook(input_path, output_path, topology_path=topology_path)

            self.assertEqual(result.kept_rows, 1)
            self.assertEqual(result.removed_rows, 2)
            out_wb = load_workbook(output_path, data_only=True)
            rows = list(out_wb.active.iter_rows(values_only=True))
            self.assertEqual(rows[1], ("木棉站", "电价(元/MWh)", 300, 300))


if __name__ == "__main__":
    unittest.main()
