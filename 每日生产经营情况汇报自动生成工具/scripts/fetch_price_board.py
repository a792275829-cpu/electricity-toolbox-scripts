#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font
from playwright.sync_api import sync_playwright

from export_online_energy import (
    BASE_URL,
    OUTPUT_DIR,
    ensure_login,
    interactive_login,
    launch_context,
    load_config,
    request_json_with_retry,
    validate_date,
)


SCRIPT_DIR = Path(__file__).resolve().parent
PRICE_DATA_TYPES = {
    "日前": "PRE",
    "实时": "RT",
}


def log(message: str) -> None:
    print(message, flush=True)


def safe_number(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def round_or_none(value: float | None, digits: int = 4) -> float | None:
    if value is None:
        return None
    return round(value, digits)


def average(values: list[float]) -> float | None:
    if not values:
        return None
    return sum(values) / len(values)


def normalize_price_type(value: str) -> tuple[str, str]:
    raw = value.strip()
    if raw in PRICE_DATA_TYPES:
        return raw, PRICE_DATA_TYPES[raw]
    upper = raw.upper()
    for name, code in PRICE_DATA_TYPES.items():
        if upper == code:
            return name, code
    raise ValueError(f"不支持的价格类型：{value}，支持：日前/PRE、实时/RT")


def fetch_extreme_price(context, date_text: str, data_type: str) -> dict[str, Any]:
    return request_json_with_retry(
        context,
        "get",
        f"{BASE_URL}/gdfire/api/cockpit/extremePrice",
        f"读取价格看板 {date_text} {data_type}",
        params={
            "startDate": date_text,
            "endDate": date_text,
            "dataType": data_type,
        },
        timeout=60000,
    )


def build_price_summary(date_text: str, type_name: str, data_type: str, payload: dict[str, Any]) -> dict[str, Any]:
    data = payload.get("data") or {}
    avg_values = [safe_number(value) for value in data.get("avgList") or []]
    max_values = [safe_number(value) for value in data.get("maxList") or []]
    min_values = [safe_number(value) for value in data.get("minList") or []]
    avg_values = [value for value in avg_values if value is not None]
    max_values = [value for value in max_values if value is not None]
    min_values = [value for value in min_values if value is not None]

    return {
        "date": date_text,
        "typeName": type_name,
        "dataType": data_type,
        "unit": "厘/kWh",
        "pointCount": len(avg_values),
        "dailyAvgPrice": round_or_none(average(avg_values), 4),
        "maxPrice": round_or_none(max(max_values), 4) if max_values else None,
        "minPrice": round_or_none(min(min_values), 4) if min_values else None,
        "avgList": avg_values,
        "maxList": max_values,
        "minList": min_values,
        "hasData": bool(avg_values),
    }


def fetch_price_summaries(context, date_text: str) -> list[dict[str, Any]]:
    summaries: list[dict[str, Any]] = []
    for type_name, data_type in PRICE_DATA_TYPES.items():
        payload = fetch_extreme_price(context, date_text, data_type)
        summaries.append(build_price_summary(date_text, type_name, data_type, payload))
    return summaries


def build_output_payload(selected_date: str, compare_date: str | None, selected: list[dict[str, Any]], compare: list[dict[str, Any]] | None) -> dict[str, Any]:
    compare_by_type = {item["dataType"]: item for item in compare or []}
    rows: list[dict[str, Any]] = []
    for item in selected:
        other = compare_by_type.get(item["dataType"])
        selected_avg = item["dailyAvgPrice"]
        compare_avg = other["dailyAvgPrice"] if other else None
        diff = None
        if selected_avg is not None and compare_avg is not None:
            diff = round(selected_avg - compare_avg, 4)
        rows.append(
            {
                "typeName": item["typeName"],
                "dataType": item["dataType"],
                "selectedDate": selected_date,
                "selectedDailyAvgPrice": selected_avg,
                "selectedPointCount": item["pointCount"],
                "compareDate": compare_date,
                "compareDailyAvgPrice": compare_avg,
                "comparePointCount": other["pointCount"] if other else None,
                "difference": diff,
                "unit": item["unit"],
            }
        )
    return {
        "source": "价格看板 /gdfire/api/cockpit/extremePrice",
        "selectedDate": selected_date,
        "compareDate": compare_date,
        "rows": rows,
        "selected": selected,
        "compare": compare or [],
        "note": "日均值取 avgList 有效点位的算术平均，结果保留 4 位小数；无实时数据时单元格留空。",
    }


def save_output(payload: dict[str, Any], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    selected_date = payload["selectedDate"]
    compare_date = payload.get("compareDate")
    suffix = f"{selected_date}_vs_{compare_date}" if compare_date else selected_date
    path = output_dir / f"价格看板_全省日均价_{suffix}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws["A1"] = "数据源"
    ws["B1"] = payload["source"]
    ws["A2"] = "说明"
    ws["B2"] = payload["note"]

    headers = [
        "价格类型",
        "选择日期",
        "选择日均价(厘/kWh)",
        "选择点数",
        "对比日期",
        "对比日均价(厘/kWh)",
        "对比点数",
        "差值(厘/kWh)",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(4, col, header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, item in enumerate(payload["rows"], start=5):
        ws.cell(row_idx, 1, item["typeName"])
        ws.cell(row_idx, 2, item["selectedDate"])
        ws.cell(row_idx, 3, item["selectedDailyAvgPrice"])
        ws.cell(row_idx, 4, item["selectedPointCount"])
        ws.cell(row_idx, 5, item["compareDate"])
        ws.cell(row_idx, 6, item["compareDailyAvgPrice"])
        ws.cell(row_idx, 7, item["comparePointCount"])
        ws.cell(row_idx, 8, item["difference"])

    for column, width in {"A": 12, "B": 14, "C": 20, "D": 12, "E": 14, "F": 20, "G": 12, "H": 16}.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    raw_ws = wb.create_sheet("原始JSON")
    raw_ws["A1"] = json.dumps(payload, ensure_ascii=False, indent=2)
    raw_ws.column_dimensions["A"].width = 120

    try:
        wb.save(path)
    except PermissionError:
        path = output_dir / f"价格看板_全省日均价_{suffix}_{datetime.now().strftime('%H%M%S')}.xlsx"
        wb.save(path)
    return path


def print_payload(payload: dict[str, Any]) -> None:
    log(f"选择日期: {payload['selectedDate']}")
    if payload.get("compareDate"):
        log(f"对比日期: {payload['compareDate']}")
    for item in payload["rows"]:
        selected_value = item["selectedDailyAvgPrice"]
        compare_value = item["compareDailyAvgPrice"]
        diff = item["difference"]
        log(
            f"{item['typeName']}: "
            f"选择={selected_value if selected_value is not None else '无数据'} "
            f"对比={compare_value if compare_value is not None else '无数据'} "
            f"差值={diff if diff is not None else '无数据'} "
            f"{item['unit']}"
        )


def run_price_board(selected_date: str, compare_date: str | None = None) -> dict[str, Any]:
    config = load_config()
    with sync_playwright() as playwright:
        context = launch_context(playwright, headless=config.headless)
        try:
            ensure_login(context, config)
            selected = fetch_price_summaries(context, selected_date)
            compare = fetch_price_summaries(context, compare_date) if compare_date else None
        finally:
            context.close()
    return build_output_payload(selected_date, compare_date, selected, compare)


def main() -> int:
    parser = argparse.ArgumentParser(description="抓取价格看板指定日期的日前/实时全省平均价格日均值")
    parser.add_argument("selected_date", nargs="?", help="选择日期，格式 YYYY-MM-DD")
    parser.add_argument("compare_date", nargs="?", help="可选：对比日期，格式 YYYY-MM-DD")
    parser.add_argument("--output-dir", default=str(OUTPUT_DIR), help="输出目录")
    parser.add_argument("--login", action="store_true", help="仅打开浏览器登录并保存会话")
    args = parser.parse_args()

    try:
        config = load_config()
        if args.login:
            with sync_playwright() as playwright:
                context = launch_context(playwright, headless=False)
                try:
                    interactive_login(context, config)
                finally:
                    context.close()
            return 0

        if not args.selected_date:
            raise ValueError("请提供选择日期，例如：python fetch_price_board.py 2026-05-16")

        selected_date = validate_date(args.selected_date)
        compare_date = validate_date(args.compare_date) if args.compare_date else None
        payload = run_price_board(selected_date, compare_date)
        print_payload(payload)
        saved_path = save_output(payload, Path(args.output_dir).expanduser())
        log(f"\n结果已保存: {saved_path}")
        return 0
    except KeyboardInterrupt:
        print("\n已取消。")
        return 130
    except Exception as exc:
        print(f"\n失败：{exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
