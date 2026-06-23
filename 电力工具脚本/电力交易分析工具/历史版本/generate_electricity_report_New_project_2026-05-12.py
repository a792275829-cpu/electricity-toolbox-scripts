import argparse
import html
import math
import pathlib
import threading
import tkinter as tk
import datetime
import re
from tkinter import filedialog, messagebox, ttk

import openpyxl


THERMAL_COMPANIES = ["汕头", "海门", "东莞"]
RENEWABLE_UNITS = ["海风", "鮀莲", "归湖"]
POINT_TYPE_ORDER = ["高价欠发亏损", "高价超发收益", "低价欠发收益", "低价超发亏损"]
DEFAULT_OUTPUT_NAME = "electricity_clearing_report.html"
DEFAULT_OUTPUT_DIR = pathlib.Path.home() / "Downloads"


def num(value, default=0.0):
    if isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(value):
        return float(value)
    return default


def fmt(value):
    if value is None or (isinstance(value, float) and not math.isfinite(value)):
        return "-"
    if isinstance(value, (int, float)):
        return f"{value:,.2f}"
    return html.escape(str(value))


def esc(value):
    return html.escape(str(value))


def normalize_pair_label(unit):
    normalized = unit.replace(",", "、")
    replacements = {
        "#1、": "#1、#2",
        "#3、": "#3、#4",
        "#5、": "#5、#6",
        "#7、": "#7、#8",
    }
    return replacements.get(normalized, normalized)


def merge_offer_segments(segments):
    merged = []
    for segment in segments:
        if (
            merged
            and abs(merged[-1]["price"] - segment["price"]) < 1e-9
            and abs(merged[-1]["end"] - segment["start"]) < 1e-9
        ):
            merged[-1]["end"] = segment["end"]
            merged[-1]["source_count"] += 1
        else:
            copied = dict(segment)
            copied["source_count"] = 1
            merged.append(copied)
    return merged


def to_hour_number(time_text):
    return int(str(time_text).split(":")[0])


def classify_point(item, compare_mode):
    if compare_mode == "dayahead_vs_ml" and item.get("cost") not in (None, 0):
        high_or_low = item["da_price"] - item["cost"]
    else:
        high_or_low = item["price_diff"]
    if high_or_low >= 0 and item["dev"] < 0:
        return POINT_TYPE_ORDER[0]
    if high_or_low >= 0 and item["dev"] > 0:
        return POINT_TYPE_ORDER[1]
    if high_or_low < 0 and item["dev"] < 0:
        return POINT_TYPE_ORDER[2]
    if high_or_low < 0 and item["dev"] > 0:
        return POINT_TYPE_ORDER[3]
    return "平价平量"

class ReportBuilder:
    def __init__(self, xlsx_path, out_path=None):
        self.xlsx_path = pathlib.Path(xlsx_path)
        self.out_path = pathlib.Path(out_path) if out_path else self.xlsx_path.with_name(f"{self.xlsx_path.stem}_分析报告.html")
        self.wb = None
        self.summary = {}
        self.hourly = {}
        self.offer_cards = {}
        self.compare_mode = "dayahead_vs_ml"
        self.companies = {
            "汕头": ["汕头#1机组", "汕头#2机组", "汕头#3机组"],
            "海门": ["海门#1机组", "海门#2机组", "海门#3机组", "海门#4机组"],
            "东莞": ["东莞#1、#2机组", "东莞#3、#4机组", "东莞#5、#6机组", "东莞#7、#8机组"],
            "海风": ["海风"],
            "鮀莲": ["鮀莲"],
            "归湖": ["归湖"],
        }

    def build(self):
        self.wb = openpyxl.load_workbook(self.xlsx_path, data_only=True, read_only=True)
        try:
            self._parse_summary()
            self.offer_cards = self._parse_offer_cards()
            self.hourly.update(self._parse_blocks("汕头", "汕头"))
            self.hourly.update(self._parse_blocks("海门", "海门"))
            self.hourly.update(self._parse_blocks("东莞", "东莞"))
            self._parse_renewables()
            self.compare_mode = "realtime_vs_dayahead" if self._has_realtime_data() else "dayahead_vs_ml"
            self._fill_dongguan_summary()
            html_doc = self._render()
            self.out_path.write_text(html_doc, encoding="utf-8")
            return self.out_path
        finally:
            self.wb.close()

    def _report_date_text(self):
        stem = self.xlsx_path.stem
        match = re.search(r"((?:19|20)\d{2})[.\-_/年](\d{1,2})[.\-_/月](\d{1,2})", stem)
        if not match:
            match = re.search(r"((?:19|20)\d{2})(\d{2})(\d{2})", stem)
        if match:
            year, month, day = match.groups()
            return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
        return datetime.date.today().isoformat()

    def _parse_summary(self):
        ws_sum = self.wb["日前"]
        summary_rows = {
            "分公司火电": 4,
            "汕头": 5,
            "汕头#1机组": 8,
            "汕头#2机组": 9,
            "汕头#3机组": 10,
            "海门": 13,
            "海门#1机组": 16,
            "海门#2机组": 17,
            "海门#3机组": 18,
            "海门#4机组": 19,
            "东莞": 21,
        }
        for name, row in summary_rows.items():
            self.summary[name] = {
                "capacity_10mw": num(ws_sum.cell(row, 3).value, None),
                "cost": ws_sum.cell(row, 4).value,
                "load_rate": ws_sum.cell(row, 5).value,
                "da_mwh": num(ws_sum.cell(row, 6).value) * 10,
                "da_price": num(ws_sum.cell(row, 7).value),
                "ml_mwh": num(ws_sum.cell(row, 8).value) * 10,
                "dev_mwh": num(ws_sum.cell(row, 9).value) * 10,
                "dev_price": num(ws_sum.cell(row, 10).value, None),
                "profit_wanyuan": num(ws_sum.cell(row, 11).value),
            }

    def _parse_offer_cards(self):
        ws_comp = self.wb["运行补偿"]
        rows = list(ws_comp.iter_rows(values_only=True))
        cards = {}
        for idx, row in enumerate(rows):
            label = row[0]
            if not isinstance(label, str):
                continue
            if label.startswith("东莞一套"):
                unit = "东莞#1、#2机组"
            elif label.startswith("东莞二套"):
                unit = "东莞#3、#4机组"
            elif label.startswith("东莞三套"):
                unit = "东莞#5、#6机组"
            elif label.startswith("东莞四套"):
                unit = "东莞#7、#8机组"
            elif label.endswith("机组"):
                unit = label
            else:
                continue
            if not any(unit.startswith(prefix) for prefix in THERMAL_COMPANIES):
                continue

            plant_rate = None
            segments = []
            for offset in range(1, 16):
                if idx + offset >= len(rows):
                    break
                candidate = rows[idx + offset]
                if candidate[16] == "厂用电率":
                    plant_rate = num(candidate[17], None)
                if candidate[16] == "段号":
                    for segment_row in rows[idx + offset + 1 : idx + offset + 11]:
                        start = num(segment_row[17], None)
                        end = num(segment_row[18], None)
                        price = num(segment_row[19], None)
                        if start is not None and end is not None and price is not None:
                            segments.append({"start": start, "end": end, "price": price})
                    break
            if plant_rate is not None:
                cards[unit] = {"plant_rate": plant_rate, "segments": merge_offer_segments(segments)}
        return cards

    def _parse_blocks(self, sheet_name, company_name):
        ws = self.wb[sheet_name]
        row2 = [ws.cell(2, c).value for c in range(1, 22)]
        blocks = {}
        if any(isinstance(value, str) and "实时电量" in value for value in row2):
            for start_col in range(2, 18, 4):
                label = ws.cell(1, start_col).value
                if not isinstance(label, str) or "合计" in label:
                    continue
                clean = label.strip()
                if clean.startswith("#"):
                    unit_name = f"{company_name}{normalize_pair_label(clean.replace('机组', ''))}机组"
                else:
                    unit_name = f"{company_name}#{clean[0]}机组"
                data = []
                for row_idx in range(3, 27):
                    time_value = ws.cell(row_idx, 1).value
                    if isinstance(time_value, datetime.time):
                        time_text = f"{time_value.hour}:00"
                    else:
                        time_text = str(time_value)
                    data.append(
                        {
                            "time": time_text,
                            "ml": 0.0,
                            "ml_price": 0.0,
                            "da": num(ws.cell(row_idx, start_col).value),
                            "da_price": num(ws.cell(row_idx, start_col + 1).value),
                            "rt": num(ws.cell(row_idx, start_col + 2).value, 0.0),
                            "rt_price": num(ws.cell(row_idx, start_col + 3).value, 0.0),
                            "dev_ml": 0.0,
                            "price_diff_ml": 0.0,
                            "profit_ml": 0.0,
                            "cost": 0.0,
                        }
                    )
                blocks[unit_name] = data

        rows = list(ws.iter_rows(values_only=True))
        for idx in range(len(rows) - 30):
            label = rows[idx][0] if rows[idx] else None
            if not isinstance(label, str):
                continue
            clean = label.strip()
            if not (clean in ["1号机", "2号机", "3号机", "4号机"] or (clean.startswith("#") and "机组" in clean)):
                continue
            data = []
            ok = True
            for ridx in range(idx + 3, idx + 27):
                row = rows[ridx]
                time_value = row[0]
                if not (isinstance(time_value, str) and ":" in time_value):
                    ok = False
                    break
                ml = num(row[1])
                ml_price = num(row[2])
                da = num(row[4])
                da_price = num(row[5])
                rt = num(row[6], 0.0)
                rt_price = num(row[7], 0.0)
                cost = num(row[9])
                dev_ml = num(row[10], da - ml)
                profit_ml = dev_ml * (da_price - cost) / 10000
                data.append(
                    {
                        "time": time_value,
                        "ml": ml,
                        "ml_price": ml_price,
                        "da": da,
                        "da_price": da_price,
                        "rt": rt,
                        "rt_price": rt_price,
                        "dev_ml": dev_ml,
                        "price_diff_ml": da_price - ml_price,
                        "profit_ml": profit_ml,
                        "cost": cost,
                    }
                )
            if ok and len(data) == 24:
                if clean.endswith("号机"):
                    unit_name = f"{company_name}#{clean[0]}机组"
                else:
                    unit_name = f"{company_name}{clean}"
                blocks[unit_name] = data
        return blocks

    def _parse_renewables(self):
        for sheet_name in RENEWABLE_UNITS:
            ws = self.wb[sheet_name]
            data = []
            for row in range(4, 28):
                time_value = ws.cell(row, 1).value
                if not (isinstance(time_value, str) and ":" in time_value):
                    continue
                da = num(ws.cell(row, 2).value)
                da_price = num(ws.cell(row, 3).value)
                rt = num(ws.cell(row, 4).value, 0.0)
                rt_price = num(ws.cell(row, 5).value, 0.0)
                ml = num(ws.cell(row, 12).value)
                ml_price = num(ws.cell(row, 13).value)
                data.append(
                    {
                        "time": time_value,
                        "ml": ml,
                        "ml_price": ml_price,
                        "da": da,
                        "da_price": da_price,
                        "rt": rt,
                        "rt_price": rt_price,
                        "dev_ml": da - ml,
                        "price_diff_ml": da_price - ml_price,
                        "profit_ml": (da - ml) * (da_price - ml_price) / 10000,
                    }
                )
            self.hourly[sheet_name] = data
            da_total = sum(item["da"] for item in data)
            rt_total = sum(item.get("rt", 0.0) for item in data)
            ml_total = sum(item["ml"] for item in data)
            self.summary[sheet_name] = {
                "da_mwh": da_total,
                "da_price": sum(item["da"] * item["da_price"] for item in data) / da_total if da_total else 0,
                "rt_mwh": rt_total,
                "rt_price_avg": sum(item.get("rt", 0.0) * item.get("rt_price", 0.0) for item in data) / rt_total if rt_total else 0,
                "ml_mwh": ml_total,
                "ml_price_avg": sum(item["ml"] * item["ml_price"] for item in data) / ml_total if ml_total else 0,
                "dev_mwh": da_total - ml_total,
                "dev_ml_mwh": da_total - ml_total,
                "profit_wanyuan": sum(item["profit_ml"] for item in data),
                "profit_ml_wanyuan": sum(item["profit_ml"] for item in data),
                "dev_rt_mwh": rt_total - da_total,
                "profit_rt_wanyuan": sum((item.get("rt", 0.0) - item["da"]) * (item.get("rt_price", 0.0) - item["da_price"]) / 10000 for item in data),
                "load_rate": None,
                "full_price": num(ws["AE28"].value, None),
            }

    def _has_realtime_data(self):
        for points in self.hourly.values():
            for item in points:
                if abs(item.get("rt", 0.0)) > 1e-6 or abs(item.get("rt_price", 0.0)) > 1e-6:
                    return True
        return False

    def _metric_fields(self):
        if self.compare_mode == "realtime_vs_dayahead":
            return {
                "base_mwh": "da_mwh",
                "base_price": "da_price",
                "compare_mwh": "rt_mwh",
                "compare_price": "rt_price_avg",
                "dev_mwh": "dev_rt_mwh",
                "profit": "profit_rt_wanyuan",
            }
        return {
            "base_mwh": "ml_mwh",
            "base_price": "ml_price_avg",
            "compare_mwh": "da_mwh",
            "compare_price": "da_price",
            "dev_mwh": "dev_ml_mwh",
            "profit": "profit_ml_wanyuan",
        }

    def _fill_dongguan_summary(self):
        all_units = [unit for units in self.companies.values() for unit in units if unit not in RENEWABLE_UNITS]
        for unit in all_units:
            if unit in self.hourly:
                data = self.hourly[unit]
                da = sum(item["da"] for item in data)
                rt = sum(item.get("rt", 0.0) for item in data)
                ml = sum(item["ml"] for item in data)
                base = dict(self.summary.get(unit, {}))
                base.update(
                    {
                        "da_mwh": da,
                        "da_price": sum(item["da"] * item["da_price"] for item in data) / da if da else 0,
                        "rt_mwh": rt,
                        "rt_price_avg": sum(item.get("rt", 0.0) * item.get("rt_price", 0.0) for item in data) / rt if rt else 0,
                        "ml_mwh": ml,
                        "ml_price_avg": sum(item["ml"] * item["ml_price"] for item in data) / ml if ml else 0,
                        "dev_mwh": da - ml,
                        "dev_ml_mwh": da - ml,
                        "profit_wanyuan": sum(item["profit_ml"] for item in data),
                        "profit_ml_wanyuan": sum(item["profit_ml"] for item in data),
                        "dev_rt_mwh": rt - da,
                        "profit_rt_wanyuan": sum((item.get("rt", 0.0) - item["da"]) * (item.get("rt_price", 0.0) - item["da_price"]) / 10000 for item in data),
                        "load_rate": None,
                    }
                )
                self.summary[unit] = base

    def _evaluate_offer_match(self, unit, clearing_mwh, clearing_price):
        card = self.offer_cards.get(unit)
        if not card:
            return None, None
        if abs(clearing_mwh) <= 0.05:
            return 0.0, "停机"
        plant_output = clearing_mwh / (1 - card["plant_rate"])
        if not card.get("segments"):
            return plant_output, "无报价单"
        failures = []
        first_segment = card["segments"][0] if card.get("segments") else None
        if first_segment and plant_output + 0.05 < first_segment["start"]:
            failures.append(f"机端应≥{fmt(first_segment['start'])}MW")
        for segment in card["segments"]:
            price = segment["price"]
            start = segment["start"]
            end = segment["end"]
            if segment.get("source_count", 1) > 1:
                if clearing_price <= price and plant_output > start + 0.05:
                    failures.append(f"价≤{fmt(price)}时机组出力应≤{fmt(start)}MW")
                if clearing_price > price and plant_output + 0.05 < end:
                    failures.append(f"价>{fmt(price)}时机组出力应≥{fmt(end)}MW")
                continue

            if clearing_price < price and plant_output > start + 0.05:
                failures.append(f"价<{fmt(price)}时机组出力应≤{fmt(start)}MW")
            elif abs(clearing_price - price) <= 0.05 and (plant_output + 0.05 < start or plant_output > end + 0.05):
                failures.append(f"价={fmt(price)}时机组出力应在{fmt(start)}-{fmt(end)}MW")
            elif clearing_price > price and plant_output + 0.05 < end:
                failures.append(f"价>{fmt(price)}时机组出力应≥{fmt(end)}MW")
        if failures:
            return plant_output, "不匹配：" + "；".join(failures)
        return plant_output, "匹配"

    def _offer_match_direction(self, unit, clearing_mwh, clearing_price):
        card = self.offer_cards.get(unit)
        if not card:
            return None
        if abs(clearing_mwh) <= 0.05:
            return "停机"

        plant_output = clearing_mwh / (1 - card["plant_rate"])
        if not card.get("segments"):
            return "-"
        first_segment = card["segments"][0] if card.get("segments") else None
        if first_segment and plant_output + 0.05 < first_segment["start"]:
            return "偏低"

        for segment in card["segments"]:
            price = segment["price"]
            start = segment["start"]
            end = segment["end"]
            if segment.get("source_count", 1) > 1:
                if clearing_price <= price and plant_output > start + 0.05:
                    return "偏高"
                if clearing_price > price and plant_output + 0.05 < end:
                    return "偏低"
                continue

            if clearing_price < price and plant_output > start + 0.05:
                return "偏高"
            if abs(clearing_price - price) <= 0.05:
                if plant_output + 0.05 < start:
                    return "偏低"
                if plant_output > end + 0.05:
                    return "偏高"
            elif clearing_price > price and plant_output + 0.05 < end:
                return "偏低"

        return "正常"

    def _company_summary(self, company_name, units):
        if self.compare_mode == "dayahead_vs_ml" and company_name in self.summary:
            item = dict(self.summary[company_name])
            item["ml_price_avg"] = item.get("ml_price_avg", item.get("dev_price", 0) + item.get("da_price", 0))
            item["dev_ml_mwh"] = item.get("dev_ml_mwh", item.get("dev_mwh", 0))
            item["profit_ml_wanyuan"] = item.get("profit_ml_wanyuan", item.get("profit_wanyuan", 0))
            item["rt_mwh"] = item.get("rt_mwh", 0)
            item["rt_price_avg"] = item.get("rt_price_avg", 0)
            item["dev_rt_mwh"] = item.get("dev_rt_mwh", item["rt_mwh"] - item.get("da_mwh", 0))
            item["profit_rt_wanyuan"] = item.get("profit_rt_wanyuan", 0)
            return item

        da = sum(self.summary.get(unit, {}).get("da_mwh", 0) for unit in units)
        rt = sum(self.summary.get(unit, {}).get("rt_mwh", 0) for unit in units)
        ml = sum(self.summary.get(unit, {}).get("ml_mwh", 0) for unit in units)
        profit_ml = sum(self.summary.get(unit, {}).get("profit_ml_wanyuan", self.summary.get(unit, {}).get("profit_wanyuan", 0)) for unit in units)
        profit_rt = sum(self.summary.get(unit, {}).get("profit_rt_wanyuan", 0) for unit in units)
        da_price = sum(self.summary.get(unit, {}).get("da_mwh", 0) * self.summary.get(unit, {}).get("da_price", 0) for unit in units) / da if da else 0
        rt_price = sum(self.summary.get(unit, {}).get("rt_mwh", 0) * self.summary.get(unit, {}).get("rt_price_avg", 0) for unit in units) / rt if rt else 0
        ml_price = sum(self.summary.get(unit, {}).get("ml_mwh", 0) * self.summary.get(unit, {}).get("ml_price_avg", 0) for unit in units) / ml if ml else 0
        item = {
            "da_mwh": da,
            "da_price": da_price,
            "rt_mwh": rt,
            "rt_price_avg": rt_price,
            "ml_mwh": ml,
            "ml_price_avg": ml_price,
            "dev_ml_mwh": da - ml,
            "profit_ml_wanyuan": profit_ml,
            "dev_rt_mwh": rt - da,
            "profit_rt_wanyuan": profit_rt,
        }
        if company_name in self.summary:
            item["full_price"] = self.summary[company_name].get("full_price")
        return item

    def _representative_points(self, units, limit_per_type=5):
        points = []
        for unit in units:
            for item in self.hourly.get(unit, []):
                copied = dict(item)
                copied["unit"] = unit
                copied["dev"] = self._display_dev(item)
                copied["price_diff"] = self._display_price_diff(item)
                copied["profit"] = self._display_profit(item)
                copied["type"] = classify_point(copied, self.compare_mode)
                points.append(copied)
        results = []
        for category in POINT_TYPE_ORDER:
            chosen = [item for item in points if item["type"] == category]
            chosen.sort(key=lambda item: abs(item["profit"]), reverse=True)
            results.extend(chosen[:limit_per_type])
        results.sort(key=lambda item: (item["unit"], to_hour_number(item["time"])))
        return results

    def _table(self, headers, rows, cls=""):
        head = "".join(f"<th>{esc(item)}</th>" for item in headers)
        rendered_rows = []
        for row in rows:
            rendered_cells = []
            for cell in row:
                if isinstance(cell, dict):
                    text = cell.get("html", cell.get("text", ""))
                    style = cell.get("style", "")
                    css_class = cell.get("class", "")
                    class_attr = f' class="{css_class}"' if css_class else ""
                    style_attr = f' style="{style}"' if style else ""
                    rendered_cells.append(f"<td{class_attr}{style_attr}>{text}</td>")
                else:
                    rendered_cells.append(f"<td>{cell}</td>")
            rendered_rows.append("<tr>" + "".join(rendered_cells) + "</tr>")
        body = "".join(rendered_rows)
        return f'<table class="{cls}"><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>'

    def _kpis(self, item):
        fields = self._metric_fields()
        default_middle_label = "中长期电量合计" if self.compare_mode == "dayahead_vs_ml" else "日前电量合计"
        default_middle_value = f"{fmt(item.get(fields['base_mwh']))} MWh"
        first_label = "日前电量合计" if self.compare_mode == "dayahead_vs_ml" else "实时电量合计"
        first_value = f"{fmt(item.get(fields['compare_mwh']))} MWh"
        middle_label = item.get("middle_label", default_middle_label)
        middle_value = item.get("middle_value", default_middle_value)
        cards = [
            (first_label, first_value),
            (middle_label, middle_value),
            ("偏差电量", f"{fmt(item.get(fields['dev_mwh']))} MWh"),
            ("偏差收益估算", f"{fmt(item.get(fields['profit']))} 万元"),
        ]
        return '<div class="kpis">' + "".join(f'<div class="kpi"><span>{label}</span><b>{value}</b></div>' for label, value in cards) + "</div>"

    def _display_dev(self, item):
        return item.get("rt", 0.0) - item["da"] if self.compare_mode == "realtime_vs_dayahead" else item["da"] - item["ml"]

    def _display_price_diff(self, item):
        return item.get("rt_price", 0.0) - item["da_price"] if self.compare_mode == "realtime_vs_dayahead" else item["da_price"] - item["ml_price"]

    def _display_profit(self, item):
        if self.compare_mode == "realtime_vs_dayahead":
            return self._display_dev(item) * self._display_price_diff(item) / 10000
        cost = item.get("cost")
        if cost is not None and cost != 0:
            return self._display_dev(item) * (item["da_price"] - cost) / 10000
        return self._display_dev(item) * self._display_price_diff(item) / 10000

    def _percentile(self, values, percentile):
        if not values:
            return 0.0
        ordered = sorted(float(value) for value in values)
        if len(ordered) == 1:
            return ordered[0]
        position = (len(ordered) - 1) * percentile
        lower = int(position)
        upper = min(lower + 1, len(ordered) - 1)
        weight = position - lower
        return ordered[lower] + (ordered[upper] - ordered[lower]) * weight

    def _blend_color(self, start_hex, end_hex, ratio):
        ratio = max(0.0, min(1.0, ratio))
        start = tuple(int(start_hex[index:index + 2], 16) for index in (1, 3, 5))
        end = tuple(int(end_hex[index:index + 2], 16) for index in (1, 3, 5))
        mixed = tuple(round(start[i] + (end[i] - start[i]) * ratio) for i in range(3))
        return "#" + "".join(f"{channel:02x}" for channel in mixed)

    def _heat_color(self, value, minimum, midpoint, maximum):
        low_color = "#63be7b"
        mid_color = "#ffeb84"
        high_color = "#f8696b"
        if maximum <= minimum:
            return mid_color
        if value <= midpoint:
            span = midpoint - minimum
            ratio = 0.5 if span <= 0 else (value - minimum) / span
            return self._blend_color(low_color, mid_color, ratio)
        span = maximum - midpoint
        ratio = 0.5 if span <= 0 else (value - midpoint) / span
        return self._blend_color(mid_color, high_color, ratio)

    def _price_cell(self, value, minimum, midpoint, maximum):
        return {
            "text": fmt(value),
            "style": f"background:{self._heat_color(value, minimum, midpoint, maximum)};font-weight:600;",
        }

    def _detail_table(self, unit):
        rows = []
        detail_points = self.hourly.get(unit, [])
        dayahead_prices = [item["da_price"] for item in detail_points]
        compare_prices = (
            [item.get("rt_price", 0.0) for item in detail_points]
            if self.compare_mode == "realtime_vs_dayahead"
            else [item["ml_price"] for item in detail_points]
        )
        da_min = min(dayahead_prices) if dayahead_prices else 0.0
        da_mid = self._percentile(dayahead_prices, 0.5)
        da_max = max(dayahead_prices) if dayahead_prices else 0.0
        compare_min = min(compare_prices) if compare_prices else 0.0
        compare_mid = self._percentile(compare_prices, 0.5)
        compare_max = max(compare_prices) if compare_prices else 0.0
        for item in detail_points:
            check_mwh = item.get("rt", 0.0) if self.compare_mode == "realtime_vs_dayahead" else item["da"]
            check_price = item.get("rt_price", 0.0) if self.compare_mode == "realtime_vs_dayahead" else item["da_price"]
            output, match = self._evaluate_offer_match(unit, check_mwh, check_price)
            direction = self._offer_match_direction(unit, check_mwh, check_price)
            if unit in RENEWABLE_UNITS:
                match = "不适用"
                direction = "不适用"
            match_text = "不匹配" if isinstance(match, str) and match.startswith("不匹配：") else match
            if match_text == "不匹配":
                match_html = f'<span class="bad">{match_text}</span>'
            elif match_text == "匹配":
                match_html = f'<span class="ok">{match_text}</span>'
            else:
                match_html = esc(match_text)
            if direction == "偏高":
                direction_html = '<span class="bad">偏高</span>'
            elif direction == "偏低":
                direction_html = '<span class="bad">偏低</span>'
            elif direction == "正常":
                direction_html = '<span class="ok">正常</span>'
            else:
                direction_html = esc(direction) if direction is not None else "-"
            rows.append(
                [
                    esc(item["time"]),
                    fmt(item["da"]),
                    self._price_cell(item["da_price"], da_min, da_mid, da_max),
                    fmt(item.get("rt", 0.0)) if self.compare_mode == "realtime_vs_dayahead" else fmt(item["ml"]),
                    self._price_cell(
                        item.get("rt_price", 0.0) if self.compare_mode == "realtime_vs_dayahead" else item["ml_price"],
                        compare_min,
                        compare_mid,
                        compare_max,
                    ),
                    fmt(self._display_dev(item)),
                    fmt(self._display_price_diff(item)),
                    fmt(self._display_profit(item)),
                    fmt(output) if output is not None else "-",
                    match_html,
                    direction_html,
                ]
            )
        headers = (
            ["时点", "日前电量", "日前价", "实时电量", "实时价", "偏差电量", "价差", "偏差收益(万元)", "机组出力", "量价匹配", "出力偏差"]
            if self.compare_mode == "realtime_vs_dayahead"
            else ["时点", "日前电量", "日前价", "中长期电量", "中长期价", "偏差电量", "价差", "偏差收益(万元)", "机组出力", "量价匹配", "出力偏差"]
        )
        return self._table(headers, rows, "detail detail-table")

    def _bid_check(self, units):
        paragraphs = []
        for unit in units:
            if unit in RENEWABLE_UNITS:
                continue
            card = self.offer_cards.get(unit, {})
            segment_text = "；".join(f"{fmt(seg['start'])}-{fmt(seg['end'])}MW @ {fmt(seg['price'])}" for seg in card.get("segments", [])) or "-"
            hourly_results = [
                self._evaluate_offer_match(
                    unit,
                    point.get("rt", 0.0) if self.compare_mode == "realtime_vs_dayahead" else point["da"],
                    point.get("rt_price", 0.0) if self.compare_mode == "realtime_vs_dayahead" else point["da_price"],
                )[1]
                for point in self.hourly.get(unit, [])
            ]
            mismatch_count = sum(1 for result in hourly_results if isinstance(result, str) and result.startswith("不匹配"))
            badge = f'<span class="ok">全部匹配</span>' if mismatch_count == 0 and hourly_results else f'<span class="bad">{mismatch_count}</span>'
            paragraphs.append(
                f'<p><b>{esc(unit)}</b>：厂用电率 {fmt(card.get("plant_rate", 0) * 100)}%；报价段（连续同价已合并）：{esc(segment_text)}；不匹配小时 {badge}</p>'
            )
        if not paragraphs:
            return '<p class="muted">新能源单元不做报价单与量价匹配校验。</p>'
        if self.compare_mode == "realtime_vs_dayahead":
            rule_text = "规则：实时价≤报价阈值，机端出力不应超过该报价段起始出力；实时价&gt;报价阈值，机端出力应达到该报价段终止出力。机端出力=实时出清电量/(1-厂用电率)。连续相同报价段合并判断。</p>"
        else:
            rule_text = "规则：日前价≤报价阈值，机端出力不应超过该报价段起始出力；日前价&gt;报价阈值，机端出力应达到该报价段终止出力。机端出力=日前出清电量/(1-厂用电率)。连续相同报价段合并判断。</p>"
        return (
            '<p class="note">' + rule_text
            + "".join(paragraphs)
        )

    def _render(self):
        sections = []
        for company_name, units in self.companies.items():
            item = self._company_summary(company_name, units)
            if company_name in RENEWABLE_UNITS:
                item["middle_label"] = "全电量均价"
                item["middle_value"] = f"{fmt(self.summary[company_name].get('full_price'))} 元/MWh"
            representative_rows = []
            for point in self._representative_points(units):
                representative_rows.append(
                    [point["type"], esc(point["unit"]), esc(point["time"]), fmt(point["dev"]), fmt(point["price_diff"]), fmt(point["profit"])]
                )
            details = "".join(f'<div class="split"><h4>{esc(unit)}（0-23点）</h4>{self._detail_table(unit)}</div>' for unit in units if unit in self.hourly)
            detail_title = "逐时出清明细（按机组分别展示）" if company_name not in RENEWABLE_UNITS else "逐时出清明细（单元0-23点）"
            bid_block = (
                f'<div class="card"><h3>报价单与量价匹配校验</h3>{self._bid_check(units)}</div>'
                if company_name not in RENEWABLE_UNITS
                else '<div class="card note">新能源单元：不做报价单及量价匹配校验，仅分析日前出清、中长期偏差与收益代表点。</div>'
            )
            sections.append(
                f"""
<h2 id="{esc(company_name)}">{esc(company_name)}</h2>
  {self._kpis(item)}
  <div class="card"><h3>{"实时-日前偏差代表点" if self.compare_mode == "realtime_vs_dayahead" else "日前-中长期偏差代表点"}</h3>
  {self._table(["类型", "单元", "时点", "偏差电量(MWh)", "价差(元/MWh)", "偏差收益(万元)"], representative_rows, "compact")}
  </div>
  {bid_block}
  <div class="card"><h3>{detail_title}</h3>
  {details}
  </div>"""
            )

        thermal_units = [unit for company in THERMAL_COMPANIES for unit in self.companies[company]]
        thermal = self._company_summary("分公司火电", thermal_units)
        new_energy = self._company_summary("新能源", RENEWABLE_UNITS)
        all_da = thermal.get("da_mwh", 0) + new_energy.get("da_mwh", 0)
        all_rt = thermal.get("rt_mwh", 0) + new_energy.get("rt_mwh", 0)
        all_ml = thermal.get("ml_mwh", 0) + new_energy.get("ml_mwh", 0)
        all_profit = thermal.get("profit_rt_wanyuan", 0) + new_energy.get("profit_rt_wanyuan", 0) if self.compare_mode == "realtime_vs_dayahead" else thermal.get("profit_wanyuan", thermal.get("profit_ml_wanyuan", 0)) + new_energy.get("profit_ml_wanyuan", 0)
        run_date = self._report_date_text()
        compare_pill = "实时数据口径" if self.compare_mode == "realtime_vs_dayahead" else "日前数据口径"
        method_text = (
            "本报告按提示词要求采用“实时 vs 日前”口径；偏差电量 = 实时电量 - 日前电量，价差 = 实时价格 - 日前价格，偏差收益估算 = 偏差电量 × 价差 / 10000，单位为万元。"
            if self.compare_mode == "realtime_vs_dayahead"
            else "本报告按提示词要求采用“日前 vs 中长期”口径；偏差电量 = 日前电量 - 中长期电量，价差 = 日前价格 - 中长期价格。火电偏差收益按“日前偏差电量 × （日前电价 - 成本）/ 10000”估算，成本取机组逐时明细中的成本列；新能源仍按偏差电量 × 价差 / 10000 估算。工作簿中的实时出清列为空或为 0，本报告不做实时出清分析。"
        )
        overview_kpis = (
            {"da_mwh": all_da, "rt_mwh": all_rt, "dev_rt_mwh": all_rt - all_da, "profit_rt_wanyuan": all_profit}
            if self.compare_mode == "realtime_vs_dayahead"
            else {"da_mwh": all_da, "ml_mwh": all_ml, "dev_ml_mwh": all_da - all_ml, "profit_ml_wanyuan": all_profit}
        )
        overview_note = (
            f"火电实时出清电量 {fmt(thermal.get('rt_mwh', 0))} MWh，较日前偏差 {fmt(thermal.get('dev_rt_mwh', 0))} MWh，偏差收益估算 {fmt(thermal.get('profit_rt_wanyuan', 0))} 万元；新能源合计实时出清电量 {fmt(new_energy.get('rt_mwh', 0))} MWh，偏差收益估算 {fmt(new_energy.get('profit_rt_wanyuan', 0))} 万元。"
            if self.compare_mode == "realtime_vs_dayahead"
            else f"火电日前出清电量 {fmt(thermal['da_mwh'])} MWh，较中长期偏差 {fmt(thermal.get('dev_mwh', thermal.get('dev_ml_mwh', 0)))} MWh，偏差收益估算 {fmt(thermal.get('profit_wanyuan', thermal.get('profit_ml_wanyuan', 0)))} 万元；新能源合计日前出清电量 {fmt(new_energy['da_mwh'])} MWh，偏差收益估算 {fmt(new_energy.get('profit_ml_wanyuan', 0))} 万元。"
        )

        css = """
*{box-sizing:border-box}body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","Microsoft YaHei",Arial,sans-serif;margin:0;background:#f6f7fb;color:#1f2937;line-height:1.55}.sidebar{position:fixed;top:0;left:0;width:220px;height:100vh;background:#fff;box-shadow:4px 0 16px #0001;padding:24px 16px;overflow-y:auto;z-index:100}.sidebar h3{margin-top:0;font-size:16px;margin-bottom:16px;border-bottom:2px solid #eef2ff;padding-bottom:8px}.sidebar ul{list-style:none;padding:0;margin:0}.sidebar li{margin-bottom:8px}.sidebar a{display:block;padding:10px 12px;color:#1f2937;text-decoration:none;border-radius:8px;transition:all .2s}.sidebar a:hover{background:#eef2ff;color:#2563eb}.wrap{margin-left:240px;max-width:1400px;padding:28px}h1{font-size:28px;margin:0 0 12px}.meta{color:#6b7280;font-size:13px;margin-bottom:10px}h2{border-left:6px solid #2563eb;padding-left:10px;margin-top:36px;font-size:22px}h3{margin-top:0;font-size:17px}h4{font-size:16px;margin:12px 0}.card{background:#fff;border-radius:14px;padding:20px;margin:14px 0;box-shadow:0 4px 16px #0001}.kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin:14px 0}.kpi{background:#f8fafc;border:1px solid #e5e7eb;border-radius:12px;padding:12px}.kpi span{display:block;color:#6b7280;font-size:13px}.kpi b{display:block;font-size:20px;margin-top:4px}table{width:100%;border-collapse:collapse;font-size:13px;margin:12px 0;background:white}th,td{border:1px solid #e5e7eb;padding:6px 8px;text-align:right;white-space:nowrap;vertical-align:top}th{background:#eef2ff;color:#111827}td:first-child,th:first-child{text-align:left}.ok{color:#047857;font-weight:700}.bad{color:#dc2626;font-weight:700}.note,.muted{color:#6b7280;font-size:13px}.split{margin-top:24px;border-top:2px dashed #d1d5db;padding-top:12px}.toc a{margin-right:10px;color:#2563eb}.detail{display:block;overflow-x:auto}.detail-table{min-width:1380px;font-size:15px}.detail-table th,.detail-table td{padding:9px 12px}.detail-table th{font-size:14px}.method{font-size:13px}.compact td,.compact th{font-size:13px}.pill{display:inline-block;background:#eef2ff;color:#1d4ed8;border-radius:999px;padding:2px 8px;font-size:12px;margin-left:8px}@media(max-width:800px){.sidebar{width:100%;height:auto;position:relative;box-shadow:none;padding:16px}.wrap{margin-left:0;padding:18px}.kpis{grid-template-columns:1fr 1fr}table{font-size:12px}.detail-table{min-width:1100px;font-size:13px}.detail-table th,.detail-table td{padding:8px 10px}}
"""

        return f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>电力交易出清分析报告</title><link rel="icon" href="data:,"><style>{css}</style></head>
<body>
<aside class="sidebar">
<h3>目录导航</h3>
<ul>
<li><a href="#总览">总览</a></li>
{''.join(f'<li><a href="#{esc(name)}">{esc(name)}</a></li>' for name in self.companies)}
</ul>
</aside>
<main class="wrap">
<h1>电力交易出清分析报告</h1>
<div class="meta">运行日期：{run_date} <span class="pill">{compare_pill}</span></div>
<div class="card note method"><b>方法说明：</b>
{method_text} 报价单取自“运行补偿”页各机组出力段报价，量价匹配按出清电量 ÷ (1 - 厂用电率) 得到机端出力后逐小时判断，连续相同报价段合并判断。新能源单元海风、鮀莲、归湖仅展示电量、价格、偏差和收益，不做报价单与量价匹配。收益测算标签只作为数据块来源，不作为机组名称展示。</div>
<h2 id="总览">总览</h2>
{self._kpis(overview_kpis)}
<div class="card note">{overview_note}</div>
{''.join(sections)}
</main></body></html>"""


def generate_report(xlsx_path, out_path=None):
    return ReportBuilder(xlsx_path, out_path=out_path).build()


def generate_batch(xlsx_paths, output_dir):
    output_dir = pathlib.Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    results = []
    for xlsx_path in xlsx_paths:
        xlsx_path = pathlib.Path(xlsx_path)
        out_path = output_dir / f"{xlsx_path.stem}_分析报告.html"
        results.append(generate_report(xlsx_path, out_path=out_path))
    return results


class ReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("电力交易出清分析")
        self.root.geometry("860x620")
        self.files = []
        self.output_dir = tk.StringVar(value=str(DEFAULT_OUTPUT_DIR))
        self.status = tk.StringVar(value="请选择一个或多个 Excel 文件。")
        self._build_ui()

    def _build_ui(self):
        frame = ttk.Frame(self.root, padding=16)
        frame.pack(fill="both", expand=True)

        title = ttk.Label(frame, text="电力交易出清分析", font=("Microsoft YaHei UI", 16, "bold"))
        title.pack(anchor="w")
        ttk.Label(frame, text="可一次选择多个出清 Excel，批量生成 HTML 报告。", foreground="#667085").pack(anchor="w", pady=(4, 14))

        toolbar = ttk.Frame(frame)
        toolbar.pack(fill="x")
        ttk.Button(toolbar, text="添加文件", command=self.add_files).pack(side="left")
        ttk.Button(toolbar, text="移除选中", command=self.remove_selected).pack(side="left", padx=(8, 0))
        ttk.Button(toolbar, text="清空列表", command=self.clear_files).pack(side="left", padx=(8, 0))

        list_frame = ttk.Frame(frame)
        list_frame.pack(fill="both", expand=True, pady=(12, 12))
        self.listbox = tk.Listbox(list_frame, selectmode=tk.EXTENDED, font=("Consolas", 10))
        self.listbox.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.listbox.configure(yscrollcommand=scrollbar.set)

        output_frame = ttk.LabelFrame(frame, text="输出目录", padding=12)
        output_frame.pack(fill="x")
        ttk.Entry(output_frame, textvariable=self.output_dir).pack(side="left", fill="x", expand=True)
        ttk.Button(output_frame, text="选择目录", command=self.choose_output_dir).pack(side="left", padx=(8, 0))

        action_frame = ttk.Frame(frame)
        action_frame.pack(fill="x", pady=(14, 8))
        self.run_button = ttk.Button(action_frame, text="开始分析", command=self.run_batch)
        self.run_button.pack(side="left")
        ttk.Button(action_frame, text="打开输出目录", command=self.open_output_dir).pack(side="left", padx=(8, 0))

        log_frame = ttk.LabelFrame(frame, text="运行日志", padding=12)
        log_frame.pack(fill="both", expand=False, pady=(8, 0))
        self.log_text = tk.Text(log_frame, height=10, wrap="word", state="disabled", font=("Consolas", 10))
        self.log_text.pack(fill="both", expand=True)

        status_bar = ttk.Label(frame, textvariable=self.status, foreground="#475467")
        status_bar.pack(anchor="w", pady=(10, 0))

    def add_files(self):
        paths = filedialog.askopenfilenames(
            title="选择出清 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if not paths:
            return
        existing = {str(path) for path in self.files}
        for path in paths:
            if path not in existing:
                self.files.append(pathlib.Path(path))
                self.listbox.insert(tk.END, path)
        self.status.set(f"已选择 {len(self.files)} 个文件。")

    def remove_selected(self):
        selected = list(self.listbox.curselection())
        if not selected:
            return
        for index in reversed(selected):
            del self.files[index]
            self.listbox.delete(index)
        self.status.set(f"剩余 {len(self.files)} 个文件。")

    def clear_files(self):
        self.files.clear()
        self.listbox.delete(0, tk.END)
        self.status.set("文件列表已清空。")

    def choose_output_dir(self):
        path = filedialog.askdirectory(title="选择输出目录", initialdir=self.output_dir.get())
        if path:
            self.output_dir.set(path)

    def open_output_dir(self):
        path = pathlib.Path(self.output_dir.get())
        path.mkdir(parents=True, exist_ok=True)
        self.root.after(0, lambda: path.resolve())
        import os

        os.startfile(path)

    def append_log(self, text):
        self.log_text.configure(state="normal")
        self.log_text.insert(tk.END, text + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state="disabled")

    def run_batch(self):
        if not self.files:
            messagebox.showwarning("没有文件", "请先添加至少一个 Excel 文件。")
            return
        output_dir = pathlib.Path(self.output_dir.get())
        output_dir.mkdir(parents=True, exist_ok=True)
        self.run_button.configure(state="disabled")
        self.status.set("正在生成报告，请稍候...")
        self.append_log(f"开始批量分析，共 {len(self.files)} 个文件。")
        threading.Thread(target=self._run_batch_worker, daemon=True).start()

    def _run_batch_worker(self):
        success = []
        errors = []
        output_dir = pathlib.Path(self.output_dir.get())
        for path in self.files:
            try:
                out_path = output_dir / f"{path.stem}_分析报告.html"
                generate_report(path, out_path=out_path)
                success.append(out_path)
                self.root.after(0, self.append_log, f"完成：{path.name} -> {out_path.name}")
            except Exception as exc:  # noqa: BLE001
                errors.append((path, exc))
                self.root.after(0, self.append_log, f"失败：{path.name} -> {exc}")
        self.root.after(0, self._finish_batch, success, errors)

    def _finish_batch(self, success, errors):
        self.run_button.configure(state="normal")
        if errors:
            self.status.set(f"完成 {len(success)} 个，失败 {len(errors)} 个。")
            messagebox.showwarning("部分完成", f"成功 {len(success)} 个，失败 {len(errors)} 个。请查看日志。")
        else:
            self.status.set(f"全部完成，共生成 {len(success)} 个报告。")
            messagebox.showinfo("完成", f"已生成 {len(success)} 个报告。")


def launch_gui():
    root = tk.Tk()
    try:
        root.iconbitmap(default="")
    except Exception:  # noqa: BLE001
        pass
    ttk.Style().theme_use("vista")
    ReportApp(root)
    root.mainloop()


def parse_args():
    parser = argparse.ArgumentParser(description="电力交易出清分析报告生成器")
    parser.add_argument("xlsx", nargs="*", help="一个或多个 Excel 文件路径")
    parser.add_argument("--output-dir", help="批量输出目录")
    parser.add_argument("--output", help="单文件输出 HTML 路径")
    parser.add_argument("--no-gui", action="store_true", help="强制命令行模式")
    return parser.parse_args()


def main():
    args = parse_args()
    if not args.no_gui and not args.xlsx:
        launch_gui()
        return
    if not args.xlsx:
        raise SystemExit("命令行模式下至少传入一个 Excel 文件。")
    if len(args.xlsx) == 1:
        out_path = pathlib.Path(args.output) if args.output else DEFAULT_OUTPUT_DIR / DEFAULT_OUTPUT_NAME
        result = generate_report(args.xlsx[0], out_path=out_path)
        print(result)
        return
    output_dir = pathlib.Path(args.output_dir) if args.output_dir else DEFAULT_OUTPUT_DIR
    for result in generate_batch(args.xlsx, output_dir):
        print(result)


if __name__ == "__main__":
    main()
