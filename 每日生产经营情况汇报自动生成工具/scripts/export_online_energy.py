#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


SCRIPT_DIR = Path(__file__).resolve().parent
APP_ROOT = SCRIPT_DIR.parent if SCRIPT_DIR.name in {"scripts", "\u811a\u672c"} else SCRIPT_DIR
BASE_URL = "https://xhxt.chng.com.cn"
FIRE_APP_URL = f"{BASE_URL}/gdfire/PrivateDataManage/PowerSituationReport"
ENERGY_APP_URL = f"{BASE_URL}/gdfire/PrivateDataManageCE/PowerSituationReport"
LOGIN_URL = f"{BASE_URL}/usercenter/#/login"
PROFILE_DIR = SCRIPT_DIR / ".browser-profile"
AUTH_STATE_PATH = SCRIPT_DIR / "auth_state.json"
CONFIG_PATH = SCRIPT_DIR / "config.json"
OUTPUT_DIR = APP_ROOT / "\u8f93\u51fa"

FALLBACK_COMPANIES = [
    {"tenantId": "e4e6eb5c80731ac70180fab1ba2f0559", "name": "汕头电厂"},
    {"tenantId": "e4e6eb5c80731ac70180fab3532d0592", "name": "海门电厂"},
    {"tenantId": "e4e6eb5c80731ac70180faa7f96904eb", "name": "谢岗电厂"},
    {"tenantId": "e4c88ecc8ec18540018eeb6d767241fe", "name": "华能广东汕头海上风电有限责任公司"},
    {"tenantId": "e4c9c04d959356250196b2b15da14f65", "name": "华能（汕头金平）新能源有限责任公司"},
    {"tenantId": "e4c17edf9942cca6019974cabef4188d", "name": "华能（潮州潮安）新能源有限责任公司"},
]

COMPANY_DISPLAY_NAME = {
    "汕头电厂": "汕头",
    "海门电厂": "海门",
    "谢岗电厂": "东莞",
    "华能广东汕头海上风电有限责任公司": "海风",
    "华能（汕头金平）新能源有限责任公司": "金平",
    "华能（潮州潮安）新能源有限责任公司": "潮安",
}

COMPANY_DISPLAY_ORDER = ["汕头", "海门", "东莞", "海风", "金平", "潮安"]


@dataclass
class Config:
    username: str = ""
    password: str = ""
    headless: bool = True
    request_delay_seconds: float = 1.0
    output_dir: Path = OUTPUT_DIR


def log(message: str) -> None:
    print(message, flush=True)


def load_config() -> Config:
    if not CONFIG_PATH.exists():
        return Config()
    raw = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    return Config(
        username=str(raw.get("username", "") or ""),
        password=str(raw.get("password", "") or ""),
        headless=bool(raw.get("headless", True)),
        request_delay_seconds=float(raw.get("requestDelaySeconds", 1.0)),
        output_dir=Path(raw.get("outputDir", str(OUTPUT_DIR))).expanduser(),
    )


def validate_date(value: str) -> str:
    value = value.strip()
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        raise ValueError("日期格式应为 YYYY-MM-DD，例如 2026-05-10")
    datetime.strptime(value, "%Y-%m-%d")
    return value


def ask_date() -> str:
    while True:
        try:
            return validate_date(input("请输入运行日 YYYY-MM-DD: "))
        except ValueError as exc:
            print(f"日期不合法：{exc}")


def parse_json_response(response, label: str) -> dict[str, Any]:
    text = response.text()
    if response.status >= 400:
        raise RuntimeError(f"{label} HTTP {response.status}: {text[:300]}")
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"{label} 返回的不是 JSON，可能登录已失效。片段：{text[:200]}") from exc
    if data.get("retCode") and data.get("retCode") != "T200":
        raise RuntimeError(f"{label} 失败：{data.get('retMsg') or data}")
    return data


def json_or_text(response) -> Any:
    text = response.text()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return text


def load_saved_auth_state(context) -> None:
    if not AUTH_STATE_PATH.exists():
        return
    try:
        state = json.loads(AUTH_STATE_PATH.read_text(encoding="utf-8"))
        cookies = state.get("cookies") or []
        if cookies:
            context.add_cookies(cookies)
    except Exception as exc:
        log(f"读取 auth_state.json 失败，将忽略：{exc}")


def save_auth_state(context) -> None:
    AUTH_STATE_PATH.write_text(json.dumps(context.storage_state(), ensure_ascii=False, indent=2), encoding="utf-8")


def launch_context(playwright, *, headless: bool):
    PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    try:
        context = playwright.chromium.launch_persistent_context(
            str(PROFILE_DIR),
            channel="chrome",
            headless=headless,
            accept_downloads=True,
        )
    except PlaywrightError:
        context = playwright.chromium.launch_persistent_context(
            str(PROFILE_DIR),
            headless=headless,
            accept_downloads=True,
        )
    load_saved_auth_state(context)
    return context


def is_logged_in(context) -> bool:
    urls = [
        f"{BASE_URL}/gdfire/api/pf/tenant/user/info",
        f"{BASE_URL}/usercenter/web/pf/tenant/user/info",
    ]
    for url in urls:
        try:
            response = context.request.get(url, timeout=15000)
        except PlaywrightError:
            continue
        if response.status != 200:
            continue
        try:
            data = response.json()
        except Exception:
            continue
        if data.get("retCode") == "T200" and bool(data.get("data", {}).get("tenantId")):
            return True
    return False


def base64_decode(value: str) -> bytes:
    import base64

    return base64.b64decode(value)


def base64_encode(value: bytes) -> str:
    import base64

    return base64.b64encode(value).decode("ascii")


def rsa_encrypt_password(public_key_b64: str, password: str) -> str:
    from cryptography.hazmat.primitives import serialization
    from cryptography.hazmat.primitives.asymmetric import padding

    public_key = serialization.load_der_public_key(base64_decode(public_key_b64))
    encrypted = public_key.encrypt(password.encode("utf-8"), padding.PKCS1v15())
    return base64_encode(encrypted)


def try_api_login(context, username: str, password: str) -> bool:
    key_response = context.request.get(f"{BASE_URL}/usercenter/web/pf/login/info/publicKey", timeout=20000)
    key_data = parse_json_response(key_response, "读取登录公钥")
    public_key = key_data.get("data")
    if not isinstance(public_key, str) or not public_key:
        raise RuntimeError("读取登录公钥失败：返回中没有公钥")

    encrypted_password = rsa_encrypt_password(public_key, password)
    login_response = context.request.post(
        f"{BASE_URL}/usercenter/web/login",
        params={"loginMode": 2, "username": username, "password": encrypted_password},
        timeout=30000,
    )
    payload = json_or_text(login_response)
    if login_response.status >= 400:
        raise RuntimeError(f"登录接口 HTTP {login_response.status}: {str(payload)[:300]}")
    if isinstance(payload, dict) and payload.get("retCode") not in (None, "T200"):
        raise RuntimeError(f"登录失败：{payload.get('retMsg') or payload}")

    page = context.new_page()
    try:
        page.goto(FIRE_APP_URL, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_load_state("networkidle", timeout=15000)
    except PlaywrightTimeoutError:
        pass
    finally:
        page.close()

    logged_in = is_logged_in(context)
    if logged_in:
        save_auth_state(context)
    return logged_in


def try_ui_login(context, username: str, password: str) -> bool:
    page = context.new_page()
    page.goto(LOGIN_URL, wait_until="domcontentloaded")
    try:
        page.locator("#username, input[placeholder*='工号'], input[placeholder*='手机号'], input[placeholder*='账号']").first.fill(
            username, timeout=10000
        )
        page.locator("#password, input[type='password']").first.fill(password, timeout=10000)
        page.locator("button:has-text('登 录'), button:has-text('登录'), button[type='submit']").first.click(timeout=10000)
    except (PlaywrightTimeoutError, PlaywrightError):
        pass

    deadline = time.time() + 45
    while time.time() < deadline:
        if is_logged_in(context):
            save_auth_state(context)
            page.close()
            return True
        time.sleep(2)
    page.close()
    return False


def try_auto_login(context, username: str, password: str) -> bool:
    try:
        log("正在尝试接口登录...")
        if try_api_login(context, username, password):
            return True
        log("接口登录后仍未检测到登录态，改用页面登录...")
    except Exception as exc:
        log(f"接口登录失败：{exc}")
        log("改用页面登录...")
    return try_ui_login(context, username, password)


def interactive_login(context, config: Config) -> None:
    if config.username and config.password and try_auto_login(context, config.username, config.password):
        save_auth_state(context)
        log("自动登录成功，登录态已保存。")
        return

    page = context.new_page()
    page.goto(FIRE_APP_URL, wait_until="domcontentloaded")
    log("请在打开的浏览器里完成登录，然后回到这里按 Enter。")
    input()
    if not is_logged_in(context):
        raise RuntimeError("仍未检测到有效登录态，请确认浏览器里已经登录成功。")
    save_auth_state(context)
    page.close()
    log("登录态已保存。")


def ensure_login(context, config: Config) -> None:
    if is_logged_in(context):
        return
    if not config.username or not config.password:
        raise RuntimeError("未检测到有效登录态，且 config.json 没有填写账号密码。")
    log("未检测到有效登录态，正在使用 config.json 的账号密码自动登录...")
    if not try_auto_login(context, config.username, config.password):
        raise RuntimeError("自动登录失败，请运行 login.bat 手动保存一次登录态。")
    log("自动登录成功。")


def get_companies(context) -> list[dict[str, str]]:
    response = context.request.get(f"{BASE_URL}/gdfire/api/pf/tenant/user/tenant/grantApplication", timeout=20000)
    data = parse_json_response(response, "读取公司列表")
    items = data.get("data") or []
    companies = [{"tenantId": item["tenantId"], "name": item["name"]} for item in items if item.get("tenantId") and item.get("name")]
    if not companies:
        companies = FALLBACK_COMPANIES

    order_lookup = {name: index for index, name in enumerate(COMPANY_DISPLAY_ORDER)}
    return sorted(
        companies,
        key=lambda item: (
            order_lookup.get(COMPANY_DISPLAY_NAME.get(item["name"], item["name"]), 999),
            item["name"],
        ),
    )


def switch_company(context, tenant_id: str) -> None:
    response = context.request.get(
        f"{BASE_URL}/usercenter/web/switchTenant",
        params={"tenantId": tenant_id},
        timeout=20000,
    )
    if response.status >= 400:
        raise RuntimeError(f"切换公司失败 HTTP {response.status}")


def get_units(context) -> list[dict[str, Any]]:
    response = context.request.get(f"{BASE_URL}/gdfire/api/unit/list", timeout=20000)
    data = parse_json_response(response, "读取机组/场站列表")
    return data.get("data") or []


def unit_display_name(unit: dict[str, Any]) -> str:
    return str(unit.get("unitName") or unit.get("unitShortName") or unit.get("unitAlias") or unit.get("id") or "").strip()


def natural_key(value: str) -> list[Any]:
    return [int(part) if part.isdigit() else part for part in re.split(r"(\d+)", value)]


def is_renewable_unit(unit: dict[str, Any]) -> bool:
    text = " ".join(
        str(unit.get(key) or "")
        for key in ("groupBusinessType", "groupSubType", "businessType", "subType", "unitName", "unitShortName")
    ).upper()
    return any(marker in text for marker in ("PHOTOVOLTAIC", "WIND", "新能源", "光伏", "风电"))


def is_renewable_company(units: list[dict[str, Any]]) -> bool:
    return bool(units) and any(is_renewable_unit(unit) for unit in units)


def number_or_zero(value: Any) -> float:
    if value in ("", None, "-"):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def fetch_fire_rows(context, run_date: str, tenant_id: str) -> list[dict[str, Any]]:
    response = context.request.get(
        f"{BASE_URL}/gdfire/api/data/unit/power/fire",
        params={"runDate": run_date, "orgIds": tenant_id, "endDate": run_date},
        timeout=30000,
    )
    data = parse_json_response(response, "读取火电上网电量")
    rows = data.get("data") or []
    if not isinstance(rows, list):
        raise RuntimeError("火电上网电量返回格式不正确")
    return rows


def fetch_energy_rows(context, run_date: str, tenant_id: str, units: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for unit in units:
        unit_id = str(unit.get("id") or "")
        if not unit_id:
            continue
        response = context.request.get(
            f"{BASE_URL}/gdfire/api/data/unit/power/energy",
            params={"runDate": run_date, "orgIds": tenant_id, "unitIds": unit_id},
            timeout=30000,
        )
        data = parse_json_response(response, f"读取新能源上网电量 {unit_display_name(unit) or unit_id}")
        items = data.get("data") or []
        if not isinstance(items, list):
            raise RuntimeError("新能源上网电量返回格式不正确")
        rows.extend(items)
    return rows


def collect_detail_rows(run_date: str, config: Config) -> list[dict[str, Any]]:
    detail_rows: list[dict[str, Any]] = []

    with sync_playwright() as playwright:
        context = launch_context(playwright, headless=config.headless)
        try:
            ensure_login(context, config)
            companies = get_companies(context)
            log(f"开始抓取 {run_date} 上网电量，共 {len(companies)} 家公司。")

            for index, company in enumerate(companies, start=1):
                tenant_id = company["tenantId"]
                company_name = company["name"]
                log(f"[{index}/{len(companies)}] 切换公司：{company_name}")
                switch_company(context, tenant_id)
                time.sleep(0.3)

                units = sorted(get_units(context), key=lambda item: natural_key(unit_display_name(item)))
                unit_names = {str(item.get("id")): unit_display_name(item) for item in units if item.get("id")}
                unit_order = {str(item.get("id")): pos for pos, item in enumerate(units) if item.get("id")}

                if is_renewable_company(units):
                    raw_rows = fetch_energy_rows(context, run_date, tenant_id, units)
                else:
                    raw_rows = fetch_fire_rows(context, run_date, tenant_id)

                power_items = [item for item in raw_rows if str(item.get("unitId") or "") != "-"]
                power_items.sort(
                    key=lambda item: (
                        unit_order.get(str(item.get("unitId") or ""), 999),
                        natural_key(unit_names.get(str(item.get("unitId") or ""), str(item.get("unitId") or ""))),
                    )
                )

                for item in power_items:
                    unit_id = str(item.get("unitId") or "")
                    unit_name = unit_names.get(unit_id, unit_id or "未知机组")
                    values = item.get("onlineEle") or []
                    if not isinstance(values, list):
                        values = []

                    padded_values = [number_or_zero(values[hour]) if hour < len(values) else 0.0 for hour in range(24)]
                    for hour, value in enumerate(padded_values):
                        detail_rows.append(
                            {
                                "company": company_name,
                                "unit": unit_name,
                                "time": f"{hour:02d}:00",
                                "onlineEle": value,
                            }
                        )

                log(f"  完成：{company_name}，记录 {len(power_items)} 条。")
                time.sleep(config.request_delay_seconds)
        finally:
            context.close()

    return detail_rows


def display_unit_header(value: str) -> str:
    match = re.search(r"#(\d+)", value)
    if match:
        return f"{match.group(1)}号机"
    return value


def display_company_name(value: str) -> str:
    return COMPANY_DISPLAY_NAME.get(value, value)


def build_company_blocks(detail_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in detail_rows:
        company = row["company"]
        target = grouped.setdefault(company, {"units": [], "rows": {}})
        if row["unit"] not in target["units"]:
            target["units"].append(row["unit"])
        target["rows"][(row["time"], row["unit"])] = row["onlineEle"]

    blocks: list[dict[str, Any]] = []
    ordered_times = [f"{hour:02d}:00" for hour in range(24)]
    for company_name in COMPANY_DISPLAY_ORDER:
        raw_name = next((key for key in grouped if display_company_name(key) == company_name), None)
        if raw_name is None:
            continue
        payload = grouped[raw_name]
        rows = []
        for time_key in ordered_times:
            rows.append([payload["rows"].get((time_key, unit), 0.0) for unit in payload["units"]])
        blocks.append(
            {
                "company": company_name,
                "units": [display_unit_header(unit) for unit in payload["units"]],
                "rows": rows,
            }
        )
    return blocks


def write_summary_sheet(ws, detail_rows: list[dict[str, Any]]) -> None:
    title_fill = PatternFill("solid", fgColor="EDEDED")
    header_fill = PatternFill("solid", fgColor="F6F6F6")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    blocks = build_company_blocks(detail_rows)
    current_row = 1
    max_columns = max((len(block["units"]) for block in blocks), default=1)

    for block in blocks:
        unit_count = max(len(block["units"]), 1)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=unit_count)
        title_cell = ws.cell(current_row, 1, f"{block['company']} 实际上网电量（万千瓦时）")
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = title_fill
        for col in range(1, unit_count + 1):
            ws.cell(current_row, col).border = border

        current_row += 1
        for col, unit in enumerate(block["units"], start=1):
            cell = ws.cell(current_row, col, unit)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
            cell.border = border

        current_row += 1
        for values in block["rows"]:
            for col, value in enumerate(values, start=1):
                cell = ws.cell(current_row, col, value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            current_row += 1

        current_row += 1

    for col in range(1, max_columns + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12


def save_excel(run_date: str, detail_rows: list[dict[str, Any]], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"上网电量汇总_{run_date}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "汇总"
    write_summary_sheet(ws, detail_rows)

    try:
        wb.save(path)
        return path
    except PermissionError:
        fallback = output_dir / f"上网电量汇总_{run_date}_{datetime.now().strftime('%H%M%S')}.xlsx"
        wb.save(fallback)
        return fallback


def export_online_energy(run_date: str, config: Config) -> Path:
    detail_rows = collect_detail_rows(run_date, config)
    return save_excel(run_date, detail_rows, config.output_dir)


def main() -> int:
    parser = argparse.ArgumentParser(description="导出指定日期各公司各机组上网电量")
    parser.add_argument("date", nargs="?", help="运行日，格式 YYYY-MM-DD")
    parser.add_argument("--login", action="store_true", help="登录并保存会话")
    args = parser.parse_args()

    config = load_config()
    try:
        if args.login:
            with sync_playwright() as playwright:
                context = launch_context(playwright, headless=False)
                try:
                    interactive_login(context, config)
                finally:
                    context.close()
            return 0

        run_date = validate_date(args.date) if args.date else ask_date()
        path = export_online_energy(run_date, config)
        log(f"\n完成：{path}")
        return 0
    except KeyboardInterrupt:
        print("\n已取消。")
        return 130
    except Exception as exc:
        print(f"\n失败：{exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
