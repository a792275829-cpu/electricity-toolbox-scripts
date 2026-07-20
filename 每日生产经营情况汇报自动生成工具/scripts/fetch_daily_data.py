#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import sync_playwright

from export_online_energy import (
    BASE_URL,
    OUTPUT_DIR,
    ensure_login,
    interactive_login,
    launch_context,
    load_config,
    request_json_with_retry,
    switch_company,
    validate_date,
)
from fetch_load_compare import (
    POINT_TO_TIME_SEGMENT,
    build_output_payload as build_load_output_payload,
    build_subject_payload,
    fetch_load_dataset,
    load_subject_map,
    metric_decimals,
    normalize_subject,
    select_load_type,
    build_summary,
)
from fetch_price_board import (
    build_output_payload as build_price_output_payload,
    fetch_price_summaries,
)


DEFAULT_POINT = 96
DEFAULT_DATA_TYPE = 1
DEFAULT_PROVINCE_AREA_ID = "044"
DEFAULT_FORECAST_ORG_ID = "e4c88ecc8ec18540018eeb6d767241fe"
DEFAULT_FORECAST_ORG_NAME = "华能广东汕头海上风电有限责任公司"
DEFAULT_FORECAST_UNIT_ID = "e4d5a0e78ecbe1f5018eeb93cff4392c"
DEFAULT_FORECAST_UNIT_NAME = "广澳海风场"


def log(message: str) -> None:
    print(message, flush=True)


def parse_subjects(value: str | None) -> list[str]:
    if not value or value.strip().lower() == "all":
        return list(load_subject_map().keys())
    return [item.strip() for item in value.split(",") if item.strip()]


def build_report_date_plan(report_date: str) -> dict[str, str]:
    report_day = datetime.strptime(report_date, "%Y-%m-%d").date()
    previous_day = report_day - timedelta(days=1)
    clearing_day = report_day - timedelta(days=2)
    forecast_start = report_day + timedelta(days=1)
    forecast_end = report_day + timedelta(days=5)
    return {
        "reportDate": report_day.strftime("%Y-%m-%d"),
        "compareDate": previous_day.strftime("%Y-%m-%d"),
        "actualOnlineEnergyDate": previous_day.strftime("%Y-%m-%d"),
        "dayAheadDate": report_day.strftime("%Y-%m-%d"),
        "loadAndPriceSelectedDate": report_day.strftime("%Y-%m-%d"),
        "loadAndPriceCompareDate": previous_day.strftime("%Y-%m-%d"),
        "dailyClearingDate": clearing_day.strftime("%Y-%m-%d"),
        "rollingForecastRunDate": report_day.strftime("%Y-%m-%d"),
        "rollingForecastStartDate": forecast_start.strftime("%Y-%m-%d"),
        "rollingForecastEndDate": forecast_end.strftime("%Y-%m-%d"),
    }


def sanitize_filename(value: str) -> str:
    cleaned = value
    for ch in '\\/:*?"<>|':
        cleaned = cleaned.replace(ch, "_")
    return cleaned


def run_load_compare_with_context(
    context,
    *,
    subjects: list[str],
    selected_date: str,
    compare_date: str,
    data_type: int,
    point: int,
    province_area_id: str,
) -> dict[str, Any]:
    selected_dataset = fetch_load_dataset(
        context,
        selected_date,
        selected_date,
        data_type=data_type,
        point=point,
        province_area_id=province_area_id,
    )
    compare_dataset = fetch_load_dataset(
        context,
        compare_date,
        compare_date,
        data_type=data_type,
        point=point,
        province_area_id=province_area_id,
    )

    subject_payloads = []
    for subject in subjects:
        subject_name, load_type = normalize_subject(subject)
        selected_series = select_load_type(selected_dataset, load_type)
        compare_series = select_load_type(compare_dataset, load_type)
        selected_summary = build_summary(selected_series, selected_date, selected_date, point=point)
        compare_summary = build_summary(compare_series, compare_date, compare_date, point=point)
        subject_payloads.append(build_subject_payload(subject_name, load_type, selected_summary, compare_summary))

    return build_load_output_payload(subject_payloads)


def date_only(value: Any) -> str:
    text = str(value or "")
    return text[:10]


def number_or_none(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def hourly_averages(values: list[Any]) -> list[float | None]:
    result: list[float | None] = []
    for hour in range(24):
        chunk = [number_or_none(value) for value in values[hour * 4 : hour * 4 + 4]]
        numbers = [value for value in chunk if value is not None]
        if len(numbers) == 4:
            result.append(round(sum(numbers) / 4, 4))
        else:
            result.append(None)
    return result


def build_forecast_row(item: dict[str, Any]) -> dict[str, Any]:
    values = item.get("modifyForecastPower") or item.get("forecastPower") or []
    numbers = [number_or_none(value) for value in values]
    valid_numbers = [value for value in numbers if value is not None]
    hours = hourly_averages(values)
    daily_total = round(sum(valid_numbers) / 4, 4) if valid_numbers else None
    daily_average = round(sum(valid_numbers) / len(valid_numbers), 4) if valid_numbers else None
    return {
        "forecastDate": date_only(item.get("date")),
        "runDate": date_only(item.get("runDate")),
        "pointCount": len(valid_numbers),
        "dailyTotal": daily_total,
        "dailyAverage": daily_average,
        "hours": hours,
        "rawValues": valid_numbers,
    }


def fetch_rolling_forecast(
    context,
    *,
    selected_date: str,
    org_id: str,
    org_name: str,
    unit_id: str,
    unit_name: str,
) -> dict[str, Any]:
    switch_company(context, org_id)
    data = request_json_with_retry(
        context,
        "post",
        f"{BASE_URL}/gdfire/api/data/personal/newEnergy/basic/query/v1",
        f"读取多日滚动预测 {selected_date}",
        data={
            "runDate": selected_date,
            "orgIds": [org_id],
            "unitIds": [unit_id],
        },
        timeout=60000,
    )
    raw_rows = data.get("data") or []
    if not isinstance(raw_rows, list):
        raise RuntimeError("多日滚动预测返回格式不正确")

    start_date = datetime.strptime(selected_date, "%Y-%m-%d").date() + timedelta(days=1)
    end_date = start_date + timedelta(days=4)
    rows = []
    for item in raw_rows:
        forecast_date_text = date_only(item.get("date"))
        if not forecast_date_text:
            continue
        forecast_date = datetime.strptime(forecast_date_text, "%Y-%m-%d").date()
        if start_date <= forecast_date <= end_date:
            rows.append(build_forecast_row(item))

    rows.sort(key=lambda item: item["forecastDate"])
    return {
        "source": "/gdfire/api/data/personal/newEnergy/basic/query/v1",
        "orgId": org_id,
        "orgName": org_name,
        "unitId": unit_id,
        "unitName": unit_name,
        "selectedDate": selected_date,
        "forecastStartDate": start_date.strftime("%Y-%m-%d"),
        "forecastEndDate": end_date.strftime("%Y-%m-%d"),
        "rows": rows,
        "rawCount": len(raw_rows),
    }


def fetch_capacity_summary(context, *, date_text: str, province_area_id: str) -> dict[str, Any]:
    payload = request_json_with_retry(
        context,
        "get",
        f"{BASE_URL}/gdfire/api/data/net/capacity",
        f"读取容量信息 {date_text}",
        params={
            "startDate": date_text,
            "endDate": date_text,
            "type": DEFAULT_DATA_TYPE,
            "provinceAreaId": province_area_id,
        },
        timeout=60000,
    )
    rows = payload.get("data", {}).get("dataNetCapacityDTOList") or []
    row = rows[0] if rows else {}
    capacity = number_or_none(row.get("capacity"))
    return {
        "date": date_text,
        "unit": "MW",
        "capacity": capacity,
        "capacityWanKw": int(capacity / 10) if capacity is not None else None,
        "versionDate": payload.get("data", {}).get("versionDate"),
        "raw": row,
    }


def fetch_startup_count_summary(context, *, date_text: str, province_area_id: str) -> dict[str, Any]:
    payload = request_json_with_retry(
        context,
        "get",
        f"{BASE_URL}/gdfire/api/spot/market/info",
        f"读取市场行情看板开机台数 {date_text}",
        params={
            "provinceAreaId": province_area_id,
            "timeSegment": POINT_TO_TIME_SEGMENT[DEFAULT_POINT],
            "startDate": date_text,
            "endDate": date_text,
        },
        timeout=60000,
    )
    rows = payload.get("data", {}).get("dataApplyConfigDOList") or []
    row = next(
        (
            item
            for item in rows
            if date_only(item.get("date")) == date_text
            and str(item.get("dataType") or "") == str(DEFAULT_DATA_TYPE)
        ),
        rows[0] if rows else {},
    )
    startup_count = number_or_none(row.get("onlineUnit"))
    return {
        "date": date_text,
        "unit": "台",
        "maxStartupCount": int(startup_count) if startup_count is not None else None,
        "unitCount": [startup_count] if startup_count is not None else [],
        "source": "/gdfire/api/spot/market/info dataApplyConfigDOList.onlineUnit",
        "raw": row,
    }


def build_operation_public_data(
    *,
    selected_capacity: dict[str, Any],
    compare_capacity: dict[str, Any],
    selected_startup: dict[str, Any],
    compare_startup: dict[str, Any],
) -> dict[str, Any]:
    selected_startup_value = selected_startup.get("maxStartupCount")
    compare_startup_value = compare_startup.get("maxStartupCount")
    startup_diff = None
    if selected_startup_value is not None and compare_startup_value is not None:
        startup_diff = selected_startup_value - compare_startup_value
    return {
        "capacity": {
            "selected": selected_capacity,
            "compare": compare_capacity,
        },
        "startupCount": {
            "selected": selected_startup,
            "compare": compare_startup,
            "difference": startup_diff,
        },
    }


def summarize_period(period: dict[str, Any] | None) -> dict[str, Any]:
    if not period:
        return {"max": None, "min": None, "avg": None, "sum": None}
    return {
        "max": number_or_none(period.get("max")),
        "min": number_or_none(period.get("min")),
        "avg": number_or_none(period.get("avg")),
        "sum": number_or_none(period.get("sum")),
        "maxDate": period.get("maxDate"),
        "minDate": period.get("minDate"),
    }


def fetch_load_info_summary(context, *, date_text: str, province_area_id: str) -> dict[str, Any]:
    payload = request_json_with_retry(
        context,
        "get",
        f"{BASE_URL}/gdfire/api/data/net/load",
        f"读取负荷信息 {date_text}",
        params={
            "startDate": date_text,
            "endDate": date_text,
            "provinceAreaId": province_area_id,
            "timeSegment": POINT_TO_TIME_SEGMENT[DEFAULT_POINT],
        },
        timeout=60000,
    )
    rows = payload.get("data", {}).get("dataNetLoadDTOList") or []
    by_type: dict[str, dict[str, Any]] = {}
    for row in rows:
        load_type = str(row.get("loadType") or "")
        if not load_type:
            continue
        forecast_periods = row.get("forecastPeriodList") or []
        actual_periods = row.get("actualPeriodList") or []
        by_type[load_type] = {
            "loadType": load_type,
            "dayAhead": summarize_period(forecast_periods[0] if forecast_periods else None),
            "actual": summarize_period(actual_periods[0] if actual_periods else None),
        }
    return {
        "date": date_text,
        "versionDate": payload.get("data", {}).get("versionDate"),
        "byType": by_type,
    }


def diff_summary(selected: dict[str, Any], compare: dict[str, Any]) -> dict[str, Any]:
    result = {"selected": selected, "compare": compare}
    for key in ("max", "min", "avg", "sum"):
        selected_value = selected.get(key)
        compare_value = compare.get(key)
        result[f"{key}Diff"] = (
            selected_value - compare_value
            if selected_value is not None and compare_value is not None
            else None
        )
    return result


def build_load_info_comparison(selected: dict[str, Any], compare: dict[str, Any]) -> dict[str, Any]:
    comparison: dict[str, Any] = {
        "selectedDate": selected["date"],
        "compareDate": compare["date"],
        "byType": {},
    }
    for load_type, selected_item in selected["byType"].items():
        compare_item = compare["byType"].get(load_type, {})
        comparison["byType"][load_type] = {
            "dayAhead": diff_summary(
                selected_item.get("dayAhead") or {},
                compare_item.get("dayAhead") or {},
            ),
            "actual": diff_summary(
                selected_item.get("actual") or {},
                compare_item.get("actual") or {},
            ),
        }
    return comparison


def build_market_analysis_data(
    *,
    load_info_selected: dict[str, Any],
    load_info_compare: dict[str, Any],
    selected_prices: list[dict[str, Any]],
    compare_prices: list[dict[str, Any]],
    selected_capacity: dict[str, Any],
    compare_capacity: dict[str, Any],
    selected_startup: dict[str, Any],
    compare_startup: dict[str, Any],
) -> dict[str, Any]:
    price_payload = build_price_output_payload(
        load_info_selected["date"],
        load_info_compare["date"],
        selected_prices,
        compare_prices,
    )
    return {
        "loadInfo": build_load_info_comparison(load_info_selected, load_info_compare),
        "priceBoard": price_payload,
        "operationPublicData": build_operation_public_data(
            selected_capacity=selected_capacity,
            compare_capacity=compare_capacity,
            selected_startup=selected_startup,
            compare_startup=compare_startup,
        ),
    }


def run_daily_data(
    *,
    selected_date: str,
    compare_date: str,
    subjects: list[str],
    data_type: int = DEFAULT_DATA_TYPE,
    point: int = DEFAULT_POINT,
    province_area_id: str = DEFAULT_PROVINCE_AREA_ID,
    forecast_org_id: str = DEFAULT_FORECAST_ORG_ID,
    forecast_org_name: str = DEFAULT_FORECAST_ORG_NAME,
    forecast_unit_id: str = DEFAULT_FORECAST_UNIT_ID,
    forecast_unit_name: str = DEFAULT_FORECAST_UNIT_NAME,
) -> dict[str, Any]:
    config = load_config()
    date_plan = build_report_date_plan(selected_date)
    market_date = date_plan["dailyClearingDate"]
    market_compare_date = (
        datetime.strptime(market_date, "%Y-%m-%d").date() - timedelta(days=1)
    ).strftime("%Y-%m-%d")
    with sync_playwright() as playwright:
        context = launch_context(playwright, headless=config.headless)
        try:
            ensure_login(context, config)
            load_payload = run_load_compare_with_context(
                context,
                subjects=subjects,
                selected_date=selected_date,
                compare_date=compare_date,
                data_type=data_type,
                point=point,
                province_area_id=province_area_id,
            )
            selected_prices = fetch_price_summaries(context, selected_date)
            compare_prices = fetch_price_summaries(context, compare_date)
            selected_capacity = fetch_capacity_summary(
                context,
                date_text=selected_date,
                province_area_id=province_area_id,
            )
            compare_capacity = fetch_capacity_summary(
                context,
                date_text=compare_date,
                province_area_id=province_area_id,
            )
            selected_startup = fetch_startup_count_summary(
                context,
                date_text=selected_date,
                province_area_id=province_area_id,
            )
            compare_startup = fetch_startup_count_summary(
                context,
                date_text=compare_date,
                province_area_id=province_area_id,
            )
            rolling_forecast = fetch_rolling_forecast(
                context,
                selected_date=selected_date,
                org_id=forecast_org_id,
                org_name=forecast_org_name,
                unit_id=forecast_unit_id,
                unit_name=forecast_unit_name,
            )
            market_load_info = fetch_load_info_summary(
                context,
                date_text=market_date,
                province_area_id=province_area_id,
            )
            market_compare_load_info = fetch_load_info_summary(
                context,
                date_text=market_compare_date,
                province_area_id=province_area_id,
            )
            market_prices = fetch_price_summaries(context, market_date)
            market_compare_prices = fetch_price_summaries(context, market_compare_date)
            market_capacity = fetch_capacity_summary(
                context,
                date_text=market_date,
                province_area_id=province_area_id,
            )
            market_compare_capacity = fetch_capacity_summary(
                context,
                date_text=market_compare_date,
                province_area_id=province_area_id,
            )
            market_startup = fetch_startup_count_summary(
                context,
                date_text=market_date,
                province_area_id=province_area_id,
            )
            market_compare_startup = fetch_startup_count_summary(
                context,
                date_text=market_compare_date,
                province_area_id=province_area_id,
            )
        finally:
            context.close()

    price_payload = build_price_output_payload(selected_date, compare_date, selected_prices, compare_prices)
    operation_public_data = build_operation_public_data(
        selected_capacity=selected_capacity,
        compare_capacity=compare_capacity,
        selected_startup=selected_startup,
        compare_startup=compare_startup,
    )
    return {
        "source": "负荷信息对比 + 价格看板",
        "reportDate": selected_date,
        "datePlan": date_plan,
        "selectedDate": selected_date,
        "compareDate": compare_date,
        "subjects": subjects,
        "loadCompare": load_payload,
        "priceBoard": price_payload,
        "operationPublicData": operation_public_data,
        "marketAnalysisData": build_market_analysis_data(
            load_info_selected=market_load_info,
            load_info_compare=market_compare_load_info,
            selected_prices=market_prices,
            compare_prices=market_compare_prices,
            selected_capacity=market_capacity,
            compare_capacity=market_compare_capacity,
            selected_startup=market_startup,
            compare_startup=market_compare_startup,
        ),
        "rollingForecast": rolling_forecast,
    }


def style_header(row) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_price_sheet(
    wb: openpyxl.Workbook,
    price_payload: dict[str, Any],
    *,
    title: str = "价格看板",
) -> None:
    ws = wb.create_sheet(title)
    headers = ["价格类型", "选择日期", "选择日均价(厘/kWh)", "选择点数", "对比日期", "对比日均价(厘/kWh)", "对比点数", "差值(厘/kWh)"]
    ws.append(headers)
    style_header(ws[1])
    for item in price_payload["rows"]:
        ws.append(
            [
                item["typeName"],
                item["selectedDate"],
                item["selectedDailyAvgPrice"],
                item["selectedPointCount"],
                item["compareDate"],
                item["compareDailyAvgPrice"],
                item["comparePointCount"],
                item["difference"],
            ]
        )
    for column, width in {"A": 12, "B": 14, "C": 20, "D": 12, "E": 14, "F": 20, "G": 12, "H": 16}.items():
        ws.column_dimensions[column].width = width


def write_rolling_forecast_sheet(wb: openpyxl.Workbook, forecast_payload: dict[str, Any]) -> None:
    ws = wb.create_sheet("多日滚动预测")
    ws["A1"] = "公司"
    ws["B1"] = forecast_payload["orgName"]
    ws["A2"] = "场站"
    ws["B2"] = forecast_payload["unitName"]
    ws["A3"] = "选择日期"
    ws["B3"] = forecast_payload["selectedDate"]
    ws["A4"] = "预测日期范围"
    ws["B4"] = f"{forecast_payload['forecastStartDate']} ~ {forecast_payload['forecastEndDate']}"

    headers = ["预测日期", "日合计", "日平均"] + [f"{hour:02d}:00" for hour in range(24)]
    for col, header in enumerate(headers, start=1):
        ws.cell(6, col, header)
    style_header(ws[6])

    for row_idx, item in enumerate(forecast_payload["rows"], start=7):
        ws.cell(row_idx, 1, item["forecastDate"])
        ws.cell(row_idx, 2, item["dailyTotal"])
        ws.cell(row_idx, 3, item["dailyAverage"])
        for hour_idx, value in enumerate(item["hours"], start=4):
            ws.cell(row_idx, hour_idx, value)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    for col in range(4, 28):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 11


def write_load_sheet(wb: openpyxl.Workbook, load_payload: dict[str, Any]) -> None:
    ws = wb.create_sheet("负荷信息对比")
    headers = ["负荷类型", "指标", "选择值", "对比值", "差值", "同比(%)", "选择时间", "对比时间"]
    ws.append(headers)
    style_header(ws[1])
    for subject_payload in load_payload["subjects"]:
        for metric_name, item in subject_payload["comparison"].items():
            decimals = metric_decimals(metric_name)
            ws.append(
                [
                    subject_payload["subject"],
                    metric_name,
                    round(item["selected"], decimals),
                    round(item["compare"], decimals),
                    round(item["difference"], decimals),
                    round(item["yoyPercent"], 4) if item.get("yoyPercent") is not None else None,
                    item.get("selectedTime"),
                    item.get("compareTime"),
                ]
            )
    for column, width in {"A": 18, "B": 18, "C": 16, "D": 16, "E": 16, "F": 14, "G": 18, "H": 18}.items():
        ws.column_dimensions[column].width = width


def write_summary_sheet(wb: openpyxl.Workbook, payload: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "汇总"
    ws["A1"] = "数据源"
    ws["B1"] = payload["source"]
    ws["A2"] = "选择日期"
    ws["B2"] = payload["selectedDate"]
    ws["A3"] = "对比日期"
    ws["B3"] = payload["compareDate"]
    ws["A4"] = "负荷指标数量"
    ws["B4"] = len(payload["loadCompare"]["subjects"])
    ws["A5"] = "滚动预测"
    ws["B5"] = f"{payload['rollingForecast']['orgName']} / {payload['rollingForecast']['unitName']}"
    ws["A6"] = "检修容量/开机台数"
    ws["B6"] = (
        f"{payload['operationPublicData']['capacity']['selected'].get('capacityWanKw')}万千瓦 / "
        f"{payload['operationPublicData']['startupCount']['selected'].get('maxStartupCount')}台"
    )

    ws["A8"] = "价格类型"
    ws["B8"] = "选择日均价(厘/kWh)"
    ws["C8"] = "对比日均价(厘/kWh)"
    ws["D8"] = "差值(厘/kWh)"
    style_header(ws[8])
    row = 9
    for item in payload["priceBoard"]["rows"]:
        ws.cell(row, 1, item["typeName"])
        ws.cell(row, 2, item["selectedDailyAvgPrice"])
        ws.cell(row, 3, item["compareDailyAvgPrice"])
        ws.cell(row, 4, item["difference"])
        row += 1

    row += 1
    market_price_payload = payload["marketAnalysisData"]["priceBoard"]
    ws.cell(row, 1, "日清分市场价格")
    ws.cell(
        row,
        2,
        f"{market_price_payload['selectedDate']} vs {market_price_payload['compareDate']}",
    )
    row += 1
    ws.cell(row, 1, "价格类型")
    ws.cell(row, 2, "选择日均价(厘/kWh)")
    ws.cell(row, 3, "对比日均价(厘/kWh)")
    ws.cell(row, 4, "差值(厘/kWh)")
    style_header(ws[row])
    row += 1
    for item in market_price_payload["rows"]:
        ws.cell(row, 1, item["typeName"])
        ws.cell(row, 2, item["selectedDailyAvgPrice"])
        ws.cell(row, 3, item["compareDailyAvgPrice"])
        ws.cell(row, 4, item["difference"])
        row += 1

    row += 1
    ws.cell(row, 1, "预测日期")
    ws.cell(row, 2, "日合计")
    ws.cell(row, 3, "日平均")
    style_header(ws[row])
    row += 1
    for item in payload["rollingForecast"]["rows"]:
        ws.cell(row, 1, item["forecastDate"])
        ws.cell(row, 2, item["dailyTotal"])
        ws.cell(row, 3, item["dailyAverage"])
        row += 1

    row += 1
    header_row = row
    headers = ["负荷类型", "指标", "选择值", "对比值", "差值", "同比(%)"]
    for col, header in enumerate(headers, start=1):
        ws.cell(header_row, col, header)
    style_header(ws[header_row])
    row = header_row + 1
    for subject_payload in payload["loadCompare"]["subjects"]:
        for metric_name, item in subject_payload["comparison"].items():
            decimals = metric_decimals(metric_name)
            ws.cell(row, 1, subject_payload["subject"])
            ws.cell(row, 2, metric_name)
            ws.cell(row, 3, round(item["selected"], decimals))
            ws.cell(row, 4, round(item["compare"], decimals))
            ws.cell(row, 5, round(item["difference"], decimals))
            ws.cell(row, 6, round(item["yoyPercent"], 4) if item.get("yoyPercent") is not None else None)
            row += 1

    for column, width in {"A": 18, "B": 22, "C": 18, "D": 18, "E": 18, "F": 14}.items():
        ws.column_dimensions[column].width = width


def save_output(payload: dict[str, Any], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    selected_date = payload["selectedDate"]
    compare_date = payload["compareDate"]
    subject_label = f"{len(payload['loadCompare']['subjects'])}项指标"
    path = output_dir / f"每日数据汇总_{sanitize_filename(subject_label)}_{selected_date}_vs_{compare_date}.xlsx"

    wb = openpyxl.Workbook()
    write_summary_sheet(wb, payload)
    write_load_sheet(wb, payload["loadCompare"])
    write_price_sheet(wb, payload["priceBoard"])
    write_price_sheet(
        wb,
        payload["marketAnalysisData"]["priceBoard"],
        title="日清分市场价格",
    )
    write_rolling_forecast_sheet(wb, payload["rollingForecast"])
    raw_ws = wb.create_sheet("原始JSON")
    raw_ws["A1"] = json.dumps(payload, ensure_ascii=False, indent=2)
    raw_ws.column_dimensions["A"].width = 120

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")

    try:
        wb.save(path)
    except PermissionError:
        path = output_dir / f"每日数据汇总_{sanitize_filename(subject_label)}_{selected_date}_vs_{compare_date}_{datetime.now().strftime('%H%M%S')}.xlsx"
        wb.save(path)
    return path


def print_result(payload: dict[str, Any], path: Path) -> None:
    plan = payload.get("datePlan") or {}
    log(f"报告日期: {payload.get('reportDate') or payload['selectedDate']}")
    log(f"负荷/价格选择日期: {payload['selectedDate']}")
    log(f"负荷/价格对比日期: {payload['compareDate']}")
    if plan:
        log(
            "日期口径: "
            f"上网电量={plan['actualOnlineEnergyDate']} "
            f"日前出清={plan['dayAheadDate']} "
            f"日清分={plan['dailyClearingDate']} "
            f"海风预测={plan['rollingForecastStartDate']}~{plan['rollingForecastEndDate']}"
        )
    log(f"负荷指标: {len(payload['loadCompare']['subjects'])} 项")
    log(
        "检修容量/市场开机台数: "
        f"{payload['operationPublicData']['capacity']['selected'].get('capacityWanKw')}万千瓦 / "
        f"{payload['operationPublicData']['startupCount']['selected'].get('maxStartupCount')}台 "
        f"环比{payload['operationPublicData']['startupCount'].get('difference')}台"
    )
    log(
        "多日滚动预测: "
        f"{payload['rollingForecast']['orgName']} / {payload['rollingForecast']['unitName']} "
        f"{payload['rollingForecast']['forecastStartDate']} ~ {payload['rollingForecast']['forecastEndDate']} "
        f"{len(payload['rollingForecast']['rows'])} 行"
    )
    price_groups = [
        ("报告日价格", payload["priceBoard"]),
        ("日清分市场价格", payload["marketAnalysisData"]["priceBoard"]),
    ]
    for group_name, price_payload in price_groups:
        log(
            f"{group_name}: "
            f"{price_payload['selectedDate']} vs {price_payload['compareDate']}"
        )
        for item in price_payload["rows"]:
            selected_value = item["selectedDailyAvgPrice"] if item["selectedDailyAvgPrice"] is not None else "无数据"
            compare_value = item["compareDailyAvgPrice"] if item["compareDailyAvgPrice"] is not None else "无数据"
            diff = item["difference"] if item["difference"] is not None else "无数据"
            log(
                f"  {item['typeName']}日均价: "
                f"{item['selectedDate']}={selected_value} "
                f"{item['compareDate']}={compare_value} "
                f"差值={diff} 厘/kWh"
            )
    log(f"\n结果已保存: {path}")


def main() -> int:
    parser = argparse.ArgumentParser(description="统一抓取负荷信息对比和价格看板日均价")
    parser.add_argument("selected_date", nargs="?", help="报告日期/选择日期，格式 YYYY-MM-DD")
    parser.add_argument("compare_date", nargs="?", help="可选：对比日期；不填时自动取报告日期前一日")
    parser.add_argument("--subjects", default="all", help="负荷指标，逗号分隔；默认 all 表示映射表全部指标")
    parser.add_argument("--data-type", type=int, default=DEFAULT_DATA_TYPE, help="负荷接口数据类型，默认 1")
    parser.add_argument("--point", type=int, choices=[24, 96], default=DEFAULT_POINT, help="负荷点位数，默认 96")
    parser.add_argument("--province-area-id", default=DEFAULT_PROVINCE_AREA_ID, help="省区编码，默认 044")
    parser.add_argument("--forecast-org-id", default=DEFAULT_FORECAST_ORG_ID, help="多日滚动预测公司 orgId")
    parser.add_argument("--forecast-org-name", default=DEFAULT_FORECAST_ORG_NAME, help="多日滚动预测公司名称")
    parser.add_argument("--forecast-unit-id", default=DEFAULT_FORECAST_UNIT_ID, help="多日滚动预测场站 unitId")
    parser.add_argument("--forecast-unit-name", default=DEFAULT_FORECAST_UNIT_NAME, help="多日滚动预测场站名称")
    parser.add_argument("--output-dir", default=str(OUTPUT_DIR), help="输出目录")
    parser.add_argument("--login", action="store_true", help="仅打开浏览器登录并保存会话")
    args = parser.parse_args()

    try:
        config = load_config()
        if args.login:
            with sync_playwright() as playwright:
                context = launch_context(playwright, headless=False)
                try:
                    interactive_login(context, config)
                finally:
                    context.close()
            return 0

        if not args.selected_date:
            raise ValueError("请提供报告日期，例如：python fetch_daily_data.py 2026-05-13")

        selected_date = validate_date(args.selected_date)
        compare_date = validate_date(args.compare_date) if args.compare_date else build_report_date_plan(selected_date)["compareDate"]
        subjects = parse_subjects(args.subjects)
        payload = run_daily_data(
            selected_date=selected_date,
            compare_date=compare_date,
            subjects=subjects,
            data_type=args.data_type,
            point=args.point,
            province_area_id=args.province_area_id,
            forecast_org_id=args.forecast_org_id,
            forecast_org_name=args.forecast_org_name,
            forecast_unit_id=args.forecast_unit_id,
            forecast_unit_name=args.forecast_unit_name,
        )
        path = save_output(payload, Path(args.output_dir).expanduser())
        print_result(payload, path)
        return 0
    except KeyboardInterrupt:
        print("\n已取消。")
        return 130
    except Exception as exc:
        print(f"\n失败：{exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
