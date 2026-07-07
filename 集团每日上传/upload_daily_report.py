#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import json
import os
import queue
import re
import subprocess
import sys
import threading
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Callable

from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import Page
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


SCRIPT_DIR = Path(__file__).resolve().parent
WORKSPACE_DIR = SCRIPT_DIR.parent
LOGIN_SOURCE_DIR = WORKSPACE_DIR / "上网电量抓取"

BASE_URL = "https://xhxt.chng.com.cn"
HUANENG_GROUP_URL = f"{BASE_URL}/huaneng/group"
REPORT_URL = f"{BASE_URL}/huaneng/dataManagement/Report"
DAY_REPORT_URL = f"{BASE_URL}/huaneng/report/dayReport"
LOGIN_URL = f"{BASE_URL}/usercenter/#/login"
USERCENTER_URL = f"{BASE_URL}/usercenter/#/"
DEFAULT_PROVINCE_ID = "44"
DEFAULT_PROVINCE_NAME = "广东省"
DEFAULT_PUBLIC_PROVINCE_AREA_ID = "044"
DEFAULT_DAILY_REPORT_TEMPLATE = "广东模板（修改）"
DAY_REPORT_TENANT_NAMES = ("广东分公司", "中国华能集团有限公司广东分公司", "华能广东分公司")

PROFILE_DIR = SCRIPT_DIR / ".browser-profile"
AUTH_STATE_PATH = SCRIPT_DIR / "auth_state.json"
AUTH_LOCK_PATH = Path(str(AUTH_STATE_PATH) + ".lock")
CONFIG_PATH = SCRIPT_DIR / "config.json"
FALLBACK_CONFIG_PATH = LOGIN_SOURCE_DIR / "config.json"

UPLOAD_TYPES = {
    "省内": {
        "name": "省内数据",
        "filename_keywords": ("省内",),
        "page_keywords": ("省内",),
        "data_type": "DAY_MARKET_INFO",
    },
    "能销": {
        "name": "能销数据",
        "filename_keywords": ("能销",),
        "page_keywords": ("能销",),
        "data_type": "DAY_ENERGY_SCALE",
    },
}


@dataclass
class Config:
    username: str = ""
    password: str = ""
    headless: bool = False


@dataclass
class UploadFile:
    path: Path
    kind: str
    date: str
    auto_template: bool = False


LOG_SINK: Callable[[str], None] | None = None


def log(message: str) -> None:
    print(message, flush=True)
    if LOG_SINK:
        LOG_SINK(message)


def short_error(exc: BaseException) -> str:
    return str(exc).split("Call log:")[0].strip()


def process_is_alive(pid: int) -> bool:
    if pid <= 0:
        return False
    if os.name != "nt":
        try:
            os.kill(pid, 0)
            return True
        except ProcessLookupError:
            return False
        except PermissionError:
            return True
    try:
        import ctypes

        handle = ctypes.windll.kernel32.OpenProcess(0x1000, False, pid)
        if not handle:
            return False
        ctypes.windll.kernel32.CloseHandle(handle)
        return True
    except Exception:
        return False


def parse_lock_timestamp(value: Any) -> float | None:
    if not isinstance(value, str):
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).timestamp()
    except ValueError:
        return None


class AuthStateLock:
    def __init__(self, timeout_seconds: int = 15 * 60) -> None:
        self.timeout_seconds = timeout_seconds
        self.path = AUTH_LOCK_PATH
        self.acquired = False

    def __enter__(self) -> "AuthStateLock":
        started = time.time()
        notified = False
        while True:
            try:
                fd = os.open(str(self.path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            except FileExistsError:
                owner: dict[str, Any] = {}
                try:
                    owner = json.loads(self.path.read_text(encoding="utf-8"))
                except Exception:
                    pass
                created_at = parse_lock_timestamp(owner.get("startedAt"))
                lock_age = time.time() - created_at if created_at else 0
                owner_pid = int(owner.get("pid") or 0)
                if lock_age > 6 * 60 * 60 or (owner_pid and not process_is_alive(owner_pid)):
                    try:
                        self.path.unlink()
                    except FileNotFoundError:
                        pass
                    continue
                if time.time() - started > self.timeout_seconds:
                    raise RuntimeError(f"等待登录态锁超时，请确认没有其他上传/抓取脚本仍在运行：{self.path}")
                if not notified:
                    log(f"检测到其他脚本正在使用同一登录态，等待其完成：{self.path}")
                    notified = True
                time.sleep(2)
                continue

            with os.fdopen(fd, "w", encoding="utf-8") as lock_file:
                json.dump(
                    {
                        "pid": os.getpid(),
                        "startedAt": datetime.now().isoformat(),
                        "script": "upload_daily_report",
                    },
                    lock_file,
                    ensure_ascii=False,
                    indent=2,
                )
            self.acquired = True
            log(f"已获得登录态锁：{self.path}")
            return self

    def __exit__(self, exc_type, exc, tb) -> None:
        if not self.acquired:
            return
        try:
            self.path.unlink()
        except FileNotFoundError:
            pass
        except PermissionError as cleanup_error:
            if exc_type is None:
                raise
            log(f"清理登录态锁失败，将保留原始错误：{cleanup_error}")
        self.acquired = False


def load_config() -> Config:
    path = CONFIG_PATH if CONFIG_PATH.exists() else FALLBACK_CONFIG_PATH
    if not path.exists():
        return Config()
    raw = json.loads(path.read_text(encoding="utf-8"))
    return Config(
        username=str(raw.get("username", "") or ""),
        password=str(raw.get("password", "") or ""),
        headless=False,
    )


def validate_date(value: str) -> str:
    value = value.strip()
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        raise ValueError("日期格式应为 YYYY-MM-DD")
    datetime.strptime(value, "%Y-%m-%d")
    return value


def daily_report_name(run_date: str) -> str:
    validated = validate_date(run_date)
    compact_date = validated.replace("-", "")
    return f"广东电力现货市场监测评估日报（{compact_date}）"


def extract_date_from_name(path: Path) -> str:
    name = path.stem
    compact = re.search(r"(20\d{2})([01]\d)([0-3]\d)", name)
    if compact:
        y, m, d = compact.groups()
        return validate_date(f"{y}-{m}-{d}")

    separated = re.search(r"(20\d{2})[-_.年 ]([01]?\d)[-_.月 ]([0-3]?\d)", name)
    if separated:
        y, m, d = separated.groups()
        return validate_date(f"{int(y):04d}-{int(m):02d}-{int(d):02d}")

    raise ValueError(f"文件名里没有识别到日期：{path.name}")


def classify_file(path: Path) -> str:
    name = path.name
    matches = [
        kind
        for kind, spec in UPLOAD_TYPES.items()
        if any(keyword in name for keyword in spec["filename_keywords"])
    ]
    if len(matches) == 1:
        return matches[0]
    if not matches:
        raise ValueError(f"无法从文件名判断上传类型，请包含“省内”或“能销”：{path.name}")
    raise ValueError(f"文件名同时匹配多个上传类型：{path.name}")


def choose_files_interactively() -> list[Path]:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception as exc:
        raise RuntimeError("无法打开文件选择窗口，请改用命令行传入两个文件路径。") from exc

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    selected = filedialog.askopenfilenames(
        title="请选择能销数据 Excel；也可同时选择省内数据",
        initialdir=str(SCRIPT_DIR),
        filetypes=[
            ("Excel 文件", "*.xlsx *.xls"),
            ("所有文件", "*.*"),
        ],
    )
    root.destroy()
    return [Path(item) for item in selected]


def probe_standard_province_templates(folder: Path, target_date: str, *, days: int = 45) -> list[Path]:
    center = datetime.strptime(target_date, "%Y-%m-%d").date()
    candidates: list[Path] = []
    for offset in range(days + 1):
        offsets = [0] if offset == 0 else [-offset, offset]
        for day_offset in offsets:
            day = center + timedelta(days=day_offset)
            path = folder / f"广东-省内数据-{day:%Y%m%d}.xlsx"
            if path.exists() and not path.name.startswith("~$") and "自动更新" not in path.stem:
                candidates.append(path)
    return candidates


def find_province_template(folder: Path, target_date: str) -> UploadFile:
    candidates: list[UploadFile] = []
    visible_excel_files: list[str] = []
    paths = list(folder.glob("*.xlsx"))
    if not paths:
        paths = probe_standard_province_templates(folder, target_date)
    for path in paths:
        if path.name.startswith("~$") or "自动更新" in path.stem:
            continue
        visible_excel_files.append(path.name)
        if "省内" not in path.name:
            continue
        try:
            candidates.append(
                UploadFile(
                    path=path.resolve(),
                    kind="省内",
                    date=extract_date_from_name(path),
                    auto_template=True,
                )
            )
        except ValueError:
            continue
    if not candidates:
        visible = "、".join(sorted(visible_excel_files)) or "无"
        raise ValueError(
            f"没有在 {folder} 找到可作为模板的省内数据 .xlsx 文件。"
            f"实际看到的 Excel 文件：{visible}"
        )
    target_day = datetime.strptime(target_date, "%Y-%m-%d").date()
    selected = min(
        candidates,
        key=lambda item: (
            abs((datetime.strptime(item.date, "%Y-%m-%d").date() - target_day).days),
            datetime.strptime(item.date, "%Y-%m-%d").date() > target_day,
            -item.path.stat().st_mtime,
        ),
    )
    log(f"自动选择省内模板：{selected.path.name}")
    return selected


def prepare_upload_files(paths: list[Path]) -> tuple[str, list[UploadFile]]:
    if len(paths) not in (1, 2):
        raise ValueError("请选择一个能销数据文件，或同时选择一个省内数据和一个能销数据文件。")

    uploads: list[UploadFile] = []
    seen: set[str] = set()
    for raw_path in paths:
        path = raw_path.expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(f"文件不存在：{path}")
        kind = classify_file(path)
        if kind in seen:
            raise ValueError(f"同一类型文件只能选一个，重复类型：{UPLOAD_TYPES[kind]['name']}")
        seen.add(kind)
        uploads.append(UploadFile(path=path, kind=kind, date=extract_date_from_name(path)))

    if len(uploads) == 1:
        only_upload = uploads[0]
        if only_upload.kind != "能销":
            raise ValueError("只选择一个文件时必须选择能销数据文件，省内数据会自动从同目录模板生成。")
        uploads.append(find_province_template(only_upload.path.parent, only_upload.date))
        seen.add("省内")

    missing = set(UPLOAD_TYPES) - seen
    if missing:
        names = "、".join(UPLOAD_TYPES[kind]["name"] for kind in sorted(missing))
        raise ValueError(f"缺少文件：{names}")

    ordered = sorted(uploads, key=lambda item: 0 if item.kind == "省内" else 1)
    energy_upload = next(item for item in ordered if item.kind == "能销")
    return energy_upload.date, ordered


def parse_json_response(response, label: str) -> dict[str, Any]:
    text = response.text()
    if response.status >= 400:
        raise RuntimeError(f"{label} HTTP {response.status}: {text[:300]}")
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"{label} 返回的不是 JSON，可能登录已失效。片段：{text[:200]}") from exc
    if data.get("retCode") and data.get("retCode") != "T200":
        raise RuntimeError(f"{label} 失败：{data.get('retMsg') or data}")
    return data


def json_or_text(response) -> Any:
    text = response.text()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return text


def load_saved_auth_state(context) -> None:
    if not AUTH_STATE_PATH.exists():
        return
    try:
        state = json.loads(AUTH_STATE_PATH.read_text(encoding="utf-8"))
        cookies = state.get("cookies") or []
        if cookies:
            context.add_cookies(cookies)
    except Exception as exc:
        log(f"读取 auth_state.json 失败，将忽略：{exc}")


def save_auth_state(context) -> None:
    AUTH_STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    AUTH_STATE_PATH.write_text(
        json.dumps(context.storage_state(), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def playwright_proxy() -> dict[str, str] | None:
    for name in ("HTTPS_PROXY", "HTTP_PROXY", "ALL_PROXY", "https_proxy", "http_proxy", "all_proxy"):
        value = os.environ.get(name)
        if value:
            return {"server": value}
    return None


def launch_context(playwright, *, headless: bool):
    PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    launch_options = {
        "headless": headless,
        "accept_downloads": True,
    }
    proxy = playwright_proxy()
    if proxy is not None:
        launch_options["proxy"] = proxy
    try:
        context = playwright.chromium.launch_persistent_context(
            str(PROFILE_DIR),
            channel="chrome",
            **launch_options,
        )
    except PlaywrightError:
        context = playwright.chromium.launch_persistent_context(
            str(PROFILE_DIR),
            **launch_options,
        )
    load_saved_auth_state(context)
    return context


def is_logged_in(context) -> bool:
    urls = [
        f"{BASE_URL}/gdfire/api/pf/tenant/user/info",
        f"{BASE_URL}/usercenter/web/pf/tenant/user/info",
    ]
    for url in urls:
        try:
            response = context.request.get(url, timeout=30000)
        except PlaywrightError:
            continue
        if response.status != 200:
            continue
        try:
            data = response.json()
        except Exception:
            continue
        if data.get("retCode") == "T200" and bool(data.get("data", {}).get("tenantId")):
            return True
    return False


def rsa_encrypt_password(public_key_b64: str, password: str) -> str:
    from cryptography.hazmat.primitives import serialization
    from cryptography.hazmat.primitives.asymmetric import padding

    public_key = serialization.load_der_public_key(base64.b64decode(public_key_b64))
    encrypted = public_key.encrypt(password.encode("utf-8"), padding.PKCS1v15())
    return base64.b64encode(encrypted).decode("ascii")


def try_api_login(context, username: str, password: str) -> bool:
    key_response = context.request.get(f"{BASE_URL}/usercenter/web/pf/login/info/publicKey", timeout=60000)
    key_data = parse_json_response(key_response, "读取登录公钥")
    public_key = key_data.get("data")
    if not isinstance(public_key, str) or not public_key:
        raise RuntimeError("读取登录公钥失败：返回中没有公钥")

    encrypted_password = rsa_encrypt_password(public_key, password)
    login_response = context.request.post(
        f"{BASE_URL}/usercenter/web/login",
        params={"loginMode": 2, "username": username, "password": encrypted_password},
        timeout=60000,
    )
    payload = json_or_text(login_response)
    if login_response.status >= 400:
        raise RuntimeError(f"登录接口 HTTP {login_response.status}: {str(payload)[:300]}")
    if isinstance(payload, dict) and payload.get("retCode") not in (None, "T200"):
        raise RuntimeError(f"登录失败：{payload.get('retMsg') or payload}")

    page = context.new_page()
    try:
        page.goto(USERCENTER_URL, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_load_state("networkidle", timeout=15000)
    except PlaywrightTimeoutError:
        pass
    finally:
        page.close()

    logged_in = is_logged_in(context)
    if logged_in:
        save_auth_state(context)
    return logged_in


def try_ui_login(context, username: str, password: str) -> bool:
    page = context.new_page()
    page.goto(LOGIN_URL, wait_until="domcontentloaded")
    try:
        page.locator("#username, input[placeholder*='工号'], input[placeholder*='手机号'], input[placeholder*='账号']").first.fill(
            username, timeout=10000
        )
        page.locator("#password, input[type='password']").first.fill(password, timeout=10000)
        page.locator("button:has-text('登 录'), button:has-text('登录'), button[type='submit']").first.click(timeout=10000)
    except (PlaywrightTimeoutError, PlaywrightError):
        pass

    deadline = time.time() + 45
    while time.time() < deadline:
        if is_logged_in(context):
            save_auth_state(context)
            page.close()
            return True
        time.sleep(2)
    page.close()
    return False


def try_auto_login(context, username: str, password: str) -> bool:
    try:
        log("正在尝试接口登录...")
        if try_api_login(context, username, password):
            return True
        log("接口登录后仍未检测到登录态，改用页面登录...")
    except Exception as exc:
        log(f"接口登录失败：{short_error(exc)}")
        log("改用页面登录...")
    return try_ui_login(context, username, password)


def ensure_login(context, config: Config) -> None:
    if is_logged_in(context):
        return
    if not config.username or not config.password:
        raise RuntimeError("未检测到有效登录态，且 config.json 没有填写账号密码。")
    log("未检测到有效登录态，正在使用 config.json 的账号密码自动登录...")
    if not try_auto_login(context, config.username, config.password):
        raise RuntimeError("自动登录失败，请运行“登录日报系统.bat”手动保存一次登录态。")
    log("自动登录成功。")


def interactive_login(context, config: Config) -> None:
    if not is_logged_in(context):
        if config.username and config.password and try_auto_login(context, config.username, config.password):
            log("自动登录成功。")
        else:
            page = context.new_page()
            page.goto(LOGIN_URL, wait_until="domcontentloaded")
            log("请在打开的浏览器里完成登录，然后回到这里按 Enter。")
            input()
            if not is_logged_in(context):
                raise RuntimeError("仍未检测到有效登录态，请确认浏览器里已经登录成功。")
            page.close()

    page = context.new_page()
    page.goto(REPORT_URL, wait_until="domcontentloaded")
    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except PlaywrightTimeoutError:
        pass
    if page_needs_manual_navigation(page):
        log("已登录，但当前会话还没有进入“日报数据管理”页面。")
        log("请在打开的浏览器里从首页进入“日报数据管理”，确认页面正常显示后回到这里按 Enter。")
        input()
        page = find_report_page(context, page)
        if page_needs_manual_navigation(page):
            raise RuntimeError("仍未进入可访问的“日报数据管理”页面，登录态未保存。")

    save_auth_state(context)
    log("日报系统登录态已保存。")


def body_text(page: Page) -> str:
    try:
        return page.locator("body").inner_text(timeout=5000)
    except PlaywrightError:
        return ""


def page_needs_manual_navigation(page: Page) -> bool:
    text = body_text(page)
    url = page.url
    title = page.title()
    return (
        "/Exception/403" in url
        or "403" == title
        or "没有访问权限" in text
        or "去登录" in text
        or "/login" in url
    )


def find_report_page(context, fallback: Page) -> Page:
    for page in context.pages:
        if "/huaneng/dataManagement/Report" in page.url and not page_needs_manual_navigation(page):
            return page
    return fallback


def select_day_report_tenant(tenants: list[dict[str, Any]]) -> dict[str, Any] | None:
    for target_name in DAY_REPORT_TENANT_NAMES:
        for tenant in tenants:
            if str(tenant.get("name") or "").strip() == target_name and tenant.get("tenantId"):
                return tenant
    for tenant in tenants:
        name = str(tenant.get("name") or "").strip()
        if "广东分公司" in name and tenant.get("tenantId"):
            return tenant
    return None


def load_day_report_application_tenants(context) -> list[dict[str, Any]]:
    response = context.request.get(
        f"{BASE_URL}/usercenter/web/pf/tenant/user/application",
        timeout=20000,
    )
    data = parse_json_response(response, "读取应用大厅授权")
    tenants = data.get("data") or []
    if not isinstance(tenants, list):
        raise RuntimeError("读取应用大厅授权返回格式不正确。")
    return tenants


def switch_to_day_report_tenant(context) -> None:
    tenant = select_day_report_tenant(load_day_report_application_tenants(context))
    if not tenant:
        response = context.request.get(
            f"{BASE_URL}/gdfire/api/pf/tenant/user/tenant/grantApplication",
            timeout=20000,
        )
        data = parse_json_response(response, "读取授权单位")
        tenants = data.get("data") or []
        if not isinstance(tenants, list):
            raise RuntimeError("读取授权单位返回格式不正确。")
        tenant = select_day_report_tenant(tenants)
    if not tenant:
        log("应用大厅授权和授权单位里都没有找到“广东分公司”，跳过租户切换。")
        return
    tenant_id = str(tenant["tenantId"])
    tenant_name = str(tenant.get("name") or tenant_id)
    log(f"正在切换到日报管理单位：{tenant_name}")
    switch_response = context.request.get(
        f"{BASE_URL}/usercenter/web/switchTenant",
        params={"tenantId": tenant_id},
        timeout=20000,
    )
    if switch_response.status >= 400:
        raise RuntimeError(f"切换日报管理单位失败 HTTP {switch_response.status}")


def open_report_page(context, *, allow_manual: bool) -> Page:
    page = context.new_page()
    log(f"正在打开日报数据管理页面：{REPORT_URL}")
    page.goto(REPORT_URL, wait_until="domcontentloaded", timeout=60000)
    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except PlaywrightTimeoutError:
        pass

    if page_needs_manual_navigation(page):
        if not allow_manual:
            raise RuntimeError("未能直接进入“日报数据管理”页面。请先运行“登录日报系统.bat”，并确认该浏览器能访问日报数据管理。")
        log("\n当前页面未直接进入日报数据管理。")
        log("请在打开的浏览器窗口中登录，或从首页进入“日报数据管理”页面；完成后回到这里按 Enter。")
        input()
        page = find_report_page(context, page)
        if page_needs_manual_navigation(page):
            raise RuntimeError("仍未进入可访问的“日报数据管理”页面，无法继续上传。")

    return page


def enter_huaneng_marketing_platform(page: Page) -> None:
    log(f"正在打开应用大厅：{USERCENTER_URL}")
    page.goto(USERCENTER_URL, wait_until="domcontentloaded", timeout=60000)
    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except PlaywrightTimeoutError:
        pass

    clicked = page.evaluate(
        """
        () => {
          const visible = el => !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
          const compact = s => String(s || '').replace(/\\s+/g, '').trim();
          const targetText = '华能集团电力营销管控平台';
          const nodes = [...document.querySelectorAll('a,button,[role="button"],div,span,li')].filter(visible);
          const target = nodes.find(el => compact(el.innerText || el.textContent || '').includes(targetText));
          if (!target) return false;
          let clickable = target;
          for (let node = target; node && node !== document.body; node = node.parentElement) {
            const style = window.getComputedStyle(node);
            if (
              node.tagName === 'A' ||
              node.tagName === 'BUTTON' ||
              node.getAttribute('role') === 'button' ||
              typeof node.onclick === 'function' ||
              style.cursor === 'pointer'
            ) {
              clickable = node;
              break;
            }
          }
          clickable.click();
          return true;
        }
        """
    )
    if not clicked:
        raise RuntimeError("应用大厅里没有找到“华能集团电力营销管控平台”入口。")
    try:
        page.wait_for_load_state("networkidle", timeout=20000)
    except PlaywrightTimeoutError:
        pass
    page.wait_for_timeout(1000)


def open_day_report_page(context, *, allow_manual: bool) -> Page:
    page = context.new_page()
    try:
        switch_to_day_report_tenant(context)
    except Exception as exc:
        log(f"切换日报管理单位失败，将继续尝试直接打开页面：{short_error(exc)}")
    log(f"正在打开日报管理页面：{DAY_REPORT_URL}")
    page.goto(DAY_REPORT_URL, wait_until="domcontentloaded", timeout=60000)
    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except PlaywrightTimeoutError:
        pass

    if page_needs_manual_navigation(page):
        log("直接进入日报管理失败，从应用大厅进入华能集团电力营销管控平台后重试。")
        try:
            enter_huaneng_marketing_platform(page)
            log(f"正在重新打开日报管理页面：{DAY_REPORT_URL}")
            page.goto(DAY_REPORT_URL, wait_until="domcontentloaded", timeout=60000)
            page.wait_for_load_state("networkidle", timeout=15000)
        except (PlaywrightTimeoutError, PlaywrightError):
            pass
        except RuntimeError as exc:
            log(str(exc))
    if page_needs_manual_navigation(page):
        if not allow_manual:
            text = body_text(page)
            raise RuntimeError(
                "已进入同站点日报数据管理，但未能切换到“日报管理”页面。"
                f"当前URL：{page.url}；页面标题：{page.title()}；页面文本片段：{text[:300]}"
            )
        log("\n当前页面未直接进入日报管理。")
        log("请在打开的浏览器窗口中登录，或手动进入“日报管理”页面；完成后回到这里按 Enter。")
        input()
        for candidate in context.pages:
            if "/huaneng/report/dayReport" in candidate.url and not page_needs_manual_navigation(candidate):
                return candidate
        if page_needs_manual_navigation(page):
            raise RuntimeError("仍未进入可访问的“日报管理”页面，无法继续新建报表。")

    return page


def click_button_by_text(page: Page, text: str, *, timeout: int = 15000) -> None:
    locators = [
        page.get_by_role("button", name=re.compile(rf"^\s*{re.escape(text)}\s*$")),
        page.locator(f"button:has-text('{text}')"),
        page.locator(f"text={text}"),
    ]
    last_error: BaseException | None = None
    for locator in locators:
        try:
            locator.first.click(timeout=timeout)
            return
        except (PlaywrightTimeoutError, PlaywrightError) as exc:
            last_error = exc
    raise RuntimeError(f"没有找到“{text}”按钮。") from last_error


def set_form_input_by_label(page: Page, label: str, value: str) -> None:
    updated = page.evaluate(
        """
        ({ label, value }) => {
          const visible = el => !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
          const compact = s => String(s || '').replace(/\\s+/g, ' ').trim();
          const normalize = s => compact(s).replace(/模版/g, '模板').replace(/[：:*\\s]/g, '');
          const target = normalize(label);
          const labels = [...document.querySelectorAll('label,.el-form-item__label,.ant-form-item-label,[class*="form-item-label"]')]
            .filter(el => visible(el) && normalize(el.innerText || el.textContent).includes(target))
            .sort((a, b) => {
              const aText = normalize(a.innerText || a.textContent);
              const bText = normalize(b.innerText || b.textContent);
              return (aText === target ? 0 : 1) - (bText === target ? 0 : 1) || aText.length - bText.length;
            });
          for (const labelEl of labels) {
            let node = labelEl;
            for (let depth = 0; node && depth < 5; depth++, node = node.parentElement) {
              const input = [...node.querySelectorAll('input,textarea')]
                .find(el => visible(el) && !el.disabled && !el.readOnly);
              if (!input) continue;
              const setter = Object.getOwnPropertyDescriptor(
                input instanceof HTMLTextAreaElement ? HTMLTextAreaElement.prototype : HTMLInputElement.prototype,
                'value'
              ).set;
              setter.call(input, value);
              input.dispatchEvent(new Event('input', { bubbles: true }));
              input.dispatchEvent(new Event('change', { bubbles: true }));
              input.blur();
              return true;
            }
          }
          return false;
        }
        """,
        {"label": label, "value": value},
    )
    if not updated:
        raise RuntimeError(f"没有找到“{label}”输入框。")
    page.wait_for_timeout(300)


def select_form_option_by_label(page: Page, label: str, option_text: str) -> None:
    marked = page.evaluate(
        """
        ({ label }) => {
          const visible = el => !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
          const compact = s => String(s || '').replace(/\\s+/g, ' ').trim();
          const normalize = s => compact(s).replace(/模版/g, '模板').replace(/[：:*\\s]/g, '');
          const target = normalize(label);
          document.querySelectorAll('[data-codex-form-select]').forEach(el => el.removeAttribute('data-codex-form-select'));
          const labels = [...document.querySelectorAll('label,.el-form-item__label,.ant-form-item-label,[class*="form-item-label"]')]
            .filter(el => visible(el) && normalize(el.innerText || el.textContent).includes(target))
            .sort((a, b) => {
              const aText = normalize(a.innerText || a.textContent);
              const bText = normalize(b.innerText || b.textContent);
              return (aText === target ? 0 : 1) - (bText === target ? 0 : 1) || aText.length - bText.length;
            });
          for (const labelEl of labels) {
            let node = labelEl;
            for (let depth = 0; node && depth < 5; depth++, node = node.parentElement) {
              const field = [...node.querySelectorAll('input,[role="combobox"],.el-select,.ant-select,[class*="select"]')]
                .find(el => visible(el) && !el.disabled);
              if (!field) continue;
              const selectRoot = field.closest('.ant-select,.el-select') || field;
              selectRoot.setAttribute('data-codex-form-select', 'true');
              return compact(selectRoot.value || selectRoot.innerText || selectRoot.textContent || field.value || '');
            }
          }
          return null;
        }
        """,
        {"label": label},
    )
    if marked is None:
        raise RuntimeError(f"没有找到“{label}”下拉框。")
    if option_text in str(marked):
        return
    select_locator = page.locator('[data-codex-form-select="true"]').first
    select_locator.click(timeout=10000)
    search_input = page.locator(
        '[data-codex-form-select="true"] input[role="combobox"], '
        '[data-codex-form-select="true"] input, '
        '[data-codex-form-select="true"][role="combobox"], '
        '[data-codex-form-select="true"]'
    ).first
    try:
        search_input.fill(option_text, timeout=5000, force=True)
    except (PlaywrightTimeoutError, PlaywrightError):
        pass
    page.wait_for_timeout(500)
    visible_dropdown_option = page.locator(
        ".ant-select-dropdown:not(.ant-select-dropdown-hidden) .ant-select-item-option"
    ).filter(has_text=option_text)
    try:
        visible_dropdown_option.last.click(timeout=10000)
    except (PlaywrightTimeoutError, PlaywrightError):
        try:
            search_input.press("Enter", timeout=5000)
        except (PlaywrightTimeoutError, PlaywrightError) as exc:
            raise RuntimeError(f"没有找到“{option_text}”选项。") from exc
    page.wait_for_timeout(300)


def find_daily_report_template_id(context, template_name: str = DEFAULT_DAILY_REPORT_TEMPLATE) -> str:
    response = context.request.get(
        f"{BASE_URL}/huaneng/group/api/report/queryTemplates",
        timeout=20000,
    )
    data = parse_json_response(response, "读取日报模板列表")
    templates = data.get("data") or []
    if not isinstance(templates, list):
        raise RuntimeError("读取日报模板列表返回格式不正确。")
    for template in templates:
        if str(template.get("name") or "").strip() == template_name and template.get("id"):
            return str(template["id"])
    available = "、".join(str(item.get("name") or "") for item in templates if "广东" in str(item.get("name") or ""))[:300]
    raise RuntimeError(f"日报模板里没有找到“{template_name}”。可见广东模板：{available or '无'}")


def daily_report_exists(context, report_name: str, report_date: str) -> bool:
    response = context.request.get(
        f"{BASE_URL}/huaneng/group/api/report",
        params={
            "startDate": report_date,
            "endDate": report_date,
            "keyword": report_name,
            "numPerPage": 10,
            "pageNum": 1,
        },
        timeout=20000,
    )
    data = parse_json_response(response, "查询已有集团日报")
    payload = data.get("data") or {}
    rows = payload.get("datas") or []
    if not isinstance(rows, list):
        return False
    return any(str(row.get("reportName") or "") == report_name for row in rows)


def create_daily_report(context, run_date: str, *, headless: bool) -> None:
    report_date = validate_date(run_date)
    report_name = daily_report_name(report_date)
    page = open_day_report_page(context, allow_manual=not headless)
    log(f"正在新建集团日报：{report_name}")
    if daily_report_exists(context, report_name, report_date):
        log(f"集团日报已存在，跳过新建：{report_name}")
        return
    template_id = find_daily_report_template_id(context)
    response = context.request.post(
        f"{BASE_URL}/huaneng/group/api/report",
        data={
            "reportName": report_name,
            "startDate": report_date,
            "endDate": report_date,
            "templateId": template_id,
            "provinceId": None,
        },
        timeout=120000,
    )
    parse_json_response(response, "新建集团日报")
    log("集团日报新建操作已提交。")


def dump_page_diagnostics(page: Page, prefix: str) -> str:
    safe_prefix = re.sub(r"[^A-Za-z0-9_.-]+", "_", prefix).strip("_") or "page"
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_path = SCRIPT_DIR / f"{safe_prefix}_{stamp}"
    data: dict[str, Any] = {"url": page.url, "title": "", "bodyText": ""}
    try:
        data["title"] = page.title()
    except PlaywrightError:
        pass
    try:
        data["bodyText"] = body_text(page)[:3000]
    except PlaywrightError:
        pass
    json_path = base_path.with_suffix(".json")
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    try:
        page.screenshot(path=str(base_path.with_suffix(".png")), full_page=True, timeout=10000)
    except PlaywrightError:
        pass
    return str(json_path)


def smoke_daily_report_page(*, headless: bool = True) -> None:
    config = load_config()
    config.headless = headless
    with AuthStateLock():
        with sync_playwright() as playwright:
            context = launch_context(playwright, headless=config.headless)
            try:
                ensure_login(context, config)
                page = open_day_report_page(context, allow_manual=not headless)
                try:
                    count = page.get_by_text("新建报表", exact=True).count()
                except PlaywrightError:
                    count = 0
                if count <= 0:
                    diagnostics = dump_page_diagnostics(page, "daily_report_smoke_failed")
                    raise RuntimeError(f"已打开页面但没有找到“新建报表”入口。诊断已保存：{diagnostics}")
                save_auth_state(context)
                log("日报管理页面自检通过：已找到“新建报表”。")
            finally:
                context.close()


def current_date_inputs(page: Page) -> list[dict[str, Any]]:
    return page.evaluate(
        """
        () => {
          const visible = el => !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
          const compact = s => String(s || '').replace(/\\s+/g, ' ').trim();
          return [...document.querySelectorAll('input')].map((el, index) => ({
            index,
            type: el.getAttribute('type') || '',
            value: compact(el.value),
            placeholder: compact(el.getAttribute('placeholder')),
            aria: compact(el.getAttribute('aria-label')),
            className: compact(el.className),
            visible: visible(el),
            score: (
              /date|picker|日期|时间/.test(`${el.type} ${el.className} ${el.placeholder} ${el.getAttribute('aria-label') || ''}`) ? 10 : 0
            ) + (/^\\d{4}[-/]\\d{1,2}[-/]\\d{1,2}$/.test(el.value) ? 20 : 0)
          })).filter(item => item.visible).sort((a, b) => b.score - a.score);
        }
        """
    )


def set_report_date(page: Page, target_date: str) -> None:
    inputs = current_date_inputs(page)
    candidates = [
        item
        for item in inputs
        if item["score"] > 0
        or "日期" in item["placeholder"]
        or "时间" in item["placeholder"]
        or "date" in item["type"].lower()
    ]
    if not candidates:
        raise RuntimeError("没有找到网页日期输入框。")

    chosen = candidates[0]
    current = str(chosen.get("value") or "").replace("/", "-")
    if current == target_date:
        log(f"网页日期已是 {target_date}。")
        return

    log(f"正在把网页日期从“{current or '空'}”调整为 {target_date}。")
    locator = page.locator("input").nth(int(chosen["index"]))
    try:
        locator.click(timeout=8000)
        locator.press("Control+A", timeout=5000)
        locator.fill(target_date, timeout=8000)
        locator.press("Enter", timeout=5000)
        page.wait_for_timeout(800)
    except PlaywrightError:
        page.evaluate(
            """
            ({ index, value }) => {
              const input = document.querySelectorAll('input')[index];
              if (!input) throw new Error('date input not found');
              const setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value').set;
              setter.call(input, value);
              input.dispatchEvent(new Event('input', { bubbles: true }));
              input.dispatchEvent(new Event('change', { bubbles: true }));
              input.blur();
            }
            """,
            {"index": int(chosen["index"]), "value": target_date},
        )
        page.wait_for_timeout(800)


def mark_upload_target(page: Page, keywords: tuple[str, ...]) -> str | None:
    return page.evaluate(
        """
        ({ keywords }) => {
          const visible = el => !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
          const textOf = el => String(el?.innerText || el?.textContent || '').replace(/\\s+/g, ' ').trim();
          const hasKeyword = text => keywords.some(keyword => text.includes(keyword));
          const uploadWords = ['选择文件', '上传', '导入', '请选择文件'];
          const hasUploadWord = text => uploadWords.some(word => text.includes(word));

          document.querySelectorAll('[data-codex-upload-target]').forEach(el => el.removeAttribute('data-codex-upload-target'));

          const fileInputs = [...document.querySelectorAll('input[type="file"]')];
          let bestInput = null;
          let bestInputScore = -1;
          for (const input of fileInputs) {
            let score = 0;
            let node = input;
            for (let depth = 0; node && depth < 8; depth++, node = node.parentElement) {
              const text = textOf(node);
              if (hasKeyword(text)) score += 100 - depth * 8;
              if (hasUploadWord(text)) score += 20 - depth;
            }
            if (score > bestInputScore) {
              bestInput = input;
              bestInputScore = score;
            }
          }
          if (bestInput && bestInputScore >= 80) {
            bestInput.setAttribute('data-codex-upload-target', 'file-input');
            return 'file-input';
          }

          const clickables = [...document.querySelectorAll('button,[role="button"],.ant-upload,.el-upload,label')].filter(visible);
          let best = null;
          let bestScore = -1;
          for (const el of clickables) {
            const own = textOf(el);
            let score = hasUploadWord(own) ? 40 : 0;
            if (!score && el.tagName !== 'LABEL') continue;
            let node = el;
            for (let depth = 0; node && depth < 8; depth++, node = node.parentElement) {
              const text = textOf(node);
              if (hasKeyword(text)) score += 100 - depth * 8;
              if (hasUploadWord(text)) score += 10 - depth;
            }
            if (score > bestScore) {
              best = el;
              bestScore = score;
            }
          }
          if (best && bestScore >= 80) {
            best.setAttribute('data-codex-upload-target', 'chooser');
            return 'chooser';
          }

          if (fileInputs.length === 2) {
            const index = keywords.some(keyword => keyword.includes('能销')) ? 1 : 0;
            fileInputs[index].setAttribute('data-codex-upload-target', 'file-input');
            return 'file-input';
          }
          return null;
        }
        """,
        {"keywords": list(keywords)},
    )


def dump_upload_diagnostics(page: Page) -> str:
    data = page.evaluate(
        """
        () => {
          const compact = s => String(s || '').replace(/\\s+/g, ' ').trim().slice(0, 180);
          const fileInputs = [...document.querySelectorAll('input[type="file"]')].map((el, index) => {
            let node = el;
            const parts = [];
            for (let depth = 0; node && depth < 5; depth++, node = node.parentElement) {
              const text = compact(node.innerText || node.textContent || '');
              if (text) parts.push(text);
            }
            return { index, accept: el.accept || '', name: el.name || '', text: parts.join(' | ') };
          });
          const buttons = [...document.querySelectorAll('button,[role="button"],label')].map((el, index) => ({
            index,
            text: compact(el.innerText || el.textContent || ''),
            className: compact(el.className || ''),
          })).filter(item => item.text).slice(0, 80);
          return { url: location.href, title: document.title, fileInputs, buttons };
        }
        """
    )
    path = SCRIPT_DIR / f"upload_page_diagnostics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return str(path)


def upload_one_file(page: Page, upload: UploadFile) -> None:
    spec = UPLOAD_TYPES[upload.kind]
    log(f"正在上传 {spec['name']}：{upload.path.name}")
    target_kind = mark_upload_target(page, spec["page_keywords"])
    if not target_kind:
        diagnostics = dump_upload_diagnostics(page)
        raise RuntimeError(f"没有找到“{spec['name']}”对应的选择文件按钮。诊断已保存：{diagnostics}")

    if target_kind == "file-input":
        page.locator('[data-codex-upload-target="file-input"]').set_input_files(str(upload.path), timeout=15000)
    else:
        with page.expect_file_chooser(timeout=15000) as chooser_info:
            page.locator('[data-codex-upload-target="chooser"]').click(timeout=10000)
        chooser_info.value.set_files(str(upload.path))

    try:
        page.wait_for_load_state("networkidle", timeout=20000)
    except PlaywrightTimeoutError:
        pass
    page.wait_for_timeout(2000)
    log(f"{spec['name']} 已选择/上传。")


def parse_api_payload(response, label: str) -> dict[str, Any]:
    text = response.text()
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"{label} 返回的不是 JSON：HTTP {response.status}，{text[:300]}") from exc
    if response.status >= 400:
        raise RuntimeError(f"{label} HTTP {response.status}: {json.dumps(payload, ensure_ascii=False)[:500]}")
    ret_code = payload.get("retCode")
    if ret_code and ret_code != "T200":
        raise RuntimeError(f"{label} 失败：{payload.get('retMsg') or payload}")
    return payload


def number_or_none(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def format_formula_number(value: float) -> str:
    if abs(value - round(value)) < 0.000001:
        return str(int(round(value)))
    return f"{value:.6f}".rstrip("0").rstrip(".")


def fetch_capacity_average(context, run_date: str, province_area_id: str) -> float:
    response = context.request.get(
        f"{BASE_URL}/gdfire/api/data/net/capacity",
        params={
            "startDate": run_date,
            "endDate": run_date,
            "type": "1",
            "provinceAreaId": province_area_id,
        },
        timeout=30000,
    )
    payload = parse_api_payload(response, "读取容量信息")
    rows = payload.get("data", {}).get("dataNetCapacityDTOList") or []
    values = [number_or_none(row.get("capacity")) for row in rows if isinstance(row, dict)]
    values = [value for value in values if value is not None]
    if not values:
        raise RuntimeError(f"容量信息没有返回 {run_date} 的“日前/省内机组预测检修”数据。")
    return sum(values) / len(values)


def fetch_startup_max(context, run_date: str, province_area_id: str) -> int:
    groups_by_type = {
        "日前": ["DAY_AHEAD_CLEARANCE_COAL", "DAY_AHEAD_CLEARANCE_GAS"],
        "实时": ["REAL_TIME_CLEARANCE_COAL", "REAL_TIME_CLEARANCE_GAS"],
    }
    max_values: dict[str, int] = {}
    for label, groups in groups_by_type.items():
        response = context.request.post(
            f"{BASE_URL}/gdfire/api/spot/clear/timeType/data/list",
            data=json.dumps(
                {
                    "provinceAreaId": province_area_id,
                    "spotClearEleGroupList": groups,
                    "startDate": run_date,
                    "endDate": run_date,
                }
            ),
            headers={"Content-Type": "application/json"},
            timeout=30000,
        )
        payload = parse_api_payload(response, f"读取{label}出清电量开机台数")
        rows = payload.get("data") or []
        if not rows:
            raise RuntimeError(f"{label}出清电量没有返回 {run_date} 的燃煤/燃气开机台数。")

        hourly_totals: list[float] = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            unit_count = row.get("unitCount") or []
            for index, raw_value in enumerate(unit_count):
                value = number_or_none(raw_value)
                if value is None:
                    continue
                if index == len(hourly_totals):
                    hourly_totals.append(0.0)
                if index < len(hourly_totals):
                    hourly_totals[index] += value
        if not hourly_totals:
            raise RuntimeError(f"{label}出清电量没有可用的开机台数明细。")
        max_values[label] = int(round(max(hourly_totals)))

    detail = "，".join(f"{label}最大{value}" for label, value in max_values.items())
    final_value = max(max_values.values())
    log(f"火电机组台数：{detail}，取 {final_value}")
    return final_value


def province_output_path(source_path: Path, run_date: str) -> Path:
    compact_date = run_date.replace("-", "")
    name = source_path.name
    new_name, replacements = re.subn(r"(20\d{2})([01]\d)([0-3]\d)", compact_date, name, count=1)
    if replacements == 0:
        new_name, replacements = re.subn(
            r"(20\d{2})[-_.年 ]([01]?\d)[-_.月 ]([0-3]?\d)",
            compact_date,
            name,
            count=1,
        )
    if replacements == 0:
        new_name = f"{source_path.stem}-{compact_date}{source_path.suffix}"

    target_path = source_path.with_name(new_name)
    lock_path = target_path.with_name(f"~${target_path.name}")
    if lock_path.exists():
        raise RuntimeError(f"省内数据目标文件正在被 Excel 打开，请先关闭：{target_path.name}")

    if target_path.exists() and target_path.resolve() != source_path.resolve():
        log(f"省内数据目标文件已存在，将覆盖：{target_path.name}")
    return target_path


def replace_capacity_in_formula(formula: str, capacity_value: float) -> str:
    match = re.match(r"^(=.*?-\s*)(\d+(?:\.\d+)?)(\s*-.*)$", formula)
    if not match:
        raise RuntimeError(f"C2 公式格式无法自动替换容量值：{formula}")
    return f"{match.group(1)}{format_formula_number(capacity_value)}{match.group(3)}"


def evaluate_simple_excel_formula(formula: str) -> float:
    expression = formula.strip()
    if expression.startswith("="):
        expression = expression[1:]
    if not re.fullmatch(r"\s*\d+(?:\.\d+)?(?:\s*[+-]\s*\d+(?:\.\d+)?)*\s*", expression):
        raise RuntimeError(f"C2 公式超出自动计算范围：{formula}")

    total = 0.0
    for sign, number_text in re.findall(r"([+-]?)\s*(\d+(?:\.\d+)?)", expression):
        value = float(number_text)
        total += -value if sign == "-" else value
    return total


def recalculate_with_excel(workbook_path: Path) -> bool:
    ps_script = f"""
$path = ConvertFrom-Json @'
{json.dumps(str(workbook_path), ensure_ascii=False)}
'@
$excel = $null
$workbook = $null
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open($path)
    $excel.CalculateFullRebuild()
    $workbook.Save()
    exit 0
}} catch {{
    Write-Error $_.Exception.Message
    exit 1
}} finally {{
    if ($workbook -ne $null) {{
        $workbook.Close($false) | Out-Null
        [System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) | Out-Null
    }}
    if ($excel -ne $null) {{
        $excel.Quit() | Out-Null
        [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
    }}
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}}
"""
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
            capture_output=True,
            text=True,
            timeout=90,
        )
    except Exception as exc:
        log(f"Excel 后台重算失败，将改写 C2 为数值：{exc}")
        return False

    if result.returncode != 0:
        detail = (result.stderr or result.stdout or "").strip()
        log(f"Excel 后台重算失败，将改写 C2 为数值：{detail[:300]}")
        return False
    return True


def cached_cell_value(workbook_path: Path, sheet_name: str, cell: str) -> Any:
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        return worksheet[cell].value
    finally:
        workbook.close()


def write_formula_result_as_value(workbook_path: Path, sheet_name: str, cell: str, value: float) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        worksheet[cell].value = int(round(value)) if abs(value - round(value)) < 0.000001 else value
        workbook.save(workbook_path)
    finally:
        workbook.close()


def update_province_excel_for_upload(
    context,
    run_date: str,
    uploads: list[UploadFile],
    province_area_id: str = DEFAULT_PUBLIC_PROVINCE_AREA_ID,
) -> list[UploadFile]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("当前 Python 环境缺少 openpyxl，无法自动更新省内数据 Excel。") from exc

    province_upload = next(item for item in uploads if item.kind == "省内")
    energy_upload = next(item for item in uploads if item.kind == "能销")
    if province_upload.path.suffix.lower() != ".xlsx":
        raise RuntimeError("省内数据自动更新只支持 .xlsx 文件。")

    log("开始自动更新省内数据 Excel...")
    if province_upload.date != run_date:
        log(f"省内文件日期 {province_upload.date} 与能销日期 {run_date} 不一致，将按能销日期另存。")

    startup_max = fetch_startup_max(context, run_date, province_area_id)
    capacity_average = fetch_capacity_average(context, run_date, province_area_id)
    log(f"市场化运行容量公式替换值：{format_formula_number(capacity_average)}")

    target_path = province_output_path(province_upload.path, run_date)
    workbook = load_workbook(province_upload.path, data_only=False)
    try:
        worksheet = workbook["省内数据"] if "省内数据" in workbook.sheetnames else workbook.active
        worksheet["A2"].value = datetime.strptime(run_date, "%Y-%m-%d")
        worksheet["B2"].value = startup_max
        current_formula = worksheet["C2"].value
        if not isinstance(current_formula, str) or not current_formula.startswith("="):
            formula_result = number_or_none(current_formula)
            if formula_result is None:
                raise RuntimeError(f"C2 不是公式，无法按模板更新：{current_formula!r}")
            log(f"C2 已是数值，保留现有结果：{current_formula}")
        else:
            updated_formula = replace_capacity_in_formula(current_formula, capacity_average)
            formula_result = evaluate_simple_excel_formula(updated_formula)
            worksheet["C2"].value = updated_formula
            if hasattr(workbook, "calculation"):
                workbook.calculation.fullCalcOnLoad = True
                workbook.calculation.forceFullCalc = True
        workbook.save(target_path)
    finally:
        workbook.close()

    if isinstance(current_formula, str) and current_formula.startswith("="):
        if recalculate_with_excel(target_path):
            cached_value = cached_cell_value(target_path, worksheet.title, "C2")
            if cached_value is None:
                log("Excel 已重算但 C2 缓存仍为空，将改写 C2 为数值。")
                write_formula_result_as_value(target_path, worksheet.title, "C2", formula_result)
        else:
            write_formula_result_as_value(target_path, worksheet.title, "C2", formula_result)

    log(f"省内数据已更新：{target_path}")
    return [
        UploadFile(path=target_path.resolve(), kind="省内", date=run_date),
        UploadFile(path=energy_upload.path, kind="能销", date=run_date),
    ]


def get_daily_status(context, run_date: str, data_type: str, province_id: str) -> dict[str, Any] | None:
    response = context.request.get(
        f"{HUANENG_GROUP_URL}/api/group/private/data/page",
        params={"date": run_date, "groupDataType": data_type, "provinceId": province_id},
        timeout=30000,
    )
    payload = parse_api_payload(response, "查询日报上传状态")
    rows = payload.get("data") or []
    if not isinstance(rows, list):
        raise RuntimeError("查询日报上传状态返回格式不正确。")
    return rows[0] if rows else None


def file_payload(path: Path) -> dict[str, Any]:
    return {
        "name": path.name,
        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "buffer": path.read_bytes(),
    }


def import_daily_file(context, upload: UploadFile, run_date: str, province_id: str) -> dict[str, Any]:
    spec = UPLOAD_TYPES[upload.kind]
    data_type = str(spec["data_type"])
    response = context.request.post(
        f"{HUANENG_GROUP_URL}/api/group/data/import/create/multi",
        multipart={
            "files": file_payload(upload.path),
            "date": run_date,
            "dataType": data_type,
            "type": data_type,
            "provinceId": province_id,
        },
        timeout=120000,
    )
    return parse_api_payload(response, f"导入{spec['name']}")


def run_api_upload(
    paths: list[Path],
    *,
    force: bool = False,
    province_id: str = DEFAULT_PROVINCE_ID,
    create_report: bool = True,
) -> None:
    target_date, uploads = prepare_upload_files(paths)
    log(f"本次上传日期：{target_date}")
    log(f"目标省份：{DEFAULT_PROVINCE_NAME}（provinceId={province_id}）")

    config = load_config()
    config.headless = True

    with AuthStateLock():
        with sync_playwright() as playwright:
            context = launch_context(playwright, headless=config.headless)
            try:
                login_refreshed = False
                try:
                    uploads = update_province_excel_for_upload(context, target_date, uploads)
                except Exception:
                    if login_refreshed:
                        raise
                    log("读取网页接口失败，正在刷新登录态后重试...")
                    ensure_login(context, config)
                    login_refreshed = True
                    uploads = update_province_excel_for_upload(context, target_date, uploads)

                for item in uploads:
                    log(f"- {UPLOAD_TYPES[item.kind]['name']}：{item.path}")

                for upload in uploads:
                    spec = UPLOAD_TYPES[upload.kind]
                    data_type = str(spec["data_type"])
                    try:
                        status = get_daily_status(context, target_date, data_type, province_id)
                    except Exception:
                        if login_refreshed:
                            raise
                        log("日报接口暂不可用，正在刷新登录态后重试...")
                        ensure_login(context, config)
                        login_refreshed = True
                        status = get_daily_status(context, target_date, data_type, province_id)
                    uploaded = bool(status and status.get("dataStatus") == 1)
                    if uploaded and not force:
                        update_time = status.get("updateTime") or "未知时间"
                        log(f"{spec['name']} 已上传，跳过。更新时间：{update_time}")
                        continue

                    if uploaded and force:
                        log(f"{spec['name']} 已上传，按 --force 要求重新导入。")
                    else:
                        log(f"{spec['name']} 未上传，开始导入。")

                    payload = import_daily_file(context, upload, target_date, province_id)
                    errors = payload.get("data") or []
                    if errors:
                        raise RuntimeError(f"{spec['name']} 导入返回校验信息：{json.dumps(errors, ensure_ascii=False)[:1000]}")
                    log(f"{spec['name']} 导入成功。")
                if create_report:
                    create_daily_report(context, target_date, headless=config.headless)
                save_auth_state(context)
            finally:
                context.close()

    log("\n处理完成。")


def run_browser_upload(paths: list[Path], *, headless: bool, create_report: bool = True) -> None:
    target_date, uploads = prepare_upload_files(paths)
    log(f"本次上传日期：{target_date}")

    config = load_config()
    config.headless = headless

    with AuthStateLock():
        with sync_playwright() as playwright:
            context = launch_context(playwright, headless=config.headless)
            try:
                ensure_login(context, config)
                uploads = update_province_excel_for_upload(context, target_date, uploads)
                for item in uploads:
                    log(f"- {UPLOAD_TYPES[item.kind]['name']}：{item.path}")
                page = open_report_page(context, allow_manual=not headless)
                set_report_date(page, target_date)
                for upload in uploads:
                    upload_one_file(page, upload)
                if create_report:
                    create_daily_report(context, target_date, headless=headless)
                save_auth_state(context)
                log("\n两个文件处理完成，请在网页上确认上传结果。")
                if not headless:
                    log("浏览器窗口会保持打开；确认无误后可手动关闭。")
                    input("按 Enter 结束脚本...")
            finally:
                context.close()


def run_gui() -> None:
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox, ttk
    except Exception as exc:
        raise RuntimeError("当前 Python 环境不可用 tkinter，无法打开图形界面。") from exc

    root = tk.Tk()
    root.title("集团日报上传")
    root.geometry("760x520")
    root.minsize(680, 460)

    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except tk.TclError:
        pass

    selected_paths: list[Path] = []
    log_events: queue.Queue[tuple[str, str]] = queue.Queue()
    force_var = tk.BooleanVar(value=False)
    summary_var = tk.StringVar(value="请选择能销数据 Excel；省内数据会自动从同目录模板生成")
    file_lines_var = tk.StringVar(value="未选择文件")

    def gui_log(message: str) -> None:
        log_events.put(("log", message))

    def append_log(message: str) -> None:
        log_box.configure(state="normal")
        log_box.insert("end", message + "\n")
        log_box.see("end")
        log_box.configure(state="disabled")

    def refresh_selection_text() -> None:
        if not selected_paths:
            file_lines_var.set("未选择文件")
            summary_var.set("请选择能销数据 Excel；省内数据会自动从同目录模板生成")
            return
        try:
            target_date, uploads = prepare_upload_files(selected_paths)
            lines = [f"{UPLOAD_TYPES[item.kind]['name']}：{item.path.name}" for item in uploads]
            file_lines_var.set("\n".join(lines))
            summary_var.set(f"已识别日期：{target_date}；目标省份：{DEFAULT_PROVINCE_NAME}")
        except Exception as exc:
            file_lines_var.set("\n".join(path.name for path in selected_paths))
            summary_var.set(f"文件校验失败：{exc}")

    def choose_files() -> None:
        nonlocal selected_paths
        selected = filedialog.askopenfilenames(
            title="请选择能销数据 Excel；也可同时选择省内数据",
            initialdir=str(SCRIPT_DIR),
            filetypes=[
                ("Excel 文件", "*.xlsx *.xls"),
                ("所有文件", "*.*"),
            ],
        )
        if not selected:
            return
        selected_paths = [Path(item) for item in selected]
        refresh_selection_text()

    def set_busy(is_busy: bool) -> None:
        state = "disabled" if is_busy else "normal"
        choose_button.configure(state=state)
        upload_button.configure(state=state)
        login_button.configure(state=state)
        force_check.configure(state=state)

    def start_upload() -> None:
        if not selected_paths:
            messagebox.showwarning("缺少文件", "请先选择能销数据 Excel 文件。")
            return
        try:
            prepare_upload_files(selected_paths)
        except Exception as exc:
            messagebox.showerror("文件校验失败", str(exc))
            return

        set_busy(True)
        append_log("")
        append_log("开始处理...")

        def worker() -> None:
            global LOG_SINK
            old_sink = LOG_SINK
            LOG_SINK = gui_log
            try:
                run_api_upload(selected_paths, force=force_var.get())
                log_events.put(("done", "处理完成。"))
            except Exception as exc:
                log_events.put(("error", str(exc)))
            finally:
                LOG_SINK = old_sink
                log_events.put(("idle", ""))

        threading.Thread(target=worker, daemon=True).start()

    def start_login() -> None:
        set_busy(True)
        append_log("")
        append_log("打开浏览器登录窗口...")

        def worker() -> None:
            global LOG_SINK
            old_sink = LOG_SINK
            LOG_SINK = gui_log
            try:
                config = load_config()
                with sync_playwright() as playwright:
                    context = launch_context(playwright, headless=False)
                    try:
                        interactive_login(context, config)
                    finally:
                        context.close()
                log_events.put(("done", "登录态已保存。"))
            except Exception as exc:
                log_events.put(("error", str(exc)))
            finally:
                LOG_SINK = old_sink
                log_events.put(("idle", ""))

        threading.Thread(target=worker, daemon=True).start()

    def poll_log_events() -> None:
        try:
            while True:
                kind, message = log_events.get_nowait()
                if kind == "log":
                    append_log(message)
                elif kind == "done":
                    append_log(message)
                    messagebox.showinfo("完成", message)
                elif kind == "error":
                    append_log(f"失败：{message}")
                    messagebox.showerror("失败", message)
                elif kind == "idle":
                    set_busy(False)
        except queue.Empty:
            pass
        root.after(120, poll_log_events)

    outer = ttk.Frame(root, padding=16)
    outer.pack(fill="both", expand=True)

    header = ttk.Frame(outer)
    header.pack(fill="x")
    ttk.Label(header, text="集团日报上传", font=("Microsoft YaHei UI", 15, "bold")).pack(side="left")
    ttk.Label(header, text=f"{DEFAULT_PROVINCE_NAME}  接口模式", foreground="#555").pack(side="right")

    file_panel = ttk.LabelFrame(outer, text="文件", padding=12)
    file_panel.pack(fill="x", pady=(14, 10))
    ttk.Label(file_panel, textvariable=summary_var).pack(anchor="w")
    ttk.Label(file_panel, textvariable=file_lines_var, foreground="#333", justify="left").pack(anchor="w", pady=(8, 0))

    actions = ttk.Frame(outer)
    actions.pack(fill="x", pady=(0, 10))
    choose_button = ttk.Button(actions, text="选择文件", command=choose_files)
    choose_button.pack(side="left")
    upload_button = ttk.Button(actions, text="开始上传", command=start_upload)
    upload_button.pack(side="left", padx=(8, 0))
    login_button = ttk.Button(actions, text="登录/刷新登录态", command=start_login)
    login_button.pack(side="left", padx=(8, 0))
    force_check = ttk.Checkbutton(actions, text="已上传也覆盖", variable=force_var)
    force_check.pack(side="right")

    log_panel = ttk.LabelFrame(outer, text="日志", padding=8)
    log_panel.pack(fill="both", expand=True)
    log_box = tk.Text(log_panel, height=12, wrap="word", state="disabled", font=("Consolas", 10))
    log_scroll = ttk.Scrollbar(log_panel, orient="vertical", command=log_box.yview)
    log_box.configure(yscrollcommand=log_scroll.set)
    log_box.pack(side="left", fill="both", expand=True)
    log_scroll.pack(side="right", fill="y")

    append_log("操作：选择能销 Excel -> 开始上传。上传完成后会自动新建集团日报。")
    poll_log_events()
    root.mainloop()


def main() -> int:
    parser = argparse.ArgumentParser(description="上传集团日报省内数据和能销数据")
    parser.add_argument("files", nargs="*", help="可选：两个 Excel 文件路径。不传则弹出文件选择窗口。")
    parser.add_argument("--login", action="store_true", help="登录并保存会话")
    parser.add_argument("--headless", action="store_true", help="无界面运行")
    parser.add_argument("--browser", action="store_true", help="使用浏览器页面上传模式")
    parser.add_argument("--force", action="store_true", help="接口模式下即使当天已上传也重新导入")
    parser.add_argument("--skip-report", action="store_true", help="只上传数据，不自动新建集团日报")
    parser.add_argument("--smoke-report", action="store_true", help="只测试能否进入日报管理页面，不上传也不新建报表")
    parser.add_argument("--cli", action="store_true", help="不打开图形界面，使用命令行/文件选择模式")
    args = parser.parse_args()

    config = load_config()
    try:
        if args.login:
            with AuthStateLock():
                with sync_playwright() as playwright:
                    context = launch_context(playwright, headless=False)
                    try:
                        interactive_login(context, config)
                    finally:
                        context.close()
            return 0

        if args.smoke_report:
            smoke_daily_report_page(headless=args.headless)
            return 0

        if not args.files and not args.cli and not args.browser and not args.headless:
            run_gui()
            return 0

        paths = [Path(item) for item in args.files] if args.files else choose_files_interactively()
        if args.browser:
            run_browser_upload(paths, headless=args.headless, create_report=not args.skip_report)
        else:
            run_api_upload(paths, force=args.force, create_report=not args.skip_report)
        return 0
    except KeyboardInterrupt:
        print("\n已取消。")
        return 130
    except Exception as exc:
        print(f"\n失败：{exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
