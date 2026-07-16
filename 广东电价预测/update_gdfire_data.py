from __future__ import annotations

import argparse
import base64
from copy import copy
from datetime import date
import hashlib
import http.cookiejar
import json
from pathlib import Path
import ssl
import subprocess
import tempfile
import urllib.parse
import urllib.request

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter


BASE_URL = "https://xhxt.chng.com.cn"
PROVINCE_AREA_ID = "044"
SCRIPT_DIR = Path(__file__).resolve().parent
CONFIG_PATH = SCRIPT_DIR / "config.json"
SHARED_CONFIG_PATH = SCRIPT_DIR.parent / "上网电量抓取" / "config.json"
DEFAULT_MASTER_WORKBOOK = SCRIPT_DIR / "输出" / "广东电价数据总表.xlsx"
DEFAULT_BASE_WORKBOOK = (
    Path.home() / "Downloads" / "广东数据" / "广东数据整合_20240101至20260625.xlsx"
)


def load_config(path: Path = CONFIG_PATH) -> dict[str, object]:
    selected = path if path.is_file() else SHARED_CONFIG_PATH
    if not selected.is_file():
        raise FileNotFoundError(
            f"找不到配置文件：{path} 或 {SHARED_CONFIG_PATH}。"
            "请从 config.example.json 复制并填写账号密码。"
        )
    config = json.loads(selected.read_text(encoding="utf-8"))
    if not str(config.get("username") or "").strip() or not str(
        config.get("password") or ""
    ):
        raise ValueError(f"配置文件缺少 username 或 password：{selected}")
    return config

FORECAST_LOAD_COLUMNS = {
    "1": "日前统调负荷(MW)",
    "2": "日前A类电源出力(MW)",
    "3": "日前地方电源出力(MW)",
    "4": "日前西电东送电力(MW)",
    "5": "日前粤港联络线(MW)",
    "6": "日前B类竞价空间(MW)",
    "18": "日前发电总出力(MW)",
    "19": "日前光伏出力(MW)",
    "20": "日前风电出力(MW)",
    "22": "日前水电(含抽)蓄总出力(MW)",
}

ACTUAL_LOAD_COLUMNS = {
    "1": "实际统调负荷(MW)",
    "2": "实际A类电源出力(MW)",
    "3": "实际地方电源出力(MW)",
    "4": "实际西电东送电力(MW)",
    "5": "实际粤港联络线(MW)",
    "6": "实际B类竞价空间(MW)",
    "12": "实际风电(MW)",
    "13": "实际光伏(MW)",
    "14": "实际水电(MW)",
    "15": "实际火电(MW)",
    "16": "实际生物质(MW)",
    "17": "实际其他(MW)",
    "18": "实际发电总出力(MW)",
    "21": "实际新能源总出力(MW)",
    "22": "实际水电(含抽)蓄总出力(MW)",
}


class GDFireClient:
    def __init__(self) -> None:
        context = ssl.create_default_context()
        self.cookies = http.cookiejar.CookieJar()
        self.opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(self.cookies),
            urllib.request.HTTPSHandler(context=context),
        )

    def _json_request(
        self,
        path: str,
        *,
        params: dict[str, object] | None = None,
        method: str = "GET",
        data: dict[str, object] | None = None,
    ) -> dict[str, object]:
        url = BASE_URL + path
        if params:
            url += "?" + urllib.parse.urlencode(params, doseq=True)
        body = None if data is None else json.dumps(data).encode("utf-8")
        request = urllib.request.Request(
            url,
            data=body,
            method=method,
            headers={
                "Accept": "application/json, text/plain, */*",
                "Content-Type": "application/json;charset=UTF-8",
                "Origin": BASE_URL,
                "Referer": BASE_URL + "/gdfire/SpotDecisionSupport/InfoCompare",
            },
        )
        with self.opener.open(request, timeout=120) as response:
            payload = json.load(response)
        if payload.get("retCode") != "T200":
            raise RuntimeError(
                f"{path}: {payload.get('retCode')} {payload.get('retMsg')} "
                f"{payload.get('data')}"
            )
        return payload

    def login(self, username: str, password: str) -> None:
        public_key = self._json_request(
            "/usercenter/web/pf/login/info/publicKey"
        )["data"]
        pem = (
            "-----BEGIN PUBLIC KEY-----\n"
            + "\n".join(
                public_key[index : index + 64]
                for index in range(0, len(public_key), 64)
            )
            + "\n-----END PUBLIC KEY-----\n"
        )
        with tempfile.NamedTemporaryFile("w", delete=False) as handle:
            handle.write(pem)
            key_path = Path(handle.name)
        try:
            encrypted = subprocess.run(
                [
                    "openssl",
                    "pkeyutl",
                    "-encrypt",
                    "-pubin",
                    "-inkey",
                    str(key_path),
                    "-pkeyopt",
                    "rsa_padding_mode:pkcs1",
                ],
                input=password.encode("utf-8"),
                capture_output=True,
                check=True,
            ).stdout
        finally:
            key_path.unlink(missing_ok=True)
        self._json_request(
            "/usercenter/web/login",
            method="POST",
            params={
                "loginMode": 2,
                "username": username,
                "password": base64.b64encode(encrypted).decode("ascii"),
            },
        )

    def get(self, path: str, params: dict[str, object]) -> dict[str, object]:
        return self._json_request(path, params=params)


def save_json(path: Path, payload: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def empty_time_frame(start_date: str, end_date: str) -> pd.DataFrame:
    dates = pd.date_range(start_date, end_date, freq="D")
    times = pd.timedelta_range("00:00:00", "23:45:00", freq="15min")
    index = pd.MultiIndex.from_product([dates, range(96)], names=["日期", "时点序号"])
    frame = index.to_frame(index=False)
    frame["时刻"] = frame["时点序号"].map(
        lambda value: str(times[value]).split()[-1][:5]
    )
    return frame


def latest_period_records(records: list[dict[str, object]]) -> list[dict[str, object]]:
    valid = [
        record
        for record in records
        if record.get("date") and len(record.get("loadVal") or []) == 96
    ]
    valid.sort(
        key=lambda record: (
            str(record.get("date")),
            str(record.get("versionDate") or ""),
            int(record.get("versionNo") or -1),
        )
    )
    latest: dict[str, dict[str, object]] = {}
    for record in valid:
        latest[str(record["date"])[:10]] = record
    return list(latest.values())


def apply_period_list(
    frame: pd.DataFrame,
    records: list[dict[str, object]],
    column: str,
) -> None:
    for record in latest_period_records(records):
        date = pd.Timestamp(str(record["date"])[:10])
        values = [float("nan") if value is None else value for value in record["loadVal"]]
        mask = frame["日期"].eq(date)
        frame.loc[mask, column] = values


def apply_price_list(
    frame: pd.DataFrame,
    records: list[dict[str, object]],
    value_key: str,
    column: str,
) -> None:
    for record in records or []:
        values = record.get(value_key) or []
        if not record.get("date") or len(values) != 96:
            continue
        date = pd.Timestamp(str(record["date"])[:10])
        numeric_values = [float("nan") if value is None else value for value in values]
        frame.loc[frame["日期"].eq(date), column] = numeric_values


def build_increment(
    start_date: str,
    end_date: str,
    load_payload: dict[str, object],
    price_payload: dict[str, object],
) -> pd.DataFrame:
    frame = empty_time_frame(start_date, end_date)
    load_data = load_payload["data"]
    for item in load_data.get("dataNetLoadDTOList", []):
        load_type = str(item.get("loadType"))
        if load_type in FORECAST_LOAD_COLUMNS:
            apply_period_list(
                frame,
                item.get("forecastPeriodList") or [],
                FORECAST_LOAD_COLUMNS[load_type],
            )
        if load_type in ACTUAL_LOAD_COLUMNS:
            apply_period_list(
                frame,
                item.get("actualPeriodList") or [],
                ACTUAL_LOAD_COLUMNS[load_type],
            )

    frame["日前西电东送电力与B类竞价空间总加(MW)"] = (
        frame.get("日前西电东送电力(MW)") + frame.get("日前B类竞价空间(MW)")
    )

    price_data = price_payload["data"]
    apply_price_list(
        frame,
        price_data.get("preAvgNodePriceDTOList") or [],
        "avgNodePrice",
        "全省日前平均电价",
    )
    apply_price_list(
        frame,
        price_data.get("rtAvgNodePriceDTOList") or [],
        "avgNodePrice",
        "全省实时平均电价",
    )
    return frame.drop(columns=["时点序号"])


def read_data_workbook(path: Path) -> pd.DataFrame:
    with pd.ExcelFile(path) as workbook:
        sheet = (
            "整合数据" if "整合数据" in workbook.sheet_names else workbook.sheet_names[0]
        )
        return pd.read_excel(workbook, sheet_name=sheet)


def duplicate_data_columns(frame: pd.DataFrame) -> list[tuple[str, str]]:
    """Find fully identical, non-empty data columns while preserving semantics."""
    candidates = [column for column in frame.columns if column not in {"日期", "时刻"}]
    buckets: dict[str, list[str]] = {}
    duplicates: list[tuple[str, str]] = []
    for column in candidates:
        series = frame[column]
        if not series.notna().any():
            continue
        digest = hashlib.sha256(
            pd.util.hash_pandas_object(series, index=False).values.tobytes()
        ).hexdigest()
        for original in buckets.get(digest, []):
            if frame[original].equals(series):
                duplicates.append((original, column))
                break
        else:
            buckets.setdefault(digest, []).append(column)
    return duplicates


def deduplicate_columns(frame: pd.DataFrame) -> tuple[pd.DataFrame, list[tuple[str, str]]]:
    """Keep the first repeated header/value column and report what was removed."""
    cleaned = frame.loc[:, ~frame.columns.duplicated(keep="first")].copy()
    duplicates = duplicate_data_columns(cleaned)
    if duplicates:
        cleaned = cleaned.drop(columns=[duplicate for _, duplicate in duplicates])
    return cleaned, duplicates


def align_to_base_schema(base_path: Path, frame: pd.DataFrame) -> pd.DataFrame:
    """Use the existing integrated workbook's columns and order as the contract."""
    base = read_data_workbook(base_path).iloc[:0]
    base, _ = deduplicate_columns(base)
    removed_columns = ["日前风电(MW)", "日前光伏(MW)", "日前火电(MW)", "日前水电(MW)"]
    columns = [column for column in base.columns if column not in removed_columns]
    return frame.drop(columns=removed_columns, errors="ignore").reindex(columns=columns)


def merge_with_base(base_path: Path, increment: pd.DataFrame) -> pd.DataFrame:
    base = read_data_workbook(base_path)
    removed_columns = ["日前风电(MW)", "日前光伏(MW)", "日前火电(MW)", "日前水电(MW)"]
    base = base.drop(columns=removed_columns, errors="ignore")
    increment = increment.drop(columns=removed_columns, errors="ignore")
    base, _ = deduplicate_columns(base)
    base["日期"] = pd.to_datetime(base["日期"]).dt.normalize()
    increment["日期"] = pd.to_datetime(increment["日期"]).dt.normalize()
    all_columns = list(base.columns)
    base = base.reindex(columns=all_columns)
    increment = increment.reindex(columns=all_columns)
    combined = pd.concat([base, increment], ignore_index=True)
    combined = combined.sort_values(["日期", "时刻"])
    value_columns = [column for column in all_columns if column not in ["日期", "时刻"]]
    merged = (
        combined.groupby(["日期", "时刻"], as_index=False, sort=True)[value_columns]
        .last()
        .sort_values(["日期", "时刻"])
    )
    merged, _ = deduplicate_columns(merged)
    if merged.duplicated(["日期", "时刻"]).any():
        raise RuntimeError("合并后仍存在重复的日期+时刻记录")
    return merged


def missing_dates(frame: pd.DataFrame, end_date: str) -> list[date]:
    """Return dates through D that do not have all 96 quarter-hour rows."""
    if frame.empty:
        return []
    normalized = frame.copy()
    normalized["日期"] = pd.to_datetime(normalized["日期"], errors="raise").dt.normalize()
    end = pd.Timestamp(end_date).normalize()
    start = normalized["日期"].min()
    if start > end:
        return []
    counts = (
        normalized.loc[normalized["日期"].le(end)]
        .drop_duplicates(["日期", "时刻"])
        .groupby("日期", observed=True)["时刻"]
        .nunique()
    )
    expected = pd.date_range(start, end, freq="D")
    return [value.date() for value in expected if int(counts.get(value, 0)) != 96]


def workbook_layout(path: Path, sheet_name: str = "整合数据") -> dict[str, dict[str, object]]:
    """Read widths and representative styles keyed by header name."""
    workbook = load_workbook(path, read_only=False, data_only=False)
    sheet = workbook[sheet_name if sheet_name in workbook.sheetnames else workbook.sheetnames[0]]
    layout: dict[str, dict[str, object]] = {}
    for index, cell in enumerate(sheet[1], start=1):
        if cell.value is None:
            continue
        data_cell = sheet.cell(2, index)
        layout[str(cell.value)] = {
            "width": sheet.column_dimensions[get_column_letter(index)].width,
            "header_style": copy(cell._style),
            "data_style": copy(data_cell._style),
        }
    workbook.close()
    return layout


def write_large_xlsx(
    frame: pd.DataFrame,
    path: Path,
    sheet_name: str,
    *,
    format_source: Path | None = None,
) -> None:
    """Stream a large data frame to a temporary XLSX, then replace atomically."""
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = path.with_suffix(path.suffix + ".tmp")
    layout = workbook_layout(format_source) if format_source and format_source.is_file() else {}
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title=sheet_name)
    sheet.freeze_panes = "A2"
    header_cells = []
    for index, column in enumerate(frame.columns, start=1):
        cell = WriteOnlyCell(sheet, value=column)
        column_layout = layout.get(str(column))
        if column_layout:
            cell._style = copy(column_layout["header_style"])
            sheet.column_dimensions[get_column_letter(index)].width = column_layout["width"]
        header_cells.append(cell)
    sheet.append(header_cells)
    for row in frame.itertuples(index=False, name=None):
        cleaned = []
        for column, value in zip(frame.columns, row):
            if pd.isna(value):
                output = None
            elif isinstance(value, pd.Timestamp):
                output = value.to_pydatetime()
            elif isinstance(value, str):
                output = ILLEGAL_CHARACTERS_RE.sub("", value)
            else:
                output = value
            cell = WriteOnlyCell(sheet, value=output)
            column_layout = layout.get(str(column))
            if column_layout:
                cell._style = copy(column_layout["data_style"])
            cleaned.append(cell)
        sheet.append(cleaned)
    workbook.save(temporary_path)
    temporary_path.replace(path)


def build_modeling_frame(frame: pd.DataFrame) -> pd.DataFrame:
    """Keep only columns that are safe or potentially available before D+1 clears."""
    drop_columns = [
        column
        for column in frame.columns
        if column.startswith("实际")
        or "实时" in column
        or column in {"来源文件", "来源记录数"}
    ]
    result = frame.drop(columns=drop_columns, errors="ignore").copy()
    keep = ["日期", "时刻", "全省日前平均电价"]
    empty = [
        column
        for column in result.columns
        if column not in keep and result[column].isna().all()
    ]
    return result.drop(columns=empty)


def run_update(
    *,
    username: str,
    password: str,
    start_date: str,
    end_date: str,
    base_workbook: Path,
    output_dir: Path,
) -> dict[str, Path]:
    client = GDFireClient()
    client.login(username, password)

    common = {
        "startDate": start_date,
        "endDate": end_date,
        "provinceAreaId": PROVINCE_AREA_ID,
    }
    load_payload = client.get("/gdfire/api/data/net/load", common)
    price_payload = client.get(
        "/gdfire/api/data/price/avg/node/price",
        {"startDate": start_date, "endDate": end_date},
    )
    overhaul_payload = client.get(
        "/gdfire/api/unit/overhaulInfo",
        {"startDate": start_date, "endDate": end_date},
    )

    raw_dir = output_dir / "raw"
    save_json(raw_dir / "load.json", load_payload)
    save_json(raw_dir / "price.json", price_payload)
    save_json(raw_dir / "unit_overhaul.json", overhaul_payload)

    increment = build_increment(start_date, end_date, load_payload, price_payload)
    increment = align_to_base_schema(base_workbook, increment)
    output_dir.mkdir(parents=True, exist_ok=True)
    increment_path = output_dir / (
        f"广东接口增量数据_{start_date.replace('-', '')}至"
        f"{end_date.replace('-', '')}.xlsx"
    )
    write_large_xlsx(
        increment,
        increment_path,
        "Sheet1",
        format_source=base_workbook,
    )

    merged = merge_with_base(base_workbook, increment)
    max_date = merged["日期"].max().strftime("%Y%m%d")
    merged_path = output_dir / f"广东数据整合_更新至{max_date}.xlsx"
    write_large_xlsx(
        merged,
        merged_path,
        "整合数据",
        format_source=base_workbook,
    )

    modeling_path = output_dir / f"广东数据建模清洗_更新至{max_date}.xlsx"
    write_large_xlsx(
        build_modeling_frame(merged),
        modeling_path,
        "Sheet1",
        format_source=base_workbook,
    )

    overhaul = pd.DataFrame(overhaul_payload["data"])
    overhaul_path = output_dir / (
        f"机组检修信息_{start_date.replace('-', '')}至"
        f"{end_date.replace('-', '')}.xlsx"
    )
    overhaul.to_excel(overhaul_path, index=False, sheet_name="机组检修")
    return {
        "increment": increment_path.resolve(),
        "merged": merged_path.resolve(),
        "modeling": modeling_path.resolve(),
        "overhaul": overhaul_path.resolve(),
        "raw": raw_dir.resolve(),
    }


def update_master_workbook(
    *,
    run_date: str | None = None,
    config_path: Path = CONFIG_PATH,
) -> dict[str, object]:
    """Fill missing dates through D into one persistent master workbook."""
    config = load_config(config_path)
    master_path = Path(
        str(config.get("masterWorkbook") or DEFAULT_MASTER_WORKBOOK)
    ).expanduser()
    if not master_path.is_absolute():
        master_path = SCRIPT_DIR / master_path
    base_path = Path(str(config.get("initialWorkbook") or DEFAULT_BASE_WORKBOOK)).expanduser()
    end_date = run_date or date.today().isoformat()

    if master_path.is_file():
        current = read_data_workbook(master_path)
    elif base_path.is_file():
        current = read_data_workbook(base_path)
    else:
        initial_date = str(config.get("initialDate") or end_date)
        current = empty_time_frame(initial_date, initial_date).drop(columns=["时点序号"])

    gaps = missing_dates(current, end_date)
    if not master_path.is_file() and not base_path.is_file():
        gaps = [value.date() for value in pd.date_range(initial_date, end_date, freq="D")]
    if not gaps:
        if not master_path.is_file():
            write_large_xlsx(current, master_path, "整合数据")
        return {
            "master": master_path.resolve(),
            "requested_dates": [],
            "updated": False,
        }

    # One API interval keeps the update path simple. Existing values survive because
    # merge_with_base takes the latest non-empty value per field.
    start = min(gaps).isoformat()
    end = max(gaps).isoformat()
    client = GDFireClient()
    client.login(str(config["username"]), str(config["password"]))
    common = {"startDate": start, "endDate": end, "provinceAreaId": PROVINCE_AREA_ID}
    load_payload = client.get("/gdfire/api/data/net/load", common)
    price_payload = client.get(
        "/gdfire/api/data/price/avg/node/price",
        {"startDate": start, "endDate": end},
    )
    increment = build_increment(start, end, load_payload, price_payload)
    increment = increment.reindex(columns=current.columns)

    temporary_base = master_path.with_suffix(".base.xlsx")
    write_large_xlsx(current, temporary_base, "整合数据")
    try:
        merged = merge_with_base(temporary_base, increment)
    finally:
        temporary_base.unlink(missing_ok=True)
    write_large_xlsx(merged, master_path, "整合数据")
    remaining = missing_dates(merged, end_date)
    return {
        "master": master_path.resolve(),
        "requested_dates": [value.isoformat() for value in gaps],
        "remaining_missing_dates": [value.isoformat() for value in remaining],
        "updated": True,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="补齐广东电价数据总表")
    parser.add_argument("--date", help="运行日 D，默认今天，格式 YYYY-MM-DD")
    parser.add_argument("--config", type=Path, default=CONFIG_PATH)
    args = parser.parse_args()
    print(
        json.dumps(
            update_master_workbook(run_date=args.date, config_path=args.config),
            ensure_ascii=False,
            indent=2,
            default=str,
        )
    )


if __name__ == "__main__":
    main()
