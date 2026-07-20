from __future__ import annotations

import json
import os
import queue
import subprocess
import sys
import threading
import traceback
import time as time_module
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, range_boundaries
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
except ImportError:  # pragma: no cover - surfaced in GUI
    load_workbook = None
    get_column_letter = None
    range_boundaries = None
    coordinate_from_string = None
    column_index_from_string = None

try:
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
    from playwright.sync_api import sync_playwright
except ImportError:  # pragma: no cover - surfaced in GUI
    PlaywrightTimeoutError = Exception
    sync_playwright = None


BASE_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "wps_excel_to_kdocs_config.json"
PROFILE_DIR = BASE_DIR / "wps-browser-profile"
WPS_CHROME_PID_PATH = BASE_DIR / "wps-chrome.pid"
WPS_CHROME_STATE_PATH = BASE_DIR / "wps-chrome-state.json"
WPS_CHROME_LAUNCH_VERSION = 2
WPS_CHROME_START_LOCK = threading.Lock()
LOG_DIR = BASE_DIR / "logs"
MAX_RECENT_URLS = 10
DOCUMENT_CONFIGS_MIN_HEIGHT = 220
CONFIG_EXPORT_SCHEMA = "wps_excel_to_kdocs_config_export"
CONFIG_EXPORT_VERSION = 1


def center_window(window: tk.Toplevel | tk.Tk, width: int | None = None, height: int | None = None) -> None:
    window.update_idletasks()
    actual_width = width or window.winfo_width() or window.winfo_reqwidth()
    actual_height = height or window.winfo_height() or window.winfo_reqheight()
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    actual_width = min(actual_width, max(640, screen_width - 80))
    actual_height = min(actual_height, max(480, screen_height - 80))
    x = max(0, (screen_width - actual_width) // 2)
    y = max(0, (screen_height - actual_height) // 2)
    window.geometry(f"{actual_width}x{actual_height}+{x}+{y}")


@dataclass
class MappingRow:
    kdocs_url: str
    local_file: str
    source_sheet: str
    source_start: str
    source_end: str
    target_sheet: str
    target_start: str
    target_end: str
    source_type: str = "excel"
    source_url: str = ""


@dataclass
class RegionRow:
    source_sheet: str
    source_start: str
    source_end: str
    target_sheet: str
    target_start: str
    target_end: str


@dataclass
class DocumentConfig:
    name: str
    kdocs_url: str
    local_file: str
    regions: list[RegionRow]
    source_type: str = "excel"
    source_url: str = ""


@dataclass
class PreparedMapping:
    index: int
    kdocs_url: str
    workbook_path: str
    source_sheet: str
    source_range: str
    target_sheet: str
    target_start: str
    target_range: str
    rows: int
    cols: int
    values: list[list[Any]]


def normalize_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return value


def normalize_for_compare(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text.lower() in {"none", "null"}:
        return ""
    return text


def values_equivalent(expected: Any, actual: Any) -> bool:
    expected_value = normalize_for_compare(expected)
    actual_value = normalize_for_compare(actual)
    if expected_value == actual_value:
        return True
    if isinstance(expected_value, bool) or isinstance(actual_value, bool):
        return False
    try:
        expected_number = float(expected_value)
        actual_number = float(actual_value)
        return abs(expected_number - actual_number) <= 1e-9
    except (TypeError, ValueError):
        return False


def parse_start_cell(cell: str) -> tuple[int, int]:
    if coordinate_from_string is None or column_index_from_string is None:
        raise RuntimeError("openpyxl is not installed.")
    col_text, row = coordinate_from_string(cell.strip().upper())
    return row, column_index_from_string(col_text)


def normalize_cell_address(cell: str) -> str:
    row, col = parse_start_cell(cell)
    return f"{get_column_letter(col)}{row}"


def range_from_cells(start_cell: str, end_cell: str) -> str:
    start_row, start_col = parse_start_cell(start_cell)
    end_row, end_col = parse_start_cell(end_cell)
    if end_row < start_row or end_col < start_col:
        raise ValueError(f"End cell must be down/right of start cell: {start_cell} -> {end_cell}")
    return f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"


def range_shape(start_cell: str, end_cell: str) -> tuple[int, int]:
    start_row, start_col = parse_start_cell(start_cell)
    end_row, end_col = parse_start_cell(end_cell)
    if end_row < start_row or end_col < start_col:
        raise ValueError(f"End cell must be down/right of start cell: {start_cell} -> {end_cell}")
    return end_row - start_row + 1, end_col - start_col + 1


def split_range(range_text: str) -> tuple[str, str]:
    text = range_text.strip()
    if ":" not in text:
        cell = normalize_cell_address(text)
        return cell, cell
    start, end = text.split(":", 1)
    return normalize_cell_address(start), normalize_cell_address(end)


def target_range_from_start(start_cell: str, rows: int, cols: int) -> str:
    start_row, start_col = parse_start_cell(start_cell)
    end_row = start_row + rows - 1
    end_col = start_col + cols - 1
    return f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"


def target_rect(start_cell: str, rows: int, cols: int) -> tuple[int, int, int, int]:
    start_row, start_col = parse_start_cell(start_cell)
    return start_row, start_col, start_row + rows - 1, start_col + cols - 1


def rect_from_cells(start_cell: str, end_cell: str) -> tuple[int, int, int, int]:
    start_row, start_col = parse_start_cell(start_cell)
    end_row, end_col = parse_start_cell(end_cell)
    if end_row < start_row or end_col < start_col:
        raise ValueError(f"End cell must be down/right of start cell: {start_cell} -> {end_cell}")
    return start_row, start_col, end_row, end_col


def rects_overlap(a: tuple[int, int, int, int], b: tuple[int, int, int, int]) -> bool:
    a_r1, a_c1, a_r2, a_c2 = a
    b_r1, b_c1, b_r2, b_c2 = b
    return not (a_r2 < b_r1 or b_r2 < a_r1 or a_c2 < b_c1 or b_c2 < a_c1)


def load_config() -> dict[str, Any]:
    if not CONFIG_PATH.exists():
        return {}
    try:
        return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_config(data: dict[str, Any]) -> None:
    CONFIG_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _string_list(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    return [item.strip() for item in value if isinstance(item, str) and item.strip()]


def _region_config_from_dict(row: dict[str, Any]) -> dict[str, str]:
    return {
        "source_sheet": str(row.get("source_sheet", "")),
        "source_start": str(row.get("source_start", "")),
        "source_end": str(row.get("source_end", "")),
        "target_sheet": str(row.get("target_sheet", "")),
        "target_start": str(row.get("target_start", "")),
        "target_end": str(row.get("target_end", "")),
    }


def _document_config_from_dict(row: dict[str, Any]) -> dict[str, Any]:
    regions = [
        _region_config_from_dict(region)
        for region in row.get("regions", [])
        if isinstance(region, dict)
    ]
    name = str(row.get("name", "")).strip() or Path(str(row.get("local_file", ""))).stem or "Config"
    source_type = str(row.get("source_type", "excel")).strip() or "excel"
    if source_type not in {"excel", "kdocs"}:
        source_type = "excel"
    return {
        "name": name,
        "source_type": source_type,
        "source_url": str(row.get("source_url", "")),
        "kdocs_url": str(row.get("kdocs_url", "")),
        "local_file": str(row.get("local_file", "")),
        "regions": regions,
    }


def normalize_runtime_config(data: dict[str, Any]) -> dict[str, Any]:
    return {
        "recent_urls": _string_list(data.get("recent_urls", [])),
        "browser_mode": str(data.get("browser_mode", "persistent") or "persistent"),
        "cdp_url": str(data.get("cdp_url", "http://127.0.0.1:9222") or "http://127.0.0.1:9222"),
        "configs": [
            _document_config_from_dict(row)
            for row in data.get("configs", [])
            if isinstance(row, dict)
        ],
    }


def build_config_export(runtime_config: dict[str, Any], exported_at: str | None = None) -> dict[str, Any]:
    export_time = exported_at or datetime.now().replace(microsecond=0).isoformat()
    return {
        "schema": CONFIG_EXPORT_SCHEMA,
        "version": CONFIG_EXPORT_VERSION,
        "exported_at": export_time,
        **normalize_runtime_config(runtime_config),
    }


def runtime_config_from_export_payload(payload: dict[str, Any]) -> dict[str, Any]:
    if payload.get("schema") != CONFIG_EXPORT_SCHEMA:
        raise ValueError(f"Unsupported config export schema: {payload.get('schema')!r}")
    if payload.get("version") != CONFIG_EXPORT_VERSION:
        raise ValueError(f"Unsupported config export version: {payload.get('version')!r}")
    runtime_config = normalize_runtime_config(payload)
    if not runtime_config["configs"]:
        raise ValueError("Config export does not contain any document configs.")
    return runtime_config


def load_config_export(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid JSON config export: {exc}") from exc
    if not isinstance(payload, dict):
        raise ValueError("Config export must be a JSON object.")
    return runtime_config_from_export_payload(payload)


def save_config_export(path: Path, runtime_config: dict[str, Any]) -> None:
    path.write_text(
        json.dumps(build_config_export(runtime_config), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def update_recent_urls(urls: list[str], url: str) -> list[str]:
    normalized = url.strip()
    if not normalized:
        return urls[:MAX_RECENT_URLS]
    result = [normalized]
    for item in urls:
        item = item.strip()
        if item and item != normalized and item not in result:
            result.append(item)
        if len(result) >= MAX_RECENT_URLS:
            break
    return result


def read_excel_range(workbook_path: Path, sheet_name: str, source_range: str) -> tuple[list[list[Any]], int, int]:
    if load_workbook is None or range_boundaries is None:
        raise RuntimeError("Missing dependency: openpyxl. Run: pip install -r work/requirements.txt")
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel file not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Source sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(source_range)
        rows = max_row - min_row + 1
        cols = max_col - min_col + 1
        values: list[list[Any]] = []
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            values.append([normalize_cell_value(cell) for cell in row])
        return values, rows, cols
    finally:
        wb.close()


def read_excel_sheet_names(workbook_path: Path) -> list[str]:
    if load_workbook is None:
        raise RuntimeError("Missing dependency: openpyxl. Run: pip install -r work/requirements.txt")
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel file not found: {workbook_path}")
    wb = load_workbook(workbook_path, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def prepare_mappings(
    mappings: list[MappingRow],
    browser_mode: str = "persistent",
    cdp_url: str = "http://127.0.0.1:9222",
    log=None,
) -> list[PreparedMapping]:
    if not mappings:
        raise ValueError("No range mappings configured.")

    prepared: list[PreparedMapping] = []
    target_rects: dict[tuple[str, str], list[tuple[int, tuple[int, int, int, int], str]]] = {}
    log = log or (lambda _message: None)
    for idx, mapping in enumerate(mappings, start=1):
        source_type = mapping.source_type or "excel"
        workbook_path = None
        if source_type == "excel":
            workbook_path_text = mapping.local_file.strip()
            if not workbook_path_text:
                raise ValueError(f"Mapping {idx} has no local file.")
            workbook_path = Path(workbook_path_text).expanduser()
        elif source_type == "kdocs":
            if not mapping.source_url.strip():
                raise ValueError(f"Mapping {idx} has no source KDocs URL.")
        else:
            raise ValueError(f"Mapping {idx} has unknown source type: {source_type}")
        if not all(
            [
                mapping.kdocs_url,
                mapping.source_sheet,
                mapping.source_start,
                mapping.source_end,
                mapping.target_sheet,
                mapping.target_start,
                mapping.target_end,
            ]
        ):
            raise ValueError(f"Mapping {idx} has empty fields.")
        source_range = range_from_cells(mapping.source_start, mapping.source_end)
        target_range = range_from_cells(mapping.target_start, mapping.target_end)
        source_rows, source_cols = range_shape(mapping.source_start, mapping.source_end)
        target_rows, target_cols = range_shape(mapping.target_start, mapping.target_end)
        if (source_rows, source_cols) != (target_rows, target_cols):
            raise ValueError(
                f"Mapping {idx} size mismatch: source {source_rows}x{source_cols}, "
                f"target {target_rows}x{target_cols}"
            )
        if source_type == "excel":
            values, rows, cols = read_excel_range(workbook_path, mapping.source_sheet, source_range)
            source_label = str(workbook_path)
        else:
            values = read_online_range_values(
                mapping.source_url,
                mapping.source_sheet,
                source_range,
                source_rows,
                source_cols,
                browser_mode,
                cdp_url,
                log,
            )
            rows, cols = source_rows, source_cols
            source_label = mapping.source_url
        rect = rect_from_cells(mapping.target_start, mapping.target_end)
        target_key = (mapping.kdocs_url, mapping.target_sheet)
        for other_idx, other_rect, other_range in target_rects.get(target_key, []):
            if rects_overlap(rect, other_rect):
                raise ValueError(
                    f"Target ranges overlap for URL '{mapping.kdocs_url}' on sheet '{mapping.target_sheet}': "
                    f"mapping {other_idx} {other_range} and mapping {idx} {target_range}"
                )
        target_rects.setdefault(target_key, []).append((idx, rect, target_range))
        prepared.append(
            PreparedMapping(
                index=idx,
                kdocs_url=mapping.kdocs_url,
                workbook_path=source_label,
                source_sheet=mapping.source_sheet,
                source_range=source_range,
                target_sheet=mapping.target_sheet,
                target_start=mapping.target_start,
                target_range=target_range,
                rows=rows,
                cols=cols,
                values=values,
            )
        )
    return prepared


def coerce_wps_matrix(raw: Any, rows: int, cols: int) -> list[list[Any]]:
    if rows == 1 and cols == 1:
        if isinstance(raw, list):
            if raw and isinstance(raw[0], list):
                return [[raw[0][0] if raw[0] else ""]]
            return [[raw[0] if raw else ""]]
        return [[raw]]
    if isinstance(raw, list):
        if rows == 1 and (not raw or not isinstance(raw[0], list)):
            return [raw]
        matrix = []
        for row in raw:
            matrix.append(row if isinstance(row, list) else [row])
        return matrix
    return [[raw]]


def compare_matrix(expected: list[list[Any]], actual_raw: Any, rows: int, cols: int) -> list[str]:
    actual = coerce_wps_matrix(actual_raw, rows, cols)
    mismatches: list[str] = []
    for r in range(rows):
        for c in range(cols):
            raw_expected = expected[r][c] if r < len(expected) and c < len(expected[r]) else ""
            raw_actual = actual[r][c] if r < len(actual) and c < len(actual[r]) else ""
            if not values_equivalent(raw_expected, raw_actual):
                cell = f"R{r + 1}C{c + 1}"
                expected_value = normalize_for_compare(raw_expected)
                actual_value = normalize_for_compare(raw_actual)
                mismatches.append(f"{cell}: expected {expected_value!r}, got {actual_value!r}")
                if len(mismatches) >= 20:
                    return mismatches
    return mismatches


def wait_for_wps_ready(page, log, timeout_seconds: int) -> dict[str, Any]:
    deadline = time_module.monotonic() + timeout_seconds
    last_state: dict[str, Any] = {"ready": False, "url": "", "title": ""}
    last_log_second = 0
    readiness_script = """
    async () => {
      const api = window.WPSOpenApi;
      if (!api?.EtApplication) {
        return {
          ready: false,
          url: location.href,
          title: document.title,
          hasWPSOpenApi: !!api
        };
      }

      const bounded = (value, milliseconds = 750) => Promise.race([
        Promise.resolve(value),
        new Promise(resolve => setTimeout(() => resolve(false), milliseconds))
      ]);
      let app;
      let applicationReady = false;
      let documentReady = false;
      try {
        app = api.EtApplication();
        applicationReady = Boolean(await bounded(app.Ready));
      } catch (_error) {}
      if (!applicationReady && api.documentReadyPromise) {
        try {
          documentReady = Boolean(await bounded(
            Promise.resolve(api.documentReadyPromise).then(() => true)
          ));
        } catch (_error) {}
      }
      return {
        ready: applicationReady || documentReady,
        url: location.href,
        title: document.title,
        hasSave: typeof api.save === "function"
      };
    }
    """
    while time_module.monotonic() < deadline:
        try:
            state = page.evaluate(readiness_script)
            last_state = state
            if state.get("ready"):
                return state
        except Exception as exc:
            last_state = {"ready": False, "url": page.url, "title": "", "error": str(exc)}

        remaining = int(deadline - time_module.monotonic())
        elapsed_bucket = timeout_seconds - remaining
        if elapsed_bucket // 15 > last_log_second // 15:
            last_log_second = elapsed_bucket
            log(f"Still waiting for WPSOpenApi... current URL: {page.url}")
        page.wait_for_timeout(1000)
    return last_state


def find_or_open_page(context, url: str):
    target_key = url.rstrip("/").split("/")[-1]
    for page in context.pages:
        if target_key and target_key in page.url:
            page.bring_to_front()
            return page
    for page in context.pages:
        if "kdocs.cn" in page.url:
            page.bring_to_front()
            page.goto(url, wait_until="commit", timeout=60000)
            return page
    page = next(
        (
            candidate
            for candidate in context.pages
            if candidate.url in {"", "about:blank", "chrome://newtab/"}
        ),
        None,
    )
    if page is None:
        page = context.new_page()
    page.bring_to_front()
    page.goto(url, wait_until="commit", timeout=60000)
    return page


def cdp_candidates(cdp_url: str) -> list[str]:
    raw = cdp_url.strip() or "http://localhost:9222"
    candidates = [raw]
    parsed = urllib.parse.urlparse(raw)
    if parsed.scheme in {"http", "https"} and parsed.port:
        for host in ("localhost", "[::1]", "127.0.0.1"):
            netloc = f"{host}:{parsed.port}"
            candidate = urllib.parse.urlunparse((parsed.scheme, netloc, parsed.path, "", "", ""))
            if candidate not in candidates:
                candidates.append(candidate)
    return candidates


def resolve_cdp_endpoint(cdp_url: str) -> str:
    if cdp_url.startswith("ws://") or cdp_url.startswith("wss://"):
        return cdp_url
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
    errors: list[str] = []
    for candidate in cdp_candidates(cdp_url):
        version_url = candidate.rstrip("/") + "/json/version"
        try:
            with opener.open(version_url, timeout=5) as response:
                data = json.loads(response.read().decode("utf-8"))
            ws_url = data.get("webSocketDebuggerUrl")
            if ws_url:
                return ws_url
            errors.append(f"{version_url}: missing webSocketDebuggerUrl")
        except Exception as exc:
            errors.append(f"{version_url}: {exc}")
    raise RuntimeError("Could not resolve Chrome DevTools websocket. " + " | ".join(errors))


def find_macos_chrome() -> Path | None:
    candidates = [
        Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
        Path.home() / "Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        Path("/Applications/Google Chrome for Testing.app/Contents/MacOS/Google Chrome for Testing"),
        Path.home() / "Applications/Google Chrome for Testing.app/Contents/MacOS/Google Chrome for Testing",
        Path("/Applications/Chromium.app/Contents/MacOS/Chromium"),
        Path.home() / "Applications/Chromium.app/Contents/MacOS/Chromium",
    ]
    return next((candidate for candidate in candidates if candidate.is_file()), None)


def managed_cdp_host_port(cdp_url: str) -> tuple[str, int]:
    parsed = urllib.parse.urlparse(cdp_url.strip() or "http://127.0.0.1:9222")
    host = parsed.hostname or "127.0.0.1"
    if host not in {"127.0.0.1", "localhost", "::1"}:
        raise ValueError("The managed WPS Chrome CDP address must use localhost.")
    return "127.0.0.1", parsed.port or 9222


def build_wps_chrome_command(
    chrome_path: Path,
    cdp_url: str,
    initial_url: str = "https://www.kdocs.cn/",
) -> list[str]:
    host, port = managed_cdp_host_port(cdp_url)
    return [
        str(chrome_path),
        f"--remote-debugging-address={host}",
        f"--remote-debugging-port={port}",
        f"--user-data-dir={PROFILE_DIR}",
        "--no-first-run",
        "--no-default-browser-check",
        # The Tk config dialog normally covers Chrome on macOS. Keep the WPS
        # renderer at foreground speed while that dedicated window is occluded.
        "--disable-background-timer-throttling",
        "--disable-backgrounding-occluded-windows",
        "--disable-renderer-backgrounding",
        initial_url,
    ]


def try_resolve_cdp_endpoint(cdp_url: str, timeout: float = 0.5) -> str | None:
    if cdp_url.startswith("ws://") or cdp_url.startswith("wss://"):
        return cdp_url
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
    for candidate in cdp_candidates(cdp_url):
        try:
            with opener.open(candidate.rstrip("/") + "/json/version", timeout=timeout) as response:
                data = json.loads(response.read().decode("utf-8"))
            websocket_url = data.get("webSocketDebuggerUrl")
            if websocket_url:
                return str(websocket_url)
        except Exception:
            continue
    return None


def cdp_page_targets(
    cdp_url: str, timeout: float = 0.5
) -> tuple[str | None, list[dict[str, Any]]]:
    if cdp_url.startswith("ws://") or cdp_url.startswith("wss://"):
        return None, []
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
    for candidate in cdp_candidates(cdp_url):
        try:
            with opener.open(candidate.rstrip("/") + "/json/list", timeout=timeout) as response:
                data = json.loads(response.read().decode("utf-8"))
            if isinstance(data, list):
                return candidate.rstrip("/"), [item for item in data if isinstance(item, dict)]
        except Exception:
            continue
    return None, []


def ensure_wps_cdp_page(
    cdp_url: str,
    initial_url: str = "https://www.kdocs.cn/",
    *,
    require_initial_url: bool = False,
) -> bool:
    base_url, targets = cdp_page_targets(cdp_url)
    page_targets = [target for target in targets if target.get("type") == "page"]
    target_key = initial_url.rstrip("/").split("/")[-1]
    if require_initial_url and target_key:
        if any(target_key in str(target.get("url") or "") for target in page_targets):
            return False
    elif page_targets:
        return False
    if base_url is None:
        raise RuntimeError("Could not inspect the dedicated WPS Chrome page list.")

    encoded_url = urllib.parse.quote(initial_url, safe="")
    request = urllib.request.Request(
        f"{base_url}/json/new?{encoded_url}", method="PUT"
    )
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
    with opener.open(request, timeout=5) as response:
        created = json.loads(response.read().decode("utf-8"))
    if not isinstance(created, dict) or created.get("type") != "page":
        raise RuntimeError("Dedicated WPS Chrome did not create a browser page.")
    return True


def managed_chrome_pid() -> int | None:
    try:
        pid = int(WPS_CHROME_PID_PATH.read_text(encoding="utf-8").strip())
        os.kill(pid, 0)
        if sys.platform == "darwin":
            result = subprocess.run(
                ["ps", "-p", str(pid), "-o", "command="],
                capture_output=True,
                text=True,
                timeout=2,
                check=False,
            )
            if str(PROFILE_DIR) not in result.stdout:
                return None
        return pid
    except (
        FileNotFoundError,
        ValueError,
        ProcessLookupError,
        PermissionError,
        subprocess.SubprocessError,
    ):
        return None


def managed_chrome_launch_version(pid: int) -> int:
    try:
        state = json.loads(WPS_CHROME_STATE_PATH.read_text(encoding="utf-8"))
        if int(state.get("pid") or 0) != pid:
            return 0
        return int(state.get("version") or 0)
    except (FileNotFoundError, ValueError, TypeError, json.JSONDecodeError):
        return 0


def stop_outdated_managed_chrome(log, cdp_url: str) -> None:
    pid = managed_chrome_pid()
    if pid is None or managed_chrome_launch_version(pid) >= WPS_CHROME_LAUNCH_VERSION:
        return
    log("Restarting dedicated WPS Chrome once to apply the macOS speed settings.")
    try:
        os.kill(pid, 15)
    except ProcessLookupError:
        return
    deadline = time_module.monotonic() + 8
    while time_module.monotonic() < deadline:
        if try_resolve_cdp_endpoint(cdp_url, timeout=0.2) is None:
            time_module.sleep(0.5)
            return
        time_module.sleep(0.2)
    raise RuntimeError(
        "The old dedicated WPS Chrome did not exit. Close that dedicated Chrome window and try again."
    )


def _ensure_wps_dedicated_chrome(
    cdp_url: str = "http://127.0.0.1:9222",
    log=lambda _message: None,
    timeout_seconds: float = 15,
    initial_url: str = "https://www.kdocs.cn/",
    preload_url: bool = False,
) -> dict[str, Any]:
    stop_outdated_managed_chrome(log, cdp_url)
    existing_endpoint = try_resolve_cdp_endpoint(cdp_url)
    if existing_endpoint:
        ensure_wps_cdp_page(
            cdp_url,
            initial_url,
            require_initial_url=preload_url,
        )
        log("Reusing the running dedicated WPS Chrome.")
        return {"started": False, "endpoint": existing_endpoint}
    existing_pid = managed_chrome_pid()
    if existing_pid is not None:
        # Chrome can briefly stop answering /json/version while a heavy KDocs
        # page is busy. Give the owned process time to recover instead of
        # launching a second process against the same profile.
        recovery_deadline = time_module.monotonic() + 3
        while time_module.monotonic() < recovery_deadline:
            endpoint = try_resolve_cdp_endpoint(cdp_url)
            if endpoint:
                ensure_wps_cdp_page(
                    cdp_url,
                    initial_url,
                    require_initial_url=preload_url,
                )
                log("Reusing the running dedicated WPS Chrome.")
                return {"started": False, "endpoint": endpoint, "pid": existing_pid}
            time_module.sleep(0.2)
        raise RuntimeError(
            "Dedicated WPS Chrome is running but its local debugging endpoint is not responding. "
            "Wait a moment or close that dedicated window and start it again."
        )
    if sys.platform != "darwin":
        raise RuntimeError("Automatic dedicated Chrome startup is currently available on macOS only.")

    chrome_path = find_macos_chrome()
    if chrome_path is None:
        raise RuntimeError("Google Chrome was not found in /Applications or ~/Applications.")
    PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    try:
        PROFILE_DIR.chmod(0o700)
    except OSError:
        pass
    command = build_wps_chrome_command(chrome_path, cdp_url, initial_url)
    log(f"Starting dedicated WPS Chrome: {chrome_path.name}")
    process = subprocess.Popen(
        command,
        stdin=subprocess.DEVNULL,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        start_new_session=True,
    )
    WPS_CHROME_PID_PATH.write_text(str(process.pid), encoding="utf-8")
    WPS_CHROME_STATE_PATH.write_text(
        json.dumps(
            {"pid": process.pid, "version": WPS_CHROME_LAUNCH_VERSION},
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    deadline = time_module.monotonic() + timeout_seconds
    while time_module.monotonic() < deadline:
        endpoint = try_resolve_cdp_endpoint(cdp_url)
        if endpoint:
            ensure_wps_cdp_page(
                cdp_url,
                initial_url,
                require_initial_url=preload_url,
            )
            log("Dedicated WPS Chrome is ready. Log in once if WPS asks for it.")
            return {"started": True, "endpoint": endpoint, "pid": process.pid}
        if process.poll() is not None:
            break
        time_module.sleep(0.2)
    raise RuntimeError(
        "Dedicated WPS Chrome did not expose its local debugging endpoint. "
        "Close any Chrome window using the WPS profile and try again."
    )


def ensure_wps_dedicated_chrome(
    cdp_url: str = "http://127.0.0.1:9222",
    log=lambda _message: None,
    timeout_seconds: float = 15,
    initial_url: str = "https://www.kdocs.cn/",
    preload_url: bool = False,
) -> dict[str, Any]:
    with WPS_CHROME_START_LOCK:
        return _ensure_wps_dedicated_chrome(
            cdp_url,
            log,
            timeout_seconds,
            initial_url,
            preload_url,
        )


def open_browser_context(
    playwright,
    browser_mode: str,
    cdp_url: str,
    log,
    initial_url: str | None = None,
):
    if browser_mode in {"managed_cdp", "cdp"}:
        if browser_mode == "managed_cdp":
            ensure_wps_dedicated_chrome(
                cdp_url,
                log,
                initial_url=initial_url or "https://www.kdocs.cn/",
                preload_url=bool(initial_url),
            )
        else:
            log(f"Connecting to existing Chrome through CDP: {cdp_url}")
        endpoint = resolve_cdp_endpoint(cdp_url)
        try:
            browser = playwright.chromium.connect_over_cdp(endpoint)
        except Exception as exc:
            if browser_mode != "managed_cdp" or "context management is not supported" not in str(exc):
                raise
            log("Dedicated Chrome had no open page; creating a WPS page and reconnecting.")
            ensure_wps_cdp_page(cdp_url)
            endpoint = resolve_cdp_endpoint(cdp_url)
            browser = playwright.chromium.connect_over_cdp(endpoint)
        context = browser.contexts[0] if browser.contexts else browser.new_context()
        return browser, context, False

    PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    log("Starting one-time browser with persistent WPS profile...")
    try:
        context = playwright.chromium.launch_persistent_context(
            str(PROFILE_DIR),
            headless=False,
            channel="chrome",
            viewport={"width": 1400, "height": 900},
        )
    except Exception:
        log("Chrome channel was not available; falling back to bundled Chromium.")
        context = playwright.chromium.launch_persistent_context(
            str(PROFILE_DIR),
            headless=False,
            viewport={"width": 1400, "height": 900},
        )
    return None, context, True


def write_to_wps(
    url: str,
    prepared: list[PreparedMapping],
    log,
    browser_mode: str = "persistent",
    cdp_url: str = "http://127.0.0.1:9222",
) -> dict[str, Any]:
    if sync_playwright is None:
        raise RuntimeError("Missing dependency: playwright. Run: pip install -r work/requirements.txt")

    payload = [
        {
            "index": item.index,
            "targetSheet": item.target_sheet,
            "targetRange": item.target_range,
            "rows": item.rows,
            "cols": item.cols,
            "values": item.values,
        }
        for item in prepared
    ]

    PROFILE_DIR.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        try:
            _browser, context, close_context = open_browser_context(
                p, browser_mode, cdp_url, log, initial_url=url
            )
        except Exception as exc:
            if browser_mode == "cdp":
                raise RuntimeError(
                    "Could not connect to existing Chrome. Start it with a custom --user-data-dir "
                    "and --remote-debugging-port, then log in to WPS in that window."
                ) from exc
            raise
        try:
            log(f"Opening or selecting KDocs URL: {url}")
            page = find_or_open_page(context, url)

            log("Waiting for WPSOpenApi. If a login page is shown, finish login in the browser window.")
            ready = wait_for_wps_ready(page, log, timeout_seconds=180)
            if not ready.get("ready"):
                raise RuntimeError(
                    "WPSOpenApi was not ready after waiting. "
                    f"Current page: {ready.get('title')} {ready.get('url')}"
                )
            log(f"Document ready: {ready.get('title')}")

            writer_script = """
            async (payload) => {
              const app = window.WPSOpenApi.EtApplication();
              const results = [];
              for (const item of payload) {
                const sheet = app.Worksheets(item.targetSheet);
                const range = sheet.Range(item.targetRange);
                const assignment = item.rows === 1 && item.cols === 1 ? item.values[0][0] : item.values;
                const mergeCells = await range.MergeCells;
                range.Value = assignment;
                results.push({
                  index: item.index,
                  targetSheet: item.targetSheet,
                  targetRange: item.targetRange,
                  rows: item.rows,
                  cols: item.cols,
                  mergeCells
                });
              }
              await window.WPSOpenApi.save();
              for (const result of results) {
                const sheet = app.Worksheets(result.targetSheet);
                const range = sheet.Range(result.targetRange);
                result.value = await range.Value;
                result.value2 = await range.Value2;
                result.text = await range.Text;
              }
              return { saved: true, results };
            }
            """
            log(f"Writing {len(payload)} mapping(s) through WPSOpenApi...")
            result = None
            last_error: Exception | None = None
            for attempt in range(1, 4):
                try:
                    if attempt > 1:
                        log(f"Retrying write attempt {attempt}/3 after page navigation...")
                        wait_for_wps_ready(page, log, timeout_seconds=60)
                    result = page.evaluate(writer_script, payload)
                    break
                except Exception as exc:
                    last_error = exc
                    if "Execution context was destroyed" not in str(exc) and "navigation" not in str(exc).lower():
                        raise
                    page.wait_for_timeout(1500)
            if result is None:
                raise RuntimeError(f"Write failed after retries: {last_error}")
            return result
        finally:
            if close_context:
                context.close()


def read_online_document_info(url: str, browser_mode: str, cdp_url: str, log) -> dict[str, Any]:
    if sync_playwright is None:
        raise RuntimeError("Missing dependency: playwright. Run: pip install -r work/requirements.txt")

    with sync_playwright() as p:
        _browser, context, close_context = open_browser_context(
            p, browser_mode, cdp_url, log, initial_url=url
        )
        try:
            page = find_or_open_page(context, url)
            ready = wait_for_wps_ready(page, log, timeout_seconds=120)
            if not ready.get("ready"):
                raise RuntimeError(
                    "WPSOpenApi was not ready after waiting. "
                    f"Current page: {ready.get('title')} {ready.get('url')}"
                )
            script = """
            async () => {
              const app = window.WPSOpenApi.EtApplication();
              const count = await app.Worksheets.Count;
              const sheets = [];
              for (let i = 1; i <= Number(count); i++) {
                const ws = app.Worksheets(i);
                sheets.push(await ws.Name);
              }
              return sheets;
            }
            """
            sheets = page.evaluate(script)
            return {"title": ready.get("title") or page.title(), "sheets": sheets}
        finally:
            if close_context:
                context.close()


def read_online_sheet_names(url: str, browser_mode: str, cdp_url: str, log) -> list[str]:
    return read_online_document_info(url, browser_mode, cdp_url, log)["sheets"]


def read_online_range_values(
    url: str,
    sheet_name: str,
    range_text: str,
    rows: int,
    cols: int,
    browser_mode: str,
    cdp_url: str,
    log,
) -> list[list[Any]]:
    if sync_playwright is None:
        raise RuntimeError("Missing dependency: playwright. Run: pip install -r work/requirements.txt")

    with sync_playwright() as p:
        _browser, context, close_context = open_browser_context(
            p, browser_mode, cdp_url, log, initial_url=url
        )
        try:
            page = find_or_open_page(context, url)
            ready = wait_for_wps_ready(page, log, timeout_seconds=120)
            if not ready.get("ready"):
                raise RuntimeError(
                    "WPSOpenApi was not ready after waiting. "
                    f"Current page: {ready.get('title')} {ready.get('url')}"
                )
            script = """
            async ({sheetName, rangeText}) => {
              const app = window.WPSOpenApi.EtApplication();
              const sheet = await app.Worksheets(sheetName);
              const range = await sheet.Range(rangeText);
              return await range.Value;
            }
            """
            raw = None
            last_error: Exception | None = None
            for attempt in range(1, 4):
                try:
                    if attempt > 1:
                        wait_for_wps_ready(page, log, timeout_seconds=60)
                    raw = page.evaluate(script, {"sheetName": sheet_name, "rangeText": range_text})
                    break
                except Exception as exc:
                    last_error = exc
                    if "Execution context was destroyed" not in str(exc) and "navigation" not in str(exc).lower():
                        raise
                    page.wait_for_timeout(1500)
            if raw is None and last_error is not None:
                raise RuntimeError(f"Read source range failed after retries: {last_error}")
            matrix = coerce_wps_matrix(raw, rows, cols)
            return [[normalize_cell_value(cell) for cell in row] for row in matrix]
        finally:
            if close_context:
                context.close()


class MappingDialog(tk.Toplevel):
    def __init__(
        self,
        parent: tk.Tk,
        title: str,
        initial: MappingRow | None = None,
        recent_urls: list[str] | None = None,
        browser_mode: str = "persistent",
        cdp_url: str = "http://127.0.0.1:9222",
        log=None,
    ):
        super().__init__(parent)
        self.title(title)
        self.resizable(False, False)
        self.result: MappingRow | None = None
        default_url = initial.kdocs_url if initial else ((recent_urls or [""])[0])
        self.local_sheet_combo: ttk.Combobox | None = None
        self.online_sheet_combo: ttk.Combobox | None = None
        self.kdocs_url_combo: ttk.Combobox | None = None
        self.last_online_sheet_url = ""
        self.browser_mode = browser_mode
        self.cdp_url = cdp_url
        self.log = log or (lambda _message: None)

        self.vars = {
            "kdocs_url": tk.StringVar(value=default_url),
            "local_file": tk.StringVar(value=initial.local_file if initial else ""),
            "source_sheet": tk.StringVar(value=initial.source_sheet if initial else ""),
            "source_start": tk.StringVar(value=initial.source_start if initial else ""),
            "source_end": tk.StringVar(value=initial.source_end if initial else ""),
            "target_sheet": tk.StringVar(value=initial.target_sheet if initial else ""),
            "target_start": tk.StringVar(value=initial.target_start if initial else ""),
            "target_end": tk.StringVar(value=initial.target_end if initial else ""),
        }

        labels = [
            ("KDocs URL", "kdocs_url"),
            ("Local file", "local_file"),
            ("Local sheet", "source_sheet"),
            ("Local start cell", "source_start"),
            ("Local end cell", "source_end"),
            ("Online sheet", "target_sheet"),
            ("Online start cell", "target_start"),
            ("Online end cell", "target_end"),
        ]
        for row, (label, key) in enumerate(labels):
            ttk.Label(self, text=label).grid(row=row, column=0, padx=10, pady=6, sticky="w")
            if key == "kdocs_url":
                entry = ttk.Combobox(self, textvariable=self.vars[key], values=recent_urls or [], width=64)
                self.kdocs_url_combo = entry
                entry.bind("<<ComboboxSelected>>", lambda _event: self.refresh_online_sheets(auto=True))
                entry.bind("<FocusOut>", lambda _event: self.refresh_online_sheets(auto=True))
                entry.bind("<Return>", lambda _event: self.refresh_online_sheets(auto=True))
            elif key == "source_sheet":
                entry = ttk.Combobox(self, textvariable=self.vars[key], values=[], width=36)
                self.local_sheet_combo = entry
            elif key == "target_sheet":
                entry = ttk.Combobox(self, textvariable=self.vars[key], values=[], width=36)
                self.online_sheet_combo = entry
            else:
                entry = ttk.Entry(self, textvariable=self.vars[key], width=64 if key == "local_file" else 36)
            entry.grid(row=row, column=1, padx=10, pady=6, sticky="ew")
            if key == "local_file":
                button_frame = ttk.Frame(self)
                button_frame.grid(row=row, column=2, padx=10, pady=6)
                ttk.Button(button_frame, text="Browse", command=self.browse_local_file).pack(side="left")
                ttk.Button(button_frame, text="Refresh sheets", command=self.refresh_local_sheets).pack(side="left", padx=(6, 0))
            elif key == "target_sheet":
                ttk.Button(self, text="Refresh online sheets", command=self.refresh_online_sheets).grid(
                    row=row, column=2, padx=10, pady=6
                )

        button_frame = ttk.Frame(self)
        button_frame.grid(row=len(labels), column=0, columnspan=3, padx=10, pady=10, sticky="e")
        ttk.Button(button_frame, text="OK", command=self.on_ok).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Cancel", command=self.destroy).pack(side="left", padx=4)

        self.bind("<Return>", lambda _event: self.on_ok())
        self.bind("<Escape>", lambda _event: self.destroy())
        self.refresh_local_sheets(show_errors=False)
        self.transient(parent)
        self.grab_set()
        center_window(self)
        self.wait_visibility()
        self.focus()
        self.wait_window(self)

    def on_ok(self) -> None:
        row = MappingRow(
            kdocs_url=self.vars["kdocs_url"].get().strip(),
            local_file=self.vars["local_file"].get().strip(),
            source_sheet=self.vars["source_sheet"].get().strip(),
            source_start=self.vars["source_start"].get().strip(),
            source_end=self.vars["source_end"].get().strip(),
            target_sheet=self.vars["target_sheet"].get().strip(),
            target_start=self.vars["target_start"].get().strip(),
            target_end=self.vars["target_end"].get().strip(),
        )
        if not all(
            [
                row.kdocs_url,
                row.local_file,
                row.source_sheet,
                row.source_start,
                row.source_end,
                row.target_sheet,
                row.target_start,
                row.target_end,
            ]
        ):
            messagebox.showerror("Invalid mapping", "All mapping fields are required.", parent=self)
            return
        self.result = row
        self.destroy()

    def browse_local_file(self) -> None:
        path = filedialog.askopenfilename(
            parent=self,
            title="Choose Excel file for this mapping",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.vars["local_file"].set(path)
            self.refresh_local_sheets(show_errors=True)

    def refresh_local_sheets(self, show_errors: bool = True) -> None:
        if self.local_sheet_combo is None:
            return
        path_text = self.vars["local_file"].get().strip()
        if not path_text:
            self.local_sheet_combo.configure(values=[])
            return
        try:
            names = read_excel_sheet_names(Path(path_text).expanduser())
        except Exception as exc:
            self.local_sheet_combo.configure(values=[])
            if show_errors:
                messagebox.showerror("Read sheets failed", str(exc), parent=self)
            return
        self.local_sheet_combo.configure(values=names)
        current = self.vars["source_sheet"].get().strip()
        if names and not current:
            self.vars["source_sheet"].set(names[0])

    def refresh_online_sheets(self, auto: bool = False) -> None:
        if self.online_sheet_combo is None:
            return
        url = self.vars["kdocs_url"].get().strip()
        if not url:
            if auto:
                return
            messagebox.showerror("Missing URL", "KDocs URL is required before reading online sheets.", parent=self)
            return
        if auto and url == self.last_online_sheet_url:
            return
        try:
            names = read_online_sheet_names(url, self.browser_mode, self.cdp_url, lambda _message: None)
        except Exception as exc:
            self.online_sheet_combo.configure(values=[])
            self.last_online_sheet_url = ""
            messagebox.showerror("Read online sheets failed", str(exc), parent=self)
            return
        self.online_sheet_combo.configure(values=names)
        self.last_online_sheet_url = url
        current = self.vars["target_sheet"].get().strip()
        if names and not current:
            self.vars["target_sheet"].set(names[0])


class RegionDialog(tk.Toplevel):
    def __init__(
        self,
        parent: tk.Toplevel,
        title: str,
        initial: RegionRow | None = None,
        local_sheets: list[str] | None = None,
        online_sheets: list[str] | None = None,
    ):
        super().__init__(parent)
        self.title(title)
        self.resizable(False, False)
        self.result: RegionRow | None = None

        self.vars = {
            "source_sheet": tk.StringVar(value=initial.source_sheet if initial else ""),
            "source_start": tk.StringVar(value=initial.source_start if initial else ""),
            "source_end": tk.StringVar(value=initial.source_end if initial else ""),
            "target_sheet": tk.StringVar(value=initial.target_sheet if initial else ""),
            "target_start": tk.StringVar(value=initial.target_start if initial else ""),
            "target_end": tk.StringVar(value=initial.target_end if initial else ""),
        }

        labels = [
            ("Source sheet", "source_sheet"),
            ("Source start cell", "source_start"),
            ("Source end cell", "source_end"),
            ("Online sheet", "target_sheet"),
            ("Online start cell", "target_start"),
            ("Online end cell", "target_end"),
        ]
        for row, (label, key) in enumerate(labels):
            ttk.Label(self, text=label).grid(row=row, column=0, padx=10, pady=6, sticky="w")
            if key == "source_sheet":
                entry = ttk.Combobox(self, textvariable=self.vars[key], values=local_sheets or [], width=42)
            elif key == "target_sheet":
                entry = ttk.Combobox(self, textvariable=self.vars[key], values=online_sheets or [], width=42)
            else:
                entry = ttk.Entry(self, textvariable=self.vars[key], width=44)
            entry.grid(row=row, column=1, padx=10, pady=6, sticky="ew")

        if not initial:
            if local_sheets:
                self.vars["source_sheet"].set(local_sheets[0])
            if online_sheets:
                self.vars["target_sheet"].set(online_sheets[0])

        button_frame = ttk.Frame(self)
        button_frame.grid(row=len(labels), column=0, columnspan=2, padx=10, pady=10, sticky="e")
        ttk.Button(button_frame, text="OK", command=self.on_ok).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Cancel", command=self.destroy).pack(side="left", padx=4)

        self.bind("<Return>", lambda _event: self.on_ok())
        self.bind("<Escape>", lambda _event: self.destroy())
        self.transient(parent)
        self.grab_set()
        center_window(self)
        self.wait_visibility()
        self.focus()
        self.wait_window(self)

    def on_ok(self) -> None:
        row = RegionRow(
            source_sheet=self.vars["source_sheet"].get().strip(),
            source_start=self.vars["source_start"].get().strip(),
            source_end=self.vars["source_end"].get().strip(),
            target_sheet=self.vars["target_sheet"].get().strip(),
            target_start=self.vars["target_start"].get().strip(),
            target_end=self.vars["target_end"].get().strip(),
        )
        if not all([row.source_sheet, row.source_start, row.source_end, row.target_sheet, row.target_start, row.target_end]):
            messagebox.showerror("Invalid region", "All region fields are required.", parent=self)
            return
        self.result = row
        self.destroy()


class DocumentConfigDialog(tk.Toplevel):
    def __init__(
        self,
        parent: tk.Tk,
        title: str,
        initial: DocumentConfig | None = None,
        recent_urls: list[str] | None = None,
        browser_mode: str = "persistent",
        cdp_url: str = "http://127.0.0.1:9222",
        log=None,
        autosave=None,
    ):
        super().__init__(parent)
        self.title(title)
        self.geometry("900x620")
        self.minsize(820, 520)
        self.result: DocumentConfig | None = None
        self.recent_urls = recent_urls or []
        self.browser_mode = browser_mode
        self.cdp_url = cdp_url
        self.log = log or (lambda _message: None)
        self.autosave = autosave
        self.local_sheets: list[str] = list(
            dict.fromkeys(
                region.source_sheet
                for region in (initial.regions if initial else [])
                if region.source_sheet
            )
        )
        self.online_sheets: list[str] = list(
            dict.fromkeys(
                region.target_sheet
                for region in (initial.regions if initial else [])
                if region.target_sheet
            )
        )
        self.last_online_sheet_url = ""
        self.regions: list[RegionRow] = list(initial.regions) if initial else []
        self._browser_read_lock = threading.Lock()
        self._online_results: queue.Queue[
            tuple[str, int, str, bool, dict[str, Any] | None, str | None]
        ] = queue.Queue()
        self._request_ids = {"source": 0, "target": 0}
        self._pending_online_reads: dict[str, tuple[int, str]] = {}
        self._online_poll_after_id: str | None = None
        self._online_workers: set[threading.Thread] = set()
        self._online_workers_lock = threading.Lock()

        default_url = initial.kdocs_url if initial else (self.recent_urls[0] if self.recent_urls else "")
        default_source_url = initial.source_url if initial else ""
        self.config_name = tk.StringVar(value=initial.name if initial else "")
        self.source_type = tk.StringVar(value=initial.source_type if initial else "excel")
        self.source_url = tk.StringVar(value=default_source_url)
        self.kdocs_url = tk.StringVar(value=default_url)
        self.local_file = tk.StringVar(value=initial.local_file if initial else "")
        self.source_doc_title = tk.StringVar(value="")
        self.target_doc_title = tk.StringVar(value="")
        cached_sheet_count = len(self.online_sheets)
        cached_status = (
            f"Using {cached_sheet_count} sheet name(s) from the saved config; "
            "refresh online only if the document changed."
            if cached_sheet_count
            else ""
        )
        self.read_status = tk.StringVar(value=cached_status)

        self.create_widgets()
        if self.source_type.get() == "excel":
            self.refresh_local_sheets(show_errors=False)

        self.transient(parent)
        self.grab_set()
        center_window(self, 900, 620)
        self.wait_visibility()
        self.focus()
        self.wait_window(self)

    def create_widgets(self) -> None:
        root = ttk.Frame(self, padding=10)
        root.pack(fill="both", expand=True)
        root.columnconfigure(0, weight=1)
        root.rowconfigure(1, weight=1)

        docs = ttk.LabelFrame(root, text="Documents")
        docs.grid(row=0, column=0, sticky="ew")
        ttk.Label(docs, text="Name").grid(row=0, column=0, padx=8, pady=6, sticky="w")
        ttk.Entry(docs, textvariable=self.config_name).grid(row=0, column=1, columnspan=2, padx=8, pady=6, sticky="ew")

        ttk.Label(docs, text="Source type").grid(row=1, column=0, padx=8, pady=6, sticky="w")
        source_frame = ttk.Frame(docs)
        source_frame.grid(row=1, column=1, columnspan=2, padx=8, pady=6, sticky="ew")
        ttk.Radiobutton(source_frame, text="Local Excel", variable=self.source_type, value="excel").pack(side="left")
        ttk.Radiobutton(source_frame, text="KDocs document", variable=self.source_type, value="kdocs").pack(side="left", padx=(16, 0))

        ttk.Label(docs, text="Source KDocs URL").grid(row=2, column=0, padx=8, pady=6, sticky="w")
        source_url_combo = ttk.Combobox(docs, textvariable=self.source_url, values=self.recent_urls)
        source_url_combo.grid(row=2, column=1, padx=8, pady=6, sticky="ew")
        self.refresh_source_button = ttk.Button(
            docs, text="Refresh source sheets", command=self.refresh_local_sheets
        )
        self.refresh_source_button.grid(row=2, column=2, padx=8, pady=6)

        ttk.Label(docs, text="Target KDocs URL").grid(row=3, column=0, padx=8, pady=6, sticky="w")
        url_combo = ttk.Combobox(docs, textvariable=self.kdocs_url, values=self.recent_urls)
        url_combo.grid(row=3, column=1, padx=8, pady=6, sticky="ew")
        self.refresh_target_button = ttk.Button(
            docs, text="Refresh target sheets", command=self.refresh_online_sheets
        )
        self.refresh_target_button.grid(row=3, column=2, padx=8, pady=6)

        ttk.Label(docs, text="Source document").grid(row=4, column=0, padx=8, pady=6, sticky="w")
        ttk.Entry(docs, textvariable=self.source_doc_title, state="readonly").grid(row=4, column=1, columnspan=2, padx=8, pady=6, sticky="ew")

        ttk.Label(docs, text="Target document").grid(row=5, column=0, padx=8, pady=6, sticky="w")
        ttk.Entry(docs, textvariable=self.target_doc_title, state="readonly").grid(row=5, column=1, columnspan=2, padx=8, pady=6, sticky="ew")

        ttk.Label(docs, text="Local file").grid(row=6, column=0, padx=8, pady=6, sticky="w")
        ttk.Entry(docs, textvariable=self.local_file).grid(row=6, column=1, padx=8, pady=6, sticky="ew")
        file_buttons = ttk.Frame(docs)
        file_buttons.grid(row=6, column=2, padx=8, pady=6)
        ttk.Button(file_buttons, text="Browse", command=self.browse_local_file).pack(side="left")
        ttk.Button(file_buttons, text="Refresh sheets", command=self.refresh_local_sheets).pack(side="left", padx=(6, 0))
        ttk.Label(docs, textvariable=self.read_status, style="Muted.TLabel").grid(
            row=7, column=0, columnspan=3, padx=8, pady=(0, 6), sticky="w"
        )
        docs.columnconfigure(1, weight=1)

        regions_frame = ttk.LabelFrame(root, text="Regions")
        regions_frame.grid(row=1, column=0, sticky="nsew", pady=(10, 0))
        toolbar = ttk.Frame(regions_frame)
        toolbar.pack(fill="x", padx=8, pady=(8, 0))
        ttk.Button(toolbar, text="Add region", command=self.add_region).pack(side="left")
        ttk.Button(toolbar, text="Edit", command=self.edit_region).pack(side="left", padx=8)
        ttk.Button(toolbar, text="Copy", command=self.copy_region).pack(side="left")
        ttk.Button(toolbar, text="Remove", command=self.remove_region).pack(side="left", padx=8)

        tree_frame = ttk.Frame(regions_frame)
        tree_frame.pack(fill="both", expand=True, padx=8, pady=8)
        columns = ("source_sheet", "source_start", "source_end", "target_sheet", "target_start", "target_end")
        self.region_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=10)
        for key, label, width in [
            ("source_sheet", "Source sheet", 180),
            ("source_start", "Source start", 100),
            ("source_end", "Source end", 100),
            ("target_sheet", "Online sheet", 180),
            ("target_start", "Online start", 100),
            ("target_end", "Online end", 100),
        ]:
            self.region_tree.heading(key, text=label)
            self.region_tree.column(key, width=width)
        self.region_tree.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.region_tree.yview)
        scrollbar.pack(side="left", fill="y")
        self.region_tree.configure(yscrollcommand=scrollbar.set)
        self.region_tree.bind("<Double-1>", lambda _event: self.edit_region())

        button_frame = ttk.Frame(root)
        button_frame.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        ttk.Button(button_frame, text="OK", command=self.on_ok).pack(side="right", padx=4)
        ttk.Button(button_frame, text="Cancel", command=self.destroy).pack(side="right", padx=4)
        self.refresh_regions()

    def browse_local_file(self) -> None:
        path = filedialog.askopenfilename(
            parent=self,
            title="Choose Excel file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.local_file.set(path)
            self.refresh_local_sheets(show_errors=True)

    def refresh_local_sheets(self, show_errors: bool = True) -> None:
        if self.source_type.get() == "kdocs":
            url = self.source_url.get().strip()
            if not url:
                self.local_sheets = []
                self.source_doc_title.set("")
                if show_errors:
                    messagebox.showerror("Missing URL", "Source KDocs URL is required.", parent=self)
                return
            self._start_online_read("source", url, show_errors)
            return

        self._request_ids["source"] += 1
        self._pending_online_reads.pop("source", None)
        self.refresh_source_button.configure(state="normal")
        try:
            path_text = self.local_file.get().strip()
            if not path_text:
                self.local_sheets = []
                self.source_doc_title.set("")
                return
            self.local_sheets = read_excel_sheet_names(Path(path_text).expanduser())
            self.source_doc_title.set(Path(path_text).name)
        except Exception as exc:
            self.local_sheets = []
            if show_errors:
                messagebox.showerror("Read source sheets failed", str(exc), parent=self)

    def refresh_online_sheets(self, auto: bool = False) -> None:
        url = self.kdocs_url.get().strip()
        if not url:
            if auto:
                return
            messagebox.showerror("Missing URL", "KDocs URL is required before reading online sheets.", parent=self)
            return
        if auto and url == self.last_online_sheet_url:
            return
        self._start_online_read("target", url, not auto)

    def _start_online_read(self, kind: str, url: str, show_errors: bool) -> None:
        pending = self._pending_online_reads.get(kind)
        if pending is not None and pending[1] == url:
            return

        self._request_ids[kind] += 1
        request_id = self._request_ids[kind]
        self._pending_online_reads[kind] = (request_id, url)
        button = self.refresh_source_button if kind == "source" else self.refresh_target_button
        button.configure(state="disabled")
        if kind == "source":
            self.source_doc_title.set("Loading...")
        else:
            self.target_doc_title.set("Loading...")
        self.read_status.set("Reading WPS document in the browser; this window remains usable...")

        # Keep the worker independent from this Toplevel. A browser read can outlive
        # the dialog, and retaining ``self`` from that thread lets Tk variables be
        # finalized off the main thread when the window has already been closed.
        browser_read_lock = self._browser_read_lock
        browser_mode = self.browser_mode
        cdp_url = self.cdp_url
        log = self.log
        online_results = self._online_results
        online_workers = self._online_workers
        online_workers_lock = self._online_workers_lock

        def worker() -> None:
            info: dict[str, Any] | None = None
            error: str | None = None
            try:
                # A persistent Playwright profile cannot be opened twice at once.
                with browser_read_lock:
                    info = read_online_document_info(
                        url, browser_mode, cdp_url, log
                    )
            except Exception as exc:
                error = str(exc)
            finally:
                online_results.put(
                    (kind, request_id, url, show_errors, info, error)
                )
                with online_workers_lock:
                    online_workers.discard(threading.current_thread())

        thread = threading.Thread(target=worker, daemon=True)
        with self._online_workers_lock:
            self._online_workers.add(thread)
        try:
            thread.start()
        except Exception:
            with self._online_workers_lock:
                self._online_workers.discard(thread)
            raise
        if self._online_poll_after_id is None:
            self._online_poll_after_id = self.after(50, self._drain_online_results)

    def destroy(self) -> None:
        if self._online_poll_after_id is not None:
            try:
                self.after_cancel(self._online_poll_after_id)
            except tk.TclError:
                pass
            self._online_poll_after_id = None
        self._pending_online_reads.clear()
        super().destroy()

    def _drain_online_results(self) -> None:
        self._online_poll_after_id = None
        while True:
            try:
                kind, request_id, url, show_errors, info, error = self._online_results.get_nowait()
            except queue.Empty:
                break
            if request_id != self._request_ids[kind]:
                continue

            self._pending_online_reads.pop(kind, None)
            button = self.refresh_source_button if kind == "source" else self.refresh_target_button
            button.configure(state="normal")
            if error is not None:
                if kind == "source":
                    self.local_sheets = []
                    self.source_doc_title.set("")
                    title = "Read source sheets failed"
                else:
                    self.online_sheets = []
                    self.target_doc_title.set("")
                    self.last_online_sheet_url = ""
                    title = "Read online sheets failed"
                self.read_status.set("WPS document read failed; use Refresh to try again.")
                if show_errors:
                    messagebox.showerror(title, error, parent=self)
                continue

            sheets = list((info or {}).get("sheets") or [])
            document_title = str((info or {}).get("title") or "")
            if kind == "source":
                self.local_sheets = sheets
                self.source_doc_title.set(document_title)
            else:
                self.online_sheets = sheets
                self.target_doc_title.set(document_title)
                self.last_online_sheet_url = url
            self.read_status.set(f"Loaded {len(sheets)} sheet(s) from WPS.")

        if self._pending_online_reads:
            self._online_poll_after_id = self.after(50, self._drain_online_results)

    def add_region(self) -> None:
        dialog = RegionDialog(self, "Add region", local_sheets=self.local_sheets, online_sheets=self.online_sheets)
        if dialog.result:
            self.regions.append(dialog.result)
            self.refresh_regions()
            self.autosave_current_config()

    def edit_region(self) -> None:
        selected = self.region_tree.selection()
        if not selected:
            return
        index = self.region_tree.index(selected[0])
        dialog = RegionDialog(
            self,
            "Edit region",
            self.regions[index],
            local_sheets=self.local_sheets,
            online_sheets=self.online_sheets,
        )
        if dialog.result:
            self.regions[index] = dialog.result
            self.refresh_regions()
            self.autosave_current_config()

    def copy_region(self) -> None:
        selected = self.region_tree.selection()
        if not selected:
            messagebox.showinfo("Copy region", "Select one region row to copy.", parent=self)
            return
        index = self.region_tree.index(selected[0])
        source = self.regions[index]
        copied = RegionRow(
            source_sheet=source.source_sheet,
            source_start=source.source_start,
            source_end=source.source_end,
            target_sheet=source.target_sheet,
            target_start=source.target_start,
            target_end=source.target_end,
        )
        self.regions.insert(index + 1, copied)
        self.refresh_regions()
        new_item = self.region_tree.get_children()[index + 1]
        self.region_tree.selection_set(new_item)
        self.region_tree.focus(new_item)
        self.region_tree.see(new_item)
        self.autosave_current_config()

    def remove_region(self) -> None:
        selected = self.region_tree.selection()
        if not selected:
            return
        count = len(selected)
        if not messagebox.askyesno(
            "Confirm remove",
            f"Remove {count} selected region(s)?",
            parent=self,
        ):
            return
        for item in reversed(selected):
            del self.regions[self.region_tree.index(item)]
        self.refresh_regions()
        self.autosave_current_config()

    def build_current_config(self) -> DocumentConfig | None:
        kdocs_url = self.kdocs_url.get().strip()
        local_file = self.local_file.get().strip()
        source_type = self.source_type.get()
        source_url = self.source_url.get().strip()
        if not kdocs_url or not self.regions:
            return None
        if source_type == "excel" and not local_file:
            return None
        if source_type == "kdocs" and not source_url:
            return None
        name = self.config_name.get().strip() or Path(local_file).stem or "Config"
        return DocumentConfig(
            name=name,
            kdocs_url=kdocs_url,
            local_file=local_file,
            regions=list(self.regions),
            source_type=source_type,
            source_url=source_url,
        )

    def autosave_current_config(self) -> None:
        if self.autosave is None:
            return
        config = self.build_current_config()
        if config is None:
            return
        self.autosave(config)

    def refresh_regions(self) -> None:
        for item in self.region_tree.get_children():
            self.region_tree.delete(item)
        for region in self.regions:
            self.region_tree.insert(
                "",
                "end",
                values=(
                    region.source_sheet,
                    region.source_start,
                    region.source_end,
                    region.target_sheet,
                    region.target_start,
                    region.target_end,
                ),
            )

    def on_ok(self) -> None:
        kdocs_url = self.kdocs_url.get().strip()
        local_file = self.local_file.get().strip()
        source_type = self.source_type.get()
        source_url = self.source_url.get().strip()
        if not kdocs_url:
            messagebox.showerror("Invalid config", "Target KDocs URL is required.", parent=self)
            return
        if source_type == "excel" and not local_file:
            messagebox.showerror("Invalid config", "Local file is required.", parent=self)
            return
        if source_type == "kdocs" and not source_url:
            messagebox.showerror("Invalid config", "Source KDocs URL is required.", parent=self)
            return
        if not self.regions:
            messagebox.showerror("Invalid config", "Add at least one region.", parent=self)
            return
        name = self.config_name.get().strip() or Path(local_file).stem or "Config"
        self.result = DocumentConfig(
            name=name,
            kdocs_url=kdocs_url,
            local_file=local_file,
            regions=list(self.regions),
            source_type=source_type,
            source_url=source_url,
        )
        if self.autosave is not None:
            self.autosave(self.result)
        self.destroy()


class WpsWriterFrame(ttk.Frame):
    def __init__(self, parent: tk.Misc, task_registry: Any | None = None) -> None:
        super().__init__(parent)
        self.task_registry = task_registry
        self.log_queue: queue.Queue[str] = queue.Queue()
        self.browser_event_queue: queue.Queue[tuple[bool, str]] = queue.Queue()
        self.worker: threading.Thread | None = None
        self.browser_worker: threading.Thread | None = None
        self.configs: list[DocumentConfig] = []
        self.log_window: tk.Toplevel | None = None
        self.log_window_text: tk.Text | None = None
        self._drain_after_id: str | None = None

        self.recent_urls: list[str] = []
        self.browser_mode = tk.StringVar(master=self, value="managed_cdp")
        self.cdp_url = tk.StringVar(master=self, value="http://127.0.0.1:9222")
        self.browser_status = tk.StringVar(master=self, value="Dedicated WPS Chrome not checked")

        self.create_widgets()
        self.load_initial_config()
        self.schedule_drain_logs()

    def create_widgets(self) -> None:
        root = ttk.Frame(self, padding=10)
        root.pack(fill="both", expand=True)
        root.columnconfigure(0, weight=1)
        root.rowconfigure(1, weight=4, minsize=DOCUMENT_CONFIGS_MIN_HEIGHT)
        root.rowconfigure(2, weight=1)

        file_frame = ttk.LabelFrame(root, text="Browser")
        file_frame.grid(row=0, column=0, sticky="ew")

        ttk.Label(file_frame, text="Browser mode").grid(row=0, column=0, padx=8, pady=8, sticky="w")
        mode_frame = ttk.Frame(file_frame)
        mode_frame.grid(row=0, column=1, padx=8, pady=8, sticky="ew")
        ttk.Radiobutton(
            mode_frame,
            text="Dedicated WPS Chrome (recommended)",
            variable=self.browser_mode,
            value="managed_cdp",
        ).pack(side="left")
        ttk.Radiobutton(
            mode_frame,
            text="One-time browser",
            variable=self.browser_mode,
            value="persistent",
        ).pack(side="left", padx=(16, 0))
        ttk.Radiobutton(
            mode_frame,
            text="Existing Chrome CDP",
            variable=self.browser_mode,
            value="cdp",
        ).pack(side="left", padx=(16, 0))

        ttk.Label(file_frame, text="CDP URL").grid(row=1, column=0, padx=8, pady=8, sticky="w")
        ttk.Entry(file_frame, textvariable=self.cdp_url).grid(row=1, column=1, padx=8, pady=8, sticky="ew")
        browser_actions = ttk.Frame(file_frame)
        browser_actions.grid(row=2, column=1, padx=8, pady=(0, 8), sticky="ew")
        self.start_browser_button = ttk.Button(
            browser_actions,
            text="Start/open dedicated Chrome",
            command=self.start_wps_browser,
        )
        self.start_browser_button.pack(side="left")
        self.check_browser_button = ttk.Button(
            browser_actions,
            text="Check status",
            command=self.check_wps_browser,
        )
        self.check_browser_button.pack(side="left", padx=(8, 0))
        ttk.Label(browser_actions, textvariable=self.browser_status, style="Muted.TLabel").pack(
            side="left", padx=(12, 0)
        )
        file_frame.columnconfigure(1, weight=1)

        mapping_frame = ttk.LabelFrame(root, text="Document configs")
        mapping_frame.grid(row=1, column=0, sticky="nsew", pady=(10, 0))

        self.mapping_toolbar = ttk.Frame(mapping_frame)
        self.mapping_toolbar.pack(fill="x", padx=8, pady=(8, 0))
        self.mapping_toolbar.columnconfigure(0, weight=1)

        self.mapping_primary_toolbar = ttk.Frame(self.mapping_toolbar)
        self.mapping_primary_toolbar.grid(row=0, column=0, sticky="w")
        primary_actions = [
            ("Add config", self.add_config),
            ("Edit", self.edit_config),
            ("Copy", self.copy_config),
            ("Remove", self.remove_config),
        ]
        for column, (text, command) in enumerate(primary_actions):
            ttk.Button(
                self.mapping_primary_toolbar,
                text=text,
                command=command,
            ).grid(row=0, column=column, padx=(0 if column == 0 else 4, 0))

        self.mapping_sort_toolbar = ttk.Frame(self.mapping_toolbar)
        self.mapping_sort_toolbar.grid(
            row=1, column=1, sticky="e", padx=(12, 0), pady=(6, 0)
        )
        ttk.Button(
            self.mapping_sort_toolbar, text="Move up", command=self.move_config_up
        ).pack(side="left")
        ttk.Button(
            self.mapping_sort_toolbar, text="Move down", command=self.move_config_down
        ).pack(side="left", padx=(4, 0))

        self.mapping_secondary_toolbar = ttk.Frame(self.mapping_toolbar)
        self.mapping_secondary_toolbar.grid(
            row=1, column=0, sticky="w", pady=(6, 0)
        )
        ttk.Button(
            self.mapping_secondary_toolbar, text="Import", command=self.import_config
        ).pack(side="left")
        ttk.Button(
            self.mapping_secondary_toolbar, text="Export", command=self.export_config
        ).pack(side="left", padx=(4, 0))

        tree_frame = ttk.Frame(mapping_frame)
        tree_frame.pack(fill="both", expand=True, padx=8, pady=8)

        columns = ("name", "source_type", "source", "kdocs_url", "region_count", "summary")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=8)
        self.tree.heading("name", text="Name")
        self.tree.heading("source_type", text="Source")
        self.tree.heading("source", text="Source file/URL")
        self.tree.heading("kdocs_url", text="Target KDocs URL")
        self.tree.heading("region_count", text="Regions")
        self.tree.heading("summary", text="First region")
        self.tree.column("name", width=180)
        self.tree.column("source_type", width=90)
        self.tree.column("source", width=280)
        self.tree.column("kdocs_url", width=280)
        self.tree.column("region_count", width=80)
        self.tree.column("summary", width=330)
        self.tree.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="left", fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.bind("<Double-1>", lambda _event: self.edit_config())
        self.tree.bind("<<TreeviewSelect>>", lambda _event: self.update_selection_status())

        self.execution_toolbar = ttk.Frame(mapping_frame)
        self.execution_toolbar.pack(fill="x", padx=8, pady=(0, 8))
        self.selection_status = tk.StringVar(value="Select a config to preview or write.")
        ttk.Label(
            self.execution_toolbar,
            textvariable=self.selection_status,
            style="Muted.TLabel",
        ).pack(side="left")
        ttk.Button(
            self.execution_toolbar, text="Write to WPS", command=self.write
        ).pack(side="right")
        ttk.Button(
            self.execution_toolbar, text="Preview", command=self.preview
        ).pack(side="right", padx=(0, 6))

        self.module_log_frame = ttk.LabelFrame(root, text="WPS operation log")
        self.module_log_frame.grid(row=2, column=0, sticky="nsew", pady=(10, 0))
        log_actions = ttk.Frame(self.module_log_frame)
        log_actions.pack(fill="x", padx=8, pady=(8, 0))
        ttk.Button(log_actions, text="Clear log", command=self.clear_log).pack(side="left")
        ttk.Button(log_actions, text="Save log", command=self.save_log).pack(side="left", padx=8)
        ttk.Button(log_actions, text="Open log window", command=self.open_log_window).pack(side="left")
        self.log_text = tk.Text(self.module_log_frame, height=8, wrap="word")
        self.log_text.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=8)
        log_scroll = ttk.Scrollbar(
            self.module_log_frame, orient="vertical", command=self.log_text.yview
        )
        log_scroll.pack(side="left", fill="y", pady=8)
        self.log_text.configure(yscrollcommand=log_scroll.set)

    def add_config(self) -> None:
        dialog = DocumentConfigDialog(
            self,
            "Add config",
            recent_urls=self.recent_urls,
            browser_mode=self.browser_mode.get(),
            cdp_url=self.cdp_url.get().strip(),
            log=self.log,
        )
        if dialog.result:
            self.configs.append(dialog.result)
            self.remember_mapping_urls()
            self.refresh_tree()
            self.save_current_config()

    def edit_config(self) -> None:
        selected = self.tree.selection()
        if not selected:
            return
        index = self.tree.index(selected[0])
        # Start (or reuse) the managed browser in parallel while the config
        # dialog opens, and preload this config's target document.
        self.start_wps_browser(initial_url=self.configs[index].kdocs_url)

        def autosave(config: DocumentConfig) -> None:
            self.configs[index] = config
            self.remember_mapping_urls()
            self.save_current_config()

        dialog = DocumentConfigDialog(
            self,
            "Edit config",
            self.configs[index],
            recent_urls=self.recent_urls,
            browser_mode=self.browser_mode.get(),
            cdp_url=self.cdp_url.get().strip(),
            log=self.log,
            autosave=autosave,
        )
        if dialog.result:
            self.configs[index] = dialog.result
            self.remember_mapping_urls()
            self.refresh_tree()
            self.save_current_config()

    def copy_config(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("Copy config", "Select one config row to copy.", parent=self)
            return
        index = self.tree.index(selected[0])
        source = self.configs[index]
        copied = DocumentConfig(
            name=f"{source.name} Copy",
            kdocs_url=source.kdocs_url,
            local_file=source.local_file,
            source_type=source.source_type,
            source_url=source.source_url,
            regions=[
                RegionRow(
                    source_sheet=region.source_sheet,
                    source_start=region.source_start,
                    source_end=region.source_end,
                    target_sheet=region.target_sheet,
                    target_start=region.target_start,
                    target_end=region.target_end,
                )
                for region in source.regions
            ],
        )
        self.configs.insert(index + 1, copied)
        self.refresh_tree()
        new_item = self.tree.get_children()[index + 1]
        self.tree.selection_set(new_item)
        self.tree.focus(new_item)
        self.tree.see(new_item)
        self.save_current_config()

    def remove_config(self) -> None:
        selected = self.tree.selection()
        if not selected:
            return
        count = len(selected)
        if not messagebox.askyesno(
            "Confirm remove",
            f"Remove {count} selected config(s)?",
            parent=self,
        ):
            return
        for item in reversed(selected):
            del self.configs[self.tree.index(item)]
        self.refresh_tree()
        self.save_current_config()

    def move_config_up(self) -> None:
        indexes = self.selected_config_indexes()
        if not indexes:
            return
        indexes = sorted(indexes)
        for index in indexes:
            if index == 0:
                continue
            self.configs[index - 1], self.configs[index] = self.configs[index], self.configs[index - 1]
        new_indexes = [max(0, index - 1) for index in indexes]
        self.refresh_tree()
        self.select_config_indexes(new_indexes)
        self.save_current_config()

    def move_config_down(self) -> None:
        indexes = self.selected_config_indexes()
        if not indexes:
            return
        indexes = sorted(indexes, reverse=True)
        last = len(self.configs) - 1
        for index in indexes:
            if index >= last:
                continue
            self.configs[index + 1], self.configs[index] = self.configs[index], self.configs[index + 1]
        new_indexes = [min(last, index + 1) for index in indexes]
        self.refresh_tree()
        self.select_config_indexes(new_indexes)
        self.save_current_config()

    def select_config_indexes(self, indexes: list[int]) -> None:
        children = self.tree.get_children()
        items = [children[index] for index in sorted(set(indexes)) if 0 <= index < len(children)]
        if not items:
            self.update_selection_status()
            return
        self.tree.selection_set(items)
        self.tree.focus(items[0])
        self.tree.see(items[0])
        self.update_selection_status()

    def update_selection_status(self) -> None:
        indexes = self.selected_config_indexes()
        if not indexes:
            self.selection_status.set("Select a config to preview or write.")
        elif len(indexes) == 1:
            self.selection_status.set(f"Selected: {self.configs[indexes[0]].name}")
        else:
            self.selection_status.set(f"{len(indexes)} configs selected.")

    def refresh_tree(self) -> None:
        for item in self.tree.get_children():
            self.tree.delete(item)
        for config in self.configs:
            first = config.regions[0] if config.regions else None
            summary = ""
            if first:
                try:
                    summary = (
                        f"{first.source_sheet}!{range_from_cells(first.source_start, first.source_end)} -> "
                        f"{first.target_sheet}!{range_from_cells(first.target_start, first.target_end)}"
                    )
                except Exception:
                    summary = (
                        f"{first.source_sheet}!{first.source_start}:{first.source_end} -> "
                        f"{first.target_sheet}!{first.target_start}:{first.target_end}"
                    )
            self.tree.insert(
                "",
                "end",
                values=(
                    config.name,
                    "KDocs" if config.source_type == "kdocs" else "Excel",
                    config.source_url if config.source_type == "kdocs" else config.local_file,
                    config.kdocs_url,
                    len(config.regions),
                    summary,
                ),
            )
        self.update_selection_status()

    def log(self, message: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_queue.put(f"[{timestamp}] {message}")

    def start_wps_browser(self, initial_url: str | None = None) -> None:
        self.browser_mode.set("managed_cdp")
        self.save_current_config()
        self._run_browser_action(start=True, initial_url=initial_url)

    def check_wps_browser(self) -> None:
        self._run_browser_action(start=False)

    def _run_browser_action(
        self, *, start: bool, initial_url: str | None = None
    ) -> None:
        if self.browser_worker and self.browser_worker.is_alive():
            return
        self.start_browser_button.configure(state="disabled")
        self.check_browser_button.configure(state="disabled")
        self.browser_status.set("Starting..." if start else "Checking...")
        cdp_url = self.cdp_url.get().strip()

        def worker() -> None:
            try:
                if start:
                    result = ensure_wps_dedicated_chrome(
                        cdp_url,
                        self.log,
                        initial_url=initial_url or "https://www.kdocs.cn/",
                        preload_url=bool(initial_url),
                    )
                    action = "started" if result.get("started") else "reused"
                    message = f"Dedicated WPS Chrome ready ({action})"
                else:
                    endpoint = try_resolve_cdp_endpoint(cdp_url)
                    if not endpoint:
                        raise RuntimeError("Dedicated WPS Chrome is not running")
                    message = "Dedicated WPS Chrome is running and reachable"
                self.browser_event_queue.put((True, message))
            except Exception as exc:
                self.browser_event_queue.put((False, str(exc)))

        self.browser_worker = self._start_tracked_thread(
            worker,
            name="wps-browser-start" if start else "wps-browser-check",
        )

    def _start_tracked_thread(
        self,
        target,
        *,
        args: tuple[Any, ...] = (),
        name: str,
    ) -> threading.Thread:
        thread: threading.Thread

        def tracked_target() -> None:
            try:
                target(*args)
            finally:
                if self.task_registry is not None:
                    self.task_registry.unregister_thread(thread)

        thread = threading.Thread(target=tracked_target, name=name, daemon=True)
        if self.task_registry is not None:
            self.task_registry.register_thread(thread)
        try:
            thread.start()
        except BaseException:
            if self.task_registry is not None:
                self.task_registry.unregister_thread(thread)
            raise
        return thread

    def detail_log(self, message: str) -> None:
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        now = datetime.now()
        log_path = LOG_DIR / f"wps_write_{now.strftime('%Y-%m-%d')}.log"
        line = f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] {message}\n"
        with log_path.open("a", encoding="utf-8") as handle:
            handle.write(line)

    def detail_log_prepared(self, action: str, prepared: list[PreparedMapping]) -> None:
        self.detail_log(f"{action}: {len(prepared)} region(s)")
        for item in prepared:
            sample = item.values[:3]
            self.detail_log(
                "REGION "
                f"index={item.index}; source={item.workbook_path}; "
                f"source_range={item.source_sheet}!{item.source_range}; "
                f"target={item.kdocs_url}; target_range={item.target_sheet}!{item.target_range}; "
                f"shape={item.rows}x{item.cols}; sample_first3={sample!r}"
            )

    def drain_logs(self) -> None:
        while True:
            try:
                success, status = self.browser_event_queue.get_nowait()
            except queue.Empty:
                break
            self.browser_status.set(status)
            self.start_browser_button.configure(state="normal")
            self.check_browser_button.configure(state="normal")
            self.log(("Browser ready: " if success else "Browser error: ") + status)
        while True:
            try:
                message = self.log_queue.get_nowait()
            except queue.Empty:
                break
            self.log_text.insert("end", f"{message}\n")
            self.log_text.see("end")
            if self.log_window_text is not None:
                try:
                    self.log_window_text.insert("end", f"{message}\n")
                    self.log_window_text.see("end")
                except tk.TclError:
                    self.log_window_text = None
                    self.log_window = None
        self.schedule_drain_logs()

    def schedule_drain_logs(self) -> None:
        try:
            self._drain_after_id = self.after(100, self.drain_logs)
        except tk.TclError:
            self._drain_after_id = None

    def destroy(self) -> None:
        if self._drain_after_id is not None:
            try:
                self.after_cancel(self._drain_after_id)
            except tk.TclError:
                pass
            self._drain_after_id = None
        super().destroy()

    def clear_log(self) -> None:
        self.log_text.delete("1.0", "end")
        if self.log_window_text is not None:
            try:
                self.log_window_text.delete("1.0", "end")
            except tk.TclError:
                self.log_window_text = None
                self.log_window = None

    def save_log(self) -> None:
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Save run log",
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
        )
        if not path:
            return
        Path(path).write_text(self.log_text.get("1.0", "end"), encoding="utf-8")

    def open_log_window(self) -> None:
        if self.log_window is not None:
            try:
                self.log_window.lift()
                self.log_window.focus()
                return
            except tk.TclError:
                self.log_window = None
                self.log_window_text = None

        window = tk.Toplevel(self)
        window.title("Run log monitor")
        window.geometry("900x520")
        center_window(window, 900, 520)
        text = tk.Text(window, wrap="word")
        text.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(window, orient="vertical", command=text.yview)
        scrollbar.pack(side="left", fill="y")
        text.configure(yscrollcommand=scrollbar.set)
        text.insert("end", self.log_text.get("1.0", "end"))
        text.see("end")

        def on_close() -> None:
            self.log_window = None
            self.log_window_text = None
            window.destroy()

        window.protocol("WM_DELETE_WINDOW", on_close)
        self.log_window = window
        self.log_window_text = text

    def load_initial_config(self) -> None:
        config = load_config()
        self.recent_urls = [url for url in config.get("recent_urls", []) if isinstance(url, str) and url.strip()]
        if config.get("kdocs_url"):
            self.recent_urls = update_recent_urls(self.recent_urls, config["kdocs_url"])
        saved_browser_mode = config.get("browser_mode", "managed_cdp")
        if saved_browser_mode == "persistent" and sys.platform == "darwin":
            saved_browser_mode = "managed_cdp"
        self.browser_mode.set(saved_browser_mode)
        self.cdp_url.set(config.get("cdp_url", "http://127.0.0.1:9222"))
        default_url = config.get("kdocs_url", "")
        default_file = config.get("excel_path", "")
        if isinstance(config.get("configs"), list):
            self.configs = [self.document_config_from_config(row) for row in config.get("configs", []) if isinstance(row, dict)]
        else:
            mappings = [
                self.mapping_from_config(row, default_url=default_url, default_file=default_file)
                for row in config.get("mappings", [])
                if isinstance(row, dict)
            ]
            self.configs = self.group_mappings_into_configs(mappings)
        self.remember_mapping_urls()
        self.refresh_tree()

    def document_config_from_config(self, row: dict[str, Any]) -> DocumentConfig:
        regions = []
        for region in row.get("regions", []):
            if not isinstance(region, dict):
                continue
            regions.append(
                RegionRow(
                    source_sheet=region.get("source_sheet", ""),
                    source_start=region.get("source_start", ""),
                    source_end=region.get("source_end", ""),
                    target_sheet=region.get("target_sheet", ""),
                    target_start=region.get("target_start", ""),
                    target_end=region.get("target_end", ""),
                )
            )
        return DocumentConfig(
            name=row.get("name", "") or Path(row.get("local_file", "")).stem or "Config",
            kdocs_url=row.get("kdocs_url", ""),
            local_file=row.get("local_file", ""),
            regions=regions,
            source_type=row.get("source_type", "excel"),
            source_url=row.get("source_url", ""),
        )

    def mapping_from_config(self, row: dict[str, Any], default_url: str = "", default_file: str = "") -> MappingRow:
        if "source_range" in row and ("source_start" not in row or "source_end" not in row):
            source_start, source_end = split_range(row.get("source_range", ""))
        else:
            source_start = row.get("source_start", "")
            source_end = row.get("source_end", "")

        target_start = row.get("target_start", "")
        target_end = row.get("target_end", "")
        if not target_end and target_start and "source_range" in row:
            source_start_for_shape, source_end_for_shape = split_range(row.get("source_range", ""))
            rows, cols = range_shape(source_start_for_shape, source_end_for_shape)
            target_range = target_range_from_start(target_start, rows, cols)
            _target_start, target_end = split_range(target_range)

        return MappingRow(
            kdocs_url=row.get("kdocs_url", "") or default_url,
            local_file=row.get("local_file", "") or default_file,
            source_sheet=row.get("source_sheet", ""),
            source_start=source_start,
            source_end=source_end,
            target_sheet=row.get("target_sheet", ""),
            target_start=target_start,
            target_end=target_end,
            source_type=row.get("source_type", "excel"),
            source_url=row.get("source_url", ""),
        )

    def group_mappings_into_configs(self, mappings: list[MappingRow]) -> list[DocumentConfig]:
        grouped: dict[tuple[str, str], list[RegionRow]] = {}
        for mapping in mappings:
            key = (mapping.kdocs_url, mapping.local_file)
            grouped.setdefault(key, []).append(
                RegionRow(
                    source_sheet=mapping.source_sheet,
                    source_start=mapping.source_start,
                    source_end=mapping.source_end,
                    target_sheet=mapping.target_sheet,
                    target_start=mapping.target_start,
                    target_end=mapping.target_end,
                )
            )
        configs = []
        for index, ((kdocs_url, local_file), regions) in enumerate(grouped.items(), start=1):
            name = Path(local_file).stem or f"Config {index}"
            configs.append(DocumentConfig(name=name, kdocs_url=kdocs_url, local_file=local_file, regions=regions))
        return configs

    def remember_mapping_urls(self) -> None:
        urls = self.recent_urls
        for config in self.configs:
            urls = update_recent_urls(urls, config.kdocs_url)
            if config.source_type == "kdocs":
                urls = update_recent_urls(urls, config.source_url)
        self.recent_urls = urls

    def save_current_config(self) -> None:
        self.remember_mapping_urls()
        save_config(self.current_runtime_config())

    def current_runtime_config(self) -> dict[str, Any]:
        return normalize_runtime_config(
            {
                "recent_urls": self.recent_urls,
                "browser_mode": self.browser_mode.get(),
                "cdp_url": self.cdp_url.get(),
                "configs": [
                    {
                        "name": config.name,
                        "source_type": config.source_type,
                        "source_url": config.source_url,
                        "kdocs_url": config.kdocs_url,
                        "local_file": config.local_file,
                        "regions": [region.__dict__ for region in config.regions],
                    }
                    for config in self.configs
                ],
            }
        )

    def apply_runtime_config(self, config: dict[str, Any]) -> None:
        runtime_config = normalize_runtime_config(config)
        self.recent_urls = runtime_config["recent_urls"]
        self.browser_mode.set(runtime_config["browser_mode"])
        self.cdp_url.set(runtime_config["cdp_url"])
        self.configs = [self.document_config_from_config(row) for row in runtime_config["configs"]]
        self.remember_mapping_urls()
        self.refresh_tree()
        self.save_current_config()

    def import_config(self) -> None:
        path_text = filedialog.askopenfilename(
            parent=self,
            title="Import WPS writer config",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
        )
        if not path_text:
            return
        try:
            runtime_config = load_config_export(Path(path_text))
            self.apply_runtime_config(runtime_config)
        except Exception as exc:
            messagebox.showerror("Import config failed", str(exc), parent=self)
            self.log(f"ERROR: Import config failed: {exc}")
            return
        self.log(f"Imported config: {Path(path_text).name}")
        messagebox.showinfo("Import config", "Config imported and saved.", parent=self)

    def export_config(self) -> None:
        default_name = f"wps_writer_config_export_{datetime.now().strftime('%Y%m%d')}.json"
        path_text = filedialog.asksaveasfilename(
            parent=self,
            title="Export WPS writer config",
            defaultextension=".json",
            initialfile=default_name,
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
        )
        if not path_text:
            return
        try:
            save_config_export(Path(path_text), self.current_runtime_config())
        except Exception as exc:
            messagebox.showerror("Export config failed", str(exc), parent=self)
            self.log(f"ERROR: Export config failed: {exc}")
            return
        self.log(f"Exported config: {path_text}")
        messagebox.showinfo("Export config", "Config exported.", parent=self)

    def selected_config_indexes(self) -> list[int]:
        selected = self.tree.selection()
        return sorted(self.tree.index(item) for item in selected)

    def flatten_configs(self, configs: list[DocumentConfig] | None = None) -> list[MappingRow]:
        mappings: list[MappingRow] = []
        for config in (configs if configs is not None else self.configs):
            for region in config.regions:
                mappings.append(
                    MappingRow(
                        kdocs_url=config.kdocs_url,
                        local_file=config.local_file,
                        source_sheet=region.source_sheet,
                        source_start=region.source_start,
                        source_end=region.source_end,
                        target_sheet=region.target_sheet,
                        target_start=region.target_start,
                        target_end=region.target_end,
                        source_type=config.source_type,
                        source_url=config.source_url,
                    )
                )
        return mappings

    def prepare_selected(self) -> list[PreparedMapping]:
        indexes = self.selected_config_indexes()
        if not indexes:
            raise ValueError("Select at least one config before previewing or writing.")
        configs = [self.configs[index] for index in indexes]
        return prepare_mappings(
            self.flatten_configs(configs),
            browser_mode=self.browser_mode.get(),
            cdp_url=self.cdp_url.get().strip(),
            log=lambda _message: None,
        )

    def preview(self) -> None:
        try:
            prepared = self.prepare_selected()
        except Exception as exc:
            messagebox.showerror("Preview failed", str(exc), parent=self)
            self.log(f"ERROR: Preview failed: {exc}")
            self.detail_log(f"PREVIEW ERROR: {exc}")
            self.detail_log(traceback.format_exc())
            return
        urls = len({item.kdocs_url for item in prepared})
        self.detail_log_prepared("PREVIEW OK", prepared)
        self.log(f"Preview OK: {len(prepared)} region(s), {urls} target document(s).")

    def write(self) -> None:
        if self.worker and self.worker.is_alive():
            messagebox.showinfo("Busy", "A write task is already running.", parent=self)
            return
        try:
            prepared = self.prepare_selected()
        except Exception as exc:
            messagebox.showerror("Write blocked", str(exc), parent=self)
            self.log(f"ERROR: Write blocked: {exc}")
            self.detail_log(f"WRITE BLOCKED: {exc}")
            self.detail_log(traceback.format_exc())
            return

        self.save_current_config()
        self.detail_log_prepared("WRITE PREPARED", prepared)
        self.worker = self._start_tracked_thread(
            self.write_worker,
            args=(prepared, self.browser_mode.get(), self.cdp_url.get().strip()),
            name="wps-write",
        )

    def write_worker(self, prepared: list[PreparedMapping], browser_mode: str, cdp_url: str) -> None:
        try:
            self.log(f"Write started: {len(prepared)} region(s).")
            self.detail_log(f"WRITE STARTED: browser_mode={browser_mode}; cdp_url={cdp_url}; regions={len(prepared)}")
            by_index = {item.index: item for item in prepared}
            all_ok = True

            groups: dict[str, list[PreparedMapping]] = {}
            for item in prepared:
                groups.setdefault(item.kdocs_url, []).append(item)

            for url, group in groups.items():
                self.detail_log(f"WRITE TARGET: url={url}; regions={len(group)}")
                result = write_to_wps(url, group, lambda _message: None, browser_mode=browser_mode, cdp_url=cdp_url)
                self.detail_log(f"WRITE TARGET RESULT: url={url}; saved={result.get('saved')}; returned={len(result.get('results', []))}")
                for entry in result.get("results", []):
                    item = by_index[entry["index"]]
                    mismatches = compare_matrix(item.values, entry.get("value"), item.rows, item.cols)
                    if mismatches:
                        all_ok = False
                        self.detail_log(
                            f"VERIFY MISMATCH: index={item.index}; target={item.target_sheet}!{item.target_range}; "
                            f"mismatches={mismatches!r}; read_value={entry.get('value')!r}"
                        )
                        self.log(f"#{item.index} verification failed: {item.target_sheet}!{item.target_range}")
                        for mismatch in mismatches:
                            self.log(f"  {mismatch}")
                    else:
                        self.detail_log(f"VERIFY OK: index={item.index}; target={item.target_sheet}!{item.target_range}")
            if all_ok:
                self.detail_log(f"WRITE OK: regions={len(prepared)}")
                self.log(f"Write OK: {len(prepared)} region(s) saved and verified.")
            else:
                self.detail_log("WRITE WARNING: verification mismatches detected")
                self.log("WARNING: Write finished, but read-back verification found differences.")
        except Exception as exc:
            self.log(f"ERROR: {exc}")
            self.log(traceback.format_exc())
            self.detail_log(f"WRITE ERROR: {exc}")
            self.detail_log(traceback.format_exc())


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Excel to WPS/KDocs Writer")
        self.geometry("1120x760")
        self.minsize(980, 640)
        self.writer_frame = WpsWriterFrame(self)
        self.writer_frame.pack(fill="both", expand=True)
        center_window(self, 1120, 760)


if __name__ == "__main__":
    if "--launch-browser" in sys.argv:
        launch_result = ensure_wps_dedicated_chrome(
            "http://127.0.0.1:9222", print
        )
        print(
            "WPS dedicated Chrome is ready "
            f"({'started' if launch_result.get('started') else 'already running'})."
        )
    else:
        app = App()
        app.mainloop()
