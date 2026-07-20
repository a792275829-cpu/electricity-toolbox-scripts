from __future__ import annotations

import argparse
import html
import json
import math
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

if __package__ in (None, ""):
    sys.path.append(str(Path(__file__).resolve().parents[1]))

from src.analyze_500kv_section_gaps import (
    MIN_SECTION_INTERVALS,
    match_price_rows,
    parse_topology_edges,
    read_price_workbook,
)


REGION_CENTERS = {
    "粤北片": (113.55, 24.80),
    "清远片": (113.05, 23.75),
    "广州北部片": (113.35, 23.35),
    "佛山片": (113.05, 23.05),
    "广州南部片": (113.55, 22.95),
    "肇云片": (112.25, 23.25),
    "江门片": (112.75, 22.45),
    "阳江片": (111.75, 21.85),
    "茂湛片": (110.45, 21.35),
    "中珠片": (113.45, 22.25),
    "南粤柔直": (113.55, 21.95),
    "东莞片": (113.85, 22.95),
    "深圳片": (114.25, 22.55),
    "惠州片": (114.35, 23.05),
    "梅州河源片": (115.40, 24.05),
    "粤东片": (116.60, 23.55),
}

NODE_OVERRIDES = {
    "丹霞": (113.72, 25.05),
    "曲江": (113.58, 24.75),
    "库湾": (113.10, 24.95),
    "贤令山": (112.95, 24.25),
    "东坡": (113.25, 24.55),
    "振兴": (113.08, 24.10),
    "从西": (112.95, 23.90),
    "科北": (113.15, 23.65),
    "渡水": (113.35, 23.85),
    "凤安": (113.55, 23.65),
    "花都": (113.20, 23.40),
    "北郊": (113.25, 23.25),
    "木棉": (113.45, 23.25),
    # 以下公开地图可核实点使用实际站址；其余点仍按公开工程地点/片区相对方位近似。
    "罗洞": (112.99281, 23.15686),
    "西江": (112.80, 23.05),
    "凤城": (113.08991, 22.83043),
    "广南": (113.20, 22.95),
    "顺德": (113.25, 22.75),
    "楚庭": (113.26404, 22.99141),
    "狮洋": (113.70, 22.72),
    "玉城": (111.95, 23.35),
    "砚都": (112.25, 23.05),
    "沧江": (112.55, 22.90),
    "侨乡": (112.75, 22.65),
    "江门": (113.05, 22.30),
    "卧龙": (111.85, 21.95),
    "安澜": (111.65, 21.75),
    "蝶岭": (111.35, 21.90),
    "茂名": (110.95, 21.65),
    "鹅凰B": (110.65, 21.40),
    "立安": (110.35, 21.25),
    "港城": (110.25, 21.05),
    "东海岛": (110.55, 20.95),
    "芷寮": (110.85, 21.35),
    "香山": (113.40, 22.30),
    "主峰": (113.55, 22.25),
    "桂山": (113.80, 22.10),
    "金鼎": (113.60, 22.40),
    "加林": (113.25, 22.25),
    "文山": (113.10, 22.20),
    "南粤": (113.90, 21.95),
    "水乡": (113.65, 22.90),
    "崇焕": (113.85, 23.05),
    "莞城": (113.80, 22.95),
    "横沥": (114.05, 23.00),
    "宝安": (113.90, 22.60),
    "盛丰": (114.15, 22.65),
    "鹏城": (114.25, 22.55),
    "深圳": (114.10, 22.50),
    "现代": (114.35, 22.60),
    "东方": (114.45, 22.52),
    "紫荆": (114.55, 22.62),
    "屹百": (114.65, 22.72),
    "博罗": (114.18159, 23.24922),
    "穗东": (114.05, 23.00),
    "惠州": (114.55, 23.05),
    "崇文": (114.75, 23.25),
    "上寨": (115.25, 24.10),
    "敬州": (115.70, 24.25),
    "嘉应": (116.10, 24.00),
    "承龙开关站": (116.30, 23.75),
    "陆丰": (115.65, 22.95),
    "吉康": (115.95, 23.25),
    "征程": (116.15, 23.35),
    "福园": (116.45, 23.35),
    "祯州": (116.75, 23.20),
    "岐山": (116.65, 23.55),
    "榕江": (116.35, 23.70),
    "盘龙": (116.85, 23.65),
    "汕头": (116.95, 23.40),
    "韩江": (117.10, 23.65),
}

# 坐标来源等级会直接显示在节点悬浮提示中，避免把示意坐标误当成测绘坐标。
VERIFIED_NODE_SOURCES = {
    "罗洞": "OpenStreetMap公开站址",
    "凤城": "OpenStreetMap公开工程站址",
    "楚庭": "广州市发改委公开站址（番禺石壁二村）",
    "博罗": "OpenStreetMap公开站址",
}

REGION_COLORS = {
    "粤北、清远": "#38bdf8", "广佛、江门、珠海": "#8b5cf6", "东莞、深圳、惠州": "#14b8a6",
    "粤东": "#f59e0b", "粤西": "#f97316",
    "粤北片": "#6b8e23", "清远片": "#2e8b57", "广州北部片": "#008b8b",
    "佛山片": "#3b82c4", "广州南部片": "#5b6fd8", "肇云片": "#8467c4",
    "江门片": "#a05fb4", "阳江片": "#c35a88", "茂湛片": "#c7654c",
    "中珠片": "#d08b32", "南粤柔直": "#9a7b2f", "东莞片": "#4195a6",
    "深圳片": "#2878b8", "惠州片": "#467aa1", "梅州河源片": "#7b6aa8",
    "粤东片": "#a45b72", "未分片": "#7b8794",
}

TOPOLOGY_GROUP_REGIONS = {
    "粤北、清远": "粤北、清远",
    "广佛、江门、珠海": "广佛、江门、珠海",
    "东莞、深圳、惠州": "东莞、深圳、惠州",
    "粤东": "粤东",
    "粤西": "粤西",
    "江门、粤西": "粤西",
}

# 由公开广东省行政边界简化而来，保留主要海岸线和省界转折，取代原先的示意多边形。
GUANGDONG_OUTLINE = [
    (113.0872, 22.1261), (113.2668, 21.8716), (113.5658, 22.0727),
    (113.6699, 22.4356), (113.5687, 22.4119), (113.7405, 22.5343),
    (113.6977, 22.7374), (113.8918, 22.4426), (114.3983, 22.6028),
    (114.5069, 22.4386), (114.6278, 22.5030), (114.5145, 22.6607),
    (114.7096, 22.7877), (114.7474, 22.5816), (114.8836, 22.5402),
    (115.1986, 22.8219), (115.5742, 22.6505), (115.6324, 22.8600),
    (115.8187, 22.7314), (116.5056, 22.9309), (116.5664, 23.1343),
    (116.8147, 23.2076), (116.8992, 23.5189), (117.1929, 23.5617),
    (116.9620, 23.8614), (116.9987, 24.1790), (116.7583, 24.5466),
    (116.8012, 24.6783), (116.5301, 24.6049), (116.3948, 24.8779),
    (116.2513, 24.7934), (115.8973, 24.9371), (115.7571, 24.7499),
    (115.8441, 24.5627), (115.6861, 24.5470), (115.4129, 24.7929),
    (114.4288, 24.4863), (114.1690, 24.6897), (114.7356, 25.1220),
    (114.7429, 25.2745), (114.5404, 25.4170), (114.0399, 25.2507),
    (113.9979, 25.4442), (113.5860, 25.3074), (113.3033, 25.5158),
    (113.1520, 25.4922), (112.8518, 25.3334), (113.0345, 25.2014),
    (113.0119, 24.9461), (112.8735, 24.8967), (112.6598, 25.1328),
    (112.1873, 25.1831), (112.1711, 24.8627), (111.9275, 24.6294),
    (112.0618, 24.3682), (111.8779, 24.2289), (111.9403, 23.9877),
    (111.8106, 23.8070), (111.6549, 23.8333), (111.6653, 23.7000),
    (111.3997, 23.4694), (111.3587, 22.8892), (110.7602, 22.5808),
    (110.6835, 22.4734), (110.7897, 22.2867), (110.6789, 22.1727),
    (110.3485, 22.1959), (110.3876, 21.8904), (109.9862, 21.8791),
    (109.8983, 21.6496), (109.7664, 21.6680), (109.7856, 21.4569),
    (109.9047, 21.4301), (109.7575, 21.3468), (109.6549, 20.9035),
    (109.9135, 20.2194), (110.3275, 20.2516), (110.5458, 20.4275),
    (110.3937, 20.8167), (110.1777, 20.9071), (110.4224, 21.1907),
    (111.0613, 21.4492), (111.8628, 21.5570), (112.1365, 21.7938),
    (112.2631, 21.6935), (112.6394, 21.7560), (112.8004, 21.9248),
    (112.9445, 21.8421), (113.0872, 22.1261),
]

LON_MIN, LON_MAX = 109.45, 117.45
LAT_MIN, LAT_MAX = 20.05, 25.65
SVG_W, SVG_H = 1500, 1020
DIAGRAM_PADDING_X = 92
DIAGRAM_PADDING_Y = 50


def build_map_html(
    topology_path: Path,
    original_price_path: Path,
    processed_price_path: Path,
    analysis_path: Path,
    output_path: Path,
    node_positions_path: Path | None = None,
) -> None:
    topology_path = Path(topology_path)
    original_price_path = Path(original_price_path)
    processed_price_path = Path(processed_price_path)
    analysis_path = Path(analysis_path)
    output_path = Path(output_path)

    edges, skipped_duplicates = parse_topology_edges(topology_path)
    time_headers, price_rows = read_price_workbook(processed_price_path)
    topology_nodes = sorted({edge.a for edge in edges} | {edge.b for edge in edges})
    matched_rows = match_price_rows(topology_nodes, price_rows)
    edge_summary = _read_sheet_dicts(analysis_path, "断面汇总")
    interval_details = _read_sheet_dicts(analysis_path, "逐时段明细")
    group_summary = _read_sheet_dicts(analysis_path, "片区汇总")
    node_regions = _parse_node_regions(topology_path)
    for edge in edges:
        inferred_region = TOPOLOGY_GROUP_REGIONS.get(edge.group)
        if inferred_region:
            node_regions.setdefault(edge.a, inferred_region)
            node_regions.setdefault(edge.b, inferred_region)
    raw_stats = _workbook_stats(original_price_path)
    processed_stats = _workbook_stats(processed_price_path)

    discovered_positions_path = (
        Path(node_positions_path)
        if node_positions_path is not None
        else topology_path.parent / "data" / "500kv_node_positions.json"
    )
    if discovered_positions_path.exists():
        coordinates, coordinate_metadata = _build_diagram_coordinates(
            topology_nodes,
            discovered_positions_path,
        )
    else:
        coordinates = _build_coordinates(topology_nodes, node_regions)
        coordinate_metadata = {
            "coordinate_mode": "geographic-fallback",
            "coordinate_source": "公开站址与片区相对方位",
            "node_positions_path": "",
        }
    severity_by_edge = _build_edge_payloads(edges, edge_summary, interval_details, coordinates)
    blocked_nodes = {edge["a"] for edge in severity_by_edge if edge["is_section"]} | {
        edge["b"] for edge in severity_by_edge if edge["is_section"]
    }
    nodes = _build_node_payloads(topology_nodes, node_regions, coordinates, matched_rows, time_headers, blocked_nodes)
    regions = _build_region_payloads(nodes)

    data = {
        "nodes": nodes,
        "regions": regions,
        "edges": severity_by_edge,
        "time_headers": time_headers,
        "group_summary": group_summary,
        "metadata": {
            "original_path": str(original_price_path),
            "processed_path": str(processed_price_path),
            "analysis_path": str(analysis_path),
            "topology_path": str(topology_path),
            "original_rows": raw_stats["rows"],
            "processed_rows": processed_stats["rows"],
            "skipped_duplicate_edges": skipped_duplicates,
            "missing_price_nodes": [node for node in topology_nodes if node not in matched_rows],
            "min_section_intervals": MIN_SECTION_INTERVALS,
            "severity_basis": "intervals_and_max_abs_diff",
            "layout_canvas_width": SVG_W,
            "layout_canvas_height": SVG_H,
            **coordinate_metadata,
        },
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(_render_html(data), encoding="utf-8")


def _read_sheet_dicts(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value) for value in next(rows)]
    except StopIteration:
        return []

    records = []
    for row in rows:
        if all(value is None for value in row):
            continue
        records.append({header: _clean_cell(value) for header, value in zip(headers, row)})
    return records


def _parse_node_regions(topology_path: Path) -> dict[str, str]:
    text = Path(topology_path).read_text(encoding="utf-8")
    lines = text.splitlines()
    in_section = False
    current_region = ""
    mapping = {}
    for raw_line in lines:
        line = raw_line.strip()
        if not in_section:
            in_section = bool(re.match(r"^##\s+1\.", line))
            continue
        if re.match(r"^##\s+2\.", line):
            break
        if not line:
            continue
        if line.startswith("### "):
            current_region = line[4:].strip()
            continue
        if current_region:
            for node in re.split(r"[，,、]", line):
                node = node.strip()
                if node and not node.startswith("原 "):
                    mapping[node] = current_region
    return mapping


def _workbook_stats(path: Path) -> dict[str, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = 0
    for sheet in workbook.worksheets:
        rows += max(sheet.max_row - 1, 0)
    return {"rows": rows, "sheets": len(workbook.worksheets)}


def _build_coordinates(nodes: list[str], node_regions: dict[str, str]) -> dict[str, dict[str, float]]:
    region_counts: dict[str, int] = {}
    coordinates = {}
    for node in nodes:
        if node in NODE_OVERRIDES:
            lon, lat = NODE_OVERRIDES[node]
        else:
            region = node_regions.get(node, "")
            base_lon, base_lat = REGION_CENTERS.get(region, (113.6, 23.2))
            index = region_counts.get(region, 0)
            region_counts[region] = index + 1
            angle = index * 2.3999632297
            radius = 0.10 + 0.035 * index
            lon = base_lon + math.cos(angle) * radius
            lat = base_lat + math.sin(angle) * radius
        x, y = _project(lon, lat)
        coordinates[node] = {"lon": lon, "lat": lat, "x": x, "y": y}
    return coordinates


def _build_diagram_coordinates(
    nodes: list[str],
    positions_path: Path,
) -> tuple[dict[str, dict[str, object]], dict[str, object]]:
    payload = json.loads(Path(positions_path).read_text(encoding="utf-8"))
    image = payload.get("image", {})
    image_width = float(image.get("width", 0))
    image_height = float(image.get("height", 0))
    if image_width <= 0 or image_height <= 0:
        raise ValueError(f"锚点文件缺少有效接线图尺寸: {positions_path}")

    records = {str(record.get("name", "")): record for record in payload.get("nodes", [])}
    missing = [
        node
        for node in nodes
        if node not in records
        or records[node].get("status") != "verified"
        or not records[node].get("anchor")
    ]
    if missing:
        raise ValueError(f"以下拓扑节点缺少已确认接线图锚点: {', '.join(missing)}")

    scale = min(
        (SVG_W - DIAGRAM_PADDING_X * 2) / image_width,
        (SVG_H - DIAGRAM_PADDING_Y * 2) / image_height,
    )
    offset_x = (SVG_W - image_width * scale) / 2
    offset_y = (SVG_H - image_height * scale) / 2
    coordinates: dict[str, dict[str, object]] = {}
    for node in nodes:
        anchor = records[node]["anchor"]
        diagram_x = float(anchor["x"])
        diagram_y = float(anchor["y"])
        coordinates[node] = {
            "x": round(offset_x + diagram_x * scale, 2),
            "y": round(offset_y + diagram_y * scale, 2),
            "diagram_x": round(diagram_x, 2),
            "diagram_y": round(diagram_y, 2),
            "position_label": f"接线图锚点 {diagram_x:.0f}, {diagram_y:.0f}",
        }

    return coordinates, {
        "coordinate_mode": "wiring-diagram",
        "coordinate_source": str(image.get("path", "接线图.png")),
        "node_positions_path": str(positions_path),
        "diagram_width": image_width,
        "diagram_height": image_height,
        "diagram_scale": round(scale, 8),
    }


def _build_node_payloads(
    nodes: list[str],
    node_regions: dict[str, str],
    coordinates: dict[str, dict[str, float]],
    matched_rows: dict[str, dict[str, object]],
    time_headers: list[str],
    blocked_nodes: set[str],
) -> list[dict[str, object]]:
    payload = []
    for node in nodes:
        price_row = matched_rows.get(node)
        coordinate = coordinates[node]
        prices = []
        raw_name = ""
        if price_row:
            raw_name = str(price_row["raw_name"])
            prices = [price_row["prices"].get(time_header) for time_header in time_headers]
        if "diagram_x" in coordinate:
            location_accuracy = "已确认接线图锚点"
            location_source = "人工校准接线图节点位置"
        else:
            location_accuracy = "公开站址" if node in VERIFIED_NODE_SOURCES else "片区校核近似"
            location_source = VERIFIED_NODE_SOURCES.get(node, "公开工程地点与片区相对方位")
            coordinate = {
                **coordinate,
                "position_label": f"{float(coordinate['lon']):.5f}, {float(coordinate['lat']):.5f}",
            }
        payload.append(
            {
                "name": node,
                "region": node_regions.get(node, "未分片"),
                "raw_name": raw_name,
                "prices": prices,
                "blocked": node in blocked_nodes,
                "location_accuracy": location_accuracy,
                "location_source": location_source,
                **coordinate,
            }
        )
    return payload


def _build_region_payloads(nodes: list[dict[str, object]]) -> list[dict[str, object]]:
    grouped: dict[str, list[dict[str, object]]] = {}
    for node in nodes:
        grouped.setdefault(str(node["region"]), []).append(node)

    payload = []
    for region, members in sorted(grouped.items()):
        xs = [float(node["x"]) for node in members]
        ys = [float(node["y"]) for node in members]
        payload.append(
            {
                "name": region,
                "color": REGION_COLORS.get(region, REGION_COLORS["未分片"]),
                "count": len(members),
                "x": round(sum(xs) / len(xs), 2),
                "y": round(sum(ys) / len(ys), 2),
                "rx": round(max((max(xs) - min(xs)) / 2 + 30, 34), 2),
                "ry": round(max((max(ys) - min(ys)) / 2 + 24, 28), 2),
            }
        )
    return payload


def _build_edge_payloads(
    edges,
    edge_summary: list[dict[str, object]],
    interval_details: list[dict[str, object]],
    coordinates: dict[str, dict[str, float]],
) -> list[dict[str, object]]:
    summary_by_edge = {
        str(row["edge"]): row
        for row in edge_summary
        if int(float(row.get("intervals", 0))) >= MIN_SECTION_INTERVALS
    }
    detail_counts = {}
    for row in interval_details:
        detail_counts[str(row["edge"])] = detail_counts.get(str(row["edge"]), 0) + 1

    gap_scores = [float(row["max_abs_diff"]) for row in summary_by_edge.values()]
    interval_scores = [float(row["intervals"]) for row in summary_by_edge.values()]

    def normalize(value: float, values: list[float]) -> float:
        if not values or max(values) == min(values):
            return 1.0
        return (value - min(values)) / (max(values) - min(values))

    payload = []
    for edge in edges:
        edge_name = f"{edge.a}-{edge.b}"
        summary = summary_by_edge.get(edge_name)
        score = 0.0
        severity = 0.0
        gap_severity = 0.0
        interval_severity = 0.0
        if summary:
            score = float(summary["max_abs_diff"])
            gap_severity = normalize(score, gap_scores)
            interval_severity = normalize(float(summary["intervals"]), interval_scores)
            severity = 0.15 + 0.85 * (gap_severity + interval_severity) / 2

        payload.append(
            {
                "group": edge.group,
                "edge": edge_name,
                "a": edge.a,
                "b": edge.b,
                "x1": coordinates[edge.a]["x"],
                "y1": coordinates[edge.a]["y"],
                "x2": coordinates[edge.b]["x"],
                "y2": coordinates[edge.b]["y"],
                "is_section": bool(summary),
                "intervals": int(float(summary["intervals"])) if summary else 0,
                "windows": summary.get("windows", "") if summary else "",
                "max_abs_diff": float(summary["max_abs_diff"]) if summary else 0.0,
                "max_time": summary.get("max_time", "") if summary else "",
                "max_low": summary.get("max_low", "") if summary else "",
                "max_high": summary.get("max_high", "") if summary else "",
                "main_direction": summary.get("main_direction", "") if summary else "",
                "direction_counts": summary.get("direction_counts", "") if summary else "",
                "detail_count": detail_counts.get(edge_name, 0),
                "gap_severity": gap_severity,
                "interval_severity": interval_severity,
                "severity_score": severity,
                "severity": severity,
            }
        )
    return payload


def _clean_cell(value):
    if value is None:
        return ""
    return value


def _project(lon: float, lat: float) -> tuple[float, float]:
    x = 84 + (lon - LON_MIN) / (LON_MAX - LON_MIN) * (SVG_W - 168)
    y = 72 + (LAT_MAX - lat) / (LAT_MAX - LAT_MIN) * (SVG_H - 144)
    return round(x, 2), round(y, 2)


def _outline_points() -> str:
    return " ".join(f"{x},{y}" for x, y in (_project(lon, lat) for lon, lat in GUANGDONG_OUTLINE))


def _render_html(data: dict[str, object]) -> str:
    json_data = json.dumps(data, ensure_ascii=False, allow_nan=False)
    outline = _outline_points()
    diagram_mode = data.get("metadata", {}).get("coordinate_mode") == "wiring-diagram"
    artifact_title = "广东电网 500kV 接线拓扑与电价分析" if diagram_mode else "广东省电网 500kV 节点断面地图"
    map_label = "广东500kV接线拓扑与节点电价图" if diagram_mode else "广东500kV节点断面地图"
    map_layer_title = "接线拓扑" if diagram_mode else "空间分布"
    if diagram_mode:
        substrate = f"""
            <defs>
              <pattern id="schematicGrid" width="40" height="40" patternUnits="userSpaceOnUse">
                <path d="M 40 0 L 0 0 0 40" class="schematic-grid"></path>
              </pattern>
            </defs>
            <rect class="schematic-field" x="24" y="24" width="{SVG_W - 48}" height="{SVG_H - 48}" rx="28"></rect>
            <rect class="schematic-grid-field" x="24" y="24" width="{SVG_W - 48}" height="{SVG_H - 48}" rx="28"></rect>
        """
    else:
        substrate = f'<polygon class="province" points="{outline}"></polygon><polyline class="coast" points="{outline}"></polyline>'
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{artifact_title}</title>
  <style>
    :root {{
      color-scheme: dark;
      --ink: #e9f3ff;
      --muted: #8296ad;
      --line: rgba(118, 177, 223, .16);
      --panel: #0c1726;
      --soft: #111f31;
      --accent: #ff4d6d;
      --blue: #34b7ff;
      --navy: #eef8ff;
      --cyan: #38d6ff;
      --shadow: 0 22px 55px rgba(0, 8, 22, .42);
    }}
    body[data-theme="light"] {{
      color-scheme: light; --ink:#172033; --muted:#68758a; --line:#dce3ec;
      --panel:#fff; --soft:#f5f7fb; --accent:#c92235; --blue:#1f63d3;
      --navy:#0e2747; --cyan:#1f63d3; --shadow:0 16px 40px rgba(20,35,55,.09);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      color: var(--ink);
      background: #07111f;
      background-image: linear-gradient(rgba(39,142,204,.035) 1px, transparent 1px), linear-gradient(90deg, rgba(39,142,204,.035) 1px, transparent 1px);
      background-size: 32px 32px;
    }}
    body[data-theme="light"] {{ background: #f2f5f9; }}
    .app {{
      min-height: 100vh;
      display: grid;
      grid-template-columns: minmax(0, 1fr) 420px;
    }}
    .map-pane {{
      padding: 14px 18px 18px;
      min-width: 0;
    }}
    .title-row {{
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 16px;
      margin-bottom: 10px;
    }}
    h1 {{
      margin: 0;
      font-size: clamp(24px, 2.2vw, 34px);
      line-height: 1.2;
      letter-spacing: -.03em;
      color: var(--navy);
    }}
    .eyebrow {{
      color: var(--blue);
      font-size: 12px;
      font-weight: 800;
      letter-spacing: .12em;
      margin-bottom: 6px;
    }}
    .header-actions {{ display:flex; gap:8px; flex-wrap:wrap; justify-content:flex-end; }}
    .icon-button {{
      min-height: 38px; border: 1px solid var(--line); border-radius: 10px; padding: 0 12px;
      background: color-mix(in srgb, var(--panel) 86%, transparent); color: var(--ink); cursor: pointer;
      font-weight: 700; backdrop-filter: blur(14px); transition: .2s ease;
    }}
    .icon-button:hover {{ border-color: var(--cyan); color: var(--cyan); transform: translateY(-1px); }}
    .summary-grid {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 8px;
      margin-bottom: 8px;
    }}
    .summary-card {{
      min-width: 0;
      padding: 8px 12px;
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 12px;
      box-shadow: 0 5px 18px rgba(29, 50, 78, .04);
      position: relative; overflow: hidden;
    }}
    .summary-card::after {{ content:""; position:absolute; inset:auto 0 0; height:2px; background:linear-gradient(90deg,transparent,var(--cyan),transparent); opacity:.55; }}
    .summary-card .k {{ color: var(--muted); font-size: 11px; }}
    .summary-card .v {{ margin-top: 1px; font-size: 18px; font-weight: 800; color: var(--navy); line-height:1.15; }}
    .summary-card.alert .v {{ color: var(--accent); }}
    .summary-card .unit {{ font-size: 11px; font-weight: 600; color: var(--muted); margin-left: 3px; }}
    .map-toolbar {{
      min-height: 48px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      padding: 10px 12px;
      border-bottom: 1px solid var(--line);
      background: color-mix(in srgb, var(--panel) 92%, transparent);
      backdrop-filter: blur(16px);
    }}
    .map-toolbar strong {{ font-size: 13px; color: var(--navy); }}
    .map-hint {{ color: var(--muted); font-size: 12px; }}
    .layer-toggles {{ display: flex; gap: 6px; flex-wrap: wrap; }}
    .node-search {{
      width: 190px; min-height: 32px; border:1px solid var(--line); border-radius:8px;
      padding:0 10px; color:var(--ink); background:var(--soft); outline:none; font-size:12px;
    }}
    .node-search:focus {{ border-color:var(--cyan); box-shadow:0 0 0 3px rgba(56,214,255,.10); }}
    .toggle {{
      border: 1px solid var(--line); background: var(--soft); color: var(--ink);
      border-radius: 8px; padding: 6px 9px; font-size: 12px; cursor: pointer;
    }}
    .toggle[aria-pressed="true"] {{ color: var(--cyan); border-color: color-mix(in srgb, var(--cyan) 58%, transparent); box-shadow:0 0 16px rgba(52,183,255,.12); }}
    .toggle[aria-pressed="false"] {{ color: var(--muted); background: var(--soft); }}
    .is-hidden {{ display: none; }}
    .map-shell {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 14px;
      overflow: hidden;
      position: relative;
      min-height: calc(100vh - 170px);
      box-shadow: var(--shadow);
    }}
    .zoom-controls {{
      position: absolute;
      top: 62px;
      left: 12px;
      display: flex;
      gap: 6px;
      padding: 6px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: color-mix(in srgb, var(--panel) 90%, transparent);
      z-index: 3;
    }}
    .zoom-controls button {{
      width: 34px;
      height: 34px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: var(--soft);
      color: var(--ink);
      font-size: 18px;
      font-weight: 700;
      cursor: pointer;
      line-height: 1;
    }}
    .zoom-controls button:hover {{
      background: color-mix(in srgb, var(--blue) 18%, var(--soft));
    }}
    svg {{
      display: block;
      width: 100%;
      height: calc(100vh - 218px);
      min-height: 590px;
      background: #081523;
    }}
    #map {{
      cursor: default;
      touch-action: none;
      overscroll-behavior: contain;
    }}
    #map.is-panning {{ cursor: grabbing; }}
    .province {{
      fill: #10283a;
      stroke: #2b7591;
      stroke-width: 1.4;
    }}
    .coast {{
      fill: none;
      stroke: #27566f;
      stroke-width: 1;
      stroke-dasharray: 4 7;
    }}
    .schematic-field {{ fill:#091a2a; stroke:#1b4a66; stroke-width:1.2; }}
    .schematic-grid-field {{ fill:url(#schematicGrid); opacity:.58; pointer-events:none; }}
    .schematic-grid {{ fill:none; stroke:#16425b; stroke-width:.7; opacity:.32; }}
    .region-area {{
      fill-opacity: .045;
      stroke-width: 1.3;
      stroke-dasharray: 7 6;
      pointer-events: none;
    }}
    .region-label {{
      font-size: 11px;
      font-weight: 800;
      fill: #9fb6ca;
      opacity: .72;
      text-anchor: middle;
      paint-order: stroke;
      stroke: #081523;
      stroke-width: 4px;
      pointer-events: none;
    }}
    .edge {{
      stroke: #4a7188;
      stroke-width: 1.15;
      opacity: 0.32;
      pointer-events: stroke;
      stroke-dasharray: none;
      transition: opacity .16s ease, stroke-width .16s ease, stroke .16s ease;
    }}
    .edge-hit {{
      fill: none;
      stroke: transparent;
      stroke-width: 18;
      pointer-events: stroke;
      cursor: pointer;
      vector-effect: non-scaling-stroke;
    }}
    .edge-hit:focus {{ outline: none; }}
    .edge-hit:focus-visible + .edge {{
      opacity: 1;
      stroke: #38a7d1 !important;
      stroke-width: 3.2 !important;
      filter: drop-shadow(0 0 4px rgba(56,214,255,.55));
    }}
    .edge-hit.section-hit:focus-visible + .edge.section {{
      stroke: #8b0000 !important;
      stroke-width: 10 !important;
    }}
    .edge:not(.section).hover, .edge:not(.section).active {{
      opacity: 1;
      stroke: #38a7d1 !important;
      stroke-width: 3.2 !important;
    }}
    .edge.section {{
      opacity: 1;
      cursor: pointer;
      stroke-linecap: round;
      stroke-dasharray: none;
      filter: drop-shadow(0 1px 2px rgba(120, 0, 0, .30));
    }}
    .edge.section:hover, .edge.section.active {{
      stroke: #8b0000 !important;
      stroke-width: 10 !important;
    }}
    .edge.section:focus {{
      outline: none;
    }}
    .edge.section:focus-visible {{
      stroke: #8b0000 !important;
      stroke-width: 10 !important;
    }}
    .edge.connected {{
      opacity: .95;
      stroke: #38d6ff !important;
      stroke-width: 3.2 !important;
      stroke-dasharray: none;
      filter: drop-shadow(0 0 4px rgba(56,214,255,.55));
    }}
    .edge.section.connected {{ stroke: var(--accent) !important; }}
    .edge.context-dimmed {{ opacity: .055; }}
    .node {{
      fill: #b9eaff;
      stroke: #62b7d9;
      stroke-width: 0.9;
      opacity: .72;
      transition: .16s ease;
    }}
    .node-group {{ cursor:pointer; outline:none; }}
    .node-hit {{ fill:transparent; stroke:none; pointer-events:all; }}
    .node-group:hover .node, .node-group:focus-visible .node, .node-group.active .node {{
      opacity:1; stroke:var(--cyan)!important; stroke-width:3.2; filter:drop-shadow(0 0 8px rgba(56,214,255,.85));
    }}
    .node-group.neighbor .node {{ opacity:1; stroke:#7fe4ff!important; stroke-width:2.1; }}
    .node-group.context-dimmed {{ opacity:.24; }}
    .node-group.search-dimmed {{ opacity:.10; }}
    .node-group.search-match .node {{ opacity:1; stroke:#ffe66d!important; stroke-width:3.2; filter:drop-shadow(0 0 8px rgba(255,230,109,.75)); }}
    .node.blocked {{
      fill: #ffe5e8;
      stroke: #ff3158;
      stroke-width: 2.8;
      opacity: 1;
      filter: drop-shadow(0 1px 3px rgba(120,0,0,.35));
    }}
    .node.missing {{
      fill: #f7f7f7;
      stroke: #a8b0b8;
      stroke-dasharray: 2 2;
    }}
    .label {{
      font-size: 10px;
      font-weight: 600;
      fill: #596b7d;
      paint-order: stroke;
      stroke: #081523;
      stroke-width: 3px;
      pointer-events: all;
      cursor: pointer;
      vector-effect: non-scaling-stroke;
    }}
    .label-leader {{
      fill: none;
      stroke: #7890a3;
      stroke-width: .8;
      stroke-linecap: round;
      opacity: .58;
      pointer-events: none;
      vector-effect: non-scaling-stroke;
    }}
    .label-leader.is-idle {{ display:none; }}
    .blocked-label {{
      font-size: 13px;
      font-weight: 800;
      fill: #7f0010;
      stroke-width: 4px;
    }}
    body[data-theme="light"] svg {{ background:#eaf0f5; }}
    body[data-theme="light"] .schematic-field {{ fill:#fbfdff; stroke:#b9c9d7; }}
    body[data-theme="light"] .schematic-grid {{ stroke:#c7d4df; opacity:.58; }}
    body[data-theme="light"] .schematic-grid-field {{ opacity:.72; }}
    body[data-theme="light"] .edge {{ stroke:#526b7f; opacity:.58; }}
    body[data-theme="light"] .edge.connected {{ stroke:#006f9f!important; opacity:1; filter:drop-shadow(0 0 3px rgba(0,111,159,.28)); }}
    body[data-theme="light"] .edge.context-dimmed {{ opacity:.10; }}
    body[data-theme="light"] .node {{ fill:#1478a8; stroke:#075985; stroke-width:1.2; opacity:.96; }}
    body[data-theme="light"] .node-group.neighbor .node {{ stroke:#005f8a!important; stroke-width:2.4; }}
    body[data-theme="light"] .node.missing {{ fill:#fff; stroke:#6f7f8d; stroke-width:1.4; }}
    body[data-theme="light"] .label {{ fill:#243b4d; stroke:#fbfdff; stroke-width:5px; }}
    body[data-theme="light"] .label-leader {{ stroke:#6f8799; opacity:.72; }}
    body[data-theme="light"] .blocked-label {{ fill:#991b1b; stroke:#fff7f7; }}
    .legend {{
      position: absolute;
      left: 14px;
      bottom: 14px;
      width: 310px;
      padding: 12px;
      border-radius: 10px;
      background: color-mix(in srgb, var(--panel) 90%, transparent);
      backdrop-filter: blur(16px);
      border: 1px solid var(--line);
      font-size: 12px;
      color: var(--muted);
    }}
    .ramp {{
      height: 10px;
      border-radius: 999px;
      background: linear-gradient(90deg, #ffd6d6, #fb7770, #c5161d, #64000a);
      margin: 8px 0 6px;
    }}
    .side {{
      background: var(--panel);
      border-left: 1px solid var(--line);
      padding: 20px;
      overflow: auto;
      max-height: 100vh;
      position: sticky;
      top: 0;
      box-shadow: -18px 0 48px rgba(0,0,0,.16);
    }}
    .side-header {{
      display: flex; align-items: center; justify-content: space-between; gap: 10px;
      margin-bottom: 12px;
    }}
    .side-label {{ font-size: 12px; color: var(--muted); font-weight: 700; letter-spacing: .08em; }}
    .search {{
      width: 100%; height: 42px; border: 1px solid var(--line); border-radius: 10px;
      padding: 0 12px; color: var(--ink); background: var(--soft); outline: none;
    }}
    .search:focus {{ border-color: #88ace8; box-shadow: 0 0 0 3px rgba(31,99,211,.11); }}
    .list-head {{ display: flex; align-items: center; justify-content: space-between; margin: 14px 0 8px; }}
    .sort-select {{ border: 0; color: var(--blue); background: transparent; font-size: 12px; }}
    .filter-block {{ margin:12px 0 4px; padding:12px; border:1px solid var(--line); border-radius:11px; background:var(--soft); }}
    .filter-row {{ display:flex; align-items:center; justify-content:space-between; gap:10px; font-size:12px; color:var(--muted); }}
    .range {{ width:100%; margin-top:9px; accent-color:var(--cyan); cursor:pointer; }}
    .quick-filters {{ display:flex; gap:6px; margin-top:9px; }}
    .quick-filter {{ flex:1; border:1px solid var(--line); border-radius:8px; padding:5px; color:var(--muted); background:var(--panel); cursor:pointer; font-size:11px; }}
    .quick-filter.active {{ color:var(--cyan); border-color:var(--cyan); }}
    .section-list {{ display: grid; gap: 7px; max-height: 240px; overflow: auto; padding-right: 3px; }}
    .section-item {{
      width: 100%; text-align: left; border: 1px solid var(--line); background: var(--panel);
      border-radius: 10px; padding: 10px 11px; cursor: pointer; color: var(--ink);
    }}
    .section-item:hover {{ border-color: var(--blue); background: color-mix(in srgb, var(--blue) 8%, var(--panel)); transform:translateX(-2px); }}
    .section-item.active {{ border-color: var(--accent); background: color-mix(in srgb, var(--accent) 9%, var(--panel)); box-shadow: inset 3px 0 0 var(--accent), 0 0 18px rgba(255,77,109,.09); }}
    .item-top {{ display: flex; justify-content: space-between; gap: 10px; font-size: 13px; font-weight: 750; }}
    .item-gap {{ color: var(--accent); white-space: nowrap; }}
    .item-meta {{ margin-top: 4px; color: var(--muted); font-size: 11px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
    .detail-divider {{ border: 0; border-top: 1px solid var(--line); margin: 18px 0; }}
    .metric-grid {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 10px;
      margin: 14px 0;
    }}
    .metric {{
      border: 1px solid var(--line);
      border-radius: 10px;
      padding: 10px;
      background: var(--soft);
    }}
    .metric .k {{
      font-size: 12px;
      color: var(--muted);
      margin-bottom: 4px;
    }}
    .metric .v {{
      font-size: 17px;
      font-weight: 700;
      overflow-wrap: anywhere;
    }}
    .panel-title {{
      font-size: 19px;
      font-weight: 700;
      margin: 0 0 6px;
    }}
    .panel-subtitle {{
      font-size: 13px;
      color: var(--muted);
      line-height: 1.45;
    }}
    .chart-wrap {{
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 10px;
      margin-top: 12px;
    }}
    .chart-title {{
      font-size: 13px;
      font-weight: 700;
      color: var(--ink);
      margin: 0 0 6px;
    }}
    #priceChart {{
      width: 100%;
      height: 300px;
      touch-action: pan-y;
    }}
    .price-crosshair-hit {{ cursor: crosshair; }}
    .source {{
      margin-top: 16px;
      font-size: 12px;
      color: var(--muted);
      line-height: 1.55;
      overflow-wrap: anywhere;
    }}
    .pill {{
      display: inline-block;
      padding: 2px 7px;
      border-radius: 999px;
      background: #fee3e2;
      color: #9b1117;
      font-size: 12px;
      font-weight: 600;
      margin-left: 6px;
    }}
    .focus-mode .app {{ grid-template-columns: 1fr; }}
    .focus-mode .side {{ display:none; }}
    .focus-mode svg {{ height:calc(100vh - 218px); }}
    .pulse {{ animation:pulse 1.5s ease-in-out infinite; }}
    @keyframes pulse {{ 50% {{ filter:drop-shadow(0 0 9px var(--accent)); }} }}
    @media (prefers-reduced-motion: reduce) {{ *,*::before,*::after {{ animation:none!important; transition:none!important; }} }}
    .empty-list {{ padding: 22px 8px; color: var(--muted); font-size: 13px; text-align: center; }}
    @media (max-width: 920px) {{
      .app {{ grid-template-columns: 1fr; }}
      .map-pane {{ padding: 16px; }}
      .summary-grid {{ grid-template-columns: 1fr 1fr; }}
      .side {{ border-left: 0; border-top: 1px solid var(--line); max-height: none; position: static; }}
      svg {{ height: 62vh; min-height: 480px; }}
      .map-shell {{ min-height: 0; }}
      .section-list {{ max-height: 280px; }}
    }}
    @media (max-width: 560px) {{
      .map-pane {{ padding: 12px; }}
      .title-row {{ margin: 6px 2px 14px; }}
      .title-row {{ flex-direction:column; }}
      .header-actions {{ width:100%; justify-content:flex-start; }}
      .summary-card {{ padding: 8px 10px; }}
      .summary-card .v {{ font-size: 18px; }}
      .map-toolbar {{ align-items: flex-start; flex-direction: column; }}
      .node-search {{ width:min(100%, 220px); }}
      .map-hint {{ display: none; }}
      .zoom-controls {{ top: 96px; }}
      .legend {{ display:none; }}
      svg {{ height:78vw; min-height:300px; max-height:390px; }}
    }}
  </style>
</head>
<body data-theme="light">
  <div class="app">
    <main class="map-pane">
      <div class="title-row">
        <div>
          <div class="eyebrow">GRID PRICE ANALYTICS</div>
          <h1>{artifact_title}</h1>
        </div>
        <div class="header-actions">
          <button class="icon-button" id="focusMode" type="button" aria-pressed="false">沉浸地图</button>
          <button class="icon-button" id="themeToggle" type="button">切换浅色</button>
        </div>
      </div>
      <section class="summary-grid" aria-label="分析总览">
        <div class="summary-card alert"><div class="k">识别断面</div><div class="v"><span id="summarySections">0</span><span class="unit">条</span></div></div>
        <div class="summary-card"><div class="k">覆盖节点</div><div class="v"><span id="summaryNodes">0</span><span class="unit">个</span></div></div>
        <div class="summary-card"><div class="k">最高价差</div><div class="v"><span id="summaryMaxGap">0</span><span class="unit">元/MWh</span></div></div>
        <div class="summary-card"><div class="k">电气片区</div><div class="v"><span id="summaryRegions">0</span><span class="unit">个</span></div></div>
      </section>
      <div class="map-shell">
        <div class="map-toolbar">
          <div><strong>{map_layer_title}</strong> <span class="map-hint">触控板双指移动 · 捏合缩放 · 鼠标滚轮缩放 · 点击节点查看详情</span></div>
          <div class="layer-toggles" aria-label="地图图层">
            <input class="node-search" id="nodeSearch" type="search" placeholder="搜索节点，Enter定位" aria-label="搜索节点">
            <button class="toggle" id="focusSelected" type="button">定位所选</button>
            <button class="toggle" id="toggleTopology" type="button" aria-pressed="true">拓扑</button>
            <button class="toggle" id="toggleLabels" type="button" aria-pressed="true">标签</button>
          </div>
        </div>
        <svg id="map" viewBox="0 0 {SVG_W} {SVG_H}" role="img" aria-label="{map_label}">
          <g id="mapViewport">
            {substrate}
            <g id="regions"></g>
            <g id="edges"></g>
            <g id="labelLeaders"></g>
            <g id="nodes"></g>
            <g id="nodeLabels"></g>
          </g>
        </svg>
        <div class="zoom-controls" aria-label="地图缩放控制">
          <button id="zoomIn" type="button" title="放大">+</button>
          <button id="zoomOut" type="button" title="缩小">−</button>
          <button id="zoomReset" type="button" title="重置">⌂</button>
        </div>
        <div class="legend" aria-label="断面强度颜色图例">
          <strong>断面强度</strong>
          <div class="ramp" aria-hidden="true"></div>
          <div>浅：触发时点较少、价差较小　深：累计时点更多、价差更大</div>
        </div>
      </div>
    </main>
    <aside class="side">
      <div class="side-header"><span class="side-label">断面导航</span><span class="pill" id="visibleCount">0 条</span></div>
      <input class="search" id="sectionSearch" type="search" placeholder="搜索节点、片区或方向" aria-label="搜索断面">
      <div class="filter-block">
        <div class="filter-row"><span>最低最大价差</span><strong><span id="gapValue">0</span> 元/MWh</strong></div>
        <input class="range" id="gapRange" type="range" min="0" value="0" step="50" aria-label="最低最大价差">
        <div class="quick-filters" aria-label="快速筛选">
          <button class="quick-filter active" type="button" data-percent="0">全部</button>
          <button class="quick-filter" type="button" data-percent="0.5">中高</button>
          <button class="quick-filter" type="button" data-percent="0.75">高风险</button>
        </div>
      </div>
      <div class="list-head">
        <strong>断面列表</strong>
        <select class="sort-select" id="sectionSort" aria-label="断面排序">
          <option value="severity">按严重度</option>
          <option value="gap">按最大价差</option>
          <option value="duration">按持续时段</option>
          <option value="name">按名称</option>
        </select>
      </div>
      <div class="section-list" id="sectionList"></div>
      <hr class="detail-divider">
      <p class="panel-title" id="detailTitle">点击任意节点查看电价数据</p>
      <div class="panel-subtitle" id="detailSubtitle">当前共识别 <span id="sectionCount"></span> 条价差断面；可点击节点查看分时电价和相邻线路。</div>
      <div class="metric-grid">
        <div class="metric"><div class="k" id="metricLabel1">最大价差</div><div class="v" id="maxGap">-</div></div>
        <div class="metric"><div class="k" id="metricLabel2">持续时段</div><div class="v" id="intervals">-</div></div>
        <div class="metric"><div class="k" id="metricLabel3">主方向</div><div class="v" id="direction">-</div></div>
        <div class="metric"><div class="k" id="metricLabel4">最大时点</div><div class="v" id="maxTime">-</div></div>
      </div>
      <div class="chart-wrap">
        <div class="chart-title" id="chartTitle">24小时节点电价曲线</div>
        <svg id="priceChart" viewBox="0 0 360 280" role="img" aria-label="两端节点24小时电价曲线"></svg>
      </div>
      <div class="source" id="sourceText"></div>
    </aside>
  </div>
  <script>
    const DATA = {json_data};
  </script>
  <script>
    const sourceCanvasWidth = Number(DATA.metadata.layout_canvas_width || 1340);
    const sourceCanvasHeight = Number(DATA.metadata.layout_canvas_height || 900);
    if (sourceCanvasWidth !== {SVG_W} || sourceCanvasHeight !== {SVG_H}) {{
      const oldCenterX = sourceCanvasWidth / 2;
      const oldCenterY = sourceCanvasHeight / 2;
      const targetCenterX = {SVG_W} / 2;
      const targetCenterY = {SVG_H} / 2;
      const sourceDiagramScale = Number(DATA.metadata.diagram_scale || 0);
      const targetDiagramScale = DATA.metadata.coordinate_mode === 'wiring-diagram'
        ? Math.min(
            ({SVG_W} - {DIAGRAM_PADDING_X * 2}) / Number(DATA.metadata.diagram_width || {SVG_W}),
            ({SVG_H} - {DIAGRAM_PADDING_Y * 2}) / Number(DATA.metadata.diagram_height || {SVG_H})
          )
        : 0;
      const layoutScale = sourceDiagramScale > 0 && targetDiagramScale > 0
        ? targetDiagramScale / sourceDiagramScale
        : Math.min({SVG_W} / sourceCanvasWidth, {SVG_H} / sourceCanvasHeight);
      const transformPoint = (x, y) => ({{
        x: targetCenterX + (Number(x) - oldCenterX) * layoutScale,
        y: targetCenterY + (Number(y) - oldCenterY) * layoutScale
      }});
      DATA.nodes.forEach(node => Object.assign(node, transformPoint(node.x, node.y)));
      DATA.edges.forEach(edge => {{
        const start = transformPoint(edge.x1, edge.y1);
        const end = transformPoint(edge.x2, edge.y2);
        edge.x1 = start.x;
        edge.y1 = start.y;
        edge.x2 = end.x;
        edge.y2 = end.y;
      }});
      DATA.metadata.layout_canvas_width = {SVG_W};
      DATA.metadata.layout_canvas_height = {SVG_H};
    }}
    const nodeByName = new Map(DATA.nodes.map(d => [d.name, d]));
    const regionData = DATA.regions || [];
    const sectionEdges = DATA.edges.filter(d => d.is_section);
    const absoluteMaxGap = sectionEdges.length ? Math.max(...sectionEdges.map(edge => Number(edge.max_abs_diff) || 0)) : 0;
    document.getElementById('sectionCount').textContent = sectionEdges.length;
    document.getElementById('summarySections').textContent = sectionEdges.length;
    document.getElementById('summaryNodes').textContent = DATA.nodes.length;
    document.getElementById('summaryRegions').textContent = regionData.length || new Set(DATA.nodes.map(node => node.region).filter(Boolean)).size;
    document.getElementById('summaryMaxGap').textContent = sectionEdges.length
      ? absoluteMaxGap.toLocaleString('zh-CN', {{ maximumFractionDigits: 0 }})
      : '0';
    const gapRange = document.getElementById('gapRange');
    gapRange.max = String(Math.ceil(absoluteMaxGap / 50) * 50 || 1000);
    const mapSvg = document.getElementById('map');
    const mapViewport = document.getElementById('mapViewport');
    const edgeLines = new Map();
    const sectionEdgeLines = new Map();
    const nodeGroups = new Map();
    const nodeLabelItems = [];
    const edgesByNode = new Map(DATA.nodes.map(node => [node.name, []]));
    DATA.edges.forEach(edge => {{
      edgesByNode.get(edge.a)?.push(edge);
      edgesByNode.get(edge.b)?.push(edge);
    }});
    let selectedEdgeName = '';
    let selectedNodeName = '';
    let view = {{ scale: 1, x: 0, y: 0 }};
    let drag = null;
    let trackpadUntil = 0;
    let gestureBaseScale = null;

    function clampView() {{
      const visibleMarginX = {SVG_W} * .12;
      const visibleMarginY = {SVG_H} * .12;
      view.x = Math.max(-{SVG_W} * view.scale + visibleMarginX, Math.min({SVG_W} - visibleMarginX, view.x));
      view.y = Math.max(-{SVG_H} * view.scale + visibleMarginY, Math.min({SVG_H} - visibleMarginY, view.y));
    }}

    function applyView() {{
      clampView();
      mapViewport.setAttribute('transform', `translate(${{view.x}} ${{view.y}}) scale(${{view.scale}})`);
    }}

    function zoomToScale(targetScale, clientX, clientY) {{
      const rect = mapSvg.getBoundingClientRect();
      const px = ((clientX ?? (rect.left + rect.width / 2)) - rect.left) / rect.width * {SVG_W};
      const py = ((clientY ?? (rect.top + rect.height / 2)) - rect.top) / rect.height * {SVG_H};
      const oldScale = view.scale;
      const nextScale = Math.max(0.8, Math.min(5.2, targetScale));
      view.x = px - (px - view.x) * (nextScale / oldScale);
      view.y = py - (py - view.y) * (nextScale / oldScale);
      view.scale = nextScale;
      applyView();
    }}

    function zoomBy(factor, clientX, clientY) {{
      zoomToScale(view.scale * factor, clientX, clientY);
    }}

    function panByWheel(deltaX, deltaY) {{
      const rect = mapSvg.getBoundingClientRect();
      view.x -= deltaX / rect.width * {SVG_W};
      view.y -= deltaY / rect.height * {SVG_H};
      applyView();
    }}

    function isLikelyTrackpadWheel(event) {{
      if (event.deltaMode !== 0) return false;
      const now = Date.now();
      const fineMotion = Math.abs(event.deltaX) > .1
        || !Number.isInteger(event.deltaY)
        || Math.abs(event.deltaY) < 50;
      if (fineMotion) trackpadUntil = now + 220;
      return fineMotion || now < trackpadUntil;
    }}

    document.getElementById('zoomIn').addEventListener('click', () => zoomBy(1.28));
    document.getElementById('zoomOut').addEventListener('click', () => zoomBy(1 / 1.28));
    document.getElementById('zoomReset').addEventListener('click', () => {{
      view = {{ scale: 1, x: 0, y: 0 }};
      applyView();
    }});
    mapSvg.addEventListener('wheel', event => {{
      event.preventDefault();
      if (event.ctrlKey || event.metaKey) {{
        const factor = Math.exp(Math.max(-.24, Math.min(.24, -event.deltaY * .012)));
        zoomBy(factor, event.clientX, event.clientY);
        return;
      }}
      if (isLikelyTrackpadWheel(event)) {{
        const deltaX = event.shiftKey && Math.abs(event.deltaX) < .1 ? event.deltaY : event.deltaX;
        const deltaY = event.shiftKey && Math.abs(event.deltaX) < .1 ? 0 : event.deltaY;
        panByWheel(deltaX, deltaY);
        return;
      }}
      zoomBy(event.deltaY < 0 ? 1.16 : 1 / 1.16, event.clientX, event.clientY);
    }}, {{ passive: false }});
    mapSvg.addEventListener('gesturestart', event => {{
      event.preventDefault();
      gestureBaseScale = view.scale;
    }}, {{ passive: false }});
    mapSvg.addEventListener('gesturechange', event => {{
      event.preventDefault();
      if (gestureBaseScale == null) return;
      zoomToScale(gestureBaseScale * Number(event.scale || 1), event.clientX, event.clientY);
    }}, {{ passive: false }});
    mapSvg.addEventListener('gestureend', event => {{
      event.preventDefault();
      gestureBaseScale = null;
    }}, {{ passive: false }});
    mapSvg.addEventListener('pointerdown', event => {{
      if (event.button !== 0) return;
      if (event.target.closest?.('.node-group, .edge, .edge-hit')) return;
      mapSvg.setPointerCapture(event.pointerId);
      drag = {{ id: event.pointerId, x: event.clientX, y: event.clientY, startX: view.x, startY: view.y, moved: false }};
    }});
    mapSvg.addEventListener('pointermove', event => {{
      if (!drag || drag.id !== event.pointerId) return;
      const moveX = event.clientX - drag.x;
      const moveY = event.clientY - drag.y;
      if (!drag.moved && Math.hypot(moveX, moveY) < 4) return;
      if (!drag.moved) {{
        drag.moved = true;
        mapSvg.classList.add('is-panning');
      }}
      const rect = mapSvg.getBoundingClientRect();
      view.x = drag.startX + moveX / rect.width * {SVG_W};
      view.y = drag.startY + moveY / rect.height * {SVG_H};
      applyView();
    }});
    mapSvg.addEventListener('pointerup', event => {{
      if (drag && drag.id === event.pointerId) {{
        const shouldClearSelection = !drag.moved;
        drag = null;
        mapSvg.classList.remove('is-panning');
        if (shouldClearSelection) clearSelection();
      }}
    }});
    mapSvg.addEventListener('pointercancel', () => {{
      drag = null;
      mapSvg.classList.remove('is-panning');
    }});
    mapSvg.addEventListener('lostpointercapture', () => mapSvg.classList.remove('is-panning'));

    function severityColor(s) {{
      if (!s) return '#dbe3eb';
      const stops = [
        [255, 214, 214],
        [251, 119, 112],
        [197, 22, 29],
        [100, 0, 10]
      ];
      const scaled = Math.max(0, Math.min(1, s)) * (stops.length - 1);
      const i = Math.floor(scaled);
      const t = scaled - i;
      const a = stops[i];
      const b = stops[Math.min(i + 1, stops.length - 1)];
      const rgb = a.map((v, idx) => Math.round(v + (b[idx] - v) * t));
      return `rgb(${{rgb[0]}},${{rgb[1]}},${{rgb[2]}})`;
    }}

    const regionsG = document.getElementById('regions');
    const edgesG = document.getElementById('edges');
    DATA.edges.forEach(edge => {{
      const hitLine = document.createElementNS('http://www.w3.org/2000/svg', 'line');
      const line = document.createElementNS('http://www.w3.org/2000/svg', 'line');
      [hitLine, line].forEach(element => {{
        element.setAttribute('x1', edge.x1);
        element.setAttribute('y1', edge.y1);
        element.setAttribute('x2', edge.x2);
        element.setAttribute('y2', edge.y2);
      }});
      hitLine.setAttribute('class', edge.is_section ? 'edge-hit section-hit' : 'edge-hit');
      hitLine.setAttribute('tabindex', '0');
      hitLine.setAttribute('role', 'button');
      hitLine.setAttribute('aria-label', edge.is_section
        ? `${{edge.edge}}，最大价差 ${{edge.max_abs_diff.toFixed(0)}} 元每兆瓦时`
        : `${{edge.edge}}，500千伏拓扑线路`);
      line.setAttribute('class', edge.is_section ? 'edge section' : 'edge');
      line.dataset.a = edge.a;
      line.dataset.b = edge.b;
      line.setAttribute('stroke', severityColor(edge.severity));
      line.setAttribute('stroke-width', edge.is_section ? String(4.8 + edge.severity * 4.4) : '1.1');
      const activateEdge = event => {{
        event.preventDefault();
        event.stopPropagation();
        selectEdge(edge, line);
      }};
      hitLine.addEventListener('pointerdown', event => {{
        event.preventDefault();
        event.stopPropagation();
      }});
      hitLine.addEventListener('click', activateEdge);
      hitLine.addEventListener('keydown', event => {{
        if (event.key === 'Enter' || event.key === ' ') activateEdge(event);
      }});
      hitLine.addEventListener('pointerenter', () => line.classList.add('hover'));
      hitLine.addEventListener('pointerleave', () => line.classList.remove('hover'));
      hitLine.addEventListener('focus', () => line.classList.add('hover'));
      hitLine.addEventListener('blur', () => line.classList.remove('hover'));
      line.addEventListener('pointerdown', event => event.stopPropagation());
      line.addEventListener('click', activateEdge);
      edgesG.appendChild(hitLine);
      edgesG.appendChild(line);
      edgeLines.set(edge.edge, line);
      if (edge.is_section) sectionEdgeLines.set(edge.edge, line);
    }});

    const nodesG = document.getElementById('nodes');
    const labelLeadersG = document.getElementById('labelLeaders');
    const nodeLabelsG = document.getElementById('nodeLabels');
    DATA.nodes.forEach(node => {{
      const group = document.createElementNS('http://www.w3.org/2000/svg', 'g');
      group.setAttribute('class', 'node-group');
      group.setAttribute('tabindex', '0');
      group.setAttribute('role', 'button');
      group.setAttribute('aria-label', `${{node.name}}节点，点击查看电价数据`);
      group.dataset.nodeName = node.name;
      const hit = document.createElementNS('http://www.w3.org/2000/svg', 'circle');
      hit.setAttribute('cx', node.x);
      hit.setAttribute('cy', node.y);
      hit.setAttribute('r', '12');
      hit.setAttribute('class', 'node-hit');
      group.appendChild(hit);
      const circle = document.createElementNS('http://www.w3.org/2000/svg', 'circle');
      circle.setAttribute('cx', node.x);
      circle.setAttribute('cy', node.y);
      circle.setAttribute('r', node.blocked ? 7.2 : (node.prices.length ? 3.4 : 2.8));
      circle.setAttribute('class', node.blocked ? 'node blocked' : (node.prices.length ? 'node' : 'node missing'));
      const region = regionData.find(item => item.name === node.region);
      circle.setAttribute('stroke', node.blocked ? '#a40012' : (region?.color || '#9aabbc'));
      circle.appendChild(document.createElementNS('http://www.w3.org/2000/svg', 'title')).textContent = `${{node.name}}｜${{node.region}}｜${{node.location_accuracy}}｜${{node.position_label}}｜${{node.location_source}}`;
      group.appendChild(circle);

      const label = document.createElementNS('http://www.w3.org/2000/svg', 'text');
      label.setAttribute('class', node.blocked ? 'label blocked-label' : 'label');
      label.dataset.nodeName = node.name;
      label.textContent = node.name;
      label.addEventListener('pointerdown', event => event.stopPropagation());
      label.addEventListener('click', event => {{
        event.stopPropagation();
        selectNode(node, group);
      }});
      const leader = document.createElementNS('http://www.w3.org/2000/svg', 'line');
      leader.setAttribute('class', 'label-leader is-idle');
      leader.setAttribute('x1', node.x);
      leader.setAttribute('y1', node.y);
      group.addEventListener('pointerdown', event => event.stopPropagation());
      group.addEventListener('click', event => {{ event.stopPropagation(); selectNode(node, group); }});
      group.addEventListener('keydown', event => {{
        if (event.key === 'Enter' || event.key === ' ') {{ event.preventDefault(); selectNode(node, group); }}
      }});
      nodesG.appendChild(group);
      labelLeadersG.appendChild(leader);
      nodeLabelsG.appendChild(label);
      nodeLabelItems.push({{ node, label, leader }});
      nodeGroups.set(node.name, group);
    }});

    function boxesOverlap(a, b, padding = 4) {{
      return a.x < b.x + b.width + padding
        && a.x + a.width + padding > b.x
        && a.y < b.y + b.height + padding
        && a.y + a.height + padding > b.y;
    }}

    function boxHitsNode(box, node, padding = 3) {{
      const nearestX = Math.max(box.x - padding, Math.min(node.x, box.x + box.width + padding));
      const nearestY = Math.max(box.y - padding, Math.min(node.y, box.y + box.height + padding));
      const radius = node.blocked ? 7.2 : 4.2;
      return Math.hypot(node.x - nearestX, node.y - nearestY) < radius;
    }}

    function segmentsIntersect(a, b, c, d) {{
      const cross = (p, q, r) => (q.x - p.x) * (r.y - p.y) - (q.y - p.y) * (r.x - p.x);
      const abC = cross(a, b, c);
      const abD = cross(a, b, d);
      const cdA = cross(c, d, a);
      const cdB = cross(c, d, b);
      return ((abC <= 0 && abD >= 0) || (abC >= 0 && abD <= 0))
        && ((cdA <= 0 && cdB >= 0) || (cdA >= 0 && cdB <= 0));
    }}

    function boxHitsEdge(box, edge, padding = 5) {{
      const left = box.x - padding;
      const right = box.x + box.width + padding;
      const top = box.y - padding;
      const bottom = box.y + box.height + padding;
      const start = {{ x: edge.x1, y: edge.y1 }};
      const end = {{ x: edge.x2, y: edge.y2 }};
      const inside = point => point.x >= left && point.x <= right && point.y >= top && point.y <= bottom;
      if (inside(start) || inside(end)) return true;
      const topLeft = {{ x: left, y: top }};
      const topRight = {{ x: right, y: top }};
      const bottomRight = {{ x: right, y: bottom }};
      const bottomLeft = {{ x: left, y: bottom }};
      return segmentsIntersect(start, end, topLeft, topRight)
        || segmentsIntersect(start, end, topRight, bottomRight)
        || segmentsIntersect(start, end, bottomRight, bottomLeft)
        || segmentsIntersect(start, end, bottomLeft, topLeft);
    }}

    function labelDirections(node) {{
      let awayX = 0;
      let awayY = 0;
      DATA.nodes.forEach(other => {{
        if (other.name === node.name) return;
        const dx = node.x - other.x;
        const dy = node.y - other.y;
        const distanceSquared = dx * dx + dy * dy;
        if (distanceSquared > 0 && distanceSquared < 120 * 120) {{
          awayX += dx / distanceSquared;
          awayY += dy / distanceSquared;
        }}
      }});
      const directions = [
        {{ x: 1, y: -1 }}, {{ x: -1, y: -1 }}, {{ x: 1, y: 1 }}, {{ x: -1, y: 1 }},
        {{ x: 0, y: -1 }}, {{ x: 1, y: 0 }}, {{ x: -1, y: 0 }}, {{ x: 0, y: 1 }}
      ];
      if (Math.abs(awayX) + Math.abs(awayY) < .0001) return directions;
      return directions.sort((a, b) => (b.x * awayX + b.y * awayY) - (a.x * awayX + a.y * awayY));
    }}

    function positionLabel(item, direction, distance) {{
      const {{ node, label }} = item;
      const fontSize = node.blocked ? 13 : 10;
      const diagonal = direction.x !== 0 && direction.y !== 0 ? .76 : 1;
      const dx = direction.x * distance * diagonal;
      const dy = direction.y * distance * diagonal;
      label.setAttribute('text-anchor', direction.x < 0 ? 'end' : (direction.x > 0 ? 'start' : 'middle'));
      label.setAttribute('x', node.x + dx);
      label.setAttribute('y', node.y + dy + (direction.y > 0 ? fontSize * .82 : (direction.y === 0 ? fontSize * .34 : 0)));
      return label.getBBox();
    }}

    function layoutNodeLabels() {{
      const placedBoxes = [];
      const orderedItems = [...nodeLabelItems].sort((a, b) => {{
        const blockedPriority = Number(b.node.blocked) - Number(a.node.blocked);
        if (blockedPriority) return blockedPriority;
        return (edgesByNode.get(b.node.name)?.length || 0) - (edgesByNode.get(a.node.name)?.length || 0);
      }});
      orderedItems.forEach(item => {{
        const directions = labelDirections(item.node);
        const distances = item.node.blocked ? [12, 16, 20, 24, 28, 32, 36, 40] : [8, 12, 16, 20, 24, 28, 32, 36];
        let chosen = null;
        let bestFallback = null;
        distances.some(distance => {{
          let bestAtDistance = null;
          directions.forEach((direction, directionIndex) => {{
            const box = positionLabel(item, direction, distance);
            const inBounds = box.x >= 28 && box.y >= 28
              && box.x + box.width <= {SVG_W - 28}
              && box.y + box.height <= {SVG_H - 28};
            const labelOverlap = placedBoxes.reduce((count, placed) => count + Number(boxesOverlap(box, placed)), 0);
            const nodeOverlap = DATA.nodes.reduce((count, node) => count + Number(node.name !== item.node.name && boxHitsNode(box, node)), 0);
            const edgeOverlap = DATA.edges.reduce((count, edge) => count + Number(boxHitsEdge(box, edge)), 0);
            const score = labelOverlap * 100000 + nodeOverlap * 10000 + (inBounds ? 0 : 1000000)
              + distance * 20 + edgeOverlap * 12 + directionIndex;
            const candidate = {{ direction, distance, box, score, edgeOverlap }};
            if (!bestFallback || score < bestFallback.score) bestFallback = candidate;
            if (inBounds && labelOverlap === 0 && nodeOverlap === 0
                && (!bestAtDistance || edgeOverlap < bestAtDistance.edgeOverlap)) {{
              bestAtDistance = candidate;
            }}
          }});
          if (bestAtDistance) {{
            chosen = bestAtDistance;
            return true;
          }}
          return false;
        }});
        chosen = chosen || bestFallback;
        const finalBox = positionLabel(item, chosen.direction, chosen.distance);
        placedBoxes.push(finalBox);
        item.label.dataset.distance = String(chosen.distance);
        item.label.dataset.edgeOverlaps = String(chosen.edgeOverlap || 0);
        const moved = chosen.distance >= 20 && chosen.distance <= 28;
        item.leader.classList.toggle('is-idle', !moved);
        if (moved) {{
          const targetX = Math.max(finalBox.x, Math.min(item.node.x, finalBox.x + finalBox.width));
          const targetY = Math.max(finalBox.y, Math.min(item.node.y, finalBox.y + finalBox.height));
          const dx = targetX - item.node.x;
          const dy = targetY - item.node.y;
          const length = Math.max(Math.hypot(dx, dy), 1);
          const nodeRadius = item.node.blocked ? 8.5 : 5.5;
          item.leader.setAttribute('x1', item.node.x + dx / length * nodeRadius);
          item.leader.setAttribute('y1', item.node.y + dy / length * nodeRadius);
          item.leader.setAttribute('x2', targetX);
          item.leader.setAttribute('y2', targetY);
        }}
      }});
    }}

    layoutNodeLabels();
    if (document.fonts?.ready) document.fonts.ready.then(layoutNodeLabels);

    function edgePriceStats(a, b) {{
      const comparable = DATA.time_headers.map((time, index) => {{
        const priceA = a?.prices[index];
        const priceB = b?.prices[index];
        if (typeof priceA !== 'number' || typeof priceB !== 'number') return null;
        return {{ time, priceA, priceB, gap: Math.abs(priceB - priceA) }};
      }}).filter(Boolean);
      const maximum = comparable.reduce((best, item) => !best || item.gap > best.gap ? item : best, null);
      return {{ comparable, maximum }};
    }}

    function clearSelection() {{
      document.querySelectorAll('.edge').forEach(element =>
        element.classList.remove('active', 'connected', 'context-dimmed')
      );
      document.querySelectorAll('.node-group').forEach(element =>
        element.classList.remove('active', 'neighbor', 'context-dimmed')
      );
      document.querySelectorAll('.section-item').forEach(item => item.classList.remove('active'));
      selectedEdgeName = '';
      selectedNodeName = '';
      document.getElementById('metricLabel1').textContent = '最大价差';
      document.getElementById('metricLabel2').textContent = '持续时段';
      document.getElementById('metricLabel3').textContent = '主方向';
      document.getElementById('metricLabel4').textContent = '最大时点';
      document.getElementById('detailTitle').textContent = '未选择节点或线路';
      document.getElementById('detailSubtitle').textContent = `当前共识别 ${{sectionEdges.length}} 条价差断面；点击节点或线路查看电价。`;
      ['maxGap', 'intervals', 'direction', 'maxTime'].forEach(id => {{
        document.getElementById(id).textContent = '-';
      }});
      document.getElementById('chartTitle').textContent = '24小时节点电价曲线';
      const chart = document.getElementById('priceChart');
      chart.innerHTML = '';
      const message = document.createElementNS('http://www.w3.org/2000/svg', 'text');
      message.setAttribute('x', '180');
      message.setAttribute('y', '140');
      message.setAttribute('text-anchor', 'middle');
      message.setAttribute('fill', '#8296ad');
      message.textContent = '点击节点或线路查看电价曲线';
      chart.appendChild(message);
      document.getElementById('sourceText').textContent = '';
    }}

    function selectEdge(edge, line) {{
      document.querySelectorAll('.edge.active').forEach(el => el.classList.remove('active'));
      document.querySelectorAll('.edge').forEach(el => el.classList.remove('connected', 'context-dimmed'));
      document.querySelectorAll('.node-group').forEach(el => el.classList.remove('active', 'neighbor', 'context-dimmed'));
      line.classList.add('active');
      nodeGroups.get(edge.a)?.classList.add('active');
      nodeGroups.get(edge.b)?.classList.add('active');
      selectedEdgeName = edge.edge;
      selectedNodeName = '';
      document.querySelectorAll('.section-item').forEach(item => item.classList.toggle('active', item.dataset.edge === edge.edge));
      const a = nodeByName.get(edge.a);
      const b = nodeByName.get(edge.b);
      const observed = edgePriceStats(a, b);
      const displayGap = edge.is_section ? edge.max_abs_diff : (observed.maximum?.gap ?? 0);
      const displayTime = edge.is_section ? edge.max_time : (observed.maximum?.time || '-');
      document.getElementById('metricLabel1').textContent = '最大价差';
      document.getElementById('metricLabel2').textContent = edge.is_section ? '持续时段' : '可比时点';
      document.getElementById('metricLabel3').textContent = edge.is_section ? '主方向' : '线路状态';
      document.getElementById('metricLabel4').textContent = '最大时点';
      document.getElementById('detailTitle').innerHTML = `${{edge.edge}} <span class="pill">${{edge.group}}</span>`;
      document.getElementById('detailSubtitle').textContent = edge.is_section
        ? `触发窗口：${{edge.windows || '无'}}`
        : '完整拓扑线路，未触发断面判定；以下显示两端节点电价。';
      document.getElementById('maxGap').textContent = displayGap.toFixed(2);
      document.getElementById('intervals').textContent = edge.is_section
        ? `${{edge.intervals}} × 15分钟`
        : `${{observed.comparable.length}} 个`;
      document.getElementById('direction').textContent = edge.is_section ? (edge.main_direction || '-') : '未触发断面';
      document.getElementById('maxTime').textContent = displayTime || '-';
      document.getElementById('chartTitle').textContent = `${{edge.a}} / ${{edge.b}} 24小时节点电价曲线`;
      drawChart(a, b, edge);
    }}

    function selectNode(node, group) {{
      document.querySelectorAll('.edge.active').forEach(el => el.classList.remove('active'));
      document.querySelectorAll('.node-group').forEach(el => el.classList.remove('active', 'neighbor', 'context-dimmed'));
      group.classList.add('active');
      const relatedEdges = edgesByNode.get(node.name) || [];
      const neighborNames = new Set();
      relatedEdges.forEach(edge => neighborNames.add(edge.a === node.name ? edge.b : edge.a));
      edgeLines.forEach((line, edgeName) => {{
        const connected = relatedEdges.some(edge => edge.edge === edgeName);
        line.classList.toggle('connected', connected);
        line.classList.toggle('context-dimmed', !connected);
      }});
      nodeGroups.forEach((candidateGroup, name) => {{
        candidateGroup.classList.toggle('neighbor', neighborNames.has(name));
        candidateGroup.classList.toggle('context-dimmed', name !== node.name && !neighborNames.has(name));
      }});
      selectedEdgeName = '';
      selectedNodeName = node.name;
      document.querySelectorAll('.section-item').forEach(item => item.classList.remove('active'));
      const prices = node.prices.filter(value => typeof value === 'number');
      const average = prices.length ? prices.reduce((sum, value) => sum + value, 0) / prices.length : null;
      const minimum = prices.length ? Math.min(...prices) : null;
      const maximum = prices.length ? Math.max(...prices) : null;
      const peakIndex = maximum === null ? -1 : node.prices.indexOf(maximum);
      document.getElementById('detailTitle').innerHTML = `${{escapeHtml(node.name)}} <span class="pill">${{escapeHtml(node.region || '未分片')}}</span>`;
      document.getElementById('detailSubtitle').textContent = `${{node.location_accuracy}} · ${{node.position_label}} · ${{relatedEdges.length}}条相邻线路`;
      document.getElementById('metricLabel1').textContent = '平均电价';
      document.getElementById('metricLabel2').textContent = '最低电价';
      document.getElementById('metricLabel3').textContent = '最高电价';
      document.getElementById('metricLabel4').textContent = '峰值时点';
      document.getElementById('maxGap').textContent = average === null ? '-' : average.toFixed(2);
      document.getElementById('intervals').textContent = minimum === null ? '-' : minimum.toFixed(2);
      document.getElementById('direction').textContent = maximum === null ? '-' : maximum.toFixed(2);
      document.getElementById('maxTime').textContent = peakIndex >= 0 ? (DATA.time_headers[peakIndex] || '-') : '-';
      document.getElementById('chartTitle').textContent = `${{node.name}} 24小时节点电价曲线`;
      drawNodeChart(node);
      if (window.innerWidth <= 920) document.getElementById('detailTitle').scrollIntoView({{ behavior:'smooth', block:'start' }});
    }}

    function renderSectionList() {{
      const query = document.getElementById('sectionSearch').value.trim().toLowerCase();
      const sort = document.getElementById('sectionSort').value;
      const minGap = Number(gapRange.value);
      document.getElementById('gapValue').textContent = minGap.toLocaleString('zh-CN');
      const filtered = sectionEdges.filter(edge => Number(edge.max_abs_diff) >= minGap &&
        [edge.edge, edge.group, edge.main_direction, edge.windows].join(' ').toLowerCase().includes(query));
      const visibleNames = new Set(filtered.map(edge => edge.edge));
      sectionEdgeLines.forEach((line, name) => line.style.display = visibleNames.has(name) ? '' : 'none');
      filtered.sort((a, b) => {{
        if (sort === 'name') return a.edge.localeCompare(b.edge, 'zh-CN');
        if (sort === 'gap') return b.max_abs_diff - a.max_abs_diff;
        if (sort === 'duration') return b.intervals - a.intervals;
        return b.severity_score - a.severity_score;
      }});
      const list = document.getElementById('sectionList');
      list.innerHTML = '';
      document.getElementById('visibleCount').textContent = `${{filtered.length}} 条`;
      if (!filtered.length) {{
        list.innerHTML = '<div class="empty-list">没有匹配的断面</div>';
        return;
      }}
      filtered.forEach(edge => {{
        const button = document.createElement('button');
        button.type = 'button';
        button.className = `section-item${{edge.edge === selectedEdgeName ? ' active' : ''}}`;
        button.dataset.edge = edge.edge;
        button.innerHTML = `<div class="item-top"><span>${{escapeHtml(edge.edge)}}</span><span class="item-gap">${{edge.max_abs_diff.toFixed(0)}}</span></div>
          <div class="item-meta">${{escapeHtml(edge.group)}} · ${{edge.intervals}} 个时段 · ${{escapeHtml(edge.main_direction || '方向未知')}}</div>`;
        button.addEventListener('click', () => {{
          const line = edgeLines.get(edge.edge);
          if (line) selectEdge(edge, line);
          if (window.innerWidth <= 920) document.getElementById('detailTitle').scrollIntoView({{ behavior: 'smooth', block: 'start' }});
        }});
        list.appendChild(button);
      }});
    }}

    document.getElementById('sectionSearch').addEventListener('input', renderSectionList);
    document.getElementById('sectionSort').addEventListener('change', renderSectionList);
    gapRange.addEventListener('input', () => {{
      document.querySelectorAll('.quick-filter').forEach(button => button.classList.remove('active'));
      renderSectionList();
    }});
    document.querySelectorAll('.quick-filter').forEach(button => button.addEventListener('click', () => {{
      document.querySelectorAll('.quick-filter').forEach(item => item.classList.toggle('active', item === button));
      gapRange.value = String(Math.round(absoluteMaxGap * Number(button.dataset.percent) / 50) * 50);
      renderSectionList();
    }}));
    function bindLayerToggle(buttonId, selector) {{
      document.getElementById(buttonId).addEventListener('click', event => {{
        const next = event.currentTarget.getAttribute('aria-pressed') !== 'true';
        event.currentTarget.setAttribute('aria-pressed', String(next));
        document.querySelectorAll(selector).forEach(element => element.classList.toggle('is-hidden', !next));
      }});
    }}
    bindLayerToggle('toggleTopology', '.edge:not(.section), .edge-hit:not(.section-hit)');
    bindLayerToggle('toggleLabels', '.label, .label-leader');
    const nodeSearch = document.getElementById('nodeSearch');
    function matchingNodes() {{
      const query = nodeSearch.value.trim().toLowerCase();
      return query ? DATA.nodes.filter(node => [node.name, node.region].join(' ').toLowerCase().includes(query)) : [];
    }}
    nodeSearch.addEventListener('input', () => {{
      const matches = new Set(matchingNodes().map(node => node.name));
      const hasQuery = Boolean(nodeSearch.value.trim());
      nodeGroups.forEach((group, name) => {{
        group.classList.toggle('search-match', matches.has(name));
        group.classList.toggle('search-dimmed', hasQuery && !matches.has(name));
      }});
    }});
    nodeSearch.addEventListener('keydown', event => {{
      if (event.key !== 'Enter') return;
      const node = matchingNodes()[0];
      const group = node ? nodeGroups.get(node.name) : null;
      if (!node || !group) return;
      event.preventDefault();
      selectNode(node, group);
      view.scale = 2.6;
      view.x = {SVG_W} / 2 - node.x * view.scale;
      view.y = {SVG_H} / 2 - node.y * view.scale;
      applyView();
    }});
    document.getElementById('focusSelected').addEventListener('click', () => {{
      const edge = sectionEdges.find(item => item.edge === selectedEdgeName);
      const node = nodeByName.get(selectedNodeName);
      if (!edge && !node) return;
      const centerX = node ? node.x : (edge.x1 + edge.x2) / 2;
      const centerY = node ? node.y : (edge.y1 + edge.y2) / 2;
      view.scale = 2.6;
      view.x = {SVG_W} / 2 - centerX * view.scale;
      view.y = {SVG_H} / 2 - centerY * view.scale;
      applyView();
      const target = node ? nodeGroups.get(node.name) : edgeLines.get(edge.edge);
      target?.classList.add('pulse');
      setTimeout(() => target?.classList.remove('pulse'), 2600);
    }});
    document.getElementById('focusMode').addEventListener('click', event => {{
      const active = !document.body.classList.contains('focus-mode');
      document.body.classList.toggle('focus-mode', active);
      event.currentTarget.setAttribute('aria-pressed', String(active));
      event.currentTarget.textContent = active ? '退出沉浸' : '沉浸地图';
    }});
    document.addEventListener('keydown', event => {{
      if (event.key === 'Escape' && document.body.classList.contains('focus-mode')) document.getElementById('focusMode').click();
    }});
    const themeStorageKey = '500kv-map-theme-v2';
    const savedTheme = localStorage.getItem(themeStorageKey);
    if (savedTheme === 'light' || savedTheme === 'dark') document.body.dataset.theme = savedTheme;
    function syncThemeLabel() {{
      document.getElementById('themeToggle').textContent = document.body.dataset.theme === 'dark' ? '切换浅色' : '切换深色';
    }}
    syncThemeLabel();
    document.getElementById('themeToggle').addEventListener('click', () => {{
      document.body.dataset.theme = document.body.dataset.theme === 'dark' ? 'light' : 'dark';
      localStorage.setItem(themeStorageKey, document.body.dataset.theme);
      syncThemeLabel();
    }});
    renderSectionList();

    function attachPriceCrosshair(svg, options) {{
      const {{ series, x, y, count, margin, width, height, make }} = options;
      if (!count || !series.length) return;
      const plotLeft = margin.left;
      const plotRight = width - margin.right;
      const plotTop = margin.top;
      const plotBottom = height - margin.bottom;
      const tooltipWidth = series.length > 1 ? 188 : 150;
      const tooltipHeight = 25 + series.length * 16;

      const layer = make('g', {{ class: 'price-crosshair', visibility: 'hidden', 'pointer-events': 'none' }});
      const guide = make('line', {{
        y1: plotTop,
        y2: plotBottom,
        stroke: '#60758a',
        'stroke-width': 1.2,
        'stroke-dasharray': '4 4',
        'vector-effect': 'non-scaling-stroke'
      }});
      layer.appendChild(guide);

      const markers = series.map(item => {{
        const marker = make('circle', {{
          r: 4.2,
          fill: item.color,
          stroke: '#ffffff',
          'stroke-width': 1.8,
          'vector-effect': 'non-scaling-stroke'
        }});
        layer.appendChild(marker);
        return marker;
      }});

      const tooltip = make('g', {{ class: 'price-crosshair-tooltip' }});
      tooltip.appendChild(make('rect', {{
        width: tooltipWidth,
        height: tooltipHeight,
        rx: 6,
        fill: 'var(--panel)',
        stroke: 'var(--line)',
        'stroke-width': 1,
        opacity: .96
      }}));
      const timeText = make('text', {{ x: 8, y: 15, 'font-size': 10, 'font-weight': 700, fill: 'var(--ink)' }});
      tooltip.appendChild(timeText);
      const valueTexts = series.map((item, seriesIndex) => {{
        const text = make('text', {{ x: 8, y: 31 + seriesIndex * 16, 'font-size': 10, fill: item.color }});
        tooltip.appendChild(text);
        return text;
      }});
      layer.appendChild(tooltip);
      svg.appendChild(layer);

      function showIndex(index) {{
        const safeIndex = Math.max(0, Math.min(count - 1, index));
        const guideX = x(safeIndex);
        guide.setAttribute('x1', guideX);
        guide.setAttribute('x2', guideX);
        timeText.textContent = `时间：${{DATA.time_headers[safeIndex] || String(safeIndex + 1)}}`;
        series.forEach((item, seriesIndex) => {{
          const value = item.prices[safeIndex];
          valueTexts[seriesIndex].textContent = `${{item.name}}：${{typeof value === 'number' ? value.toFixed(2) + ' 元/MWh' : '暂无'}}`;
          if (typeof value === 'number') {{
            markers[seriesIndex].setAttribute('cx', guideX);
            markers[seriesIndex].setAttribute('cy', y(value));
            markers[seriesIndex].removeAttribute('display');
          }} else {{
            markers[seriesIndex].setAttribute('display', 'none');
          }}
        }});
        const tooltipX = guideX + tooltipWidth + 10 > plotRight
          ? Math.max(plotLeft, guideX - tooltipWidth - 10)
          : guideX + 10;
        tooltip.setAttribute('transform', `translate(${{tooltipX}} ${{plotTop + 6}})`);
        layer.setAttribute('visibility', 'visible');
      }}

      function showFromPointer(event) {{
        const bounds = svg.getBoundingClientRect();
        const localX = (event.clientX - bounds.left) / bounds.width * width;
        const ratio = (Math.max(plotLeft, Math.min(plotRight, localX)) - plotLeft) / (plotRight - plotLeft || 1);
        showIndex(Math.round(ratio * (count - 1)));
      }}

      const hit = make('rect', {{
        class: 'price-crosshair-hit',
        x: plotLeft,
        y: plotTop,
        width: plotRight - plotLeft,
        height: plotBottom - plotTop,
        fill: 'transparent',
        'pointer-events': 'all'
      }});
      hit.addEventListener('pointerenter', showFromPointer);
      hit.addEventListener('pointermove', showFromPointer);
      hit.addEventListener('pointerdown', showFromPointer);
      hit.addEventListener('pointerleave', () => layer.setAttribute('visibility', 'hidden'));
      svg.appendChild(hit);
    }}

    function drawNodeChart(node) {{
      const svg = document.getElementById('priceChart');
      svg.innerHTML = '';
      const width = 360, height = 280;
      const margin = {{ left: 48, right: 12, top: 24, bottom: 34 }};
      const points = node.prices.map((value, index) => ({{ value, index }})).filter(point => typeof point.value === 'number');
      if (!points.length) {{
        const message = document.createElementNS('http://www.w3.org/2000/svg', 'text');
        message.setAttribute('x', width / 2); message.setAttribute('y', height / 2);
        message.setAttribute('text-anchor', 'middle'); message.setAttribute('fill', '#8296ad');
        message.textContent = '该节点缺少电价数据'; svg.appendChild(message);
        document.getElementById('sourceText').textContent = `节点 ${{node.name}} 暂无可用分时电价。`;
        return;
      }}
      const values = points.map(point => point.value);
      const min = Math.min(...values), max = Math.max(...values);
      const count = Math.max(DATA.time_headers.length, node.prices.length, 2);
      const x = index => margin.left + index / (count - 1) * (width - margin.left - margin.right);
      const y = value => margin.top + (max - value) / (max - min || 1) * (height - margin.top - margin.bottom);
      const make = (tag, attrs) => {{
        const element = document.createElementNS('http://www.w3.org/2000/svg', tag);
        Object.entries(attrs).forEach(([key, value]) => element.setAttribute(key, value));
        return element;
      }};
      for (let step = 0; step <= 4; step++) {{
        const gridY = margin.top + step / 4 * (height - margin.top - margin.bottom);
        svg.appendChild(make('line', {{ x1:margin.left, y1:gridY, x2:width-margin.right, y2:gridY, stroke:'#27445b', 'stroke-width':1 }}));
        const value = max - step / 4 * (max - min);
        const label = make('text', {{ x:margin.left-7, y:gridY+3, 'text-anchor':'end', 'font-size':9, fill:'#8296ad' }});
        label.textContent = Math.round(value); svg.appendChild(label);
      }}
      const areaPath = points.map((point, index) => `${{index ? 'L' : 'M'}} ${{x(point.index).toFixed(1)}} ${{y(point.value).toFixed(1)}}`).join(' ');
      const fillPath = `${{areaPath}} L ${{x(points.at(-1).index).toFixed(1)}} ${{height-margin.bottom}} L ${{x(points[0].index).toFixed(1)}} ${{height-margin.bottom}} Z`;
      svg.appendChild(make('path', {{ d:fillPath, fill:'rgba(52,183,255,.13)' }}));
      svg.appendChild(make('path', {{ d:areaPath, fill:'none', stroke:'#34b7ff', 'stroke-width':2.8, 'stroke-linejoin':'round' }}));
      const tickIndices = [0, Math.round((count-1)/4), Math.round((count-1)/2), Math.round((count-1)*3/4), count-1];
      tickIndices.forEach(index => {{
        const label = make('text', {{ x:x(index), y:height-11, 'text-anchor':'middle', 'font-size':9, fill:'#8296ad' }});
        label.textContent = DATA.time_headers[index] || String(index + 1); svg.appendChild(label);
      }});
      const last = points.at(-1);
      svg.appendChild(make('circle', {{ cx:x(last.index), cy:y(last.value), r:4, fill:'#38d6ff', stroke:'#07111f', 'stroke-width':2 }}));
      attachPriceCrosshair(svg, {{
        series: [{{ name: node.name, color: '#168ac0', prices: node.prices }}],
        x, y, count, margin, width, height, make
      }});
      document.getElementById('sourceText').innerHTML = `
        节点：${{escapeHtml(node.name)}}；片区：${{escapeHtml(node.region || '未分片')}}。<br>
        坐标精度：${{escapeHtml(node.location_accuracy)}}；来源：${{escapeHtml(node.location_source)}}。<br>
        有效电价时点：${{points.length}} / ${{node.prices.length}}。`;
    }}

    function drawChart(a, b, edge) {{
      const svg = document.getElementById('priceChart');
      svg.innerHTML = '';
      const width = 360, height = 280;
      const margin = {{ left: 42, right: 12, top: 18, bottom: 34 }};
      const values = [...a.prices, ...b.prices].filter(v => typeof v === 'number');
      if (!values.length) {{
        svg.textContent = '缺少节点电价曲线';
        return;
      }}
      const min = Math.min(...values, -50);
      const max = Math.max(...values);
      const x = i => margin.left + i / (DATA.time_headers.length - 1) * (width - margin.left - margin.right);
      const y = v => margin.top + (max - v) / (max - min || 1) * (height - margin.top - margin.bottom);
      const make = (tag, attrs) => {{
        const el = document.createElementNS('http://www.w3.org/2000/svg', tag);
        Object.entries(attrs).forEach(([k, v]) => el.setAttribute(k, v));
        return el;
      }};
      for (let t = 0; t <= 4; t++) {{
        const gy = margin.top + t / 4 * (height - margin.top - margin.bottom);
        svg.appendChild(make('line', {{ x1: margin.left, y1: gy, x2: width - margin.right, y2: gy, stroke: '#e1e8ef', 'stroke-width': 1 }}));
      }}
      svg.appendChild(make('line', {{ x1: margin.left, y1: margin.top, x2: margin.left, y2: height - margin.bottom, stroke: '#9aa8b5' }}));
      svg.appendChild(make('line', {{ x1: margin.left, y1: height - margin.bottom, x2: width - margin.right, y2: height - margin.bottom, stroke: '#9aa8b5' }}));

      function pathFor(prices) {{
        let drawing = false;
        return prices.map((value, index) => {{
          if (typeof value !== 'number') {{ drawing = false; return ''; }}
          const command = drawing ? 'L' : 'M';
          drawing = true;
          return `${{command}} ${{x(index).toFixed(1)}} ${{y(value).toFixed(1)}}`;
        }}).filter(Boolean).join(' ');
      }}
      svg.appendChild(make('path', {{ d: pathFor(a.prices), fill: 'none', stroke: '#1261d1', 'stroke-width': 2.8 }}));
      svg.appendChild(make('path', {{ d: pathFor(b.prices), fill: 'none', stroke: '#e02020', 'stroke-width': 2.8 }}));

      const ticks = [['00:00', 0], ['06:00', 24], ['12:00', 48], ['18:00', 72], ['23:45', 95]];
      ticks.forEach(([label, idx]) => {{
        const tx = x(idx);
        svg.appendChild(make('line', {{ x1: tx, y1: height - margin.bottom, x2: tx, y2: height - margin.bottom + 4, stroke: '#9aa8b5' }}));
        const text = make('text', {{ x: tx, y: height - 12, 'text-anchor': 'middle', 'font-size': 10, fill: '#657487' }});
        text.textContent = label;
        svg.appendChild(text);
      }});
      [min, max].forEach((val, idx) => {{
        const text = make('text', {{ x: margin.left - 8, y: y(val) + 3, 'text-anchor': 'end', 'font-size': 10, fill: '#657487' }});
        text.textContent = Math.round(val);
        svg.appendChild(text);
      }});

      const legendA = make('text', {{ x: margin.left, y: 12, 'font-size': 11, fill: '#2467c9' }});
      legendA.textContent = a.name;
      svg.appendChild(legendA);
      const legendB = make('text', {{ x: margin.left + 82, y: 12, 'font-size': 11, fill: '#d62f2f' }});
      legendB.textContent = b.name;
      svg.appendChild(legendB);

      attachPriceCrosshair(svg, {{
        series: [
          {{ name: a.name, color: '#1261d1', prices: a.prices }},
          {{ name: b.name, color: '#e02020', prices: b.prices }}
        ],
        x, y, count: DATA.time_headers.length, margin, width, height, make
      }});

      document.getElementById('sourceText').textContent = '';
    }}

    function escapeHtml(text) {{
      return String(text).replace(/[&<>"']/g, ch => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[ch]));
    }}

    if (sectionEdges.length) {{
      const strongest = [...sectionEdges].sort((a, b) => b.severity_score - a.severity_score)[0];
      const line = edgeLines.get(strongest.edge);
      if (line) selectEdge(strongest, line);
    }}
  </script>
</body>
</html>
"""


def main() -> None:
    parser = argparse.ArgumentParser(description="生成广东500kV节点断面交互HTML地图。")
    parser.add_argument("--topology", required=True, type=Path, help="拓扑 Markdown 文件路径")
    parser.add_argument("--original-prices", required=True, type=Path, help="原始节点电价 Excel 文件路径")
    parser.add_argument("--processed-prices", required=True, type=Path, help="处理后的500kV节点电价 Excel 文件路径")
    parser.add_argument("--analysis", required=True, type=Path, help="断面分析 Excel 文件路径")
    parser.add_argument("--output", required=True, type=Path, help="输出 HTML 文件路径")
    parser.add_argument("--node-positions", type=Path, help="已确认的接线图节点锚点 JSON；默认从拓扑文件旁的 data 目录自动发现")
    args = parser.parse_args()

    build_map_html(
        topology_path=args.topology,
        original_price_path=args.original_prices,
        processed_price_path=args.processed_prices,
        analysis_path=args.analysis,
        output_path=args.output,
        node_positions_path=args.node_positions,
    )
    print(f"HTML地图已生成: {args.output}")


if __name__ == "__main__":
    main()
