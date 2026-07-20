from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


TIME_HEADER_RE = re.compile(r"^\d{2}:\d{2}$")
SECTION_2_RE = re.compile(r"^##\s+2\.")
SECTION_NEXT_RE = re.compile(r"^##\s+3\.")
MIN_SECTION_INTERVALS = 4


@dataclass(frozen=True)
class TopologyEdge:
    group: str
    a: str
    b: str


@dataclass(frozen=True)
class AnalysisResult:
    output_path: Path
    unique_edges: int
    skipped_duplicate_edges: list[tuple[str, str, str]]
    matched_nodes: int
    missing_nodes: list[str]
    flagged_edges: int
    flagged_intervals: int


def analyze_section_gaps(
    topology_path: Path,
    workbook_path: Path,
    output_dir: Path,
    threshold: float = 100.0,
) -> AnalysisResult:
    topology_path = Path(topology_path)
    workbook_path = Path(workbook_path)
    output_dir = Path(output_dir)

    edges, skipped_duplicates = parse_topology_edges(topology_path)
    time_headers, price_rows = read_price_workbook(workbook_path)
    topology_nodes = sorted({edge.a for edge in edges} | {edge.b for edge in edges})
    matched_rows = match_price_rows(topology_nodes, price_rows)
    missing_nodes = [node for node in topology_nodes if node not in matched_rows]

    interval_records = []
    for edge in edges:
        if edge.a not in matched_rows or edge.b not in matched_rows:
            continue

        prices_a = matched_rows[edge.a]["prices"]
        prices_b = matched_rows[edge.b]["prices"]
        edge_records = []
        for time_index, time_header in enumerate(time_headers):
            if time_index == 0:
                continue

            price_a = prices_a[time_header]
            price_b = prices_b[time_header]
            previous_price_a = prices_a[time_headers[time_index - 1]]
            previous_price_b = prices_b[time_headers[time_index - 1]]
            if (
                price_a is None
                or price_b is None
                or previous_price_a is None
                or previous_price_b is None
            ):
                continue

            move_a = price_a - previous_price_a
            move_b = price_b - previous_price_b
            if move_a * move_b >= 0:
                continue

            diff_b_minus_a = price_b - price_a
            abs_diff = abs(diff_b_minus_a)
            if abs_diff <= threshold:
                continue

            high = edge.b if diff_b_minus_a > 0 else edge.a
            low = edge.a if diff_b_minus_a > 0 else edge.b
            edge_records.append(
                {
                    "group": edge.group,
                    "edge": f"{edge.a}-{edge.b}",
                    "a": edge.a,
                    "b": edge.b,
                    "time": time_header,
                    "price_a": price_a,
                    "price_b": price_b,
                    "move_a": move_a,
                    "move_b": move_b,
                    "diff_b_minus_a": diff_b_minus_a,
                    "abs_diff": abs_diff,
                    "low": low,
                    "high": high,
                    "direction": f"{low}->{high}",
                }
            )
        interval_records.extend(_sustained_records(edge_records, time_headers))

    edge_summary = summarize_edges(interval_records, time_headers)
    group_summary = summarize_groups(interval_records)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "500kv_section_price_gap_analysis.xlsx"

    write_analysis_workbook(
        output_path,
        [
            ("断面汇总", edge_summary),
            ("逐时段明细", interval_records),
            ("片区汇总", group_summary),
        ],
    )

    return AnalysisResult(
        output_path=output_path,
        unique_edges=len(edges),
        skipped_duplicate_edges=skipped_duplicates,
        matched_nodes=len(matched_rows),
        missing_nodes=missing_nodes,
        flagged_edges=len({record["edge"] for record in interval_records}),
        flagged_intervals=len(interval_records),
    )


def parse_topology_edges(topology_path: Path) -> tuple[list[TopologyEdge], list[tuple[str, str, str]]]:
    in_section = False
    current_group = ""
    edges = []
    seen: set[tuple[str, str]] = set()
    skipped_duplicates = []

    for raw_line in Path(topology_path).read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not in_section:
            in_section = bool(SECTION_2_RE.match(line))
            continue

        if SECTION_NEXT_RE.match(line):
            break
        if not line:
            continue
        if line.startswith("### "):
            current_group = line[4:].strip()
            continue
        if not current_group:
            continue

        if "：" in line:
            a, neighbors_text = (value.strip() for value in line.split("：", 1))
            parts = [f"{a}-{neighbor.strip()}" for neighbor in re.split(r"[，,、]", neighbors_text) if neighbor.strip()]
        else:
            parts = re.split(r"[，,、]", line)

        for part in parts:
            part = part.strip()
            if not part or "-" not in part:
                continue
            a, b = (value.strip() for value in part.split("-", 1))
            edge_key = tuple(sorted((a, b)))
            if edge_key in seen:
                skipped_duplicates.append((a, b, current_group))
                continue
            seen.add(edge_key)
            edges.append(TopologyEdge(current_group, a, b))

    if not edges:
        raise ValueError(f"拓扑文件未解析到“## 2.”下的边: {topology_path}")
    return edges, skipped_duplicates


def read_price_workbook(workbook_path: Path) -> tuple[list[str], list[dict[str, object]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue

        column_map = _column_map(header)
        if column_map is None:
            continue

        price_rows = []
        for row in rows:
            node_name = _cell(row, column_map["node"])
            if node_name is None or str(node_name).strip() == "":
                continue

            prices = {}
            for index, time_header in zip(column_map["time_columns"], column_map["time_headers"]):
                prices[time_header] = _to_float(_cell(row, index))

            price_rows.append(
                {
                    "raw_name": str(node_name).strip(),
                    "normalized_name": normalize_price_node_name(node_name),
                    "prices": prices,
                }
            )

        if price_rows:
            return column_map["time_headers"], price_rows

    raise ValueError(f"未找到包含“节点名称”和分时时间列的工作表: {workbook_path}")


def match_price_rows(topology_nodes: Iterable[str], price_rows: list[dict[str, object]]) -> dict[str, dict[str, object]]:
    matched = {}
    for node in topology_nodes:
        candidates = [row for row in price_rows if row["normalized_name"] == node]
        if not candidates:
            candidates = [row for row in price_rows if node in str(row["raw_name"])]
        if not candidates:
            continue

        candidates.sort(key=lambda row: _price_row_priority(str(row["raw_name"])))
        matched[node] = candidates[0]
    return matched


def normalize_price_node_name(name: object) -> str:
    text = "" if name is None else str(name).strip()
    text = re.sub(r"^其他", "", text)
    text = text.replace("开关站", "")
    text = text.replace("换流站", "")
    text = text.replace("站", "")
    text = text.replace("变", "")
    text = re.sub(r"\(500kV\)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"500\s*kV.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"220\s*kV.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"#.*$", "", text)
    text = re.sub(r"\d+母$", "", text)
    return text.strip()


def summarize_edges(records: list[dict[str, object]], time_headers: list[str]) -> list[dict[str, object]]:
    rows = []
    for edge in sorted({str(record["edge"]) for record in records}):
        edge_records = [record for record in records if record["edge"] == edge]
        max_record = max(edge_records, key=lambda record: float(record["abs_diff"]))
        direction_counts = _direction_counts(edge_records)
        main_direction = max(direction_counts.items(), key=lambda item: (item[1], item[0]))[0]
        rows.append(
            {
                "group": max_record["group"],
                "edge": edge,
                "intervals": len(edge_records),
                "windows": _windows([str(record["time"]) for record in edge_records], time_headers),
                "max_abs_diff": _round(max_record["abs_diff"]),
                "max_time": max_record["time"],
                "max_low": max_record["low"],
                "max_high": max_record["high"],
                "main_direction": main_direction,
                "direction_counts": "; ".join(
                    f"{direction}:{count}" for direction, count in direction_counts.items()
                ),
            }
        )
    return rows


def summarize_groups(records: list[dict[str, object]]) -> list[dict[str, object]]:
    rows = []
    for group in sorted({str(record["group"]) for record in records}):
        group_records = [record for record in records if record["group"] == group]
        rows.append(
            {
                "group": group,
                "flagged_edges": len({record["edge"] for record in group_records}),
                "flagged_intervals": len(group_records),
                "max_abs_diff": _round(max(record["abs_diff"] for record in group_records)),
            }
        )
    rows.sort(key=lambda row: (-float(row["max_abs_diff"]), str(row["group"])))
    return rows


def write_analysis_workbook(path: Path, sheets: list[tuple[str, list[dict[str, object]]]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, rows in sheets:
        sheet = workbook.create_sheet(sheet_name)
        if not rows:
            continue

        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])

        for column_cells in sheet.columns:
            max_length = max(len("" if cell.value is None else str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 80)
        sheet.freeze_panes = "A2"

    workbook.save(path)


def _column_map(header: tuple[object, ...]) -> dict[str, object] | None:
    cleaned = ["" if value is None else str(value).strip() for value in header]
    try:
        node_index = cleaned.index("节点名称")
    except ValueError:
        return None

    time_columns = [index for index, value in enumerate(cleaned) if TIME_HEADER_RE.match(value)]
    if not time_columns:
        return None

    return {
        "node": node_index,
        "time_columns": time_columns,
        "time_headers": [cleaned[index] for index in time_columns],
    }


def _price_row_priority(raw_name: str) -> tuple[int, str]:
    is_500kv_not_220kv = "500kv" in raw_name.lower() and "220kv" not in raw_name.lower()
    return (0 if is_500kv_not_220kv else 1, raw_name)


def _to_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text == "":
        return None
    return float(text)


def _cell(row: tuple[object, ...], index: int) -> object:
    return row[index] if index < len(row) else None


def _direction_counts(records: list[dict[str, object]]) -> dict[str, int]:
    counts = {}
    for record in records:
        direction = str(record["direction"])
        counts[direction] = counts.get(direction, 0) + 1
    return dict(sorted(counts.items()))


def _sustained_records(records: list[dict[str, object]], time_headers: list[str]) -> list[dict[str, object]]:
    if not records:
        return []

    time_index = {time_header: index for index, time_header in enumerate(time_headers)}
    sorted_records = sorted(records, key=lambda record: time_index[str(record["time"])])
    sustained = []
    run = [sorted_records[0]]

    for record in sorted_records[1:]:
        previous_index = time_index[str(run[-1]["time"])]
        current_index = time_index[str(record["time"])]
        if current_index == previous_index + 1:
            run.append(record)
            continue

        if len(run) >= MIN_SECTION_INTERVALS:
            sustained.extend(run)
        run = [record]

    if len(run) >= MIN_SECTION_INTERVALS:
        sustained.extend(run)
    return sustained


def _windows(times: list[str], time_headers: list[str]) -> str:
    time_index = {time_header: index for index, time_header in enumerate(time_headers)}
    indexes = sorted(time_index[time] for time in times)
    windows = []
    start = previous = indexes[0]
    for index in indexes[1:]:
        if index == previous + 1:
            previous = index
            continue
        windows.append((time_headers[start], time_headers[previous]))
        start = previous = index
    windows.append((time_headers[start], time_headers[previous]))
    return "; ".join(start if start == end else f"{start}-{end}" for start, end in windows)


def _round(value: object) -> str:
    return f"{float(value):.2f}"


def main() -> None:
    parser = argparse.ArgumentParser(description="按500kV拓扑相连节点价差筛选潮流断面。")
    parser.add_argument("--topology", required=True, type=Path, help="拓扑 Markdown 文件路径")
    parser.add_argument("--prices", required=True, type=Path, help="500kV节点电价 Excel 文件路径")
    parser.add_argument("--output-dir", default=Path("."), type=Path, help="Excel 输出目录")
    parser.add_argument("--threshold", default=100.0, type=float, help="价差阈值，默认 100 元/MWh")
    args = parser.parse_args()

    result = analyze_section_gaps(
        topology_path=args.topology,
        workbook_path=args.prices,
        output_dir=args.output_dir,
        threshold=args.threshold,
    )

    print(f"拓扑边数: {result.unique_edges}")
    print(f"匹配节点数: {result.matched_nodes}")
    print(f"缺失节点: {', '.join(result.missing_nodes) if result.missing_nodes else '无'}")
    print(f"触发断面边数: {result.flagged_edges}")
    print(f"触发时段数: {result.flagged_intervals}")
    print(f"分析结果: {result.output_path}")


if __name__ == "__main__":
    main()
