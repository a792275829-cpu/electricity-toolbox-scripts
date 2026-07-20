import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from src.analyze_500kv_section_gaps import TopologyEdge
from src.build_500kv_map_html import (
    SVG_H,
    SVG_W,
    _build_diagram_coordinates,
    _build_edge_payloads,
    _parse_node_regions,
    build_map_html,
)


class Build500kVMapHtmlTest(unittest.TestCase):
    def test_parse_node_regions_accepts_generic_numbered_section_heading(self):
        with tempfile.TemporaryDirectory() as tmp:
            topology_path = Path(tmp) / "topology.md"
            topology_path.write_text(
                """# test

## 1. 节点分区

### 粤北、清远
凤城站、水乡站

## 2. 已确认的接线拓扑
凤城站-水乡站
""",
                encoding="utf-8",
            )

            self.assertEqual(
                _parse_node_regions(topology_path),
                {"凤城站": "粤北、清远", "水乡站": "粤北、清远"},
            )

    def test_build_diagram_coordinates_requires_verified_anchors_and_preserves_layout(self):
        with tempfile.TemporaryDirectory() as tmp:
            positions_path = Path(tmp) / "positions.json"
            positions_path.write_text(
                json.dumps(
                    {
                        "image": {"width": 7111, "height": 5025, "path": "接线图.png"},
                        "nodes": [
                            {"name": "凤城站", "status": "verified", "anchor": {"x": 100, "y": 200}},
                            {"name": "水乡站", "status": "verified", "anchor": {"x": 7000, "y": 4900}},
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            coordinates, metadata = _build_diagram_coordinates(["凤城站", "水乡站"], positions_path)

            self.assertEqual(metadata["coordinate_mode"], "wiring-diagram")
            self.assertLess(coordinates["凤城站"]["x"], coordinates["水乡站"]["x"])
            self.assertLess(coordinates["凤城站"]["y"], coordinates["水乡站"]["y"])
            self.assertGreaterEqual(coordinates["凤城站"]["x"], 0)
            self.assertLessEqual(coordinates["水乡站"]["x"], SVG_W)
            self.assertLessEqual(coordinates["水乡站"]["y"], SVG_H)

    def test_build_diagram_coordinates_rejects_unverified_nodes(self):
        with tempfile.TemporaryDirectory() as tmp:
            positions_path = Path(tmp) / "positions.json"
            positions_path.write_text(
                json.dumps(
                    {
                        "image": {"width": 7111, "height": 5025},
                        "nodes": [{"name": "凤城站", "status": "candidate", "anchor": {"x": 100, "y": 200}}],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            with self.assertRaisesRegex(ValueError, "凤城站"):
                _build_diagram_coordinates(["凤城站"], positions_path)

    def test_build_map_html_embeds_nodes_edges_and_price_series(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            original_path = tmp_path / "original.xlsx"
            processed_path = tmp_path / "processed.xlsx"
            analysis_path = tmp_path / "analysis.xlsx"
            topology_path = tmp_path / "topology.md"
            output_path = tmp_path / "map.html"

            self._write_price_workbook(original_path, extra_rows=1)
            self._write_price_workbook(processed_path, extra_rows=0)
            self._write_analysis_workbook(analysis_path)
            topology_path.write_text(
                """# test

## 2. 跨片区主通道边

### 佛山片 - 东莞片
凤城-水乡、凤城-广南

## 3. next
""",
                encoding="utf-8",
            )

            build_map_html(
                topology_path=topology_path,
                original_price_path=original_path,
                processed_price_path=processed_path,
                analysis_path=analysis_path,
                output_path=output_path,
            )

            html = output_path.read_text(encoding="utf-8")
            self.assertIn("广东省电网 500kV 节点断面地图", html)
            self.assertIn('"edge": "凤城-水乡"', html)
            self.assertIn('"is_section": true', html)
            self.assertIn('"edge": "凤城-广南"', html)
            self.assertEqual(html.count('"is_section": true'), 2)
            self.assertIn('"name": "凤城"', html)
            self.assertIn('"prices": [100.0, 120.0]', html)
            self.assertIn('"original_rows": 3', html)
            self.assertIn('"processed_rows": 2', html)
            self.assertIn('"severity_basis": "intervals_and_max_abs_diff"', html)
            self.assertIn('"min_section_intervals": 4', html)
            self.assertIn('累计时点更多、价差更大', html)
            self.assertNotIn('原始数据行数：', html)
            self.assertNotIn('处理后500kV行数：', html)
            self.assertNotIn('缺少价格节点：', html)
            self.assertIn('id="zoomIn"', html)
            self.assertIn('id="mapViewport"', html)
            self.assertIn('function panByWheel(deltaX, deltaY)', html)
            self.assertIn('function isLikelyTrackpadWheel(event)', html)
            self.assertIn("event.ctrlKey || event.metaKey", html)
            self.assertIn("mapSvg.addEventListener('gesturechange'", html)
            self.assertIn('touch-action: none', html)
            self.assertIn("event.target.closest?.('.node-group, .edge, .edge-hit')", html)
            self.assertIn('Math.hypot(moveX, moveY) < 4', html)
            self.assertIn('const shouldClearSelection = !drag.moved;', html)
            self.assertIn('if (shouldClearSelection) clearSelection();', html)
            self.assertIn("hitLine.setAttribute('class', edge.is_section ? 'edge-hit section-hit' : 'edge-hit');", html)
            self.assertIn('.edge-hit:focus { outline: none; }', html)
            self.assertIn('.edge-hit:focus-visible + .edge {', html)
            self.assertIn("hitLine.addEventListener('pointerdown', event => {", html)
            self.assertIn('event.preventDefault();', html)
            self.assertIn("selectEdge(edge, line);", html)
            self.assertIn('function edgePriceStats(a, b)', html)
            self.assertIn('完整拓扑线路，未触发断面判定；以下显示两端节点电价。', html)
            self.assertIn("document.getElementById('chartTitle').textContent = `${edge.a} / ${edge.b} 24小时节点电价曲线`;", html)
            self.assertIn("const command = drawing ? 'L' : 'M';", html)
            self.assertIn("bindLayerToggle('toggleTopology', '.edge:not(.section), .edge-hit:not(.section-hit)');", html)
            self.assertIn('"blocked": true', html)
            self.assertIn('label.setAttribute(\'class\', node.blocked ? \'label blocked-label\' : \'label\');', html)
            self.assertIn('24小时节点电价曲线', html)
            self.assertIn('"regions":', html)
            self.assertIn('"location_accuracy": "公开站址"', html)
            self.assertIn('"lon": 113.08991', html)
            self.assertIn('id="regions"', html)
            self.assertIn('id="sectionSearch"', html)
            self.assertIn('id="sectionList"', html)
            self.assertNotIn('id="toggleRegions"', html)
            self.assertNotIn('regionData.forEach(region =>', html)
            self.assertNotIn('119个节点按已确认接线图锚点排布', html)
            self.assertNotIn('灰蓝线为完整拓扑边', html)
            self.assertIn('id="summaryMaxGap"', html)
            self.assertIn('function renderSectionList()', html)
            self.assertIn('id="gapRange"', html)
            self.assertIn('id="focusSelected"', html)
            self.assertIn('id="focusMode"', html)
            self.assertIn('id="themeToggle"', html)
            self.assertIn('<body data-theme="light">', html)
            self.assertIn("500kv-map-theme-v2", html)
            self.assertIn('body[data-theme="light"] .node', html)
            self.assertIn("group.setAttribute('class', 'node-group');", html)
            self.assertIn('function layoutNodeLabels()', html)
            self.assertIn('function boxesOverlap(a, b, padding = 4)', html)
            self.assertIn('function boxHitsEdge(box, edge, padding = 5)', html)
            self.assertIn('const distances = item.node.blocked ? [12, 16, 20, 24, 28, 32, 36, 40] : [8, 12, 16, 20, 24, 28, 32, 36];', html)
            self.assertIn('if (inBounds && labelOverlap === 0 && nodeOverlap === 0', html)
            self.assertIn('item.label.dataset.distance = String(chosen.distance);', html)
            self.assertIn('const moved = chosen.distance >= 20 && chosen.distance <= 28;', html)
            self.assertNotIn('116, 140', html)
            self.assertIn('id="labelLeaders"', html)
            self.assertIn('"layout_canvas_width": 1500', html)
            self.assertIn("label.addEventListener('click'", html)
            self.assertIn("hit.setAttribute('r', '12');", html)
            self.assertIn("bindLayerToggle('toggleLabels', '.label, .label-leader');", html)
            self.assertIn('function selectNode(node, group)', html)
            self.assertIn('function clearSelection()', html)
            self.assertIn("element.classList.remove('active', 'neighbor', 'context-dimmed')", html)
            self.assertIn('点击节点或线路查看电价曲线', html)
            self.assertIn('function drawNodeChart(node)', html)
            self.assertIn('function attachPriceCrosshair(svg, options)', html)
            self.assertIn("hit.addEventListener('pointermove', showFromPointer);", html)
            self.assertIn("hit.addEventListener('pointerdown', showFromPointer);", html)
            self.assertIn("'stroke-dasharray': '4 4'", html)
            self.assertIn("value.toFixed(2) + ' 元/MWh'", html)
            self.assertIn("series: [{ name: node.name, color: '#168ac0', prices: node.prices }]", html)
            self.assertIn('点击查看电价数据', html)

    def test_edge_color_severity_increases_with_intervals_and_gap(self):
        edges = [
            TopologyEdge("测试", "甲", "乙"),
            TopologyEdge("测试", "甲", "丙"),
            TopologyEdge("测试", "甲", "丁"),
        ]
        summaries = [
            {"edge": "甲-乙", "intervals": 4, "max_abs_diff": 200},
            {"edge": "甲-丙", "intervals": 8, "max_abs_diff": 200},
            {"edge": "甲-丁", "intervals": 8, "max_abs_diff": 400},
        ]
        coordinates = {
            "甲": {"x": 0, "y": 0},
            "乙": {"x": 1, "y": 1},
            "丙": {"x": 2, "y": 2},
            "丁": {"x": 3, "y": 3},
        }

        payloads = _build_edge_payloads(edges, summaries, [], coordinates)

        self.assertLess(payloads[0]["severity"], payloads[1]["severity"])
        self.assertLess(payloads[1]["severity"], payloads[2]["severity"])

    def _write_price_workbook(self, path: Path, extra_rows: int) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "实时节点电价查询"
        ws.append(["节点名称", "数据项", "00:00", "00:15"])
        ws.append(["其他凤城站500kV", "电价(元/MWh)", 100, 120])
        ws.append(["其他水乡站500kV", "电价(元/MWh)", 800, 900])
        for index in range(extra_rows):
            ws.append([f"其他非500kV节点{index}", "电价(元/MWh)", 1, 2])
        wb.save(path)

    def _write_analysis_workbook(self, path: Path) -> None:
        wb = Workbook()
        summary = wb.active
        summary.title = "断面汇总"
        summary.append(
            [
                "group",
                "edge",
                "intervals",
                "windows",
                "max_abs_diff",
                "max_time",
                "max_low",
                "max_high",
                "main_direction",
                "direction_counts",
            ]
        )
        summary.append(["佛山片 - 东莞片", "凤城-水乡", 2, "00:00-00:15", "780.00", "00:15", "凤城", "水乡", "凤城->水乡", "凤城->水乡:2"])
        summary.append(["佛山片 - 东莞片", "凤城-广南", 4, "00:15-01:00", "1200.00", "00:15", "凤城", "广南", "凤城->广南", "凤城->广南:4"])
        summary.append(["佛山片 - 东莞片", "凤城-水乡", 5, "00:15-01:15", "780.00", "00:15", "凤城", "水乡", "凤城->水乡", "凤城->水乡:5"])

        details = wb.create_sheet("逐时段明细")
        details.append(["group", "edge", "a", "b", "time", "price_a", "price_b", "diff_b_minus_a", "abs_diff", "low", "high", "direction"])
        details.append(["佛山片 - 东莞片", "凤城-水乡", "凤城", "水乡", "00:00", 100, 800, 700, 700, "凤城", "水乡", "凤城->水乡"])
        details.append(["佛山片 - 东莞片", "凤城-水乡", "凤城", "水乡", "00:15", 120, 900, 780, 780, "凤城", "水乡", "凤城->水乡"])

        group = wb.create_sheet("片区汇总")
        group.append(["group", "flagged_edges", "flagged_intervals", "max_abs_diff"])
        group.append(["佛山片 - 东莞片", 1, 2, "780.00"])
        wb.save(path)


if __name__ == "__main__":
    unittest.main()
