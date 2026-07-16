import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill


def write_d1_price_workbook(path: Path, prediction: pd.DataFrame) -> None:
    """Write the operational deliverable: D+1's 96 quarter-hour prices."""
    if prediction.empty:
        raise ValueError("没有可导出的 D+1 预测结果")
    first_date = pd.to_datetime(prediction["日期"]).min()
    d1 = prediction.loc[pd.to_datetime(prediction["日期"]).eq(first_date)].copy()
    d1 = d1[["日期", "时刻", "预测日前电价"]].rename(
        columns={"预测日前电价": "日前电价"}
    )
    d1 = d1.sort_values("时刻").reset_index(drop=True)
    if len(d1) != 96 or d1["时刻"].nunique() != 96:
        raise ValueError(f"D+1 输出必须包含 96 个时点，当前为 {len(d1)} 行")
    d1["日期"] = pd.to_datetime(d1["日期"]).dt.date

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        d1.to_excel(writer, sheet_name="D+1日前电价", index=False)
        sheet = writer.book["D+1日前电价"]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        widths = {"A": 14, "B": 10, "C": 16}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet["C"][1:]:
            cell.number_format = "0.00"


def write_workbook(
    path: Path,
    summary: dict,
    validation: pd.DataFrame,
    prediction: pd.DataFrame,
    daily: pd.DataFrame,
    by_slot: pd.DataFrame,
    experiments: pd.DataFrame,
    importance: pd.DataFrame,
) -> None:
    summary_frame = pd.DataFrame({"指标": list(summary), "数值": [json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v for v in summary.values()]})
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary_frame.to_excel(writer, sheet_name="模型摘要", index=False)
        validation.to_excel(writer, sheet_name="封存测试预测", index=False)
        prediction.to_excel(writer, sheet_name="待预测日期结果", index=False)
        daily.to_excel(writer, sheet_name="逐日MAE", index=False)
        by_slot.to_excel(writer, sheet_name="分时点MAE", index=False)
        experiments.to_excel(writer, sheet_name="滚动验证实验", index=False)
        importance.to_excel(writer, sheet_name="特征重要性", index=False)
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            for column_cells in sheet.columns:
                width = min(max(len(str(c.value or "")) for c in list(column_cells)[:200]) + 2, 32)
                sheet.column_dimensions[column_cells[0].column_letter].width = max(width, 10)


def model_feature_importance(model, original_columns: pd.Index) -> pd.DataFrame:
    estimator = model.named_steps["model"]
    if not hasattr(estimator, "feature_importances_"):
        return pd.DataFrame(columns=["特征", "重要性"])
    names = model.named_steps["imputer"].get_feature_names_out(original_columns)
    return pd.DataFrame({"特征": names, "重要性": estimator.feature_importances_}).sort_values("重要性", ascending=False)
