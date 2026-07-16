from __future__ import annotations

import argparse
import datetime as dt
import math
import re
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl


SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from generate_electricity_report import ReportBuilder, THERMAL_COMPANIES, fmt


DEFAULT_INPUT_DIR = Path.home() / "Downloads" / "haimen shouxian"
DEFAULT_TOP = 10
DEFAULT_UNIT_FILTERS = ["海门#1机组", "海门#2机组", "海门#3机组", "海门#4机组"]


@dataclass(frozen=True)
class NormalOutputRange:
    lower: float | None
    upper: float | None
    label: str
    usable: bool = True


@dataclass(frozen=True)
class HourDeviation:
    market_type: str
    date: str
    unit: str
    time: str
    clearing_mwh: float
    clearing_price: float
    plant_output: float
    normal_range: str
    deviation_mw: float
    source_file: str


@dataclass(frozen=True)
class PlantDayScore:
    market_type: str
    date: str
    total_deviation_mw: float
    unit_count: int
    point_count: int
    source_file: str


def parse_date_text(value: str) -> dt.date:
    match = re.fullmatch(r"\s*((?:19|20)\d{2})[-./年](\d{1,2})[-./月](\d{1,2})日?\s*", value)
    if not match:
        match = re.fullmatch(r"\s*((?:19|20)\d{2})(\d{2})(\d{2})\s*", value)
    if not match:
        raise ValueError(f"日期格式不支持：{value}，请使用 YYYY-MM-DD")
    year, month, day = (int(part) for part in match.groups())
    return dt.date(year, month, day)


def parse_date_from_filename(path: Path) -> dt.date | None:
    match = re.search(r"((?:19|20)\d{2})[.\-_/年](\d{1,2})[.\-_/月](\d{1,2})", path.stem)
    if not match:
        match = re.search(r"((?:19|20)\d{2})(\d{2})(\d{2})", path.stem)
    if not match:
        return None
    year, month, day = (int(part) for part in match.groups())
    return dt.date(year, month, day)


def discover_workbooks(input_dir: Path, start_date: dt.date | None, end_date: dt.date | None) -> list[Path]:
    if input_dir.is_file():
        candidates = [input_dir]
    else:
        candidates = [
            path
            for path in input_dir.rglob("*.xlsx")
            if "出清情况" in path.name and not path.name.startswith("~$")
        ]

    by_date: dict[dt.date, Path] = {}
    for path in candidates:
        report_date = parse_date_from_filename(path)
        if report_date is None:
            continue
        if start_date and report_date < start_date:
            continue
        if end_date and report_date > end_date:
            continue
        current = by_date.get(report_date)
        if current is None or path.stat().st_mtime >= current.stat().st_mtime:
            by_date[report_date] = path
    return [by_date[day] for day in sorted(by_date)]


def normal_output_range(card: dict, clearing_mwh: float, clearing_price: float) -> NormalOutputRange:
    if abs(clearing_mwh) <= 0.05:
        return NormalOutputRange(lower=0.0, upper=0.0, label="0")

    segments = card.get("segments", [])
    if not segments:
        return NormalOutputRange(lower=None, upper=None, label="-", usable=False)

    lower = segments[0]["start"]
    upper = None
    for segment in segments:
        price = segment["price"]
        start = segment["start"]
        end = segment["end"]
        if clearing_price < price:
            upper = start if upper is None else min(upper, start)
        elif abs(clearing_price - price) <= 0.05:
            lower = max(lower, start)
            upper = end if upper is None else min(upper, end)
        else:
            lower = max(lower, end)

    if upper is not None and lower > upper + 0.05:
        return NormalOutputRange(lower=lower, upper=upper, label="无有效范围", usable=False)
    if upper is None:
        return NormalOutputRange(lower=lower, upper=None, label=f">={fmt(lower)}MW")
    return NormalOutputRange(lower=lower, upper=upper, label=f"{fmt(lower)}-{fmt(upper)}MW")


def output_range_distance(normal_range: NormalOutputRange, plant_output: float) -> float:
    if not normal_range.usable:
        return math.nan
    if normal_range.lower is not None and plant_output < normal_range.lower:
        return normal_range.lower - plant_output
    if normal_range.upper is not None and plant_output > normal_range.upper:
        return plant_output - normal_range.upper
    return 0.0


def load_report_builder(path: Path) -> ReportBuilder:
    builder = ReportBuilder(path)
    builder.wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        builder.offer_cards = builder._parse_offer_cards()
        for company in THERMAL_COMPANIES:
            if company in builder.wb.sheetnames:
                builder.hourly.update(builder._parse_blocks(company, company))
        builder.compare_mode = "realtime_vs_dayahead" if builder._has_realtime_data() else "dayahead_vs_ml"
        return builder
    except Exception:
        builder.wb.close()
        raise


def workbook_hour_deviations(path: Path, unit_filters: list[str] | None = None) -> list[HourDeviation]:
    builder = load_report_builder(path)
    try:
        report_date = builder._report_date_text()
        rows = []
        for unit, hourly_points in builder.hourly.items():
            if unit_filters and not any(unit_filter in unit for unit_filter in unit_filters):
                continue
            card = builder.offer_cards.get(unit)
            if not card:
                continue
            plant_rate = card["plant_rate"]
            for item in hourly_points:
                for market_type, mwh_key, price_key in (
                    ("日前", "da", "da_price"),
                    ("实时", "rt", "rt_price"),
                ):
                    clearing_mwh = item.get(mwh_key, 0.0)
                    clearing_price = item.get(price_key, 0.0)
                    if market_type == "实时" and abs(clearing_mwh) <= 1e-6 and abs(clearing_price) <= 1e-6:
                        continue

                    plant_output = 0.0 if abs(clearing_mwh) <= 0.05 else clearing_mwh / (1 - plant_rate)
                    normal_range = normal_output_range(card, clearing_mwh, clearing_price)
                    deviation_mw = output_range_distance(normal_range, plant_output)
                    if math.isnan(deviation_mw):
                        continue
                    rows.append(
                        HourDeviation(
                            market_type=market_type,
                            date=report_date,
                            unit=unit,
                            time=str(item.get("time", "")),
                            clearing_mwh=clearing_mwh,
                            clearing_price=clearing_price,
                            plant_output=plant_output,
                            normal_range=normal_range.label,
                            deviation_mw=deviation_mw,
                            source_file=str(path),
                        )
                    )
        return rows
    finally:
        builder.wb.close()


def has_start_stop(rows: list[HourDeviation]) -> bool:
    by_unit: dict[str, list[HourDeviation]] = {}
    for row in rows:
        by_unit.setdefault(row.unit, []).append(row)

    for unit_rows in by_unit.values():
        statuses = [abs(row.clearing_mwh) <= 0.05 for row in unit_rows]
        if len(set(statuses)) > 1:
            return True
    return False


def rank_plant_days_with_exclusions(rows: list[HourDeviation], top: int) -> tuple[list[PlantDayScore], list[tuple[str, str]]]:
    by_date: dict[str, list[HourDeviation]] = {}
    for row in rows:
        by_date.setdefault(row.date, []).append(row)
    excluded_dates = {
        date_text
        for date_text, date_rows in by_date.items()
        if has_start_stop(date_rows)
    }

    grouped: dict[tuple[str, str], list[HourDeviation]] = {}
    for row in rows:
        grouped.setdefault((row.market_type, row.date), []).append(row)

    scores = []
    for (market_type, date_text), day_rows in grouped.items():
        if date_text in excluded_dates:
            continue
        scores.append(
            PlantDayScore(
                market_type=market_type,
                date=date_text,
                total_deviation_mw=round(sum(row.deviation_mw for row in day_rows), 6),
                unit_count=len({row.unit for row in day_rows}),
                point_count=len(day_rows),
                source_file=day_rows[0].source_file,
            )
        )

    scores.sort(key=lambda item: item.total_deviation_mw, reverse=True)
    excluded = [(date_text, "存在机组出力状态变化") for date_text in sorted(excluded_dates)]
    return top_per_market(scores, top), excluded


def rank_plant_days(rows: list[HourDeviation], top: int) -> list[PlantDayScore]:
    scores, _excluded = rank_plant_days_with_exclusions(rows, top)
    return scores


def top_per_market(scores: list[PlantDayScore], top: int) -> list[PlantDayScore]:
    grouped: dict[str, list[PlantDayScore]] = {}
    for score in scores:
        grouped.setdefault(score.market_type, []).append(score)

    limited = []
    for market_type in ("日前", "实时"):
        limited.extend(grouped.get(market_type, [])[:top])
    return limited


def write_results(output_path: Path, scores: list[PlantDayScore], excluded: list[tuple[str, str]], rows: list[HourDeviation]) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for market_type in ("日前", "实时"):
        ws = wb.create_sheet(f"{market_type}Top天")
        ws.append(["排名", "日期", "全厂全天偏差绝对值加和(MW)", "机组数", "有效时点数", "来源文件"])
        for index, score in enumerate([item for item in scores if item.market_type == market_type], start=1):
            ws.append(
                [
                    index,
                    score.date,
                    round(score.total_deviation_mw, 3),
                    score.unit_count,
                    score.point_count,
                    score.source_file,
                ]
            )

    detail = wb.create_sheet("逐小时明细")
    detail.append(["口径", "日期", "机组", "时点", "出清电量(MWh)", "出清价格(元/MWh)", "机端出力(MW)", "正常出力范围", "偏差绝对值(MW)", "来源文件"])
    for row in sorted(rows, key=lambda item: (item.market_type, item.date, item.unit, item.time)):
        detail.append(
            [
                row.market_type,
                row.date,
                row.unit,
                row.time,
                round(row.clearing_mwh, 3),
                round(row.clearing_price, 3),
                round(row.plant_output, 3),
                row.normal_range,
                round(row.deviation_mw, 3),
                row.source_file,
            ]
        )

    excluded_sheet = wb.create_sheet("剔除日期")
    excluded_sheet.append(["日期", "剔除原因"])
    for date_text, reason in excluded:
        excluded_sheet.append([date_text, reason])

    note = wb.create_sheet("说明")
    note.append(["项目", "说明"])
    note.append(["计算口径", "机端出力=出清电量/(1-厂用电率)，正常出力范围沿用电力交易分析报告的运行补偿报价段规则。"])
    note.append(["偏差计分", "机端出力在正常范围内记0；低于下限或高于上限时，按到最近边界的绝对差计入。"])
    note.append(["统计范围", "默认只统计海门#1机组、海门#2机组、海门#3机组、海门#4机组。"])
    note.append(["排序粒度", "日前和实时分开统计；每个口径按日期汇总海门4台机组全天偏差绝对值后各取前10天。"])
    note.append(["启停剔除", "任一机组在该日发生出力状态变化则剔除；全天有出力或全天无出力都不视为启停变化。"])

    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            letter = column_cells[0].column_letter
            width = min(max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells) + 2, 80)
            sheet.column_dimensions[letter].width = width

    wb.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="分日前和实时找出海门全厂正常出力范围与机端出力偏差绝对值加和最大的日期。")
    parser.add_argument("input_dir", nargs="?", default=str(DEFAULT_INPUT_DIR), help="出清情况Excel所在目录或单个Excel文件，默认读取 ~/Downloads/haimen shouxian")
    parser.add_argument("-x", "--top", type=int, default=DEFAULT_TOP, help="日前、实时分别输出偏差最大的前X天，默认10")
    parser.add_argument("--start-date", help="开始日期，格式 YYYY-MM-DD")
    parser.add_argument("--end-date", help="结束日期，格式 YYYY-MM-DD")
    parser.add_argument("--unit", action="append", help="只分析名称包含该文本的机组，可重复传入；不传时默认海门4台机组")
    parser.add_argument("--output", help="输出xlsx路径，默认写到输入目录下")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.top <= 0:
        raise SystemExit("--top 必须大于0")

    input_path = Path(args.input_dir).expanduser()
    if not input_path.exists():
        raise SystemExit(f"输入路径不存在：{input_path}")

    start_date = parse_date_text(args.start_date) if args.start_date else None
    end_date = parse_date_text(args.end_date) if args.end_date else None
    if start_date and end_date and start_date > end_date:
        raise SystemExit("--start-date 不能晚于 --end-date")

    workbooks = discover_workbooks(input_path, start_date, end_date)
    if not workbooks:
        raise SystemExit("没有找到符合日期范围的出清情况xlsx。")

    unit_filters = args.unit or DEFAULT_UNIT_FILTERS
    rows: list[HourDeviation] = []
    for workbook in workbooks:
        rows.extend(workbook_hour_deviations(workbook, unit_filters=unit_filters))
    if not rows:
        raise SystemExit("没有生成可计算的火电机组逐小时偏差，请检查机组筛选或工作簿报价段。")

    scores, excluded = rank_plant_days_with_exclusions(rows, args.top)
    if args.output:
        output_path = Path(args.output).expanduser()
    else:
        output_dir = input_path.parent if input_path.is_file() else input_path
        output_path = output_dir / f"正常出力范围偏差_全厂全天_日前实时分开Top{args.top}.xlsx"
    result = write_results(output_path, scores, excluded, rows)

    print(f"已读取 {len(workbooks)} 个出清文件，逐小时记录 {len(rows)} 条。")
    print(f"统计机组：{', '.join(unit_filters)}")
    print(f"结果已输出：{result}")
    for market_type in ("日前", "实时"):
        print(f"{market_type}前{args.top}天：")
        for index, score in enumerate([item for item in scores if item.market_type == market_type], start=1):
            print(f"{index}. {score.date} 全厂全天合计 {score.total_deviation_mw:.3f} MW")
    print(f"剔除启停日期记录 {len(excluded)} 条。")


if __name__ == "__main__":
    main()
