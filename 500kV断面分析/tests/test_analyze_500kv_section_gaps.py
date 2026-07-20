import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.analyze_500kv_section_gaps import analyze_section_gaps


class Analyze500kVSectionGapsTest(unittest.TestCase):
    def test_analyze_section_gaps_requires_opposite_moves_for_at_least_four_intervals(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            topology_path = tmp_path / "topology.md"
            workbook_path = tmp_path / "prices.xlsx"
            output_dir = tmp_path / "out"

            topology_path.write_text(
                """# test

## 2. 跨片区主通道边

### A片 - B片
甲-乙、甲-丙、甲-丁

## 3. next
""",
                encoding="utf-8",
            )

            wb = Workbook()
            ws = wb.active
            ws.title = "实时节点电价查询"
            ws.append(["节点名称", "数据项", "00:00", "00:15", "00:30", "00:45", "01:00", "01:15"])
            ws.append(["其他甲站500kV", "电价(元/MWh)", 100, 110, 120, 130, 140, 150])
            ws.append(["其他乙站500kV", "电价(元/MWh)", 100, 90, 80, 70, 60, 50])
            ws.append(["其他丙站500kV", "电价(元/MWh)", 200, 220, 240, 260, 280, 300])
            ws.append(["其他丁站500kV", "电价(元/MWh)", 100, 90, 80, 70, 60, 65])
            wb.save(workbook_path)

            result = analyze_section_gaps(
                topology_path=topology_path,
                workbook_path=workbook_path,
                output_dir=output_dir,
                threshold=0,
            )

            self.assertEqual(result.flagged_edges, 2)
            self.assertEqual(result.flagged_intervals, 9)

            out_wb = load_workbook(result.output_path, data_only=True)
            summary_rows = list(out_wb["断面汇总"].iter_rows(values_only=True))
            detail_rows = list(out_wb["逐时段明细"].iter_rows(values_only=True))
            edge_index = summary_rows[0].index("edge")
            window_index = summary_rows[0].index("windows")
            windows = {row[edge_index]: row[window_index] for row in summary_rows[1:]}
            self.assertEqual(windows["甲-乙"], "00:15-01:15")
            self.assertEqual(windows["甲-丁"], "00:15-01:00")
            detail_edges = [row[detail_rows[0].index("edge")] for row in detail_rows[1:]]
            self.assertEqual(detail_edges.count("甲-乙"), 5)
            self.assertEqual(detail_edges.count("甲-丁"), 4)

    def test_analyze_section_gaps_flags_connected_price_gaps_and_missing_nodes(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            topology_path = tmp_path / "topology.md"
            workbook_path = tmp_path / "prices.xlsx"
            output_dir = tmp_path / "out"

            topology_path.write_text(
                """# test

## 2. 跨片区主通道边

### A片 - B片
甲-乙、乙-丙

### 重复片
乙-甲

### 缺失片
甲-丁

## 3. next
""",
                encoding="utf-8",
            )

            wb = Workbook()
            ws = wb.active
            ws.title = "实时节点电价查询"
            ws.append(["节点名称", "数据项", "00:00", "00:15", "00:30", "00:45", "01:00", "01:15"])
            ws.append(["其他甲站500kV", "电价(元/MWh)", 100, 110, 120, 130, 140, 150])
            ws.append(["其他乙站500kV", "电价(元/MWh)", 100, 90, 80, 70, 60, 50])
            ws.append(["其他丙站500kV", "电价(元/MWh)", 100, 110, 120, 130, 140, 150])
            wb.save(workbook_path)

            result = analyze_section_gaps(
                topology_path=topology_path,
                workbook_path=workbook_path,
                output_dir=output_dir,
                threshold=0,
            )

            self.assertEqual(result.unique_edges, 3)
            self.assertEqual(result.skipped_duplicate_edges, [("乙", "甲", "重复片")])
            self.assertEqual(result.missing_nodes, ["丁"])
            self.assertEqual(result.flagged_edges, 2)
            self.assertEqual(result.flagged_intervals, 10)

            self.assertEqual(result.output_path, output_dir / "500kv_section_price_gap_analysis.xlsx")
            self.assertTrue(result.output_path.exists())
            self.assertFalse((output_dir / "500kv_section_price_gap_summary.csv").exists())

            out_wb = load_workbook(result.output_path, data_only=True)
            self.assertEqual(out_wb.sheetnames, ["断面汇总", "逐时段明细", "片区汇总"])

            rows = list(out_wb["断面汇总"].iter_rows(values_only=True))
            header = rows[0]
            edge_index = header.index("edge")
            direction_index = header.index("main_direction")
            windows_index = header.index("windows")
            data_rows = rows[1:]

            self.assertEqual([row[edge_index] for row in data_rows], ["乙-丙", "甲-乙"])
            self.assertEqual(data_rows[0][direction_index], "乙->丙")
            self.assertEqual(data_rows[0][windows_index], "00:15-01:15")
            self.assertEqual(data_rows[1][direction_index], "乙->甲")
            self.assertEqual(data_rows[1][windows_index], "00:15-01:15")


if __name__ == "__main__":
    unittest.main()
