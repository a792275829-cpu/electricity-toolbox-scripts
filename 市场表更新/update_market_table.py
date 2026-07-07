#!/usr/bin/env python3
from __future__ import annotations

import argparse
import importlib.util
import re
import sys
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import sync_playwright


SCRIPT_DIR = Path(__file__).resolve().parent
WORKSPACE = SCRIPT_DIR.parent
ONLINE_SCRIPTS_DIR = WORKSPACE / "上网电量抓取"
REPORT_SCRIPTS_DIR = WORKSPACE / "每日生产经营情况汇报自动生成工具" / "scripts"
if str(ONLINE_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(ONLINE_SCRIPTS_DIR))
if str(REPORT_SCRIPTS_DIR) not in sys.path:
    sys.path.append(str(REPORT_SCRIPTS_DIR))

online_export_spec = importlib.util.spec_from_file_location(
    "export_online_energy",
    ONLINE_SCRIPTS_DIR / "export_online_energy.py",
)
if online_export_spec is None or online_export_spec.loader is None:
    raise ImportError("无法加载上网电量登录模块")
online_export_module = importlib.util.module_from_spec(online_export_spec)
sys.modules["export_online_energy"] = online_export_module
online_export_spec.loader.exec_module(online_export_module)
online_export_module.PROFILE_DIR = SCRIPT_DIR / ".browser-profile"
online_export_module.AUTH_STATE_PATH = SCRIPT_DIR / "auth_state.json"
online_export_module.AUTH_LOCK_PATH = Path(str(online_export_module.AUTH_STATE_PATH) + ".lock")
AuthStateLock = online_export_module.AuthStateLock
ensure_login = online_export_module.ensure_login
interactive_login = online_export_module.interactive_login
launch_context = online_export_module.launch_context
load_config = online_export_module.load_config
validate_date = online_export_module.validate_date

from fetch_load_compare import select_load_type  # noqa: E402


MARKET_SHEET_NAME = "2026年市场情况"
COST_SHEET_NAME = "2026年运行方式及成本"
DAY_AHEAD_LOAD_START_COLUMN = "B"
DAY_AHEAD_LOAD_END_COLUMN = "U"
DAY_AHEAD_MARKET_START_COLUMN = "V"
DAY_AHEAD_MARKET_END_COLUMN = "AG"
ACTUAL_LOAD_START_COLUMN = "AH"
ACTUAL_LOAD_END_COLUMN = "BA"
REALTIME_MARKET_START_COLUMN = "BB"
REALTIME_MARKET_END_COLUMN = "BE"
GENERATION_SETTLEMENT_START_COLUMN = "BF"
GENERATION_SETTLEMENT_END_COLUMN = "BQ"
USER_SETTLEMENT_START_COLUMN = "BR"
USER_SETTLEMENT_END_COLUMN = "BX"
UNIT_COST_START_COLUMN = "N"
UNIT_COST_END_COLUMN = "X"
OPERATION_MODE_START_COLUMN = "B"
OPERATION_MODE_END_COLUMN = "M"
ACTUAL_LOAD_TYPES = (
    "1",  # 统调
    "6",  # B类
    "4",  # 西电
    "2",  # A类
    "3",  # 地方电
)
DEFAULT_POINT = 96
DEFAULT_PROVINCE_AREA_ID = "044"
UNIT_COST_ORG_IDS = (
    "e4e6eb5c80731ac70180fab1ba2f0559",  # 汕头电厂
    "e4e6eb5c80731ac70180fab3532d0592",  # 海门电厂
    "e4e6eb5c80731ac70180faa7f96904eb",  # 谢岗电厂
)
UNIT_COST_PRICE_UNIT_IDS = (
    "e4e6eb5580fa3aa00180fbcb51be0016",  # 汕头#1机组
    "e4e6eb5580fa3aa00180fbcb69a40017",  # 汕头#2机组
    "e4e6eb5580fa3aa00180fbcb89040018",  # 汕头#3机组
    "e4e6eb5580fa3aa00180fbbf9e940007",  # 海门#1机组
    "e4e6eb5580fa3aa00180fbc0d3800008",  # 海门#2机组
    "e4e6eb5580fa3aa00180fbc0f8690009",  # 海门#3机组
    "e4e6eb5580fa3aa00180fbc1276a000a",  # 海门#4机组
    "e4e6eb5580fa3aa00180fbd3ff800027",  # 谢岗#1、#2机组
    "e4e6eb5580fa3aa00180fbd457c40028",  # 谢岗#3、#4机组
    "e4e3842b93f2cbd90193f6af88e70796",  # 谢岗#5、#6机组
    "e4e3842b93f2cbd90193f6b01de40797",  # 谢岗#7、#8机组
)
DAY_AHEAD_MARKET_FIELDS = (
    "generateSideDeclareAvgPrice",
    "totalGenerateSideDealEle",
    "totalCoalDealEle",
    "totalGasDealEle",
    "totalNuclearDealEle",
    "totalNewEnergyDealEle",
    "totalPowerSideDealEle",
    "generateSideAvgPrice",
    "coalDealMaxPrice",
    "coalDealAvgPrice",
    "gasDealMaxPrice",
    "gasDealAvgPrice",
)
REALTIME_MARKET_FIELDS = (
    ("generateSideDealEle", "totalGenerateSideDealEle"),
    "generateSideAvgPrice",
    "coalDealAvgPrice",
    "gasDealAvgPrice",
)
GENERATION_SETTLEMENT_FIELDS = (
    "marketUnitOnlineEle",
    "marketUnitBaseEle",
    "marketUnitBaseEleRadio",
    "marketUnitMltEle",
    "marketUnitMltEleRadio",
    "marketUnitDeviationEle",
    "marketUnitDeviationEleRadio",
    "spotPosDeviationRadio",
    "spotNegDeviationRadio",
    "marketUnitFee",
    "generateAvgPriceWithCompensate",
    "generateAvgPrice",
)
GENERATION_SETTLEMENT_PERCENT_FIELDS = {
    "marketUnitBaseEleRadio",
    "marketUnitMltEleRadio",
    "marketUnitDeviationEleRadio",
    "spotPosDeviationRadio",
    "spotNegDeviationRadio",
}
USER_SETTLEMENT_FIELDS = (
    "marketUserConsumeEle",
    "mltEle",
    "mltEleRadio",
    "spotDeviationEle",
    "spotDeviationRadio",
    "getMarketUserConsumeFee",
    "userAvePrice",
)
USER_SETTLEMENT_PERCENT_FIELDS = {
    "mltEleRadio",
    "spotDeviationRadio",
}


def log(message: str) -> None:
    print(message, flush=True)


def parse_date(value: str) -> date:
    return datetime.strptime(validate_date(value), "%Y-%m-%d").date()


def parse_cell_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%Y年%m月%d日"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def date_matches(value: Any, target: date) -> bool:
    parsed = parse_cell_date(value)
    return parsed == target


def find_date_row(sheet, target: date) -> int:
    resolved_dates: dict[int, date] = {}
    formula_pattern = re.compile(r"^=\$?A\$?(\d+)\s*([+-])\s*(\d+)$", re.IGNORECASE)

    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row=row, column=1).value
        parsed = parse_cell_date(value)
        if parsed is None and isinstance(value, str):
            match = formula_pattern.match(value.strip())
            if match:
                ref_row = int(match.group(1))
                delta = int(match.group(3))
                if match.group(2) == "-":
                    delta = -delta
                ref_date = resolved_dates.get(ref_row)
                if ref_date is not None:
                    parsed = ref_date + timedelta(days=delta)
        if parsed is not None:
            resolved_dates[row] = parsed
        if parsed == target:
            return row
    raise RuntimeError(f"{sheet.title} 未找到日期行：{target.isoformat()}")


def workbook_permission_error(path: Path) -> RuntimeError:
    return RuntimeError(
        "无法打开 Excel 文件，macOS 拒绝了当前程序访问该文件。"
        f"请关闭正在占用的 Excel/WPS，并在“系统设置 > 隐私与安全性 > 文件和文件夹”或"
        f"“完全磁盘访问权限”中允许启动工具箱的程序访问桌面/文稿，或把文件移到当前程序可访问的目录后重试：{path}"
    )


def load_workbook_checked(workbook_path: Path | str, **kwargs):
    path = Path(workbook_path).expanduser()
    try:
        return openpyxl.load_workbook(path, **kwargs)
    except PermissionError as exc:
        raise workbook_permission_error(path) from exc


def ensure_workbook_accessible(workbook_path: Path | str) -> None:
    workbook = load_workbook_checked(workbook_path, read_only=True, data_only=False)
    workbook.close()


def find_date_row_in_workbook(
    workbook_path: Path,
    target: date,
    sheet_name: str = MARKET_SHEET_NAME,
) -> int:
    workbook = load_workbook_checked(workbook_path, read_only=True, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"Excel 中未找到工作表：{sheet_name}")
        return find_date_row(workbook[sheet_name], target)
    finally:
        workbook.close()


def actual_load_range(row: int) -> str:
    return f"{ACTUAL_LOAD_START_COLUMN}{row}:{ACTUAL_LOAD_END_COLUMN}{row}"


def day_ahead_load_range(row: int) -> str:
    return f"{DAY_AHEAD_LOAD_START_COLUMN}{row}:{DAY_AHEAD_LOAD_END_COLUMN}{row}"


def day_ahead_market_range(row: int) -> str:
    return f"{DAY_AHEAD_MARKET_START_COLUMN}{row}:{DAY_AHEAD_MARKET_END_COLUMN}{row}"


def realtime_market_range(row: int) -> str:
    return f"{REALTIME_MARKET_START_COLUMN}{row}:{REALTIME_MARKET_END_COLUMN}{row}"


def generation_settlement_range(row: int) -> str:
    return f"{GENERATION_SETTLEMENT_START_COLUMN}{row}:{GENERATION_SETTLEMENT_END_COLUMN}{row}"


def user_settlement_range(row: int) -> str:
    return f"{USER_SETTLEMENT_START_COLUMN}{row}:{USER_SETTLEMENT_END_COLUMN}{row}"


def unit_cost_price_range(row: int) -> str:
    return f"{UNIT_COST_START_COLUMN}{row}:{UNIT_COST_END_COLUMN}{row}"


def operation_mode_range(row: int) -> str:
    return f"{OPERATION_MODE_START_COLUMN}{row}:{OPERATION_MODE_END_COLUMN}{row}"


def actual_load_source_date(base_date: date) -> date:
    return base_date - timedelta(days=2)


def day_ahead_market_source_date(base_date: date) -> date:
    return base_date - timedelta(days=1)


def realtime_market_source_date(base_date: date) -> date:
    return base_date - timedelta(days=2)


def generation_settlement_source_date(base_date: date) -> date:
    return base_date - timedelta(days=6)


def user_settlement_source_date(base_date: date) -> date:
    return base_date - timedelta(days=6)


def normalize_number(value: float) -> int | float:
    rounded = round(float(value), 4)
    if rounded.is_integer():
        return int(rounded)
    return rounded


def normalize_report_number(value: Any, *, convert_percent: bool = False) -> int | float:
    if convert_percent:
        if isinstance(value, str):
            text = value.strip()
            if text.endswith("%"):
                value = text[:-1].strip()
        return normalize_number(float(value) / 100)
    if isinstance(value, str):
        text = value.strip()
    return normalize_number(value)


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def cell_is_non_empty(cell) -> bool:
    return cell.value not in (None, "")


def write_count_text(summary: dict[str, Any]) -> str:
    count = summary.get("count")
    if count is None:
        return "未知个数"
    skipped = summary.get("skipped", 0)
    if skipped:
        return f"{count} 个数，跳过 {skipped} 个非空单元格"
    return f"{count} 个数"


def summarize_load_type(
    dataset: dict[str, Any],
    load_type: str,
    run_date: date,
    period_list_key: str,
    label: str,
) -> list[int | float]:
    series = select_load_type(dataset, load_type)
    periods = series.get(period_list_key) or []
    if not periods:
        raise RuntimeError(
            f"{run_date.isoformat()} loadType={load_type} 没有返回{label}数据"
        )
    summary = periods[0]
    fields = ("max", "min", "avg", "sum")
    if any(summary.get(field) in (None, "") for field in fields):
        raise RuntimeError(
            f"{run_date.isoformat()} loadType={load_type} 没有返回{label}数据"
        )
    return [
        normalize_number(summary["max"]),
        normalize_number(summary["min"]),
        normalize_number(summary["avg"]),
        normalize_number(summary["sum"]),
    ]


def summarize_actual_load_type(
    dataset: dict[str, Any],
    load_type: str,
    run_date: date,
) -> list[int | float]:
    return summarize_load_type(
        dataset,
        load_type,
        run_date,
        "actualPeriodList",
        "实际运行",
    )


def summarize_day_ahead_load_type(
    dataset: dict[str, Any],
    load_type: str,
    run_date: date,
) -> list[int | float]:
    return summarize_load_type(
        dataset,
        load_type,
        run_date,
        "forecastPeriodList",
        "日前",
    )


def extract_actual_load_values(
    dataset: dict[str, Any],
    run_date: date,
) -> list[int | float]:
    values: list[int | float] = []
    for load_type in ACTUAL_LOAD_TYPES:
        values.extend(summarize_actual_load_type(dataset, load_type, run_date))
    if len(values) != 20:
        raise RuntimeError(f"市场负荷信息应为 20 个数，实际得到 {len(values)} 个")
    return values


def extract_day_ahead_load_values(
    dataset: dict[str, Any],
    run_date: date,
) -> list[int | float]:
    values: list[int | float] = []
    for load_type in ACTUAL_LOAD_TYPES:
        values.extend(summarize_day_ahead_load_type(dataset, load_type, run_date))
    if len(values) != 20:
        raise RuntimeError(f"日前市场负荷信息应为 20 个数，实际得到 {len(values)} 个")
    return values


def fetch_actual_load_values(
    context,
    run_date: date,
    *,
    province_area_id: str = DEFAULT_PROVINCE_AREA_ID,
) -> list[int | float]:
    response = context.request.get(
        f"{online_export_module.BASE_URL}/gdfire/api/data/net/load",
        params={
            "startDate": run_date.isoformat(),
            "endDate": run_date.isoformat(),
            "provinceAreaId": province_area_id,
            "timeSegment": 2,
        },
        timeout=30000,
    )
    dataset = online_export_module.parse_json_response(
        response,
        f"读取市场负荷信息 {run_date.isoformat()}",
    )
    return extract_actual_load_values(dataset, run_date)


def fetch_day_ahead_load_values(
    context,
    run_date: date,
    *,
    province_area_id: str = DEFAULT_PROVINCE_AREA_ID,
) -> list[int | float]:
    response = context.request.get(
        f"{online_export_module.BASE_URL}/gdfire/api/data/net/load",
        params={
            "startDate": run_date.isoformat(),
            "endDate": run_date.isoformat(),
            "type": 1,
            "provinceAreaId": province_area_id,
            "timeSegment": 2,
        },
        timeout=30000,
    )
    dataset = online_export_module.parse_json_response(
        response,
        f"读取日前市场负荷信息 {run_date.isoformat()}",
    )
    return extract_day_ahead_load_values(dataset, run_date)


def extract_day_ahead_market_values(
    payload: list[dict[str, Any]],
    run_date: date,
) -> list[int | float]:
    return extract_report_market_values(
        payload,
        run_date,
        fields=DAY_AHEAD_MARKET_FIELDS,
        expected_count=12,
        label="日前市场",
    )


def extract_realtime_market_values(
    payload: list[dict[str, Any]],
    run_date: date,
) -> list[int | float]:
    return extract_report_market_values(
        payload,
        run_date,
        fields=REALTIME_MARKET_FIELDS,
        expected_count=4,
        label="实时市场",
    )


def extract_generation_settlement_values(
    payload: list[dict[str, Any]],
    run_date: date,
) -> list[int | float]:
    return extract_report_market_values(
        payload,
        run_date,
        fields=GENERATION_SETTLEMENT_FIELDS,
        expected_count=12,
        label="发电侧结算",
        percent_fields=GENERATION_SETTLEMENT_PERCENT_FIELDS,
    )


def extract_user_settlement_values(
    payload: list[dict[str, Any]],
    run_date: date,
) -> list[int | float]:
    return extract_report_market_values(
        payload,
        run_date,
        fields=USER_SETTLEMENT_FIELDS,
        expected_count=7,
        label="用电侧结算",
        percent_fields=USER_SETTLEMENT_PERCENT_FIELDS,
    )


def extract_report_market_values(
    payload: list[dict[str, Any]],
    run_date: date,
    *,
    fields: tuple[Any, ...],
    expected_count: int,
    label: str,
    percent_fields: set[str] | None = None,
) -> list[int | float]:
    rows = [row for row in payload if date_matches(row.get("date"), run_date)]
    if not rows:
        raise RuntimeError(f"{label}没有返回日期 {run_date.isoformat()} 的数据")

    info = rows[0].get("info") or {}
    values: list[int | float] = []
    percent_fields = percent_fields or set()
    for field in fields:
        field_names = field if isinstance(field, tuple) else (field,)
        value = next(
            (
                info.get(field_name)
                for field_name in field_names
                if info.get(field_name) not in (None, "")
            ),
            None,
        )
        if value in (None, ""):
            raise RuntimeError(f"{label}缺少字段：{'/'.join(field_names)}")
        values.append(
            normalize_report_number(
                value,
                convert_percent=any(field_name in percent_fields for field_name in field_names),
            )
        )

    if len(values) != expected_count:
        raise RuntimeError(f"{label}应为 {expected_count} 个数，实际得到 {len(values)} 个")
    return values


def fetch_day_ahead_market_values(
    context,
    run_date: date,
) -> list[int | float]:
    response = context.request.get(
        f"{online_export_module.BASE_URL}/gdfire/api/pub/data/report",
        params={
            "startDate": run_date.isoformat(),
            "endDate": run_date.isoformat(),
            "type": 1,
        },
        timeout=30000,
    )
    payload = online_export_module.parse_json_response(
        response,
        f"读取日前市场 {run_date.isoformat()}",
    )
    return extract_day_ahead_market_values(payload.get("data") or [], run_date)


def fetch_realtime_market_values(
    context,
    run_date: date,
) -> list[int | float]:
    response = context.request.get(
        f"{online_export_module.BASE_URL}/gdfire/api/pub/data/report",
        params={
            "startDate": run_date.isoformat(),
            "endDate": run_date.isoformat(),
            "type": 2,
        },
        timeout=30000,
    )
    payload = online_export_module.parse_json_response(
        response,
        f"读取实时市场 {run_date.isoformat()}",
    )
    return extract_realtime_market_values(payload.get("data") or [], run_date)


def fetch_generation_settlement_values(
    context,
    run_date: date,
) -> list[int | float]:
    response = context.request.get(
        f"{online_export_module.BASE_URL}/gdfire/api/pub/data/report",
        params={
            "startDate": run_date.isoformat(),
            "endDate": run_date.isoformat(),
            "type": 3,
        },
        timeout=30000,
    )
    payload = online_export_module.parse_json_response(
        response,
        f"读取发电侧结算 {run_date.isoformat()}",
    )
    return extract_generation_settlement_values(payload.get("data") or [], run_date)


def fetch_user_settlement_values(
    context,
    run_date: date,
) -> list[int | float]:
    response = context.request.get(
        f"{online_export_module.BASE_URL}/gdfire/api/pub/data/report",
        params={
            "startDate": run_date.isoformat(),
            "endDate": run_date.isoformat(),
            "type": 4,
        },
        timeout=30000,
    )
    payload = online_export_module.parse_json_response(
        response,
        f"读取用电侧结算 {run_date.isoformat()}",
    )
    return extract_user_settlement_values(payload.get("data") or [], run_date)


def extract_unit_cost_price_values(payload: list[dict[str, Any]]) -> list[int | float]:
    rows_by_unit_id: dict[str, dict[str, Any]] = {}
    for org_item in payload:
        for row in org_item.get("unitCostInfoDTOList") or []:
            unit_id = row.get("unitId")
            if unit_id:
                rows_by_unit_id[str(unit_id)] = row

    values: list[int | float] = []
    for unit_id in UNIT_COST_PRICE_UNIT_IDS:
        row = rows_by_unit_id.get(unit_id)
        if row is None:
            raise RuntimeError(f"机组成本信息未找到机组 ID：{unit_id}")
        value = row.get("priceOfDyCost")
        if value in (None, ""):
            raise RuntimeError(f"机组成本信息缺少单位变动成本对应电价：{unit_id}")
        values.append(normalize_number(value))

    if len(values) != 11:
        raise RuntimeError(f"单位变动成本对应电价应为 11 个数，实际得到 {len(values)} 个")
    return values


def fetch_unit_cost_price_values(
    context,
    run_date: date,
) -> list[int | float]:
    response = context.request.get(
        f"{online_export_module.BASE_URL}/gdgroup/fire/api/data/personal/approve/report/data",
        params={
            "runDate": run_date.isoformat(),
            "orgIds": ",".join(UNIT_COST_ORG_IDS),
        },
        timeout=30000,
    )
    payload = online_export_module.parse_json_response(
        response,
        f"读取机组成本信息 {run_date.isoformat()}",
    )
    return extract_unit_cost_price_values(payload.get("data") or [])


def write_actual_load_values(
    workbook_path: Path | str,
    run_date: date,
    values: list[int | float],
) -> dict[str, Any]:
    return write_market_load_values(
        workbook_path,
        run_date,
        values,
        start_column=ACTUAL_LOAD_START_COLUMN,
        label="市场负荷",
    )


def write_day_ahead_load_values(
    workbook_path: Path | str,
    run_date: date,
    values: list[int | float],
) -> dict[str, Any]:
    return write_market_load_values(
        workbook_path,
        run_date,
        values,
        start_column=DAY_AHEAD_LOAD_START_COLUMN,
        label="日前市场负荷",
    )


def write_day_ahead_market_values(
    workbook_path: Path | str,
    run_date: date,
    values: list[int | float],
) -> dict[str, Any]:
    return write_market_row_values(
        workbook_path,
        run_date,
        values,
        start_column=DAY_AHEAD_MARKET_START_COLUMN,
        expected_count=12,
        label="日前市场",
    )


def write_realtime_market_values(
    workbook_path: Path | str,
    run_date: date,
    values: list[int | float],
) -> dict[str, Any]:
    return write_market_row_values(
        workbook_path,
        run_date,
        values,
        start_column=REALTIME_MARKET_START_COLUMN,
        expected_count=4,
        label="实时市场",
    )


def write_generation_settlement_values(
    workbook_path: Path | str,
    run_date: date,
    values: list[int | float],
) -> dict[str, Any]:
    return write_market_row_values(
        workbook_path,
        run_date,
        values,
        start_column=GENERATION_SETTLEMENT_START_COLUMN,
        expected_count=12,
        label="发电侧结算",
        overwrite=True,
    )


def write_user_settlement_values(
    workbook_path: Path | str,
    run_date: date,
    values: list[int | float],
) -> dict[str, Any]:
    return write_market_row_values(
        workbook_path,
        run_date,
        values,
        start_column=USER_SETTLEMENT_START_COLUMN,
        expected_count=7,
        label="用电侧结算",
        overwrite=True,
    )


def write_market_load_values(
    workbook_path: Path | str,
    run_date: date,
    values: list[int | float],
    *,
    start_column: str,
    label: str,
) -> dict[str, Any]:
    return write_market_row_values(
        workbook_path,
        run_date,
        values,
        start_column=start_column,
        expected_count=20,
        label=label,
    )


def write_market_row_values(
    workbook_path: Path | str,
    run_date: date,
    values: list[int | float],
    *,
    start_column: str,
    expected_count: int,
    label: str,
    overwrite: bool = False,
) -> dict[str, Any]:
    if len(values) != expected_count:
        raise ValueError(f"需要写入 {expected_count} 个{label}数值，实际收到 {len(values)} 个")

    path = Path(workbook_path).expanduser()
    row = find_date_row_in_workbook(path, run_date)
    workbook = load_workbook_checked(path)
    if MARKET_SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(f"Excel 中未找到工作表：{MARKET_SHEET_NAME}")
    sheet = workbook[MARKET_SHEET_NAME]
    start_column_index = openpyxl.utils.column_index_from_string(start_column)

    written_count = 0
    skipped_count = 0
    for offset, value in enumerate(values):
        cell = sheet.cell(row=row, column=start_column_index + offset)
        if cell_is_non_empty(cell) and not overwrite:
            skipped_count += 1
            continue
        if row > 1:
            copy_cell_style(sheet.cell(row=row - 1, column=start_column_index + offset), cell)
        cell.value = value
        written_count += 1

    workbook.save(path)
    end_column = start_column_index + len(values) - 1
    return {
        "workbook": str(path),
        "sheet": sheet.title,
        "row": row,
        "range": f"{start_column}{row}:{get_column_letter(end_column)}{row}",
        "count": written_count,
        "skipped": skipped_count,
    }


def write_unit_cost_price_values(
    workbook_path: Path | str,
    run_date: date,
    values: list[int | float],
) -> dict[str, Any]:
    if len(values) != 11:
        raise ValueError(f"需要写入 11 个单位变动成本对应电价，实际收到 {len(values)} 个")

    path = Path(workbook_path).expanduser()
    row = find_date_row_in_workbook(path, run_date, COST_SHEET_NAME)
    workbook = load_workbook_checked(path)
    if COST_SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(f"Excel 中未找到工作表：{COST_SHEET_NAME}")
    sheet = workbook[COST_SHEET_NAME]
    start_column = openpyxl.utils.column_index_from_string(UNIT_COST_START_COLUMN)

    written_count = 0
    skipped_count = 0
    for offset, value in enumerate(values):
        cell = sheet.cell(row=row, column=start_column + offset)
        if cell_is_non_empty(cell):
            skipped_count += 1
            continue
        if row > 1:
            copy_cell_style(sheet.cell(row=row - 1, column=start_column + offset), cell)
        cell.value = value
        written_count += 1

    workbook.save(path)
    end_column = start_column + len(values) - 1
    return {
        "workbook": str(path),
        "sheet": sheet.title,
        "row": row,
        "range": f"{UNIT_COST_START_COLUMN}{row}:{get_column_letter(end_column)}{row}",
        "count": written_count,
        "skipped": skipped_count,
    }


def copy_previous_day_unit_operation_mode(
    workbook_path: Path | str,
    run_date: date,
) -> dict[str, Any]:
    path = Path(workbook_path).expanduser()
    source_date = run_date - timedelta(days=1)
    target_row = find_date_row_in_workbook(path, run_date, COST_SHEET_NAME)
    source_row = find_date_row_in_workbook(path, source_date, COST_SHEET_NAME)

    workbook = load_workbook_checked(path)
    if COST_SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(f"Excel 中未找到工作表：{COST_SHEET_NAME}")
    sheet = workbook[COST_SHEET_NAME]
    start_column = openpyxl.utils.column_index_from_string(OPERATION_MODE_START_COLUMN)
    end_column = openpyxl.utils.column_index_from_string(OPERATION_MODE_END_COLUMN)

    written_count = 0
    skipped_count = 0
    for column in range(start_column, end_column + 1):
        source = sheet.cell(row=source_row, column=column)
        target = sheet.cell(row=target_row, column=column)
        if cell_is_non_empty(target):
            skipped_count += 1
            continue
        copy_cell_style(source, target)
        target.value = source.value
        written_count += 1

    workbook.save(path)
    return {
        "workbook": str(path),
        "sheet": sheet.title,
        "sourceDate": source_date.isoformat(),
        "sourceRow": source_row,
        "row": target_row,
        "range": operation_mode_range(target_row),
        "count": written_count,
        "skipped": skipped_count,
    }


def fetch_values_or_skip(label: str, fetcher, *, attempts: int = 3):
    for attempt in range(1, attempts + 1):
        try:
            return fetcher()
        except PlaywrightError as exc:
            if attempt < attempts:
                log(f"{label}读取失败，正在重试 {attempt + 1}/{attempts}：{exc}")
                continue
            log(f"跳过{label}：连续 {attempts} 次读取失败：{exc}")
            return None
        except (RuntimeError, ValueError) as exc:
            log(f"跳过{label}：{exc}")
            return None


def run_update(
    workbook_path: Path,
    run_date: date,
    *,
    headless: bool,
) -> dict[str, Any]:
    ensure_workbook_accessible(workbook_path)
    config = load_config()
    config.headless = headless

    with AuthStateLock():
        with sync_playwright() as playwright:
            context = launch_context(playwright, headless=config.headless)
            try:
                ensure_login(context, config)
                log(
                    "正在读取现货运行日报市场负荷信息-日前："
                    f"{run_date.isoformat()}（基准日期 D）"
                )
                day_ahead_values = fetch_values_or_skip(
                    "现货运行日报市场负荷信息-日前",
                    lambda: fetch_day_ahead_load_values(
                        context,
                        run_date,
                    ),
                )
                day_ahead_summary = None
                if day_ahead_values is not None:
                    log("已读取 20 个日前市场负荷数值。")
                    day_ahead_summary = write_day_ahead_load_values(
                        workbook_path,
                        run_date,
                        day_ahead_values,
                    )
                    log(
                        f"已写入：{day_ahead_summary['sheet']}!{day_ahead_summary['range']} "
                        f"（{write_count_text(day_ahead_summary)}）"
                    )
                market_date = day_ahead_market_source_date(run_date)
                log(
                    "正在读取现货运行日报日前市场："
                    f"{market_date.isoformat()}（基准日期 {run_date.isoformat()} 的 D-1）"
                )
                day_ahead_market_values = fetch_values_or_skip(
                    "现货运行日报日前市场",
                    lambda: fetch_day_ahead_market_values(
                        context,
                        market_date,
                    ),
                )
                day_ahead_market_summary = None
                if day_ahead_market_values is not None:
                    log("已读取发电侧申报均价起 12 个日前市场数值。")
                    day_ahead_market_summary = write_day_ahead_market_values(
                        workbook_path,
                        market_date,
                        day_ahead_market_values,
                    )
                    log(
                        f"已写入：{day_ahead_market_summary['sheet']}!{day_ahead_market_summary['range']} "
                        f"（{write_count_text(day_ahead_market_summary)}）"
                    )
                realtime_market_date = realtime_market_source_date(run_date)
                log(
                    "正在读取现货运行日报实时市场："
                    f"{realtime_market_date.isoformat()}（基准日期 {run_date.isoformat()} 的 D-2）"
                )
                realtime_market_values = fetch_values_or_skip(
                    "现货运行日报实时市场",
                    lambda: fetch_realtime_market_values(
                        context,
                        realtime_market_date,
                    ),
                )
                realtime_market_summary = None
                if realtime_market_values is not None:
                    log("已读取发电侧成交电量等 4 个实时市场数值。")
                    realtime_market_summary = write_realtime_market_values(
                        workbook_path,
                        realtime_market_date,
                        realtime_market_values,
                    )
                    log(
                        f"已写入：{realtime_market_summary['sheet']}!{realtime_market_summary['range']} "
                        f"（{write_count_text(realtime_market_summary)}）"
                    )
                generation_settlement_date = generation_settlement_source_date(run_date)
                log(
                    "正在读取现货运行日报发电侧结算："
                    f"{generation_settlement_date.isoformat()}（基准日期 {run_date.isoformat()} 的 D-6）"
                )
                generation_settlement_values = fetch_values_or_skip(
                    "现货运行日报发电侧结算",
                    lambda: fetch_generation_settlement_values(
                        context,
                        generation_settlement_date,
                    ),
                )
                generation_settlement_summary = None
                if generation_settlement_values is not None:
                    log("已读取 12 个发电侧结算数值。")
                    generation_settlement_summary = write_generation_settlement_values(
                        workbook_path,
                        generation_settlement_date,
                        generation_settlement_values,
                    )
                    log(
                        f"已写入：{generation_settlement_summary['sheet']}!{generation_settlement_summary['range']} "
                        f"（{write_count_text(generation_settlement_summary)}）"
                    )
                user_settlement_date = user_settlement_source_date(run_date)
                log(
                    "正在读取现货运行日报用电侧结算："
                    f"{user_settlement_date.isoformat()}（基准日期 {run_date.isoformat()} 的 D-6）"
                )
                user_settlement_values = fetch_values_or_skip(
                    "现货运行日报用电侧结算",
                    lambda: fetch_user_settlement_values(
                        context,
                        user_settlement_date,
                    ),
                )
                user_settlement_summary = None
                if user_settlement_values is not None:
                    log("已读取 7 个用电侧结算数值。")
                    user_settlement_summary = write_user_settlement_values(
                        workbook_path,
                        user_settlement_date,
                        user_settlement_values,
                    )
                    log(
                        f"已写入：{user_settlement_summary['sheet']}!{user_settlement_summary['range']} "
                        f"（{write_count_text(user_settlement_summary)}）"
                    )
                load_date = actual_load_source_date(run_date)
                log(
                    "正在读取现货运行日报市场负荷信息："
                    f"{load_date.isoformat()}（基准日期 {run_date.isoformat()} 的 D-2）"
                )
                values = fetch_values_or_skip(
                    "现货运行日报市场负荷信息-实际",
                    lambda: fetch_actual_load_values(
                        context,
                        load_date,
                    ),
                )
                summary = None
                if values is not None:
                    log("已读取 20 个市场负荷数值。")
                    summary = write_actual_load_values(workbook_path, load_date, values)
                    log(
                        f"已写入：{summary['sheet']}!{summary['range']} "
                        f"（{write_count_text(summary)}）"
                    )
                log(f"正在读取现货基础信息报送机组成本信息：{run_date.isoformat()}")
                cost_values = fetch_values_or_skip(
                    "现货基础信息报送机组成本信息",
                    lambda: fetch_unit_cost_price_values(
                        context,
                        run_date,
                    ),
                )
                cost_summary = None
                if cost_values is not None:
                    log("已读取 11 个单位变动成本对应电价。")
                    cost_summary = write_unit_cost_price_values(
                        workbook_path,
                        run_date,
                        cost_values,
                    )
                    log(
                        f"已写入：{cost_summary['sheet']}!{cost_summary['range']} "
                        f"（{write_count_text(cost_summary)}）"
                    )
                log(
                    "正在复制机组运行方式："
                    f"{(run_date - timedelta(days=1)).isoformat()} -> {run_date.isoformat()}"
                )
                operation_summary = copy_previous_day_unit_operation_mode(
                    workbook_path,
                    run_date,
                )
                log(
                    f"已写入：{operation_summary['sheet']}!{operation_summary['range']} "
                    f"（{write_count_text(operation_summary)}，来自第 {operation_summary['sourceRow']} 行）"
                )
                return {
                    "dayAheadLoad": day_ahead_summary,
                    "dayAheadMarket": day_ahead_market_summary,
                    "realtimeMarket": realtime_market_summary,
                    "generationSettlement": generation_settlement_summary,
                    "userSettlement": user_settlement_summary,
                    "actualLoad": summary,
                    "unitCostPrices": cost_summary,
                    "unitOperationMode": operation_summary,
                }
            finally:
                context.close()


def login_only(*, headless: bool) -> None:
    config = load_config()
    config.headless = headless
    with AuthStateLock():
        with sync_playwright() as playwright:
            context = launch_context(playwright, headless=config.headless)
            try:
                interactive_login(context, config)
            finally:
                context.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="抓取网页市场数据并更新市场 Excel。")
    parser.add_argument("--date", required=False, help="基准运行日期，格式 YYYY-MM-DD")
    parser.add_argument("--workbook", type=Path, help="要写入的 Excel 文件")
    parser.add_argument("--login", action="store_true", help="打开登录窗口并保存登录态")
    parser.add_argument("--headed", action="store_true", help="使用可见浏览器窗口")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    headless = not args.headed
    if args.login:
        login_only(headless=headless)
        return 0

    if not args.date or not args.workbook:
        raise SystemExit("请同时提供 --date 和 --workbook，或使用 --login。")

    run_update(
        args.workbook,
        parse_date(args.date),
        headless=headless,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
