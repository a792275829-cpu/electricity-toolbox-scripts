from __future__ import annotations

import subprocess
import sys
import tempfile
import time
import tkinter as tk
import unittest
import runpy
import os
import shutil
from contextlib import ExitStack
from datetime import date
from datetime import timedelta
from pathlib import Path
from unittest import mock


def destroy_tk(root: tk.Misc) -> None:
    try:
        root.update_idletasks()
    except tk.TclError:
        pass
    root.destroy()


class RuntimeTests(unittest.TestCase):
    def test_tool_paths_resolve_from_workspace(self) -> None:
        from toolbox.runtime import ToolPaths

        with tempfile.TemporaryDirectory() as directory:
            workspace = Path(directory)
            paths = ToolPaths(workspace)

            self.assertEqual(paths.workspace, workspace.resolve())
            self.assertEqual(
                paths.online_energy,
                workspace.resolve() / "上网电量抓取" / "export_online_energy.py",
            )
            self.assertEqual(
                paths.private_uploader,
                workspace.resolve()
                / "电力工具脚本"
                / "private-data-uploader-tool"
                / "scripts"
                / "upload-private-data.mjs",
            )
            self.assertEqual(
                paths.wps_writer,
                workspace.resolve() / "wps自动" / "wps_excel_to_kdocs_gui.py",
            )
            self.assertEqual(
                paths.market_table_update,
                workspace.resolve() / "市场表更新" / "update_market_table.py",
            )

    def test_load_module_rejects_missing_file(self) -> None:
        from toolbox.runtime import load_module

        with tempfile.TemporaryDirectory() as directory:
            missing = Path(directory) / "missing.py"
            with self.assertRaisesRegex(FileNotFoundError, "missing.py"):
                load_module("missing_tool", missing)

    def test_report_export_proxy_reads_environment(self) -> None:
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        module = load_module("report_export_online_energy_test", paths.report_scripts_dir / "export_online_energy.py")

        with mock.patch.dict(os.environ, {"HTTPS_PROXY": "http://127.0.0.1:7890"}, clear=True):
            self.assertEqual(module.playwright_proxy(), {"server": "http://127.0.0.1:7890"})

    def test_login_state_paths_are_separate_for_parallel_tools(self) -> None:
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        online = load_module("online_energy_state_paths_test", paths.online_energy)
        market = load_module("market_table_state_paths_test", paths.market_table_update)
        group = load_module("group_upload_state_paths_test", paths.group_upload)

        self.assertEqual(online.AUTH_STATE_PATH, paths.online_energy_dir / "auth_state.json")
        self.assertEqual(online.PROFILE_DIR, paths.online_energy_dir / ".browser-profile")
        self.assertEqual(market.online_export_module.AUTH_STATE_PATH, paths.market_table_update_dir / "auth_state.json")
        self.assertEqual(market.online_export_module.PROFILE_DIR, paths.market_table_update_dir / ".browser-profile")
        self.assertEqual(group.AUTH_STATE_PATH, paths.group_upload_dir / "auth_state.json")
        self.assertEqual(group.PROFILE_DIR, paths.group_upload_dir / ".browser-profile")

        auth_paths = {
            online.AUTH_STATE_PATH,
            market.online_export_module.AUTH_STATE_PATH,
            group.AUTH_STATE_PATH,
        }
        profile_paths = {
            online.PROFILE_DIR,
            market.online_export_module.PROFILE_DIR,
            group.PROFILE_DIR,
        }
        self.assertEqual(len(auth_paths), 3)
        self.assertEqual(len(profile_paths), 3)

    def test_online_energy_process_is_alive_detects_current_process(self) -> None:
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        module = load_module("online_energy_process_alive_test", paths.online_energy)

        self.assertTrue(module.process_is_alive(os.getpid()))
        self.assertFalse(module.process_is_alive(-1))

    def test_online_energy_lock_cleanup_does_not_mask_body_exception(self) -> None:
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        module = load_module("online_energy_lock_cleanup_test", paths.online_energy)

        with tempfile.TemporaryDirectory() as directory:
            lock_path = Path(directory) / "auth_state.json.lock"
            with mock.patch.object(module, "AUTH_LOCK_PATH", lock_path):
                with mock.patch.object(Path, "unlink", side_effect=PermissionError("locked")):
                    with self.assertRaisesRegex(ValueError, "original failure"):
                        with module.AuthStateLock():
                            raise ValueError("original failure")

    def test_group_upload_process_is_alive_detects_current_process(self) -> None:
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        module = load_module("group_upload_process_alive_test", paths.group_upload)

        self.assertTrue(module.process_is_alive(os.getpid()))
        self.assertFalse(module.process_is_alive(-1))

    def test_group_upload_lock_cleanup_does_not_mask_body_exception(self) -> None:
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        module = load_module("group_upload_lock_cleanup_test", paths.group_upload)

        with tempfile.TemporaryDirectory() as directory:
            lock_path = Path(directory) / "auth_state.json.lock"
            with mock.patch.object(module, "AUTH_LOCK_PATH", lock_path):
                with mock.patch.object(Path, "unlink", side_effect=PermissionError("locked")):
                    with self.assertRaisesRegex(ValueError, "original failure"):
                        with module.AuthStateLock():
                            raise ValueError("original failure")

    def test_group_upload_daily_report_name_uses_energy_date(self) -> None:
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        module = load_module("group_upload_daily_report_name_test", paths.group_upload)

        self.assertEqual(
            module.daily_report_name("2026-07-03"),
            "广东电力现货市场监测评估日报（20260703）",
        )

    def test_group_upload_opens_day_report_from_authenticated_report_entry(self) -> None:
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        module = load_module("group_upload_day_report_entry_test", paths.group_upload)

        class DummyPage:
            def __init__(self) -> None:
                self.urls: list[str] = []

            def goto(self, url, **_kwargs):
                self.urls.append(url)

            def wait_for_load_state(self, *_args, **_kwargs):
                return None

        class DummyContext:
            def __init__(self) -> None:
                self.page = DummyPage()
                self.pages = [self.page]

            def new_page(self):
                return self.page

        context = DummyContext()
        with mock.patch.object(module, "page_needs_manual_navigation", return_value=False):
            page = module.open_day_report_page(context, allow_manual=False)

        self.assertIs(page, context.page)
        self.assertEqual(context.page.urls, [module.DAY_REPORT_URL])

    def test_group_upload_returns_to_report_entry_before_menu_fallback(self) -> None:
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        module = load_module("group_upload_day_report_fallback_test", paths.group_upload)

        class DummyText:
            @property
            def first(self):
                return self

            def click(self, **_kwargs):
                return None

        class DummyPage:
            url = "about:blank"

            def __init__(self) -> None:
                self.urls: list[str] = []

            def goto(self, url, **_kwargs):
                self.urls.append(url)
                self.url = url

            def wait_for_load_state(self, *_args, **_kwargs):
                return None

            def get_by_text(self, *_args, **_kwargs):
                return DummyText()

        class DummyContext:
            def __init__(self) -> None:
                self.page = DummyPage()
                self.pages = [self.page]

            def new_page(self):
                return self.page

        context = DummyContext()
        with (
            mock.patch.object(module, "page_needs_manual_navigation", side_effect=[True, False]),
            mock.patch.object(module, "enter_huaneng_marketing_platform", side_effect=lambda page: page.goto(module.USERCENTER_URL)),
            mock.patch.object(module, "click_button_by_text"),
        ):
            page = module.open_day_report_page(context, allow_manual=False)

        self.assertIs(page, context.page)
        self.assertEqual(
            context.page.urls,
            [module.DAY_REPORT_URL, module.USERCENTER_URL, module.DAY_REPORT_URL],
        )

    def test_group_upload_accepts_numeric_c2_when_rerunning_same_day_output(self) -> None:
        from openpyxl import Workbook, load_workbook
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        module = load_module("group_upload_numeric_c2_rerun_test", paths.group_upload)

        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            province_path = folder / "广东-省内数据-20260703.xlsx"
            energy_path = folder / "广东-能销数据-20260703.xlsx"
            energy_path.write_text("x", encoding="utf-8")
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "省内数据"
            sheet["A2"] = date(2026, 7, 3)
            sheet["B2"] = 18
            sheet["C2"] = 108314.96
            workbook.save(province_path)

            uploads = [
                module.UploadFile(path=province_path, kind="省内", date="2026-07-03"),
                module.UploadFile(path=energy_path, kind="能销", date="2026-07-03"),
            ]
            with (
                mock.patch.object(module, "fetch_startup_max", return_value=20),
                mock.patch.object(module, "fetch_capacity_average", return_value=50000.0),
            ):
                updated_uploads = module.update_province_excel_for_upload(
                    object(),
                    "2026-07-03",
                    uploads,
                )

            updated = load_workbook(province_path, data_only=True)
            updated_sheet = updated["省内数据"]

        self.assertEqual(updated_uploads[0].path, province_path.resolve())
        self.assertEqual(updated_sheet["B2"].value, 20)
        self.assertEqual(updated_sheet["C2"].value, 108314.96)

    def test_group_upload_accepts_numeric_c2_from_nearest_template(self) -> None:
        from openpyxl import Workbook, load_workbook
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        module = load_module("group_upload_numeric_c2_nearest_template_test", paths.group_upload)

        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            province_path = folder / "广东-省内数据-20260704.xlsx"
            energy_path = folder / "广东-能销数据-20260705.xlsx"
            energy_path.write_text("x", encoding="utf-8")
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "省内数据"
            sheet["A2"] = date(2026, 7, 4)
            sheet["B2"] = 18
            sheet["C2"] = 108314.96
            workbook.save(province_path)

            uploads = [
                module.UploadFile(path=province_path, kind="省内", date="2026-07-04", auto_template=True),
                module.UploadFile(path=energy_path, kind="能销", date="2026-07-05"),
            ]
            with (
                mock.patch.object(module, "fetch_startup_max", return_value=20),
                mock.patch.object(module, "fetch_capacity_average", return_value=50000.0),
            ):
                updated_uploads = module.update_province_excel_for_upload(
                    object(),
                    "2026-07-05",
                    uploads,
                )

            target_path = folder / "广东-省内数据-20260705.xlsx"
            updated = load_workbook(target_path, data_only=True)
            updated_sheet = updated["省内数据"]

        self.assertEqual(updated_uploads[0].path, target_path.resolve())
        self.assertEqual(updated_sheet["A2"].value.date(), date(2026, 7, 5))
        self.assertEqual(updated_sheet["B2"].value, 20)
        self.assertEqual(updated_sheet["C2"].value, 108314.96)

    def test_group_upload_prefers_guangdong_branch_tenant(self) -> None:
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        module = load_module("group_upload_tenant_selection_test", paths.group_upload)

        tenants = [
            {"tenantId": "plant", "name": "华能广东汕头海上风电有限责任公司"},
            {"tenantId": "branch", "name": "广东分公司"},
            {"tenantId": "other", "name": "华能集团"},
        ]

        self.assertEqual(module.select_day_report_tenant(tenants), tenants[1])

    def test_group_upload_switches_using_application_tenant(self) -> None:
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        module = load_module("group_upload_application_tenant_test", paths.group_upload)

        class FakeResponse:
            status = 200

            def __init__(self, body: dict[str, object]) -> None:
                self.body = body

            def text(self) -> str:
                import json

                return json.dumps(self.body, ensure_ascii=False)

        class FakeRequest:
            def __init__(self) -> None:
                self.calls: list[tuple[str, dict[str, object] | None]] = []

            def get(self, url, params=None, **_kwargs):
                self.calls.append((url, params))
                if url.endswith("/usercenter/web/pf/tenant/user/application"):
                    return FakeResponse(
                        {
                            "retCode": "T200",
                            "data": [
                                {"tenantId": "root", "name": "华能集团"},
                                {"tenantId": "branch", "name": "广东分公司"},
                            ],
                        }
                    )
                if url.endswith("/usercenter/web/switchTenant"):
                    return FakeResponse({"retCode": "T200", "retMsg": "切换租户成功"})
                raise AssertionError(f"unexpected request: {url}")

        class FakeContext:
            def __init__(self) -> None:
                self.request = FakeRequest()

        context = FakeContext()
        module.switch_to_day_report_tenant(context)

        self.assertEqual(
            context.request.calls,
            [
                (f"{module.BASE_URL}/usercenter/web/pf/tenant/user/application", None),
                (f"{module.BASE_URL}/usercenter/web/switchTenant", {"tenantId": "branch"}),
            ],
        )

    def test_group_upload_create_daily_report_posts_api_payload(self) -> None:
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        module = load_module("group_upload_create_report_api_test", paths.group_upload)

        class FakeResponse:
            status = 200

            def __init__(self, body: dict[str, object]) -> None:
                self.body = body

            def text(self) -> str:
                import json

                return json.dumps(self.body, ensure_ascii=False)

        class FakeRequest:
            def __init__(self) -> None:
                self.posts: list[tuple[str, dict[str, object]]] = []

            def get(self, url, params=None, **_kwargs):
                if url.endswith("/huaneng/group/api/report"):
                    return FakeResponse({"retCode": "T200", "data": {"datas": []}})
                if url.endswith("/huaneng/group/api/report/queryTemplates"):
                    return FakeResponse(
                        {
                            "retCode": "T200",
                            "data": [
                                {"name": "广东模板（修改）", "id": "template-id"},
                            ],
                        }
                    )
                raise AssertionError(f"unexpected request: {url}")

            def post(self, url, data=None, **_kwargs):
                self.posts.append((url, data or {}))
                return FakeResponse({"retCode": "T200", "data": "file-id"})

        class FakeContext:
            def __init__(self) -> None:
                self.request = FakeRequest()

        context = FakeContext()
        with (
            mock.patch.object(module, "open_day_report_page", return_value=object()),
        ):
            module.create_daily_report(context, "2026-07-04", headless=True)

        self.assertEqual(
            context.request.posts,
            [
                (
                    f"{module.BASE_URL}/huaneng/group/api/report",
                    {
                        "reportName": "广东电力现货市场监测评估日报（20260704）",
                        "startDate": "2026-07-04",
                        "endDate": "2026-07-04",
                        "templateId": "template-id",
                        "provinceId": None,
                    },
                )
            ],
        )

    def test_group_upload_smoke_report_cli_skips_file_selection(self) -> None:
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        module = load_module("group_upload_smoke_report_cli_test", paths.group_upload)

        with (
            mock.patch.object(sys, "argv", ["upload_daily_report.py", "--smoke-report"]),
            mock.patch.object(module, "smoke_daily_report_page") as smoke,
            mock.patch.object(module, "choose_files_interactively") as choose_files,
        ):
            exit_code = module.main()

        self.assertEqual(exit_code, 0)
        smoke.assert_called_once()
        choose_files.assert_not_called()

    def test_report_summary_reads_shifted_renewable_rows(self) -> None:
        from openpyxl import Workbook
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        sys.path.insert(0, str(paths.report_scripts_dir))
        self.addCleanup(lambda: sys.path.remove(str(paths.report_scripts_dir)))
        module = load_module(
            "report_generate_red_marked_shifted_renewables_test",
            paths.report_scripts_dir / "generate_red_marked_report.py",
        )

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "出清情况（分公司2026.7.2).xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "快报  (日前)"
            for cell, value in {
                "F7": 10116.2898,
                "G7": 422.0653,
                "F8": 1317.6129,
                "G8": 379.0106,
                "F13": 8042.6026,
                "G13": 419.1546,
                "F19": 756.0743,
                "G19": 528.0599,
            }.items():
                sheet[cell] = value
            sheet["A28"] = "新能源"
            sheet["G29"] = "日前出清"
            sheet["G30"] = "电量"
            sheet["H30"] = "电价"
            for row, name, energy, price in [
                (31, "分公司", 247.5663, 332.4490),
                (32, "汕头海风", 189.3145, 285.4164),
                (33, "鮀莲光伏", 24.4543, 512.7376),
                (34, "归湖光伏", 33.7975, 465.4510),
            ]:
                sheet.cell(row, 1).value = name
                sheet.cell(row, 7).value = energy
                sheet.cell(row, 8).value = price
            workbook.save(workbook_path)

            clearing = module.load_day_ahead_clearing_summary("2026-07-02", workbook_path)
            prices = module.load_online_price_summary("2026-07-02", workbook_path)

        self.assertEqual(clearing["companies"]["海上风电"]["energy"], 189.3145)
        self.assertEqual(clearing["companies"]["鮀莲"]["price"], 512.7376)
        self.assertEqual(clearing["companies"]["归湖"]["energy"], 33.7975)
        self.assertEqual(prices["companies"]["海上风电"], 285.4164)

    def test_report_energy_text_includes_qingyuan_descriptions(self) -> None:
        from toolbox.runtime import ToolPaths, load_module

        paths = ToolPaths(Path(__file__).resolve().parents[2])
        sys.path.insert(0, str(paths.report_scripts_dir))
        self.addCleanup(lambda: sys.path.remove(str(paths.report_scripts_dir)))
        module = load_module(
            "report_generate_red_marked_qingyuan_text_test",
            paths.report_scripts_dir / "generate_red_marked_report.py",
        )

        online_text = module.ensure_qingyuan_online_energy_text(
            "7月1日分公司上网电量9003万千瓦时，其中，汕头1526万、海门6120万、"
            "东莞1232万、海上风电58万、光伏67万千瓦时；分公司日前均价404厘/千瓦时，"
            "其中汕头387厘、海门406厘、东莞411厘、海上风电347厘、鮀莲361厘、归湖358厘/千瓦时。"
        )
        self.assertIn("光伏67万千瓦时、清远0万千瓦时；", online_text)

        clearing_text = module.build_day_ahead_clearing_text(
            "2026-07-02",
            {
                "totalEnergy": 10116.2898,
                "averagePrice": 422.0653,
                "companies": {
                    "汕头": {"energy": 1317.6129, "price": 379.0106},
                    "海门": {"energy": 8042.6026, "price": 419.1546},
                    "东莞": {"energy": 756.0743, "price": 528.0599},
                    "海上风电": {"energy": 189.3145, "price": 285.4164},
                    "鮀莲": {"energy": 24.4543, "price": 512.7376},
                    "归湖": {"energy": 33.7975, "price": 465.4510},
                },
            },
        )
        self.assertIn("归湖34万千瓦时、清远（计划）0万千瓦时；", clearing_text)

    def test_utf8_environment_sets_python_encoding(self) -> None:
        from toolbox.runtime import utf8_environment

        environment = utf8_environment()

        self.assertEqual(environment["PYTHONUTF8"], "1")
        self.assertEqual(environment["PYTHONIOENCODING"], "utf-8")

    def test_python_executable_uses_current_interpreter_on_macos(self) -> None:
        if sys.platform != "darwin":
            self.skipTest("macOS-specific runtime expectation")

        from toolbox.runtime import python_executable, pythonw_executable

        self.assertEqual(Path(python_executable()).resolve(), Path(sys.executable).resolve())
        self.assertEqual(Path(pythonw_executable()).resolve(), Path(sys.executable).resolve())

    def test_node_executable_uses_plain_node_on_macos(self) -> None:
        if sys.platform != "darwin":
            self.skipTest("macOS-specific runtime expectation")

        from toolbox.runtime import node_executable

        self.assertEqual(node_executable(), "node")

    def test_task_registry_tracks_and_terminates_owned_process(self) -> None:
        from toolbox.runtime import TaskRegistry

        registry = TaskRegistry()
        process = subprocess.Popen(
            [sys.executable, "-c", "import time; time.sleep(30)"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            text=True,
        )
        self.addCleanup(lambda: process.poll() is None and process.kill())

        registry.register_process(process)
        self.assertTrue(registry.has_running_tasks())

        registry.terminate_all()
        deadline = time.time() + 5
        while process.poll() is None and time.time() < deadline:
            time.sleep(0.05)

        self.assertIsNotNone(process.poll())
        self.assertFalse(registry.has_running_tasks())


class ToolboxAppTests(unittest.TestCase):
    def create_root(self) -> tk.Tk:
        try:
            root = tk.Tk()
            root.withdraw()
            return root
        except tk.TclError as exc:
            self.skipTest(f"Tk unavailable: {exc}")

    def test_app_creates_pages_and_switches_without_recreating(self) -> None:
        from toolbox.app import PAGE_NAMES, ToolboxApp

        created: dict[str, tk.Frame] = {}

        def make_factory(name: str):
            def factory(parent, _paths, _registry):
                frame = tk.Frame(parent, name=f"page_{len(created)}")
                frame.marker = tk.StringVar(master=frame, value=name)  # type: ignore[attr-defined]
                created[name] = frame
                return frame

            return factory

        factories = {name: make_factory(name) for name in PAGE_NAMES}
        app = ToolboxApp(page_factories=factories)
        app.withdraw()
        self.addCleanup(destroy_tk, app)

        self.assertEqual(list(app.pages), list(PAGE_NAMES))
        first_page = app.pages["导出上网电量"]
        app.show_page("电量汇总")
        app.show_page("导出上网电量")

        self.assertEqual(app.current_page, "导出上网电量")
        self.assertIs(app.pages["导出上网电量"], first_page)
        self.assertEqual(first_page.marker.get(), "导出上网电量")  # type: ignore[attr-defined]

    def test_tool_page_keeps_its_own_log_and_busy_state(self) -> None:
        from toolbox.runtime import TaskRegistry
        from toolbox.widgets import ToolPage

        root = self.create_root()
        self.addCleanup(destroy_tk, root)
        page = ToolPage(root, registry=TaskRegistry())
        page.append_log("第一条日志")
        page.set_busy(True, "运行中")
        root.update()

        self.assertIn("第一条日志", page.log_text.get("1.0", "end"))
        self.assertTrue(page.busy)
        self.assertEqual(page.status_var.get(), "运行中")


class PageAdapterTests(unittest.TestCase):
    def setUp(self) -> None:
        try:
            self.root = tk.Tk()
            self.root.withdraw()
        except tk.TclError as exc:
            self.skipTest(f"Tk unavailable: {exc}")
        from toolbox.runtime import TaskRegistry, ToolPaths

        self.paths = ToolPaths(Path(__file__).resolve().parents[2])
        self.registry = TaskRegistry()

    def tearDown(self) -> None:
        if hasattr(self, "root"):
            destroy_tk(self.root)

    def test_core_pages_build_expected_paths_and_commands(self) -> None:
        from toolbox.pages import (
            OnlineEnergyPage,
            SummaryPage,
            TradeAnalysisPage,
        )

        online = OnlineEnergyPage(self.root, self.paths, self.registry)
        trade = TradeAnalysisPage(self.root, self.paths, self.registry)
        summary = SummaryPage(self.root, self.paths, self.registry)

        export_command = online.build_export_command("2026-06-09")
        self.assertEqual(export_command[-1], "2026-06-09")
        self.assertEqual(Path(export_command[-2]), self.paths.online_energy)
        self.assertTrue(trade.output_dir.get())
        self.assertEqual(summary.default_output_for(Path("C:/data")).suffix, ".xlsx")

    def test_online_energy_defaults_to_previous_day(self) -> None:
        from toolbox.pages import OnlineEnergyPage

        page = OnlineEnergyPage(self.root, self.paths, self.registry)
        expected = date.today() - timedelta(days=1)

        self.assertEqual(page.run_date.get(), expected.isoformat())

    def test_trade_auto_match_finds_latest_d_and_d_minus_two_files(self) -> None:
        from toolbox.pages import TradeAnalysisPage

        with tempfile.TemporaryDirectory() as directory:
            downloads = Path(directory)
            older_d = downloads / "出清情况（分公司2026.6.17).xlsx"
            latest_d = downloads / "出清情况（分公司2026.6.17) (1).xlsx"
            latest_d_minus_two = downloads / "出清情况（分公司2026.6.15) (1).xlsx"
            unrelated = downloads / "市场代购电量调整系数(20260617).xlsx"
            for index, path in enumerate(
                [older_d, latest_d, latest_d_minus_two, unrelated], start=1
            ):
                path.write_text("x", encoding="utf-8")
                os.utime(path, (1_700_000_000 + index, 1_700_000_000 + index))
            os.utime(latest_d, (1_700_000_100, 1_700_000_100))

            matches = TradeAnalysisPage.find_clearing_files_for_dates(
                date(2026, 6, 17),
                downloads,
            )

        self.assertEqual(matches, [latest_d, latest_d_minus_two])

    def test_review_folder_match_uses_d_plus_one_then_latest_fallback(self) -> None:
        from toolbox.pages import default_review_folder_for_date

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            target = root / "6-18"
            fallback = root / "6-17"
            target.mkdir()
            fallback.mkdir()
            os.utime(fallback, (1_700_000_100, 1_700_000_100))

            self.assertEqual(
                default_review_folder_for_date(date(2026, 6, 17), root),
                target,
            )

            target.rmdir()
            self.assertEqual(
                default_review_folder_for_date(date(2026, 6, 17), root),
                fallback,
            )

    def test_summary_and_private_pages_auto_fill_d_plus_one_folder(self) -> None:
        from toolbox.pages import PrivateUploadPage, SummaryPage

        with tempfile.TemporaryDirectory() as directory:
            review_root = Path(directory)
            target = review_root / "6-18"
            target.mkdir()

            summary = SummaryPage(
                self.root,
                self.paths,
                self.registry,
                today=date(2026, 6, 17),
                review_root=review_root,
            )
            private = PrivateUploadPage(
                self.root,
                self.paths,
                self.registry,
                today=date(2026, 6, 17),
                review_root=review_root,
            )

        self.assertEqual(Path(summary.input_dir.get()), target)
        self.assertEqual(Path(private.source_folder.get()), target)

    def test_report_page_maps_source_dates_and_builds_word_command(self) -> None:
        from toolbox.pages import ReportPage

        page = ReportPage(self.root, self.paths, self.registry)
        self.assertEqual(
            page.source_dates("2026-06-09"),
            (date(2026, 6, 8), date(2026, 6, 9), date(2026, 6, 7)),
        )

        command = page.build_word_command(
            "2026-06-09",
            Path("C:/online.xlsx"),
            Path("C:/day-ahead.xlsx"),
            Path("C:/daily.xlsx"),
            None,
        )
        self.assertIn("--online-workbook", command)
        self.assertIn("--day-ahead-workbook", command)
        self.assertIn("--daily-clearing-workbook", command)

    def test_upload_pages_build_non_shell_commands(self) -> None:
        from toolbox.pages import GroupUploadPage, MarketTableUpdatePage, PrivateUploadPage

        private = PrivateUploadPage(self.root, self.paths, self.registry)
        group = GroupUploadPage(self.root, self.paths, self.registry)
        market = MarketTableUpdatePage(self.root, self.paths, self.registry)

        plan_command = private.build_command("--plan", Path(r"C:\review"))
        self.assertEqual(plan_command[-3:], ["--plan", "--source", r"C:\review"])
        self.assertIn("--execute", private.build_command("--execute", Path(r"C:\review")))

        group_command = group.build_upload_command(
            [Path(r"C:\data\广东-能销数据-20260609.xlsx")],
            force=True,
        )
        self.assertEqual(Path(group_command[1]), self.paths.group_upload)
        self.assertIn("--force", group_command)

        market_command = market.build_update_command(
            date(2026, 7, 2),
            Path(r"C:\data\现货统调市场出清、结算数据2026.xlsx"),
        )
        self.assertEqual(Path(market_command[1]), self.paths.market_table_update)
        self.assertIn("--date", market_command)
        self.assertIn("2026-07-02", market_command)

    def test_wps_writer_page_embeds_writer_frame(self) -> None:
        from toolbox.pages import WpsWriterPage

        page = WpsWriterPage(self.root, self.paths, self.registry)

        self.assertTrue(hasattr(page, "writer_frame"))
        self.assertFalse(hasattr(page, "open_button"))
        self.assertFalse(hasattr(page, "build_open_command"))

    def test_wps_writer_keeps_document_configs_visible_when_resized(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("toolbox_wps_writer_layout_test", self.paths.wps_writer)
        frame = module.WpsWriterFrame(self.root)
        frame.pack(fill="both", expand=True)
        self.root.geometry("520x360")
        self.root.update_idletasks()

        content_root = frame.winfo_children()[0]
        self.assertGreaterEqual(content_root.grid_rowconfigure(1)["minsize"], 220)

    def test_wps_writer_imports_exported_config_payload(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("toolbox_wps_writer_import_test", self.paths.wps_writer)
        payload = {
            "schema": "wps_excel_to_kdocs_config_export",
            "version": 1,
            "exported_at": "2026-06-29T16:45:11",
            "recent_urls": ["https://www.kdocs.cn/l/target"],
            "browser_mode": "cdp",
            "cdp_url": "http://127.0.0.1:9222",
            "configs": [
                {
                    "name": "上网电量写入",
                    "source_type": "excel",
                    "source_url": "",
                    "kdocs_url": "https://www.kdocs.cn/l/target",
                    "local_file": "C:/Users/lllg/Desktop/source.xlsx",
                    "regions": [
                        {
                            "source_sheet": "汇总",
                            "source_start": "A3",
                            "source_end": "C26",
                            "target_sheet": "汕头",
                            "target_start": "T3",
                            "target_end": "V26",
                        }
                    ],
                }
            ],
        }

        runtime_config = module.runtime_config_from_export_payload(payload)

        self.assertEqual(runtime_config["browser_mode"], "cdp")
        self.assertEqual(runtime_config["cdp_url"], "http://127.0.0.1:9222")
        self.assertEqual(runtime_config["configs"][0]["name"], "上网电量写入")
        self.assertEqual(runtime_config["configs"][0]["regions"][0]["target_sheet"], "汕头")

    def test_wps_writer_exports_runtime_config_with_schema(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("toolbox_wps_writer_export_test", self.paths.wps_writer)
        runtime_config = {
            "recent_urls": ["https://www.kdocs.cn/l/target"],
            "browser_mode": "persistent",
            "cdp_url": "http://127.0.0.1:9222",
            "configs": [
                {
                    "name": "Config",
                    "source_type": "kdocs",
                    "source_url": "https://www.kdocs.cn/l/source",
                    "kdocs_url": "https://www.kdocs.cn/l/target",
                    "local_file": "",
                    "regions": [
                        {
                            "source_sheet": "Sheet1",
                            "source_start": "B4",
                            "source_end": "C27",
                            "target_sheet": "Sheet2",
                            "target_start": "B3",
                            "target_end": "C26",
                        }
                    ],
                }
            ],
        }

        exported = module.build_config_export(runtime_config, exported_at="2026-06-29T16:45:11")

        self.assertEqual(exported["schema"], "wps_excel_to_kdocs_config_export")
        self.assertEqual(exported["version"], 1)
        self.assertEqual(exported["exported_at"], "2026-06-29T16:45:11")
        self.assertEqual(exported["configs"][0]["source_type"], "kdocs")
        self.assertEqual(module.runtime_config_from_export_payload(exported), runtime_config)

    def test_open_directory_uses_macos_open_command(self) -> None:
        if sys.platform != "darwin":
            self.skipTest("macOS-specific directory opener")

        from toolbox.pages import _open_directory

        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory) / "output"
            with mock.patch("subprocess.run") as run:
                _open_directory(target)

            self.assertTrue(target.is_dir())
            run.assert_called_once_with(["open", str(target)], check=False)

    def test_report_gui_open_directory_uses_macos_open_command(self) -> None:
        if sys.platform != "darwin":
            self.skipTest("macOS-specific directory opener")

        from toolbox.runtime import load_module

        module = load_module("report_gui_open_directory_test", self.paths.report_gui)

        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory) / "output"
            with mock.patch("subprocess.run") as run:
                module.open_directory(target)

            self.assertTrue(target.is_dir())
            run.assert_called_once_with(["open", str(target)], check=False)

    def test_trade_analysis_open_directory_uses_macos_open_command(self) -> None:
        if sys.platform != "darwin":
            self.skipTest("macOS-specific directory opener")

        from toolbox.runtime import load_module

        module = load_module("trade_analysis_open_directory_test", self.paths.trade_analysis)

        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory) / "output"
            with mock.patch("subprocess.run") as run:
                module.open_directory(target)

            self.assertTrue(target.is_dir())
            run.assert_called_once_with(["open", str(target)], check=False)

    def test_group_upload_auto_selects_latest_d_minus_two_energy_file(self) -> None:
        from toolbox.pages import GroupUploadPage

        with tempfile.TemporaryDirectory() as directory:
            upload_dir = Path(directory)
            older = upload_dir / "广东-能销数据-20260615.xlsx"
            latest = upload_dir / "广东-能销数据-20260615-新版.xlsx"
            other_date = upload_dir / "广东-能销数据-20260616.xlsx"
            province_template = upload_dir / "广东-省内数据-20260614.xlsx"
            for index, path in enumerate(
                [older, latest, other_date, province_template], start=1
            ):
                path.write_text("x", encoding="utf-8")
                os.utime(path, (1_700_000_000 + index, 1_700_000_000 + index))
            os.utime(latest, (1_700_000_100, 1_700_000_100))

            match = GroupUploadPage.find_latest_energy_file_for_date(
                date(2026, 6, 17),
                upload_dir,
            )

        self.assertEqual(match, latest)

    def test_group_upload_page_initializes_with_d_minus_two_energy_file(self) -> None:
        from toolbox.pages import GroupUploadPage

        with tempfile.TemporaryDirectory() as directory:
            upload_dir = Path(directory)
            energy = upload_dir / "广东-能销数据-20260615.xlsx"
            province_template = upload_dir / "广东-省内数据-20260614.xlsx"
            energy.write_text("x", encoding="utf-8")
            province_template.write_text("x", encoding="utf-8")
            paths = type(
                "Paths",
                (),
                {
                    "group_upload": self.paths.group_upload,
                    "group_upload_dir": upload_dir,
                },
            )()

            page = GroupUploadPage(
                self.root,
                paths,  # type: ignore[arg-type]
                self.registry,
                today=date(2026, 6, 17),
            )

        self.assertEqual(page.selected_paths, [province_template.resolve(), energy.resolve()])
        self.assertIn("能销数据", page.selection_var.get())

    def test_group_upload_page_keeps_auto_province_template_in_selection(self) -> None:
        from toolbox.pages import GroupUploadPage

        with tempfile.TemporaryDirectory() as directory:
            upload_dir = Path(directory)
            energy = upload_dir / "广东-能销数据-20260705.xlsx"
            province_template = upload_dir / "广东-省内数据-20260704.xlsx"
            energy.write_text("x", encoding="utf-8")
            province_template.write_text("x", encoding="utf-8")
            paths = type(
                "Paths",
                (),
                {
                    "group_upload": self.paths.group_upload,
                    "group_upload_dir": upload_dir,
                },
            )()

            page = GroupUploadPage(
                self.root,
                paths,  # type: ignore[arg-type]
                self.registry,
                today=date(2026, 7, 7),
            )

        self.assertEqual(page.selected_paths, [province_template.resolve(), energy.resolve()])
        command = page.build_upload_command(page.selected_paths, force=False)
        self.assertIn(str(province_template.resolve()), command)
        self.assertIn(str(energy.resolve()), command)

    def test_group_upload_template_error_lists_visible_excel_files(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("group_upload_template_error_test", self.paths.group_upload)

        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            (folder / "广东-能销数据-20260705.xlsx").write_text("x", encoding="utf-8")
            (folder / "广东-省内数据-没有日期.xlsx").write_text("x", encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "实际看到的 Excel 文件"):
                module.find_province_template(folder, "2026-07-05")

    def test_group_upload_auto_template_prefers_nearest_province_date(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("group_upload_nearest_template_test", self.paths.group_upload)

        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            energy = folder / "广东-能销数据-20260705.xlsx"
            older = folder / "广东-省内数据-20260621.xlsx"
            nearest = folder / "广东-省内数据-20260704.xlsx"
            newer = folder / "广东-省内数据-20260710.xlsx"
            for path in [energy, older, nearest, newer]:
                path.write_text("x", encoding="utf-8")

            target_date, uploads = module.prepare_upload_files([energy])

        self.assertEqual(target_date, "2026-07-05")
        self.assertEqual(uploads[0].path, nearest.resolve())
        self.assertEqual(uploads[1].path, energy.resolve())

    def test_group_upload_auto_template_probes_standard_names_when_listing_is_empty(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("group_upload_template_probe_test", self.paths.group_upload)

        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            energy = folder / "广东-能销数据-20260705.xlsx"
            nearest = folder / "广东-省内数据-20260704.xlsx"
            energy.write_text("x", encoding="utf-8")
            nearest.write_text("x", encoding="utf-8")
            with mock.patch.object(Path, "glob", return_value=iter(())):
                target_date, uploads = module.prepare_upload_files([energy])

        self.assertEqual(target_date, "2026-07-05")
        self.assertEqual(uploads[0].path, nearest.resolve())
        self.assertEqual(uploads[1].path, energy.resolve())

    def test_default_page_factories_create_all_frames(self) -> None:
        from toolbox.app import PAGE_NAMES
        from toolbox.pages import page_factories

        factories = page_factories()
        self.assertEqual(list(factories), list(PAGE_NAMES))
        frames = [
            factories[name](self.root, self.paths, self.registry)
            for name in PAGE_NAMES
        ]
        self.assertEqual(len(frames), len(PAGE_NAMES))
        self.assertEqual(len({str(frame) for frame in frames}), len(PAGE_NAMES))
        self.assertIn("WPS写入工具", PAGE_NAMES)
        self.assertIn("市场表更新", PAGE_NAMES)

    def test_market_table_update_writes_actual_load_values_by_date(self) -> None:
        from openpyxl import Workbook, load_workbook
        from toolbox.runtime import load_module

        module = load_module("market_table_update_test", self.paths.market_table_update)
        values = [1000 + index for index in range(20)]

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "market.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "2026年市场情况"
            ws["A1"] = "日期"
            ws["AH1"] = "实际运行"
            ws["A5"] = date(2026, 7, 1)
            ws["A6"] = date(2026, 7, 2)
            ws["A7"] = date(2026, 7, 3)
            wb.save(workbook_path)

            summary = module.write_actual_load_values(
                workbook_path,
                date(2026, 7, 2),
                values,
            )

            updated = load_workbook(workbook_path, data_only=True)
            updated_ws = updated["2026年市场情况"]

        self.assertEqual(summary["sheet"], "2026年市场情况")
        self.assertEqual(summary["row"], 6)
        self.assertEqual(summary["range"], "AH6:BA6")
        self.assertEqual(
            [updated_ws.cell(6, column).value for column in range(34, 54)],
            values,
        )

    def test_market_table_update_reports_permission_error_before_fetching(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("market_table_update_permission_error_test", self.paths.market_table_update)

        with mock.patch.object(module.openpyxl, "load_workbook", side_effect=PermissionError("Operation not permitted")):
            with self.assertRaisesRegex(RuntimeError, "无法打开 Excel 文件"):
                module.ensure_workbook_accessible(Path("/Users/auren/Desktop/日常工作/market.xlsx"))

    def test_market_table_update_skips_non_empty_market_cells(self) -> None:
        from openpyxl import Workbook, load_workbook
        from toolbox.runtime import load_module

        module = load_module("market_table_update_skip_non_empty_market_test", self.paths.market_table_update)
        values = [7000 + index for index in range(4)]

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "market.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "2026年市场情况"
            ws["A1"] = "日期"
            ws["BB1"] = "实时市场"
            ws["A6"] = date(2026, 7, 2)
            ws["BB6"] = "保留"
            ws["BD6"] = 999
            wb.save(workbook_path)

            summary = module.write_realtime_market_values(
                workbook_path,
                date(2026, 7, 2),
                values,
            )

            updated = load_workbook(workbook_path, data_only=True)
            updated_ws = updated["2026年市场情况"]

        self.assertEqual(summary["range"], "BB6:BE6")
        self.assertEqual(summary["count"], 2)
        self.assertEqual(summary["skipped"], 2)
        self.assertEqual(
            [updated_ws.cell(6, column).value for column in range(54, 58)],
            ["保留", 7001, 999, 7003],
        )

    def test_market_table_update_writes_day_ahead_load_values_by_base_date(self) -> None:
        from openpyxl import Workbook, load_workbook
        from toolbox.runtime import load_module

        module = load_module("market_table_update_day_ahead_test", self.paths.market_table_update)
        values = [2000 + index for index in range(20)]

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "market.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "2026年市场情况"
            ws["A1"] = "日期"
            ws["B1"] = "日前"
            ws["A5"] = date(2026, 7, 1)
            ws["A6"] = date(2026, 7, 2)
            ws["A7"] = date(2026, 7, 3)
            wb.save(workbook_path)

            summary = module.write_day_ahead_load_values(
                workbook_path,
                date(2026, 7, 2),
                values,
            )

            updated = load_workbook(workbook_path, data_only=True)
            updated_ws = updated["2026年市场情况"]

        self.assertEqual(summary["sheet"], "2026年市场情况")
        self.assertEqual(summary["row"], 6)
        self.assertEqual(summary["range"], "B6:U6")
        self.assertEqual(
            [updated_ws.cell(6, column).value for column in range(2, 22)],
            values,
        )

    def test_market_table_update_writes_day_ahead_market_values_by_source_date(self) -> None:
        from openpyxl import Workbook, load_workbook
        from toolbox.runtime import load_module

        module = load_module("market_table_update_day_ahead_market_test", self.paths.market_table_update)
        values = [3000 + index for index in range(12)]

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "market.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "2026年市场情况"
            ws["A1"] = "日期"
            ws["V1"] = "发电侧申报均价"
            ws["A5"] = date(2026, 7, 4)
            ws["A6"] = date(2026, 7, 5)
            ws["A7"] = date(2026, 7, 6)
            wb.save(workbook_path)

            summary = module.write_day_ahead_market_values(
                workbook_path,
                date(2026, 7, 5),
                values,
            )

            updated = load_workbook(workbook_path, data_only=True)
            updated_ws = updated["2026年市场情况"]

        self.assertEqual(summary["sheet"], "2026年市场情况")
        self.assertEqual(summary["row"], 6)
        self.assertEqual(summary["range"], "V6:AG6")
        self.assertEqual(
            [updated_ws.cell(6, column).value for column in range(22, 34)],
            values,
        )

    def test_market_table_update_writes_realtime_market_values_by_source_date(self) -> None:
        from openpyxl import Workbook, load_workbook
        from toolbox.runtime import load_module

        module = load_module("market_table_update_realtime_market_write_test", self.paths.market_table_update)
        values = [4000 + index for index in range(4)]

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "market.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "2026年市场情况"
            ws["A1"] = "日期"
            ws["BB1"] = "实时市场"
            ws["A5"] = date(2026, 7, 4)
            ws["A6"] = date(2026, 7, 5)
            ws["A7"] = date(2026, 7, 6)
            wb.save(workbook_path)

            summary = module.write_realtime_market_values(
                workbook_path,
                date(2026, 7, 5),
                values,
            )

            updated = load_workbook(workbook_path, data_only=True)
            updated_ws = updated["2026年市场情况"]

        self.assertEqual(summary["sheet"], "2026年市场情况")
        self.assertEqual(summary["row"], 6)
        self.assertEqual(summary["range"], "BB6:BE6")
        self.assertEqual(
            [updated_ws.cell(6, column).value for column in range(54, 58)],
            values,
        )

    def test_market_table_update_writes_generation_settlement_values_by_source_date(self) -> None:
        from openpyxl import Workbook, load_workbook
        from toolbox.runtime import load_module

        module = load_module("market_table_update_generation_settlement_write_test", self.paths.market_table_update)
        values = [5000 + index for index in range(12)]

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "market.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "2026年市场情况"
            ws["A1"] = "日期"
            ws["BF1"] = "发电侧结算"
            ws["A5"] = date(2026, 6, 30)
            ws["A6"] = date(2026, 7, 1)
            ws["A7"] = date(2026, 7, 2)
            wb.save(workbook_path)

            summary = module.write_generation_settlement_values(
                workbook_path,
                date(2026, 7, 1),
                values,
            )

            updated = load_workbook(workbook_path, data_only=True)
            updated_ws = updated["2026年市场情况"]

        self.assertEqual(summary["sheet"], "2026年市场情况")
        self.assertEqual(summary["row"], 6)
        self.assertEqual(summary["range"], "BF6:BQ6")
        self.assertEqual(
            [updated_ws.cell(6, column).value for column in range(58, 70)],
            values,
        )

    def test_market_table_update_overwrites_generation_settlement_percent_cells(self) -> None:
        from openpyxl import Workbook, load_workbook
        from toolbox.runtime import load_module

        module = load_module("market_table_update_generation_settlement_overwrite_percent_test", self.paths.market_table_update)
        values = [
            17.25,
            -0.51,
            -0.029,
            14.41,
            0.835,
            2.02,
            0.117,
            0.485,
            0.368,
            7.28,
            421.6,
            372.6,
        ]

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "market.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "2026年市场情况"
            ws["A1"] = "日期"
            ws["BF1"] = "发电侧结算"
            ws["A6"] = date(2026, 7, 1)
            ws["BH6"] = -2.9
            ws["BH6"].number_format = "0.00%"
            ws["BJ6"] = 83.5
            ws["BJ6"].number_format = "0.00%"
            ws["BL6"] = 11.7
            ws["BL6"].number_format = "0.00%"
            wb.save(workbook_path)

            summary = module.write_generation_settlement_values(
                workbook_path,
                date(2026, 7, 1),
                values,
            )

            updated = load_workbook(workbook_path, data_only=True)
            updated_ws = updated["2026年市场情况"]

        self.assertEqual(summary["count"], 12)
        self.assertEqual(summary["skipped"], 0)
        self.assertEqual(updated_ws["BH6"].value, -0.029)
        self.assertEqual(updated_ws["BJ6"].value, 0.835)
        self.assertEqual(updated_ws["BL6"].value, 0.117)

    def test_market_table_update_writes_user_settlement_values_by_source_date(self) -> None:
        from openpyxl import Workbook, load_workbook
        from toolbox.runtime import load_module

        module = load_module("market_table_update_user_settlement_write_test", self.paths.market_table_update)
        values = [6000 + index for index in range(7)]

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "market.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "2026年市场情况"
            ws["A1"] = "日期"
            ws["BR1"] = "用电侧结算"
            ws["A5"] = date(2026, 6, 30)
            ws["A6"] = date(2026, 7, 1)
            ws["A7"] = date(2026, 7, 2)
            wb.save(workbook_path)

            summary = module.write_user_settlement_values(
                workbook_path,
                date(2026, 7, 1),
                values,
            )

            updated = load_workbook(workbook_path, data_only=True)
            updated_ws = updated["2026年市场情况"]

        self.assertEqual(summary["sheet"], "2026年市场情况")
        self.assertEqual(summary["row"], 6)
        self.assertEqual(summary["range"], "BR6:BX6")
        self.assertEqual(
            [updated_ws.cell(6, column).value for column in range(70, 77)],
            values,
        )

    def test_market_table_update_writes_unit_cost_prices_by_formula_date(self) -> None:
        from openpyxl import Workbook, load_workbook
        from toolbox.runtime import load_module

        module = load_module("market_table_update_cost_write_test", self.paths.market_table_update)
        values = [300 + index for index in range(11)]

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "market.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "2026年运行方式及成本"
            ws["A1"] = "日期"
            ws["N1"] = "火电成本"
            ws["A4"] = date(2026, 1, 1)
            ws["A5"] = "=A4+1"
            ws["A6"] = "=A5+1"
            wb.save(workbook_path)

            summary = module.write_unit_cost_price_values(
                workbook_path,
                date(2026, 1, 3),
                values,
            )

            updated = load_workbook(workbook_path, data_only=True)
            updated_ws = updated["2026年运行方式及成本"]

        self.assertEqual(summary["sheet"], "2026年运行方式及成本")
        self.assertEqual(summary["row"], 6)
        self.assertEqual(summary["range"], "N6:X6")
        self.assertEqual(
            [updated_ws.cell(6, column).value for column in range(14, 25)],
            values,
        )
        self.assertEqual(updated_ws["N6"].number_format, updated_ws["N5"].number_format)

    def test_market_table_update_skips_non_empty_unit_cost_cells(self) -> None:
        from openpyxl import Workbook, load_workbook
        from toolbox.runtime import load_module

        module = load_module("market_table_update_skip_non_empty_cost_test", self.paths.market_table_update)
        values = [800 + index for index in range(11)]

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "market.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "2026年运行方式及成本"
            ws["A1"] = "日期"
            ws["N1"] = "火电成本"
            ws["A4"] = date(2026, 1, 1)
            ws["A5"] = "=A4+1"
            ws["A6"] = "=A5+1"
            ws["N6"] = "保留"
            ws["P6"] = 999
            wb.save(workbook_path)

            summary = module.write_unit_cost_price_values(
                workbook_path,
                date(2026, 1, 3),
                values,
            )

            updated = load_workbook(workbook_path, data_only=True)
            updated_ws = updated["2026年运行方式及成本"]

        self.assertEqual(summary["range"], "N6:X6")
        self.assertEqual(summary["count"], 9)
        self.assertEqual(summary["skipped"], 2)
        expected = values[:]
        expected[0] = "保留"
        expected[2] = 999
        self.assertEqual(
            [updated_ws.cell(6, column).value for column in range(14, 25)],
            expected,
        )

    def test_market_table_update_preserves_previous_row_column_formats(self) -> None:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill
        from toolbox.runtime import load_module

        module = load_module("market_table_update_cost_format_test", self.paths.market_table_update)

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "market.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "2026年运行方式及成本"
            ws["A4"] = date(2026, 1, 1)
            ws["A5"] = "=A4+1"
            for column in range(14, 25):
                source = ws.cell(4, column)
                source.number_format = "0.0000"
                source.font = Font(name="Arial", bold=True, color="FF0000")
                source.fill = PatternFill("solid", fgColor="FFF2CC")
            wb.save(workbook_path)

            module.write_unit_cost_price_values(
                workbook_path,
                date(2026, 1, 2),
                [400 + index for index in range(11)],
            )

            updated = load_workbook(workbook_path)
            updated_ws = updated["2026年运行方式及成本"]

        self.assertEqual(updated_ws["N5"].number_format, "0.0000")
        self.assertEqual(updated_ws["N5"].font.name, "Arial")
        self.assertTrue(updated_ws["N5"].font.bold)
        self.assertEqual(updated_ws["N5"].font.color.rgb, "00FF0000")
        self.assertEqual(updated_ws["N5"].fill.fgColor.rgb, "00FFF2CC")

    def test_market_table_update_copies_previous_day_unit_operation_mode(self) -> None:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill
        from toolbox.runtime import load_module

        module = load_module("market_table_update_operation_mode_test", self.paths.market_table_update)
        source_values = [
            "运行", "运行", "备用", "二分列",
            "运行", "运行", "运行", "运行",
            "备用", "调峰", "调峰", "备用",
        ]

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "market.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "2026年运行方式及成本"
            ws["A4"] = date(2026, 7, 3)
            ws["A5"] = "=A4+1"
            ws["B5"] = "已有运行"
            ws["D5"] = "已有备用"
            ws["N5"] = 999
            for offset, value in enumerate(source_values, start=2):
                cell = ws.cell(4, offset)
                cell.value = value
                cell.font = Font(name="Arial", bold=True, color="FF0000")
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
            wb.save(workbook_path)

            summary = module.copy_previous_day_unit_operation_mode(
                workbook_path,
                date(2026, 7, 4),
            )

            updated = load_workbook(workbook_path)
            updated_ws = updated["2026年运行方式及成本"]

        self.assertEqual(summary["sheet"], "2026年运行方式及成本")
        self.assertEqual(summary["sourceRow"], 4)
        self.assertEqual(summary["row"], 5)
        self.assertEqual(summary["range"], "B5:M5")
        expected_values = source_values[:]
        expected_values[0] = "已有运行"
        expected_values[2] = "已有备用"
        self.assertEqual([updated_ws.cell(5, column).value for column in range(2, 14)], expected_values)
        self.assertEqual(summary["count"], 10)
        self.assertEqual(summary["skipped"], 2)
        self.assertIsNotNone(updated_ws["A5"].value)
        self.assertEqual(updated_ws["N5"].value, 999)
        self.assertEqual(updated_ws["C5"].font.name, "Arial")
        self.assertTrue(updated_ws["C5"].font.bold)
        self.assertEqual(updated_ws["C5"].fill.fgColor.rgb, "00FFF2CC")

    def test_market_table_update_extracts_page_actual_load_order(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("market_table_update_order_test", self.paths.market_table_update)
        dataset = {
            "data": {
                "dataNetLoadDTOList": [
                    {"loadType": "2", "actualPeriodList": [{"max": 41, "min": 42, "avg": 43, "sum": 44}]},
                    {"loadType": "1", "actualPeriodList": [{"max": 11, "min": 12, "avg": 13, "sum": 14}]},
                    {"loadType": "3", "actualPeriodList": [{"max": 51, "min": 52, "avg": 53, "sum": 54}]},
                    {"loadType": "6", "actualPeriodList": [{"max": 21, "min": 22, "avg": 23, "sum": 24}]},
                    {"loadType": "4", "actualPeriodList": [{"max": 31, "min": 32, "avg": 33, "sum": 34}]},
                ]
            }
        }

        self.assertEqual(
            module.extract_actual_load_values(dataset, date(2026, 7, 2)),
            [
                11, 12, 13, 14,
                21, 22, 23, 24,
                31, 32, 33, 34,
                41, 42, 43, 44,
                51, 52, 53, 54,
            ],
        )

    def test_market_table_update_treats_none_actual_load_metrics_as_unavailable(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("market_table_update_none_actual_load_test", self.paths.market_table_update)
        dataset = {
            "data": {
                "dataNetLoadDTOList": [
                    {
                        "loadType": "1",
                        "actualPeriodList": [{"max": None, "min": None, "avg": None, "sum": None}],
                    }
                ]
            }
        }

        with self.assertRaisesRegex(RuntimeError, "没有返回实际运行数据"):
            module.extract_actual_load_values(dataset, date(2026, 7, 5))

    def test_market_table_update_extracts_day_ahead_market_field_order(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("market_table_update_market_order_test", self.paths.market_table_update)
        payload = [
            {
                "date": "2026-07-05 00:00:00",
                "info": {
                    "generateSideDeclareAvgPrice": 376.6,
                    "totalGenerateSideDealEle": 15.64,
                    "totalCoalDealEle": 9.26,
                    "totalGasDealEle": 1.33,
                    "totalNuclearDealEle": 2.42,
                    "totalNewEnergyDealEle": 2.58,
                    "totalPowerSideDealEle": 25.27,
                    "generateSideAvgPrice": 295.63,
                    "coalDealMaxPrice": 1157.53,
                    "coalDealAvgPrice": 318.36,
                    "gasDealMaxPrice": 1376.96,
                    "gasDealAvgPrice": 460.69,
                },
            }
        ]

        self.assertEqual(
            module.extract_day_ahead_market_values(payload, date(2026, 7, 5)),
            [
                376.6,
                15.64,
                9.26,
                1.33,
                2.42,
                2.58,
                25.27,
                295.63,
                1157.53,
                318.36,
                1376.96,
                460.69,
            ],
        )

    def test_market_table_update_extracts_realtime_market_field_order(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("market_table_update_realtime_market_order_test", self.paths.market_table_update)
        payload = [
            {
                "date": "2026-07-05 00:00:00",
                "info": {
                    "generateSideDeclareAvgPrice": 376.6,
                    "totalGenerateSideDealEle": 15.64,
                    "totalCoalDealEle": 9.26,
                    "totalGasDealEle": 1.33,
                    "totalNuclearDealEle": 2.42,
                    "totalNewEnergyDealEle": 2.58,
                    "totalPowerSideDealEle": 25.27,
                    "generateSideAvgPrice": 295.63,
                    "coalDealMaxPrice": 1157.53,
                    "coalDealAvgPrice": 318.36,
                    "gasDealMaxPrice": 1376.96,
                    "gasDealAvgPrice": 460.69,
                },
            }
        ]

        self.assertEqual(
            module.extract_realtime_market_values(payload, date(2026, 7, 5)),
            [
                15.64,
                295.63,
                318.36,
                460.69,
            ],
        )

    def test_market_table_update_extracts_realtime_market_live_deal_energy_field(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("market_table_update_realtime_market_live_field_test", self.paths.market_table_update)
        payload = [
            {
                "date": "2026-06-30 00:00:00",
                "info": {
                    "generateSideDealEle": 17.01,
                    "generateSideAvgPrice": 341.2,
                    "coalDealAvgPrice": 320.5,
                    "gasDealAvgPrice": 442.8,
                },
            }
        ]

        self.assertEqual(
            module.extract_realtime_market_values(payload, date(2026, 6, 30)),
            [
                17.01,
                341.2,
                320.5,
                442.8,
            ],
        )

    def test_market_table_update_extracts_generation_settlement_field_order(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("market_table_update_generation_settlement_order_test", self.paths.market_table_update)
        payload = [
            {
                "date": "2026-06-30 00:00:00",
                "info": {
                    "marketUnitMltEleRadio": 83.5,
                    "spotNegDeviationRadio": 36.8,
                    "marketUnitBaseEle": -0.51,
                    "marketUnitDeviationEle": 2.02,
                    "spotPosDeviationRadio": 48.5,
                    "generateAvgPrice": 372.6,
                    "marketUnitFee": 7.28,
                    "marketUnitDeviationEleRadio": 11.7,
                    "marketUnitMltEle": 14.41,
                    "generateAvgPriceWithCompensate": 421.6,
                    "marketUnitBaseEleRadio": -2.9,
                    "marketUnitOnlineEle": 17.25,
                },
            }
        ]

        self.assertEqual(
            module.extract_generation_settlement_values(payload, date(2026, 6, 30)),
            [
                17.25,
                -0.51,
                -0.029,
                14.41,
                0.835,
                2.02,
                0.117,
                0.485,
                0.368,
                7.28,
                421.6,
                372.6,
            ],
        )

    def test_market_table_update_extracts_user_settlement_field_order(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("market_table_update_user_settlement_order_test", self.paths.market_table_update)
        payload = [
            {
                "date": "2026-06-30 00:00:00",
                "info": {
                    "marketUserConsumeEle": 25.77,
                    "userAvePrice": 418.7,
                    "mltEle": 21.1,
                    "mltEleRadio": 81.9,
                    "spotDeviationEle": 4.67,
                    "spotDeviationRadio": 18.1,
                    "getMarketUserConsumeFee": 10.79,
                },
            }
        ]

        self.assertEqual(
            module.extract_user_settlement_values(payload, date(2026, 6, 30)),
            [
                25.77,
                21.1,
                0.819,
                4.67,
                0.181,
                10.79,
                418.7,
            ],
        )

    def test_market_table_update_converts_generation_settlement_percent_strings(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("market_table_update_generation_settlement_percent_test", self.paths.market_table_update)
        payload = [
            {
                "date": "2026-06-30 00:00:00",
                "info": {
                    "marketUnitOnlineEle": "17.25",
                    "marketUnitBaseEle": "-0.51",
                    "marketUnitBaseEleRadio": "-2.9%",
                    "marketUnitMltEle": "14.41",
                    "marketUnitMltEleRadio": 83.5,
                    "marketUnitDeviationEle": "2.02",
                    "marketUnitDeviationEleRadio": "11.7%",
                    "spotPosDeviationRadio": 48.5,
                    "spotNegDeviationRadio": "36.8%",
                    "marketUnitFee": "7.28",
                    "generateAvgPriceWithCompensate": "421.6",
                    "generateAvgPrice": "372.6",
                },
            }
        ]

        values = module.extract_generation_settlement_values(payload, date(2026, 6, 30))

        self.assertEqual(
            values,
            [
                17.25,
                -0.51,
                -0.029,
                14.41,
                0.835,
                2.02,
                0.117,
                0.485,
                0.368,
                7.28,
                421.6,
                372.6,
            ],
        )

    def test_market_table_update_converts_user_settlement_percent_strings(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("market_table_update_user_settlement_percent_test", self.paths.market_table_update)
        payload = [
            {
                "date": "2026-06-30 00:00:00",
                "info": {
                    "marketUserConsumeEle": "25.77",
                    "userAvePrice": "418.7",
                    "mltEle": "21.1",
                    "mltEleRadio": 81.9,
                    "spotDeviationEle": "4.67",
                    "spotDeviationRadio": "18.1%",
                    "getMarketUserConsumeFee": "10.79",
                },
            }
        ]

        values = module.extract_user_settlement_values(payload, date(2026, 6, 30))

        self.assertEqual(
            values,
            [
                25.77,
                21.1,
                0.819,
                4.67,
                0.181,
                10.79,
                418.7,
            ],
        )

    def test_market_table_update_skips_mapping_when_values_are_unavailable(self) -> None:
        from types import SimpleNamespace
        from toolbox.runtime import load_module

        module = load_module("market_table_update_skip_test", self.paths.market_table_update)

        class DummyLock:
            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return None

        class DummyPlaywright:
            def __enter__(self):
                return object()

            def __exit__(self, exc_type, exc, tb):
                return None

        context = SimpleNamespace(close=mock.Mock())
        with ExitStack() as stack:
            stack.enter_context(mock.patch.object(module, "load_config", return_value=SimpleNamespace(headless=True)))
            stack.enter_context(mock.patch.object(module, "ensure_workbook_accessible"))
            stack.enter_context(mock.patch.object(module, "AuthStateLock", return_value=DummyLock()))
            stack.enter_context(mock.patch.object(module, "sync_playwright", return_value=DummyPlaywright()))
            stack.enter_context(mock.patch.object(module, "launch_context", return_value=context))
            ensure_login = stack.enter_context(mock.patch.object(module, "ensure_login"))
            stack.enter_context(mock.patch.object(module, "fetch_day_ahead_load_values", side_effect=RuntimeError("没有返回日前数据")))
            write_day_ahead_load = stack.enter_context(mock.patch.object(module, "write_day_ahead_load_values"))
            stack.enter_context(mock.patch.object(module, "fetch_day_ahead_market_values", return_value=[1] * 12))
            stack.enter_context(mock.patch.object(module, "write_day_ahead_market_values", return_value={"range": "V6:AG6", "sheet": "2026年市场情况", "count": 12}))
            stack.enter_context(mock.patch.object(module, "fetch_realtime_market_values", return_value=[4] * 4))
            stack.enter_context(mock.patch.object(module, "write_realtime_market_values", return_value={"range": "BB5:BE5", "sheet": "2026年市场情况", "count": 4}))
            stack.enter_context(mock.patch.object(module, "fetch_generation_settlement_values", return_value=[5] * 12))
            stack.enter_context(mock.patch.object(module, "write_generation_settlement_values", return_value={"range": "BF1:BQ1", "sheet": "2026年市场情况", "count": 12}))
            stack.enter_context(mock.patch.object(module, "fetch_user_settlement_values", return_value=[6] * 7))
            stack.enter_context(mock.patch.object(module, "write_user_settlement_values", return_value={"range": "BR1:BX1", "sheet": "2026年市场情况", "count": 7}))
            stack.enter_context(mock.patch.object(module, "fetch_actual_load_values", return_value=[2] * 20))
            stack.enter_context(mock.patch.object(module, "write_actual_load_values", return_value={"range": "AH5:BA5", "sheet": "2026年市场情况", "count": 20}))
            stack.enter_context(mock.patch.object(module, "fetch_unit_cost_price_values", return_value=[3] * 11))
            stack.enter_context(mock.patch.object(module, "write_unit_cost_price_values", return_value={"range": "N7:X7", "sheet": "2026年运行方式及成本", "count": 11}))
            stack.enter_context(mock.patch.object(module, "copy_previous_day_unit_operation_mode", return_value={"range": "B7:M7", "sheet": "2026年运行方式及成本", "sourceRow": 6}))
            result = module.run_update(Path("market.xlsx"), date(2026, 7, 6), headless=True)

        self.assertIsNone(result["dayAheadLoad"])
        self.assertEqual(result["dayAheadMarket"]["range"], "V6:AG6")
        self.assertEqual(result["realtimeMarket"]["range"], "BB5:BE5")
        self.assertEqual(result["generationSettlement"]["range"], "BF1:BQ1")
        self.assertEqual(result["userSettlement"]["range"], "BR1:BX1")
        self.assertEqual(result["actualLoad"]["range"], "AH5:BA5")
        self.assertEqual(result["unitCostPrices"]["range"], "N7:X7")
        self.assertEqual(result["unitOperationMode"]["range"], "B7:M7")
        write_day_ahead_load.assert_not_called()
        ensure_login.assert_called_once()
        context.close.assert_called_once()

    def test_market_table_update_retries_transient_playwright_fetch_errors(self) -> None:
        from playwright.sync_api import Error as PlaywrightError
        from toolbox.runtime import load_module

        module = load_module("market_table_update_retry_fetch_test", self.paths.market_table_update)
        fetcher = mock.Mock(side_effect=[PlaywrightError("socket disconnected"), [1, 2, 3]])

        self.assertEqual(
            module.fetch_values_or_skip("现货运行日报日前市场", fetcher),
            [1, 2, 3],
        )
        self.assertEqual(fetcher.call_count, 2)

    def test_market_table_update_skips_after_repeated_playwright_fetch_errors(self) -> None:
        from playwright.sync_api import Error as PlaywrightError
        from toolbox.runtime import load_module

        module = load_module("market_table_update_skip_fetch_error_test", self.paths.market_table_update)
        fetcher = mock.Mock(side_effect=PlaywrightError("socket disconnected"))

        self.assertIsNone(
            module.fetch_values_or_skip("现货运行日报日前市场", fetcher)
        )
        self.assertEqual(fetcher.call_count, 3)

    def test_market_table_update_uses_base_date_minus_two_for_realtime_market(self) -> None:
        from types import SimpleNamespace
        from toolbox.runtime import load_module

        module = load_module("market_table_update_realtime_market_date_test", self.paths.market_table_update)

        class DummyLock:
            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return None

        class DummyPlaywright:
            def __enter__(self):
                return object()

            def __exit__(self, exc_type, exc, tb):
                return None

        context = SimpleNamespace(close=mock.Mock())
        with ExitStack() as stack:
            stack.enter_context(mock.patch.object(module, "load_config", return_value=SimpleNamespace(headless=True)))
            stack.enter_context(mock.patch.object(module, "ensure_workbook_accessible"))
            stack.enter_context(mock.patch.object(module, "AuthStateLock", return_value=DummyLock()))
            stack.enter_context(mock.patch.object(module, "sync_playwright", return_value=DummyPlaywright()))
            stack.enter_context(mock.patch.object(module, "launch_context", return_value=context))
            stack.enter_context(mock.patch.object(module, "ensure_login"))
            stack.enter_context(mock.patch.object(module, "fetch_day_ahead_load_values", return_value=[1] * 20))
            stack.enter_context(mock.patch.object(module, "write_day_ahead_load_values", return_value={"range": "B7:U7", "sheet": "2026年市场情况", "count": 20}))
            stack.enter_context(mock.patch.object(module, "fetch_day_ahead_market_values", return_value=[2] * 12))
            stack.enter_context(mock.patch.object(module, "write_day_ahead_market_values", return_value={"range": "V6:AG6", "sheet": "2026年市场情况", "count": 12}))
            fetch_realtime_market = stack.enter_context(mock.patch.object(module, "fetch_realtime_market_values", return_value=[3] * 4))
            write_realtime_market = stack.enter_context(mock.patch.object(module, "write_realtime_market_values", return_value={"range": "BB5:BE5", "sheet": "2026年市场情况", "count": 4}))
            stack.enter_context(mock.patch.object(module, "fetch_generation_settlement_values", return_value=[6] * 12))
            stack.enter_context(mock.patch.object(module, "write_generation_settlement_values", return_value={"range": "BF1:BQ1", "sheet": "2026年市场情况", "count": 12}))
            stack.enter_context(mock.patch.object(module, "fetch_user_settlement_values", return_value=[7] * 7))
            stack.enter_context(mock.patch.object(module, "write_user_settlement_values", return_value={"range": "BR1:BX1", "sheet": "2026年市场情况", "count": 7}))
            stack.enter_context(mock.patch.object(module, "fetch_actual_load_values", return_value=[4] * 20))
            stack.enter_context(mock.patch.object(module, "write_actual_load_values", return_value={"range": "AH5:BA5", "sheet": "2026年市场情况", "count": 20}))
            stack.enter_context(mock.patch.object(module, "fetch_unit_cost_price_values", return_value=[5] * 11))
            stack.enter_context(mock.patch.object(module, "write_unit_cost_price_values", return_value={"range": "N7:X7", "sheet": "2026年运行方式及成本", "count": 11}))
            stack.enter_context(mock.patch.object(module, "copy_previous_day_unit_operation_mode", return_value={"range": "B7:M7", "sheet": "2026年运行方式及成本", "sourceRow": 6}))
            result = module.run_update(Path("market.xlsx"), date(2026, 7, 6), headless=True)

        self.assertEqual(result["realtimeMarket"]["range"], "BB5:BE5")
        fetch_realtime_market.assert_called_once_with(context, date(2026, 7, 4))
        write_realtime_market.assert_called_once_with(Path("market.xlsx"), date(2026, 7, 4), [3] * 4)

    def test_market_table_update_uses_base_date_minus_six_for_generation_settlement(self) -> None:
        from types import SimpleNamespace
        from toolbox.runtime import load_module

        module = load_module("market_table_update_generation_settlement_date_test", self.paths.market_table_update)

        class DummyLock:
            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return None

        class DummyPlaywright:
            def __enter__(self):
                return object()

            def __exit__(self, exc_type, exc, tb):
                return None

        context = SimpleNamespace(close=mock.Mock())
        with ExitStack() as stack:
            stack.enter_context(mock.patch.object(module, "load_config", return_value=SimpleNamespace(headless=True)))
            stack.enter_context(mock.patch.object(module, "ensure_workbook_accessible"))
            stack.enter_context(mock.patch.object(module, "AuthStateLock", return_value=DummyLock()))
            stack.enter_context(mock.patch.object(module, "sync_playwright", return_value=DummyPlaywright()))
            stack.enter_context(mock.patch.object(module, "launch_context", return_value=context))
            stack.enter_context(mock.patch.object(module, "ensure_login"))
            stack.enter_context(mock.patch.object(module, "fetch_day_ahead_load_values", return_value=[1] * 20))
            stack.enter_context(mock.patch.object(module, "write_day_ahead_load_values", return_value={"range": "B7:U7", "sheet": "2026年市场情况", "count": 20}))
            stack.enter_context(mock.patch.object(module, "fetch_day_ahead_market_values", return_value=[2] * 12))
            stack.enter_context(mock.patch.object(module, "write_day_ahead_market_values", return_value={"range": "V6:AG6", "sheet": "2026年市场情况", "count": 12}))
            stack.enter_context(mock.patch.object(module, "fetch_realtime_market_values", return_value=[3] * 4))
            stack.enter_context(mock.patch.object(module, "write_realtime_market_values", return_value={"range": "BB5:BE5", "sheet": "2026年市场情况", "count": 4}))
            fetch_generation_settlement = stack.enter_context(mock.patch.object(module, "fetch_generation_settlement_values", return_value=[4] * 12))
            write_generation_settlement = stack.enter_context(mock.patch.object(module, "write_generation_settlement_values", return_value={"range": "BF1:BQ1", "sheet": "2026年市场情况", "count": 12}))
            stack.enter_context(mock.patch.object(module, "fetch_user_settlement_values", return_value=[7] * 7))
            stack.enter_context(mock.patch.object(module, "write_user_settlement_values", return_value={"range": "BR1:BX1", "sheet": "2026年市场情况", "count": 7}))
            stack.enter_context(mock.patch.object(module, "fetch_actual_load_values", return_value=[5] * 20))
            stack.enter_context(mock.patch.object(module, "write_actual_load_values", return_value={"range": "AH5:BA5", "sheet": "2026年市场情况", "count": 20}))
            stack.enter_context(mock.patch.object(module, "fetch_unit_cost_price_values", return_value=[6] * 11))
            stack.enter_context(mock.patch.object(module, "write_unit_cost_price_values", return_value={"range": "N7:X7", "sheet": "2026年运行方式及成本", "count": 11}))
            stack.enter_context(mock.patch.object(module, "copy_previous_day_unit_operation_mode", return_value={"range": "B7:M7", "sheet": "2026年运行方式及成本", "sourceRow": 6}))
            result = module.run_update(Path("market.xlsx"), date(2026, 7, 6), headless=True)

        self.assertEqual(result["generationSettlement"]["range"], "BF1:BQ1")
        fetch_generation_settlement.assert_called_once_with(context, date(2026, 6, 30))
        write_generation_settlement.assert_called_once_with(Path("market.xlsx"), date(2026, 6, 30), [4] * 12)

    def test_market_table_update_uses_base_date_minus_six_for_user_settlement(self) -> None:
        from types import SimpleNamespace
        from toolbox.runtime import load_module

        module = load_module("market_table_update_user_settlement_date_test", self.paths.market_table_update)

        class DummyLock:
            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return None

        class DummyPlaywright:
            def __enter__(self):
                return object()

            def __exit__(self, exc_type, exc, tb):
                return None

        context = SimpleNamespace(close=mock.Mock())
        with ExitStack() as stack:
            stack.enter_context(mock.patch.object(module, "load_config", return_value=SimpleNamespace(headless=True)))
            stack.enter_context(mock.patch.object(module, "ensure_workbook_accessible"))
            stack.enter_context(mock.patch.object(module, "AuthStateLock", return_value=DummyLock()))
            stack.enter_context(mock.patch.object(module, "sync_playwright", return_value=DummyPlaywright()))
            stack.enter_context(mock.patch.object(module, "launch_context", return_value=context))
            stack.enter_context(mock.patch.object(module, "ensure_login"))
            stack.enter_context(mock.patch.object(module, "fetch_day_ahead_load_values", return_value=[1] * 20))
            stack.enter_context(mock.patch.object(module, "write_day_ahead_load_values", return_value={"range": "B7:U7", "sheet": "2026年市场情况", "count": 20}))
            stack.enter_context(mock.patch.object(module, "fetch_day_ahead_market_values", return_value=[2] * 12))
            stack.enter_context(mock.patch.object(module, "write_day_ahead_market_values", return_value={"range": "V6:AG6", "sheet": "2026年市场情况", "count": 12}))
            stack.enter_context(mock.patch.object(module, "fetch_realtime_market_values", return_value=[3] * 4))
            stack.enter_context(mock.patch.object(module, "write_realtime_market_values", return_value={"range": "BB5:BE5", "sheet": "2026年市场情况", "count": 4}))
            stack.enter_context(mock.patch.object(module, "fetch_generation_settlement_values", return_value=[4] * 12))
            stack.enter_context(mock.patch.object(module, "write_generation_settlement_values", return_value={"range": "BF1:BQ1", "sheet": "2026年市场情况", "count": 12}))
            fetch_user_settlement = stack.enter_context(mock.patch.object(module, "fetch_user_settlement_values", return_value=[5] * 7))
            write_user_settlement = stack.enter_context(mock.patch.object(module, "write_user_settlement_values", return_value={"range": "BR1:BX1", "sheet": "2026年市场情况", "count": 7}))
            stack.enter_context(mock.patch.object(module, "fetch_actual_load_values", return_value=[6] * 20))
            stack.enter_context(mock.patch.object(module, "write_actual_load_values", return_value={"range": "AH5:BA5", "sheet": "2026年市场情况", "count": 20}))
            stack.enter_context(mock.patch.object(module, "fetch_unit_cost_price_values", return_value=[7] * 11))
            stack.enter_context(mock.patch.object(module, "write_unit_cost_price_values", return_value={"range": "N7:X7", "sheet": "2026年运行方式及成本", "count": 11}))
            stack.enter_context(mock.patch.object(module, "copy_previous_day_unit_operation_mode", return_value={"range": "B7:M7", "sheet": "2026年运行方式及成本", "sourceRow": 6}))
            result = module.run_update(Path("market.xlsx"), date(2026, 7, 6), headless=True)

        self.assertEqual(result["userSettlement"]["range"], "BR1:BX1")
        fetch_user_settlement.assert_called_once_with(context, date(2026, 6, 30))
        write_user_settlement.assert_called_once_with(Path("market.xlsx"), date(2026, 6, 30), [5] * 7)

    def test_market_table_update_uses_base_date_minus_two_for_actual_load(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("market_table_update_actual_load_date_test", self.paths.market_table_update)

        self.assertEqual(
            module.actual_load_source_date(date(2026, 7, 4)),
            date(2026, 7, 2),
        )

    def test_market_table_update_extracts_unit_cost_price_order(self) -> None:
        from toolbox.runtime import load_module

        module = load_module("market_table_update_cost_order_test", self.paths.market_table_update)
        payload = [
            {
                "unitCostInfoDTOList": [
                    {"unitId": "hm2", "priceOfDyCost": 344.8668888},
                    {"unitId": "hm1", "priceOfDyCost": 341.4059744},
                    {"unitId": "hm4", "priceOfDyCost": 342.5596126},
                    {"unitId": "hm3", "priceOfDyCost": 343.7132507},
                ]
            },
            {
                "unitCostInfoDTOList": [
                    {"unitId": "st3", "priceOfDyCost": 424.7904072},
                    {"unitId": "st1", "priceOfDyCost": 394.3277667},
                    {"unitId": "st2", "priceOfDyCost": 395.4994067},
                ]
            },
            {
                "unitCostInfoDTOList": [
                    {"unitId": "xg56", "priceOfDyCost": 620.36619563},
                    {"unitId": "xg78", "priceOfDyCost": 620.36619563},
                    {"unitId": "xg12", "priceOfDyCost": 620.36619563},
                    {"unitId": "xg34", "priceOfDyCost": 620.36619563},
                ]
            },
        ]

        with mock.patch.object(
            module,
            "UNIT_COST_PRICE_UNIT_IDS",
            ("st1", "st2", "st3", "hm1", "hm2", "hm3", "hm4", "xg12", "xg34", "xg56", "xg78"),
        ):
            self.assertEqual(
                module.extract_unit_cost_price_values(payload),
                [
                    394.3278,
                    395.4994,
                    424.7904,
                    341.406,
                    344.8669,
                    343.7133,
                    342.5596,
                    620.3662,
                    620.3662,
                    620.3662,
                    620.3662,
                ],
            )


class LauncherTests(unittest.TestCase):
    def setUp(self) -> None:
        self.workspace = Path(__file__).resolve().parent.parent
        self.script_root = self.workspace.parent

    def test_pythonw_launcher_can_be_loaded_without_starting_mainloop(self) -> None:
        old_cwd = Path.cwd()
        old_sys_path = list(sys.path)
        with tempfile.TemporaryDirectory() as directory:
            try:
                os.chdir(directory)
                sys.path[:] = [
                    item
                    for item in sys.path
                    if item and Path(item).resolve() != self.workspace.resolve()
                ]
                for name in list(sys.modules):
                    if name == "toolbox" or name.startswith("toolbox."):
                        sys.modules.pop(name)
                namespace = runpy.run_path(
                    str(self.workspace / "toolbox_launcher.pyw"),
                    run_name="toolbox_smoke",
                )
            finally:
                sys.path[:] = old_sys_path
                os.chdir(old_cwd)
        self.assertIn("main", namespace)

    def test_default_app_workspace_is_script_collection_root(self) -> None:
        from toolbox.app import default_workspace

        workspace = default_workspace()

        self.assertEqual(self.workspace.parent.resolve(), workspace)
        self.assertTrue((workspace / "wps自动" / "wps_excel_to_kdocs_gui.py").is_file())

    def test_root_keeps_only_hidden_vbs_launcher_as_file(self) -> None:
        if sys.platform != "win32":
            self.skipTest("Windows launcher layout expectation")
        root_files = [item for item in self.script_root.iterdir() if item.is_file()]
        self.assertEqual(1, len(root_files), [item.name for item in root_files])
        self.assertEqual(".vbs", root_files[0].suffix.lower())

    def test_toolbox_bundle_contains_python_entrypoint(self) -> None:
        self.assertTrue((self.workspace / "toolbox_launcher.pyw").is_file())
        self.assertTrue((self.workspace / "toolbox" / "app.py").is_file())
    def test_vbs_launcher_runs_powershell_hidden_and_can_move(self) -> None:
        if sys.platform != "win32":
            self.skipTest("Windows launcher smoke test")
        source = next(item for item in self.script_root.iterdir() if item.is_file() and item.suffix.lower() == ".vbs")
        content = source.read_text(encoding="utf-8")

        self.assertIn("TOOLBOX_SMOKE", content)
        self.assertIn("waitOnReturn = False", content)
        self.assertIn("shell.Run command, 0, waitOnReturn", content)
        self.assertNotIn("C:\\Users\\lllg\\Desktop\\脚本汇总", content)

        with tempfile.TemporaryDirectory() as directory:
            copied = Path(directory) / "moved-toolbox.vbs"
            shutil.copy2(source, copied)
            result = subprocess.run(
                ["cscript.exe", "//nologo", str(copied)],
                cwd=directory,
                text=True,
                encoding="utf-8",
                errors="replace",
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                env={**os.environ, "TOOLBOX_SMOKE": "1"},
                timeout=30,
            )

        self.assertEqual(result.returncode, 0, result.stdout)
        self.assertIn("TOOLBOX_SCRIPT=", result.stdout)
        self.assertIn("toolbox_launcher.pyw", result.stdout)

    def test_macos_command_launcher_runs_from_moved_copy(self) -> None:
        if sys.platform != "darwin":
            self.skipTest("macOS launcher smoke test")

        source = self.script_root / "电力工具箱.command"
        self.assertTrue(source.is_file())
        content = source.read_text(encoding="utf-8")
        self.assertIn("TOOLBOX_SMOKE", content)
        self.assertNotIn("C:\\Users\\lllg", content)

        with tempfile.TemporaryDirectory() as directory:
            copied = Path(directory) / "moved-toolbox.command"
            shutil.copy2(source, copied)
            copied.chmod(0o755)
            result = subprocess.run(
                [str(copied)],
                cwd=directory,
                text=True,
                encoding="utf-8",
                errors="replace",
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                env={**os.environ, "TOOLBOX_SMOKE": "1"},
                timeout=30,
            )

        self.assertEqual(result.returncode, 0, result.stdout)
        self.assertIn("TOOLBOX_SCRIPT=", result.stdout)
        self.assertIn("toolbox_launcher.pyw", result.stdout)

    def test_readme_documents_unified_and_legacy_launchers(self) -> None:
        readme = next(self.script_root.glob("*/README_*.txt"))
        content = readme.read_text(encoding="utf-8")

        self.assertIn("电力工具箱.bat", content)
        self.assertIn("原有", content)
        self.assertIn("保留", content)


if __name__ == "__main__":
    unittest.main()
